from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .bootstrap import resolve_repo_path

from src.config import get_config, load_mechanism_spec, load_spec, normalize_audit_replace_factory_codes


_ENUM_RE = re.compile(r"^enum\[(?P<enum_name>[^\]]+)\]$")


class FormMetadataService:
    def __init__(self) -> None:
        self.spec = load_spec()
        self.config = get_config()

    def build_form_schema(self) -> dict[str, Any]:
        params_cfg = self.spec.doc_generation.get("params", {})
        sections: list[dict[str, Any]] = []

        for section_id, section in params_cfg.items():
            if not isinstance(section, dict):
                continue

            fields: list[dict[str, Any]] = []
            for field_key, rule in section.items():
                if not isinstance(rule, dict) or rule.get("source") != "frontend":
                    continue
                fields.append(self._build_field_schema(field_key, rule))

            if fields:
                sections.append(
                    {
                        "id": section_id,
                        "title": section_id,
                        "fields": fields,
                    },
                )

        return {
            "schema_version": "frontend-form@1",
            "upload_limits": {
                "max_files": self.config.upload_limits.max_files,
                "allowed_exts": self.config.upload_limits.allowed_exts,
                "max_total_mb": self.config.upload_limits.max_total_mb,
            },
            "deliverable": {
                "sections": sections,
            },
            "audit_check": {
                "unit_consistency": {
                    "enabled": self.config.audit_check.unit_consistency.enabled,
                    "project_units": self.config.audit_check.unit_consistency.project_units,
                    "allow_unlisted_unit_no": self.config.audit_check.unit_consistency.allow_unlisted_unit_no,
                    "unit_no_pattern": self.config.audit_check.unit_consistency.unit_no_pattern,
                },
            },
            "audit_replace": {
                "project_options": self._resolve_project_options(),
                "project_units": self._resolve_project_units(),
                "source_unit_options": self._resolve_source_unit_options(),
                "target_unit_options": self._resolve_target_unit_options(),
                "unit_factory_codes": normalize_audit_replace_factory_codes(
                    load_mechanism_spec().audit_replace.unit_factory_codes,
                ),
                "factory_index_maps": {
                    "source_variant_options": {
                        project_no: list(variants.keys())
                        for project_no, variants in self.config.factory_index_maps.source_variant_rules.items()
                    },
                    "target_variant_options": {
                        project_no: list(variants.keys())
                        for project_no, variants in self.config.factory_index_maps.island_templates.items()
                    },
                },
            },
            "calculation_book": {
                **self.spec.calculation_book,
                "project_options": [
                    {
                        "value": str(item.get("id") or ""),
                        "label": str(item.get("name") or item.get("id") or ""),
                    }
                    for item in self.spec.enums.get("project_no", [])
                    if isinstance(item, dict) and str(item.get("id") or "").strip()
                ],
            },
            "management": self._build_management_schema(),
        }

    def _build_field_schema(self, field_key: str, rule: dict[str, Any]) -> dict[str, Any]:
        options = self._resolve_options(field_key, rule)
        field_type = "text"
        ui = rule.get("ui") if isinstance(rule.get("ui"), dict) else {}
        if ui.get("widget") == "checkbox" or rule.get("type") == "bool" or isinstance(rule.get("default"), bool):
            field_type = "checkbox"
        elif ui.get("widget") == "combobox":
            field_type = "combobox"
        elif options:
            field_type = "select"
        elif rule.get("format") == "YYYY-MM-DD":
            field_type = "date"

        payload = {
            "key": field_key,
            "label": field_key,
            "type": field_type,
            "required": bool(rule.get("required", False)),
            "required_when": rule.get("required_when"),
            "source": rule.get("source"),
            "default": rule.get("default"),
            "format": rule.get("format"),
            "desc": rule.get("desc"),
            "options": options,
        }
        if ui:
            payload["ui"] = ui
            payload["allow_custom_input"] = bool(ui.get("allow_custom_input", False))
            payload["filterable"] = bool(ui.get("filterable", False))
        if isinstance(rule.get("option_source"), dict):
            payload["option_source"] = rule["option_source"]
        return payload

    def _resolve_options(self, field_key: str, rule: dict[str, Any]) -> list[str]:
        inline_options = rule.get("options")
        if isinstance(inline_options, list):
            return [str(item) for item in inline_options]

        option_source = rule.get("option_source")
        if isinstance(option_source, dict):
            source_type = str(option_source.get("type") or "").strip().lower()
            if source_type == "workbook_range":
                workbook = option_source.get("workbook")
                sheet_index = option_source.get("sheet_index")
                cell_range = option_source.get("range")
                if workbook and sheet_index and cell_range:
                    return self._load_range_options(
                        resolve_repo_path(str(workbook)),
                        sheet_index=int(sheet_index) - 1,
                        cell_range=str(cell_range),
                    )

        if field_key in self.spec.enums:
            enum_values = self.spec.enums.get(field_key, [])
            if isinstance(enum_values, list):
                if enum_values and isinstance(enum_values[0], dict):
                    return [str(item["id"]) for item in enum_values if "id" in item]
                return [str(item) for item in enum_values]
        if field_key == "file_category":
            return self._load_first_column_options(
                resolve_repo_path(self.spec.get_template_path("design", "2016")),
                sheet_index=4,
            )
        if field_key == "ied_person_qual_category":
            return self._load_data_validation_list_options(
                resolve_repo_path(self.spec.get_template_path("ied", "2016")),
                sheet_index=0,
                column_letter="M",
            )
        if field_key == "ied_discipline_office":
            path = resolve_repo_path("documents_bin/responsible_unit.json")
            return list(json.loads(path.read_text(encoding="utf-8-sig")))

        enum_name = self._enum_name(rule.get("type"))
        if enum_name is None:
            return []

        enum_values = self.spec.enums.get(enum_name, [])
        if isinstance(enum_values, list):
            if enum_values and isinstance(enum_values[0], dict):
                return [str(item["id"]) for item in enum_values if "id" in item]
            return [str(item) for item in enum_values]
        return []

    def _build_management_schema(self) -> dict[str, Any]:
        management_cfg = self.spec.get_management_features()
        account_cfg = dict(management_cfg.get("account") or {})
        workflow_cfg = dict(management_cfg.get("workflow") or {})
        workload_cfg = dict(management_cfg.get("workload") or {})
        factor_cfg = dict(workflow_cfg.get("factor") or {})
        mechanism = load_mechanism_spec()

        return {
            "account": {
                "fields": {
                    str(key): str(value)
                    for key, value in dict(account_cfg.get("fields") or {}).items()
                    if str(key).strip() and str(value).strip()
                },
                "valid_roles": [
                    str(role)
                    for role in account_cfg.get("valid_roles") or []
                    if str(role).strip()
                ],
                "admin_roles": list(mechanism.permissions.account_admin_roles),
                "admin_created_default_password": str(
                    account_cfg.get("admin_created_default_password") or ""
                ),
            },
            "workflow": {
                "terminal_status": mechanism.workflow_runtime.approval_terminal_status,
                "archive_trigger_status": mechanism.workflow_runtime.archive_trigger_status,
                "factor": {
                    "default": float(factor_cfg["default"]),
                    "min": float(factor_cfg["min"]),
                    "max": float(factor_cfg["max"]),
                    "precision": int(factor_cfg["precision"]),
                },
                "nodes": list(workflow_cfg.get("nodes") or []),
                "status_labels": dict(mechanism.management_ui.workflow_status_labels),
                "node_labels": {
                    str(node.get("key")): str(node.get("label"))
                    for node in workflow_cfg.get("nodes") or []
                    if str(node.get("key") or "").strip()
                },
                "empty_current_node_label": mechanism.management_ui.empty_current_node_label,
            },
            "workload": {
                "settlement_trigger": str(workload_cfg["settlement_trigger"]),
                "scope_roles": {
                    str(scope): [str(role) for role in roles]
                    for scope, roles in mechanism.permissions.workload_scope_roles.items()
                },
                "scope_labels": dict(mechanism.management_ui.workload_scope_labels),
                "status_options": [
                    option.model_dump(mode="json")
                    for option in mechanism.workload_runtime.status_options
                ],
            },
            "archive": {
                "status_labels": dict(mechanism.management_ui.archive_status_labels),
            },
        }

    def _resolve_project_options(self) -> list[str]:
        enum_values = self.spec.enums.get("project_no", [])
        if isinstance(enum_values, list):
            if enum_values and isinstance(enum_values[0], dict):
                return [str(item["id"]) for item in enum_values if "id" in item]
            return [str(item) for item in enum_values]
        return []

    def _resolve_project_units(self) -> dict[str, list[str]]:
        return {
            str(project_no): [str(unit_no) for unit_no in unit_nos if str(unit_no).strip()]
            for project_no, unit_nos in self.config.audit_check.unit_consistency.project_units.items()
            if unit_nos
        }

    def _resolve_source_unit_options(self) -> dict[str, list[dict[str, str]]]:
        factory_config = self.config.factory_index_maps
        unit_options = self._build_unit_option_map(
            self.config.audit_check.unit_consistency.project_units,
            default_suffix=self.config.audit_check.unit_consistency.label_suffix,
        )
        factory_options = self._build_unit_option_map(
            {
                str(project_no): list(variant_rules.keys())
                for project_no, variant_rules in factory_config.source_variant_rules.items()
                if variant_rules
            },
            default_suffix=factory_config.variant_label_suffix,
        )
        return self._merge_unit_option_maps(unit_options, factory_options)

    def _resolve_target_unit_options(self) -> dict[str, list[dict[str, str]]]:
        factory_config = self.config.factory_index_maps
        unit_options = self._build_unit_option_map(
            self.config.audit_check.unit_consistency.project_units,
            default_suffix=self.config.audit_check.unit_consistency.label_suffix,
        )
        factory_options = self._build_unit_option_map(
            {
                str(project_no): list(templates.keys())
                for project_no, templates in factory_config.island_templates.items()
                if templates
            },
            default_suffix=factory_config.variant_label_suffix,
        )
        return self._merge_unit_option_maps(unit_options, factory_options)

    @staticmethod
    def _build_unit_option_map(
        units_by_project: dict[str, list[str]],
        *,
        default_suffix: str,
    ) -> dict[str, list[dict[str, str]]]:
        options: dict[str, list[dict[str, str]]] = {}
        for project_no, unit_nos in units_by_project.items():
            normalized_project_no = str(project_no).strip()
            if not normalized_project_no:
                continue
            project_options: list[dict[str, str]] = []
            seen_values: set[str] = set()
            for unit_no in unit_nos:
                value = str(unit_no).strip()
                if not value or value in seen_values:
                    continue
                seen_values.add(value)
                project_options.append({"value": value, "label": f"{value}{default_suffix}"})
            if project_options:
                options[normalized_project_no] = project_options
        return options

    @staticmethod
    def _merge_unit_option_maps(
        base: dict[str, list[dict[str, str]]],
        override: dict[str, list[dict[str, str]]],
    ) -> dict[str, list[dict[str, str]]]:
        merged = {project_no: list(options) for project_no, options in base.items()}
        for project_no, options in override.items():
            by_value = {option["value"]: option for option in merged.get(project_no, [])}
            for option in options:
                by_value[option["value"]] = option
            merged[project_no] = list(by_value.values())
        return merged

    @staticmethod
    def _enum_name(type_value: Any) -> str | None:
        if not isinstance(type_value, str):
            return None
        match = _ENUM_RE.match(type_value)
        if match:
            return match.group("enum_name")
        if type_value == "enum":
            return None
        return None

    @staticmethod
    def _load_first_column_options(workbook_path: Path, *, sheet_index: int) -> list[str]:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[sheet_index]]
        values: list[str] = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            value = row[0]
            if value is None:
                continue
            text = str(value).strip()
            if text and text not in values:
                values.append(text)
        return values

    @staticmethod
    def _load_range_options(workbook_path: Path, *, sheet_index: int, cell_range: str) -> list[str]:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[sheet_index]]
        values: list[str] = []
        for row in ws[cell_range]:
            cells = row if isinstance(row, tuple) else (row,)
            for cell in cells:
                value = cell.value
                if value is None:
                    continue
                text = str(value).strip()
                if text and text not in values:
                    values.append(text)
        return values

    @staticmethod
    def _load_data_validation_list_options(
        workbook_path: Path,
        *,
        sheet_index: int,
        column_letter: str,
    ) -> list[str]:
        wb = load_workbook(workbook_path, read_only=False, data_only=False)
        ws = wb[wb.sheetnames[sheet_index]]
        matches: list[str] = []

        for validation in ws.data_validations.dataValidation:
            formula = str(validation.formula1 or "").strip()
            if validation.type != "list" or not formula:
                continue

            refs = str(validation.sqref).split()
            if not any(
                ref.startswith(column_letter)
                or f":{column_letter}" in ref
                or f"{column_letter}:" in ref
                for ref in refs
            ):
                continue

            matches.extend(FormMetadataService._parse_list_formula(formula))

        deduped: list[str] = []
        for value in matches:
            if value and value not in deduped:
                deduped.append(value)
        return deduped

    @staticmethod
    def _parse_list_formula(formula: str) -> list[str]:
        if formula.startswith('"') and formula.endswith('"'):
            return [item.strip() for item in formula[1:-1].split(",") if item.strip()]

        return []
