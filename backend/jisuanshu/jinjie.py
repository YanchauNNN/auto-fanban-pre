# -*- coding: utf-8 -*-
import os
import re
import sys
import math
from datetime import date
from concurrent.futures import as_completed, ThreadPoolExecutor

from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
from PIL import Image
import pytesseract
from openpyxl import load_workbook

from tkinter import Tk, filedialog, messagebox, StringVar, Canvas
from tkinter import ttk


CURRENT_CHAPTER = "7.1"

# ========================
# 模板选项
# ========================
TEMPLATE_OPTIONS = [
    "内部结构计算书",
    "核岛厂房计算书"
]

TEMPLATE_FILE_MAP = {
    "内部结构计算书": "内部结构计算书.docx",
    "核岛厂房计算书": "核岛厂房计算书.docx"
}

# ========================
# 厂房列表
# ========================
ALL_PLANTS = ["RX", "NH", "SD", "SU", "KA"]

# ========================
# 下拉选项
# ========================
PROJECT_NAME_OPTIONS = [
    "漳州核电厂3、4号机组",
    "浙江金七门核电厂1、2号机组",
    "巴基斯坦恰希玛核电厂五号机组"
]

SUBPROJECT_CODE_OPTIONS = [
    "RX", "KA", "NH", "SD", "SU", "DA", "DB"
]

SUBPROJECT_NAME_OPTIONS = [
    "内部结构",
    "燃料厂房",
    "核辅助厂房",
    "安全厂房SD区",
    "安全厂房SU区",
    "DA应急柴油机发电厂房",
    "DB应急柴油机发电厂房"
]

DESIGN_PHASE_OPTIONS = [
    "施工图设计"
]

# ========================
# 获取资源路径
# 适配源码运行 / PyInstaller onefile / PyInstaller onedir
# ========================
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# ========================
# 获取程序所在目录
# 用于 output 输出文件夹
# ========================
def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(".")

# ========================
# 根据界面选择获取模板路径
# ========================
def get_selected_template_path(template_type):
    filename = TEMPLATE_FILE_MAP.get(template_type)

    if not filename:
        return ""

    bundled_path = resource_path(filename)
    if os.path.exists(bundled_path):
        return bundled_path

    local_path = os.path.join(os.path.abspath("."), filename)
    if os.path.exists(local_path):
        return local_path

    return bundled_path

# ========================
# 钢筋自动选型配置
# ========================
REBAR_TABLE_PATH = resource_path("钢筋的公称直径、公称面积表.xlsx")

# OCR 计算面积放大系数：sm_value × 1.20
REBAR_EXTRA_RATIO = 0.20

# 自动适配钢筋排数：只允许 1 排或 2 排
REBAR_ROW_COUNT_OPTIONS = [1, 2]

# 自动适配钢筋间距：只允许 200mm 或 250mm
REBAR_SPACING_OPTIONS = [200, 250]

# 钢筋最大直径限制
REBAR_MAX_DIAMETER = 40

# ========================
# 配置 Tesseract-OCR 路径
# ========================
def get_tesseract_path():
    bundled_path = resource_path(os.path.join("Tesseract-OCR", "tesseract.exe"))
    if os.path.exists(bundled_path):
        return bundled_path

    local_path = os.path.join(os.path.abspath("."), "Tesseract-OCR", "tesseract.exe")
    if os.path.exists(local_path):
        return local_path

    return r"D:\Program Files\Tesseract-OCR\tesseract.exe"


pytesseract.pytesseract.tesseract_cmd = get_tesseract_path()

tessdata_dir = resource_path(os.path.join("Tesseract-OCR", "tessdata"))
if os.path.exists(tessdata_dir):
    os.environ["TESSDATA_PREFIX"] = tessdata_dir

print("OCR路径:", pytesseract.pytesseract.tesseract_cmd)

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp'}
DIRECTION_MAP = {"X": "水平向", "Y": "竖向", "Z": "拉筋"}

# ========================
# 字段
# ========================
FIELD_LABELS = {
    "internal_code": "内部编码",
    "project_name": "项目名称",
    "version": "版本",
    "subproject_code": "子项号或系统号",
    "subproject_name": "子项或系统名称",
    "design_phase": "设计阶段",
    "document_name": "图册（文件）名称",
    "workshop_length": "厂房外轮廓长度",
    "workshop_width": "厂房外轮廓宽度",
    "raft_slab_top_elevation": "筏板顶标高",
    "roof_top_elevation": "屋顶标高",
    "factory_extreme_min_temperature": "厂址的极端最低温度",
    "factory_extreme_max_temperature": "厂址的极端最高温度",
    "site_soil_temperature": "场地土壤温度"
}

FIELD_UNITS = {
    "document_name": "配筋计算书",
    "workshop_length": "m",
    "workshop_width": "m",
    "raft_slab_top_elevation": "m",
    "roof_top_elevation": "m",
    "factory_extreme_min_temperature": "℃",
    "factory_extreme_max_temperature": "℃",
    "site_soil_temperature": "℃"
}

COMBOBOX_FIELDS = {
    "project_name": PROJECT_NAME_OPTIONS,
    "subproject_code": SUBPROJECT_CODE_OPTIONS,
    "subproject_name": SUBPROJECT_NAME_OPTIONS,
    "design_phase": DESIGN_PHASE_OPTIONS
}

BASIC_INFO_FIELDS = [
    "internal_code",
    "project_name",
    "version",
    "subproject_code",
    "subproject_name",
    "design_phase",
    "document_name"
]

WORKSHOP_INFO_FIELDS = [
    "workshop_length",
    "workshop_width",
    "raft_slab_top_elevation",
    "roof_top_elevation",
    "factory_extreme_min_temperature",
    "factory_extreme_max_temperature",
    "site_soil_temperature"
]

# ========================
# OCR 识别 SM 值
# ========================
def extract_sm_value_from_image(image_path):
    try:
        img = Image.open(image_path)
        width, height = img.size

        img = img.resize((int(width * 0.8), int(height * 0.8)))
        crop = img.crop((0, 0, int(width * 0.5), int(height * 0.4)))
        crop = crop.convert('L')
        crop = crop.point(lambda x: 0 if x < 160 else 255)

        config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=SMX=0123456789'
        text = pytesseract.image_to_string(crop, config=config)

        match = re.search(r'SM[Xx][^0-9]{0,3}(\d+)', text)
        value = match.group(1) if match else "0"

        if not value.isdigit():
            value = "0"

        return value

    except:
        return "0"

# ========================
# 图片排序
# ========================
def extract_sort_key(filename):
    base = os.path.splitext(filename)[0].upper()
    match = re.match(r'^([A-Za-z]+)(\d+)-([XYZ])$', base)

    if not match:
        return None

    prefix, num_str, direction = match.groups()
    num = int(num_str)
    dir_order = {"X": 0, "Y": 1, "Z": 2}.get(direction, 99)

    return (num, dir_order, prefix, direction, filename)

# ========================
# 并行加载配筋图
# ========================
def load_reinforcement_figures_parallel(picture_dir):
    files = [
        f for f in os.listdir(picture_dir)
        if os.path.isfile(os.path.join(picture_dir, f))
        and f.lower().endswith(tuple(IMAGE_EXTENSIONS))
    ]

    results = []

    with ThreadPoolExecutor(max_workers=4) as executor:
        future_map = {}

        for f in files:
            key = extract_sort_key(f)

            if key:
                num, dir_order, prefix, direction, filename = key
                full_path = os.path.join(picture_dir, filename)

                future = executor.submit(extract_sm_value_from_image, full_path)

                future_map[future] = {
                    "wall_id": f"{prefix}{num}",
                    "direction": direction,
                    "full_path": full_path,
                    "sort_key": (num, dir_order)
                }

        for future in as_completed(future_map):
            info = future_map[future]

            try:
                sm_value = future.result()
            except:
                sm_value = "0"

            info["sm_value"] = sm_value
            results.append(info)

    results.sort(key=lambda x: x["sort_key"])
    return results

# ========================
# 数字解析
# ========================
def parse_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    match = re.search(r'\d+(?:\.\d+)?', text)

    if not match:
        return None

    try:
        return float(match.group())
    except:
        return None

# ========================
# 读取钢筋公称面积表
# ========================
def load_rebar_area_table(excel_path):
    """
    读取《钢筋的公称直径、公称面积表.xlsx》。

    兼容常见结构：
    - 第1列为钢筋公称直径；
    - 某一行为“根数”表头，例如 1、2、3、4、5；
    - 数据区为对应根数下的公称截面面积 mm²/m。

    程序会自动寻找根数表头行。
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"未找到钢筋面积表：{excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = None
    root_count_to_col = {}

    max_scan_rows = min(ws.max_row, 20)

    for row in range(1, max_scan_rows + 1):
        current_map = {}

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            number = parse_number(value)

            if number is None:
                continue

            if abs(number - int(number)) < 1e-6:
                root_count = int(number)

                if 1 <= root_count <= 20:
                    current_map[root_count] = col

        if len(current_map) >= 3:
            header_row = row
            root_count_to_col = current_map
            break

    if header_row is None:
        header_row = 2
        root_count_to_col = {}

        for col in range(2, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col).value
            number = parse_number(value)

            if number is not None and abs(number - int(number)) < 1e-6:
                root_count = int(number)

                if 1 <= root_count <= 20:
                    root_count_to_col[root_count] = col

    rebar_rows = []
    diameter_col = 1

    for row in range(header_row + 1, ws.max_row + 1):
        diameter_value = ws.cell(row=row, column=diameter_col).value
        diameter_number = parse_number(diameter_value)

        if diameter_number is None:
            continue

        diameter = int(round(diameter_number))
        area_by_root_count = {}

        for root_count, col in root_count_to_col.items():
            area_value = ws.cell(row=row, column=col).value
            area_number = parse_number(area_value)

            if area_number is None:
                continue

            area_by_root_count[root_count] = float(area_number)

        if area_by_root_count:
            rebar_rows.append({
                "diameter": diameter,
                "area_by_root_count": area_by_root_count
            })

    rebar_rows.sort(key=lambda x: x["diameter"])

    if not rebar_rows:
        raise ValueError("钢筋面积表读取失败：未识别到有效的钢筋直径和面积数据。")

    return rebar_rows

# ========================
# 根据 OCR 面积自动选择钢筋
# ========================
def select_rebar_by_sm_value(
    sm_value,
    rebar_table,
    row_count_options=REBAR_ROW_COUNT_OPTIONS,
    spacing_options=REBAR_SPACING_OPTIONS,
    max_diameter=REBAR_MAX_DIAMETER,
    extra_ratio=REBAR_EXTRA_RATIO
):
    """
    自动选筋逻辑：

    1. 目标面积 = sm_value × 1.20；
    2. 自动在以下范围内搜索：
       - 排数：1排、2排；
       - 间距：@200、@250；
       - 直径：不大于40；
    3. 实际配筋面积 = 单排面积 × 排数；
    4. 选择满足 实际配筋面积 >= 目标面积 的组合；
    5. 在满足条件的组合中，选择实际配筋面积最小的组合。
    """
    try:
        sm = float(sm_value)
    except:
        sm = 0.0

    if sm <= 0:
        return {
            "spec": "未识别",
            "area": "0",
            "target_area": "0",
            "diameter": "",
            "row_count": "",
            "spacing": ""
        }

    target_area = sm * (1 + extra_ratio)
    candidates = []

    for row in rebar_table:
        diameter = row["diameter"]

        if diameter > max_diameter:
            continue

        area_by_root_count = row["area_by_root_count"]

        for spacing in spacing_options:
            bars_per_meter = int(round(1000 / spacing))

            if bars_per_meter not in area_by_root_count:
                continue

            single_row_area = area_by_root_count[bars_per_meter]

            for row_count in row_count_options:
                total_area = single_row_area * row_count

                if total_area >= target_area:
                    candidates.append({
                        "diameter": diameter,
                        "spacing": spacing,
                        "row_count": row_count,
                        "bars_per_meter": bars_per_meter,
                        "single_row_area": single_row_area,
                        "total_area": total_area,
                        "target_area": target_area,
                        "excess_area": total_area - target_area
                    })

    if candidates:
        candidates.sort(
            key=lambda x: (
                x["total_area"],
                x["row_count"],
                x["diameter"],
                -x["spacing"]
            )
        )

        best = candidates[0]

        spec = f"{best['row_count']}排{best['diameter']}@{best['spacing']}"
        area = int(round(best["total_area"]))

        return {
            "spec": spec,
            "area": str(area),
            "target_area": str(int(math.ceil(target_area))),
            "diameter": str(best["diameter"]),
            "row_count": str(best["row_count"]),
            "spacing": str(best["spacing"])
        }

    # 如果没有任何组合满足要求，则取直径不超过40范围内的最大可用组合兜底
    available = []

    for row in rebar_table:
        diameter = row["diameter"]

        if diameter > max_diameter:
            continue

        area_by_root_count = row["area_by_root_count"]

        for spacing in spacing_options:
            bars_per_meter = int(round(1000 / spacing))

            if bars_per_meter not in area_by_root_count:
                continue

            single_row_area = area_by_root_count[bars_per_meter]

            for row_count in row_count_options:
                total_area = single_row_area * row_count

                available.append({
                    "diameter": diameter,
                    "spacing": spacing,
                    "row_count": row_count,
                    "total_area": total_area
                })

    if not available:
        return {
            "spec": "未匹配",
            "area": "0",
            "target_area": str(int(math.ceil(target_area))),
            "diameter": "",
            "row_count": "",
            "spacing": ""
        }

    available.sort(
        key=lambda x: (
            x["total_area"],
            x["row_count"],
            x["diameter"],
            -x["spacing"]
        ),
        reverse=True
    )

    best = available[0]

    spec = f"{best['row_count']}排{best['diameter']}@{best['spacing']}"
    area = int(round(best["total_area"]))

    return {
        "spec": spec,
        "area": str(area),
        "target_area": str(int(math.ceil(target_area))),
        "diameter": str(best["diameter"]),
        "row_count": str(best["row_count"]),
        "spacing": str(best["spacing"])
    }

# ========================
# 计算配筋裕度
# ========================
def calculate_rebar_margin(calc_value, actual_area):
    """
    裕度计算：
    (实际配筋面积 - 计算配筋面积) / 计算配筋面积 × 100%
    """
    try:
        calc = float(calc_value)
        actual = float(actual_area)
    except:
        return ""

    if calc <= 0:
        return ""

    margin = (actual - calc) / calc * 100
    return f"{margin:.1f}%"

# ========================
# 生成表格行
# ========================
def generate_table_rows_from_figures(figures):
    """
    根据每张配筋云图生成墙体配筋表。

    row.x_calc   = OCR识别的X向计算配筋面积
    row.x_actual = 自动选筋后的X向实际配筋面积
    row.x_margin = (实际配筋面积 - 计算配筋面积) / 计算配筋面积 × 100%

    Y、Z方向同理。
    """
    wall_dict = {}

    for fig in figures:
        wid = fig["wall_id"]
        direction = fig["direction"]
        sm_value = fig.get("sm_value", "0")
        rebar_area = fig.get("rebar_area", "0")

        if wid not in wall_dict:
            wall_dict[wid] = {
                "id": wid,
                "x_calc": "0", "x_actual": "", "x_margin": "",
                "y_calc": "0", "y_actual": "", "y_margin": "",
                "z_calc": "0", "z_actual": "", "z_margin": ""
            }

        margin_text = calculate_rebar_margin(sm_value, rebar_area)

        if direction == "X":
            wall_dict[wid]["x_calc"] = sm_value
            wall_dict[wid]["x_actual"] = rebar_area
            wall_dict[wid]["x_margin"] = margin_text

        elif direction == "Y":
            wall_dict[wid]["y_calc"] = sm_value
            wall_dict[wid]["y_actual"] = rebar_area
            wall_dict[wid]["y_margin"] = margin_text

        elif direction == "Z":
            wall_dict[wid]["z_calc"] = sm_value
            wall_dict[wid]["z_actual"] = rebar_area
            wall_dict[wid]["z_margin"] = margin_text

    return sorted(
        wall_dict.values(),
        key=lambda x: int(re.match(r'[A-Za-z]+(\d+)', x["id"]).group(1))
    )

# ========================
# 获取文件夹首图
# ========================
def get_first_image_from_folder(folder_path):
    files = [
        f for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
        and f.lower().endswith(tuple(IMAGE_EXTENSIONS))
    ]

    if not files:
        raise FileNotFoundError(f"文件夹中未找到图片：{folder_path}")

    files.sort()
    return os.path.join(folder_path, files[0])

# ========================
# 根据当前厂房获取其他厂房
# ========================
def get_other_plants(subproject_code):
    current_plant = subproject_code.strip().upper()
    return [plant for plant in ALL_PLANTS if plant.upper() != current_plant]

# ========================
# 获取可用输出路径，避免同名文件被占用
# ========================
def get_available_output_path(output_path):
    if not os.path.exists(output_path):
        return output_path

    folder = os.path.dirname(output_path)
    filename = os.path.basename(output_path)
    name, ext = os.path.splitext(filename)

    index = 1

    while True:
        new_path = os.path.join(folder, f"{name}_{index}{ext}")

        if not os.path.exists(new_path):
            return new_path

        index += 1

# ========================
# GUI 样式
# ========================
def setup_style(root):
    style = ttk.Style(root)

    try:
        style.theme_use("clam")
    except:
        pass

    root.configure(bg="#eef2f7")

    style.configure("App.TFrame", background="#eef2f7")
    style.configure("Header.TFrame", background="#1f4e79")
    style.configure("Card.TFrame", background="#ffffff")
    style.configure("Footer.TFrame", background="#eef2f7")

    style.configure(
        "HeaderTitle.TLabel",
        background="#1f4e79",
        foreground="#ffffff",
        font=("Microsoft YaHei UI", 20, "bold")
    )

    style.configure(
        "HeaderSubTitle.TLabel",
        background="#1f4e79",
        foreground="#dbeafe",
        font=("Microsoft YaHei UI", 10)
    )

    style.configure(
        "Section.TLabelframe",
        background="#ffffff",
        foreground="#111827",
        borderwidth=1,
        relief="solid"
    )

    style.configure(
        "Section.TLabelframe.Label",
        background="#ffffff",
        foreground="#1f2937",
        font=("Microsoft YaHei UI", 11, "bold")
    )

    style.configure(
        "Form.TLabel",
        background="#ffffff",
        foreground="#374151",
        font=("Microsoft YaHei UI", 10)
    )

    style.configure(
        "Unit.TLabel",
        background="#ffffff",
        foreground="#6b7280",
        font=("Microsoft YaHei UI", 9)
    )

    style.configure(
        "Tip.TLabel",
        background="#eef2f7",
        foreground="#6b7280",
        font=("Microsoft YaHei UI", 9)
    )

    style.configure(
        "Status.TLabel",
        background="#eef2f7",
        foreground="#374151",
        font=("Microsoft YaHei UI", 10, "bold")
    )

    style.configure(
        "TEntry",
        padding=6,
        font=("Microsoft YaHei UI", 10)
    )

    style.configure(
        "TCombobox",
        padding=6,
        font=("Microsoft YaHei UI", 10)
    )

    style.configure(
        "Primary.TButton",
        font=("Microsoft YaHei UI", 12, "bold"),
        padding=(24, 10)
    )

    style.configure(
        "Secondary.TButton",
        font=("Microsoft YaHei UI", 10),
        padding=(12, 6)
    )

# ========================
# 鼠标滚轮绑定
# ========================
def bind_mousewheel(widget, canvas):
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    widget.bind_all("<MouseWheel>", _on_mousewheel)

# ========================
# 普通单列输入区
# ========================
def create_form_rows(parent, fields, entries):
    for row_index, field in enumerate(fields):
        label = ttk.Label(
            parent,
            text=FIELD_LABELS[field],
            style="Form.TLabel",
            width=22,
            anchor="w"
        )
        label.grid(row=row_index, column=0, sticky="w", padx=(0, 10), pady=6)

        if field in COMBOBOX_FIELDS:
            widget = ttk.Combobox(
                parent,
                width=42,
                values=COMBOBOX_FIELDS[field],
                state="readonly"
            )
            widget.set("")
        else:
            widget = ttk.Entry(parent, width=45)

        widget.grid(row=row_index, column=1, sticky="ew", pady=6)
        entries[field] = widget

        if field in FIELD_UNITS:
            unit_label = ttk.Label(
                parent,
                text=FIELD_UNITS[field],
                style="Unit.TLabel",
                anchor="w"
            )
            unit_label.grid(row=row_index, column=2, sticky="w", padx=(8, 0), pady=6)

    parent.columnconfigure(1, weight=1)

# ========================
# 厂房信息双列紧凑输入区
# ========================
def create_compact_workshop_rows(parent, fields, entries):
    for index, field in enumerate(fields):
        row = index // 2
        group_col = index % 2
        base_col = group_col * 3

        label = ttk.Label(
            parent,
            text=FIELD_LABELS[field],
            style="Form.TLabel",
            width=18,
            anchor="w"
        )
        label.grid(row=row, column=base_col, sticky="w", padx=(0, 6), pady=4)

        widget = ttk.Entry(parent, width=20)
        widget.grid(row=row, column=base_col + 1, sticky="ew", pady=4)
        entries[field] = widget

        unit_label = ttk.Label(
            parent,
            text=FIELD_UNITS.get(field, ""),
            style="Unit.TLabel",
            anchor="w"
        )
        unit_label.grid(row=row, column=base_col + 2, sticky="w", padx=(5, 18), pady=4)

    parent.columnconfigure(1, weight=1)
    parent.columnconfigure(4, weight=1)

# ========================
# 必填校验
# ========================
def validate_required_fields(context, picture_dir, template_type):
    required_fields = [
        "internal_code",
        "project_name",
        "subproject_code",
        "subproject_name",
        "design_phase",
        "document_name"
    ]

    missing = []

    if not template_type.strip():
        missing.append("计算书模板类型")

    for field in required_fields:
        if not context.get(field, "").strip():
            missing.append(FIELD_LABELS[field])

    if not picture_dir.strip():
        missing.append("配筋图文件夹")

    if missing:
        messagebox.showerror(
            "信息不完整",
            "请补充以下必填信息：\n\n" + "\n".join(f"· {item}" for item in missing)
        )
        return False

    return True

# ========================
# GUI
# ========================
def create_gui():
    root = Tk()
    root.title("核岛厂房计算书自动生成工具")
    root.geometry("820x810")
    root.minsize(760, 700)

    setup_style(root)

    outer_frame = ttk.Frame(root, style="App.TFrame")
    outer_frame.pack(fill="both", expand=True)

    canvas = Canvas(
        outer_frame,
        background="#eef2f7",
        highlightthickness=0
    )

    scrollbar = ttk.Scrollbar(
        outer_frame,
        orient="vertical",
        command=canvas.yview
    )

    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    content_frame = ttk.Frame(canvas, style="App.TFrame", padding=(24, 18, 24, 18))
    canvas_window = canvas.create_window((0, 0), window=content_frame, anchor="nw")

    def on_content_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def on_canvas_configure(event):
        canvas.itemconfig(canvas_window, width=event.width)

    content_frame.bind("<Configure>", on_content_configure)
    canvas.bind("<Configure>", on_canvas_configure)
    bind_mousewheel(root, canvas)

    header_frame = ttk.Frame(content_frame, style="Header.TFrame", padding=(24, 20, 24, 20))
    header_frame.pack(fill="x", pady=(0, 16))

    title_label = ttk.Label(
        header_frame,
        text="核岛厂房计算书自动生成工具",
        style="HeaderTitle.TLabel"
    )
    title_label.pack(anchor="w")

    subtitle_label = ttk.Label(
        header_frame,
        text="选择模板类型，填写工程信息、厂房信息并选择配筋图文件夹后，自动生成配筋计算书 Word 文件",
        style="HeaderSubTitle.TLabel"
    )
    subtitle_label.pack(anchor="w", pady=(6, 0))

    entries = {}

    # ========================
    # 一、模板选择
    # ========================
    template_frame = ttk.LabelFrame(
        content_frame,
        text="一、模板选择",
        style="Section.TLabelframe",
        padding=(20, 14, 20, 14)
    )
    template_frame.pack(fill="x", pady=(0, 12))

    template_label = ttk.Label(
        template_frame,
        text="计算书模板类型",
        style="Form.TLabel",
        width=22,
        anchor="w"
    )
    template_label.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)

    template_combo = ttk.Combobox(
        template_frame,
        width=42,
        values=TEMPLATE_OPTIONS,
        state="readonly"
    )
    template_combo.set("")
    template_combo.grid(row=0, column=1, sticky="ew", pady=6)

    template_tip = ttk.Label(
        template_frame,
        text="对应文件：内部结构计算书.docx / 核岛厂房计算书.docx",
        style="Unit.TLabel",
        anchor="w"
    )
    template_tip.grid(row=0, column=2, sticky="w", padx=(8, 0), pady=6)

    template_frame.columnconfigure(1, weight=1)

    # ========================
    # 二、工程基本信息
    # ========================
    form_frame = ttk.LabelFrame(
        content_frame,
        text="二、工程基本信息",
        style="Section.TLabelframe",
        padding=(20, 14, 20, 14)
    )
    form_frame.pack(fill="x", pady=(0, 12))

    create_form_rows(form_frame, BASIC_INFO_FIELDS, entries)

    # ========================
    # 三、厂房信息
    # ========================
    workshop_frame = ttk.LabelFrame(
        content_frame,
        text="三、厂房信息",
        style="Section.TLabelframe",
        padding=(20, 10, 12, 10)
    )
    workshop_frame.pack(fill="x", pady=(0, 12))

    create_compact_workshop_rows(workshop_frame, WORKSHOP_INFO_FIELDS, entries)

    # ========================
    # 四、图片文件夹
    # ========================
    folder_frame = ttk.LabelFrame(
        content_frame,
        text="四、图片文件夹",
        style="Section.TLabelframe",
        padding=(20, 14, 20, 14)
    )
    folder_frame.pack(fill="x", pady=(0, 12))

    folder_label = ttk.Label(
        folder_frame,
        text="配筋图文件夹",
        style="Form.TLabel",
        width=22,
        anchor="w"
    )
    folder_label.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)

    picture_entry = ttk.Entry(folder_frame, width=45)
    picture_entry.grid(row=0, column=1, sticky="ew", pady=6)

    def select_picture_folder():
        path = filedialog.askdirectory(title="选择配筋图文件夹")

        if path:
            picture_entry.delete(0, 'end')
            picture_entry.insert(0, path)

    select_button = ttk.Button(
        folder_frame,
        text="选择文件夹",
        style="Secondary.TButton",
        command=select_picture_folder
    )
    select_button.grid(row=0, column=2, sticky="e", padx=(10, 0), pady=6)

    folder_frame.columnconfigure(1, weight=1)

    # ========================
    # 底部状态区
    # ========================
    footer_frame = ttk.Frame(content_frame, style="Footer.TFrame", padding=(2, 4, 2, 0))
    footer_frame.pack(fill="x", pady=(0, 0))

    tip_label = ttk.Label(
        footer_frame,
        text="提示：模板、钢筋面积表和 Tesseract-OCR 可封装进 exe；运行时只需选择图片文件夹。",
        style="Tip.TLabel"
    )
    tip_label.pack(anchor="w", pady=(0, 8))

    status_var = StringVar()
    status_var.set("准备就绪")

    status_label = ttk.Label(
        footer_frame,
        textvariable=status_var,
        style="Status.TLabel"
    )
    status_label.pack(anchor="w", pady=(0, 8))

    progress = ttk.Progressbar(
        footer_frame,
        mode="indeterminate",
        length=260
    )
    progress.pack(anchor="w", pady=(0, 12))

    button_frame = ttk.Frame(footer_frame, style="Footer.TFrame")
    button_frame.pack(fill="x", pady=(0, 8))

    # ========================
    # 生成计算书
    # ========================
    def generate_doc():
        context = {f: entries[f].get().strip() for f in FIELD_LABELS}
        picture_dir = picture_entry.get().strip()
        template_type = template_combo.get().strip()

        if not validate_required_fields(context, picture_dir, template_type):
            return

        template_path = get_selected_template_path(template_type)

        if not os.path.exists(template_path):
            messagebox.showerror(
                "错误",
                f"未找到所选模板文件：\n\n{template_path}\n\n"
                f"请确认文件名是否为：\n{TEMPLATE_FILE_MAP.get(template_type, '')}"
            )
            return

        if not os.path.exists(REBAR_TABLE_PATH):
            messagebox.showerror(
                "错误",
                f"未找到钢筋面积表：\n\n{REBAR_TABLE_PATH}\n\n"
                "请确认文件名为：钢筋的公称直径、公称面积表.xlsx"
            )
            return

        if not os.path.isdir(picture_dir):
            messagebox.showerror("错误", "配筋图文件夹不存在")
            return

        folder_01 = os.path.join(picture_dir, "01")
        folder_02 = os.path.join(picture_dir, "02")

        if not os.path.isdir(folder_01):
            messagebox.showerror("错误", "配筋图文件夹中未找到 01 子文件夹")
            return

        if not os.path.isdir(folder_02):
            messagebox.showerror("错误", "配筋图文件夹中未找到 02 子文件夹")
            return

        try:
            progress.start(10)
            status_var.set("正在整理输入信息……")
            root.update_idletasks()

            internal_code = context["internal_code"]
            context["project_number"] = internal_code[:4]
            context["document_serial_number"] = internal_code[-2:]
            context["record_1_version"] = context.get("version", "A")
            context["record_1_date"] = date.today().strftime("%Y-%m-%d")

            # 根据当前厂房自动生成 other_plants
            context["other_plants"] = get_other_plants(context["subproject_code"])

            match = re.search(r'\d+(?:\.\d+)?m~\d+(?:\.\d+)?m', context["document_name"])
            context["plant_elevation_range"] = match.group() if match else "未识别"

            output_dir = os.path.join(get_app_dir(), "output")
            os.makedirs(output_dir, exist_ok=True)

            safe_template_name = template_type.replace(" ", "")

            OUTPUT_PATH = os.path.join(
                output_dir,
                f"{context['project_number']}{context['document_name']}{safe_template_name}.docx"
            )

            status_var.set(f"正在读取模板：{template_type}……")
            root.update_idletasks()

            doc = DocxTemplate(template_path)

            status_var.set("正在读取钢筋面积表……")
            root.update_idletasks()

            rebar_table = load_rebar_area_table(REBAR_TABLE_PATH)

            status_var.set("正在读取图片并识别配筋面积，请稍候……")
            root.update_idletasks()

            image_plant = get_first_image_from_folder(folder_01)
            image_wall = get_first_image_from_folder(folder_02)

            raw_figures = load_reinforcement_figures_parallel(picture_dir)

            # 先为每张配筋云图计算自动选筋结果
            for fig in raw_figures:
                rebar_result = select_rebar_by_sm_value(
                    sm_value=fig["sm_value"],
                    rebar_table=rebar_table,
                    row_count_options=REBAR_ROW_COUNT_OPTIONS,
                    spacing_options=REBAR_SPACING_OPTIONS,
                    max_diameter=REBAR_MAX_DIAMETER,
                    extra_ratio=REBAR_EXTRA_RATIO
                )

                fig["rebar_spec"] = rebar_result["spec"]
                fig["rebar_area"] = rebar_result["area"]
                fig["rebar_target_area"] = rebar_result["target_area"]
                fig["rebar_diameter"] = rebar_result["diameter"]
                fig["rebar_row_count"] = rebar_result["row_count"]
                fig["rebar_spacing"] = rebar_result["spacing"]

            # 再生成表格行：
            # row.x_actual / row.y_actual / row.z_actual 填入自动计算的实际配筋面积 rebar_area
            # row.x_margin / row.y_margin / row.z_margin 填入裕度百分比
            table_rows = generate_table_rows_from_figures(raw_figures)

            context["image_plant_elevation_layout"] = InlineImage(doc, image_plant, width=Mm(150))
            context["image_wall_fem_calculation_model"] = InlineImage(doc, image_wall, width=Mm(160))
            context["wall_table_rows"] = table_rows

            context["reinforcement_figures"] = []

            for i, fig in enumerate(raw_figures, start=1):
                context["reinforcement_figures"].append({
                    "figure_number": f"{CURRENT_CHAPTER}-{i}",
                    "image": InlineImage(doc, fig["full_path"], width=Mm(140)),
                    "caption": f"{fig['wall_id']}-{DIRECTION_MAP[fig['direction']]}",
                    "sm_value": fig["sm_value"],

                    # 自动选筋结果，供正文段落使用
                    "rebar_spec": fig["rebar_spec"],
                    "rebar_area": fig["rebar_area"],
                    "rebar_target_area": fig["rebar_target_area"],
                    "rebar_diameter": fig["rebar_diameter"],
                    "rebar_row_count": fig["rebar_row_count"],
                    "rebar_spacing": fig["rebar_spacing"]
                })

            status_var.set("正在渲染 Word 模板……")
            root.update_idletasks()

            doc.render(context)

            SAVE_PATH = get_available_output_path(OUTPUT_PATH)

            try:
                doc.save(SAVE_PATH)
            except PermissionError:
                SAVE_PATH = get_available_output_path(OUTPUT_PATH)
                doc.save(SAVE_PATH)

            status_var.set(f"生成完成：{SAVE_PATH}")
            messagebox.showinfo("完成", f"生成成功:\n{SAVE_PATH}")
            os.startfile(SAVE_PATH)

        except Exception as e:
            status_var.set("生成失败")
            messagebox.showerror("错误", f"生成失败：\n{e}")

        finally:
            progress.stop()

    generate_button = ttk.Button(
        button_frame,
        text="生成计算书",
        style="Primary.TButton",
        command=generate_doc
    )
    generate_button.pack(anchor="center", ipadx=40, ipady=4)

    root.mainloop()


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    create_gui()