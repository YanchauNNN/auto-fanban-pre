from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from ..audit_check.lexicon import normalize_text
from ..config import get_config


@dataclass(frozen=True)
class ReplaceMapping:
    source_project_no: str
    target_project_no: str
    replacements: dict[str, str] = field(default_factory=dict)
    no_op_tokens: list[str] = field(default_factory=list)
    missing_target_tokens: list[str] = field(default_factory=list)


class ReplaceMappingBuilder:
    def __init__(self) -> None:
        audit_cfg = get_config().audit_check
        self._project_no_re = re.compile(audit_cfg.project_column_header_pattern)
        self._include_rows = list(audit_cfg.include_rows)

    def build(
        self,
        *,
        workbook_path: str | Path,
        source_project_no: str,
        target_project_no: str,
    ) -> ReplaceMapping:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        project_columns: dict[str, int] = {}
        for column in range(1, worksheet.max_column + 1):
            raw = worksheet.cell(1, column).value
            if raw is None:
                continue
            project_no = str(raw).strip()
            if self._project_no_re.fullmatch(project_no):
                project_columns[project_no] = column

        if source_project_no not in project_columns:
            raise ValueError(f"unknown source_project_no: {source_project_no}")
        if target_project_no not in project_columns:
            raise ValueError(f"unknown target_project_no: {target_project_no}")

        replacements: dict[str, str] = {}
        no_op_tokens: list[str] = []
        missing_target_tokens: list[str] = []

        source_column = project_columns[source_project_no]
        target_column = project_columns[target_project_no]
        for row in range(1, worksheet.max_row + 1):
            if not self._should_include_row(row):
                continue

            source_token = self._normalize_cell(worksheet.cell(row, source_column).value)
            if not source_token:
                continue
            target_token = self._normalize_cell(worksheet.cell(row, target_column).value)
            if not target_token:
                if source_token not in missing_target_tokens:
                    missing_target_tokens.append(source_token)
                continue
            if source_token == target_token:
                if source_token not in no_op_tokens:
                    no_op_tokens.append(source_token)
                continue
            replacements[source_token] = target_token

        return ReplaceMapping(
            source_project_no=source_project_no,
            target_project_no=target_project_no,
            replacements=replacements,
            no_op_tokens=sorted(no_op_tokens),
            missing_target_tokens=sorted(missing_target_tokens),
        )

    def _should_include_row(self, row_number: int) -> bool:
        for marker in self._include_rows:
            if isinstance(marker, int) and row_number == marker:
                return True
            if isinstance(marker, str) and marker.endswith("+"):
                try:
                    base = int(marker[:-1])
                except ValueError:
                    continue
                if row_number >= base:
                    return True
            elif str(marker).isdigit() and row_number == int(marker):
                return True
        return False

    @staticmethod
    def _normalize_cell(value: object) -> str:
        if value is None:
            return ""
        return normalize_text(str(value))
