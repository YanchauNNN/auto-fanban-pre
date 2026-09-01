from __future__ import annotations

import re
import unicodedata
from itertools import chain
from pathlib import Path

from openpyxl import load_workbook

from ..config import get_config
from .models import AuditLexicon

_SPACE_RE = re.compile(r"\s+")


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = _SPACE_RE.sub(" ", text)
    return text.strip().upper()


def normalize_text_without_spaces(value: str) -> str:
    return _SPACE_RE.sub("", normalize_text(value))


class AuditLexiconLoader:
    def __init__(self) -> None:
        audit_cfg = get_config().audit_check
        self._project_no_re = re.compile(audit_cfg.project_column_header_pattern)
        self._include_rows = list(audit_cfg.include_rows)

    def load(self, workbook_path: str | Path) -> AuditLexicon:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows = worksheet.iter_rows(values_only=True)
            header = tuple(next(rows, ()))

            project_columns: list[tuple[int, str]] = []
            for column_index, raw in enumerate(header):
                if raw is None:
                    continue
                project_no = str(raw).strip()
                if self._project_no_re.fullmatch(project_no):
                    project_columns.append((column_index, project_no))

            project_options = [project_no for _, project_no in project_columns]
            allowed: dict[str, set[str]] = {
                project_no: set() for project_no in project_options
            }
            token_projects: dict[str, set[str]] = {}

            for row_number, values in enumerate(chain((header,), rows), start=1):
                if not self._should_include_row(row_number):
                    continue
                for column_index, project_no in project_columns:
                    raw = values[column_index] if column_index < len(values) else None
                    if raw is None:
                        continue
                    normalized = normalize_text(str(raw))
                    if not normalized:
                        continue
                    allowed[project_no].add(normalized)
                    token_projects.setdefault(normalized, set()).add(project_no)
        finally:
            workbook.close()

        foreign: dict[str, set[str]] = {}
        all_tokens = set(token_projects)
        for project_no in project_options:
            foreign[project_no] = all_tokens.difference(allowed[project_no])

        return AuditLexicon(
            project_options=project_options,
            allowed_texts=allowed,
            foreign_texts=foreign,
            token_projects=token_projects,
        )

    def _should_include_row(self, row_number: int) -> bool:
        for marker in self._include_rows:
            if isinstance(marker, int) and row_number == marker:
                return True
            if isinstance(marker, str) and marker.endswith("+"):
                try:
                    base = int(marker[:-1])
                except ValueError:
                    continue
                if row_number >= base:
                    return True
            elif str(marker).isdigit() and row_number == int(marker):
                return True
        return False
