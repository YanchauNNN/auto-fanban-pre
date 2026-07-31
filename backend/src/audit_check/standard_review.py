from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import AuditFinding, ScanTextItem

_SPACE_RE = re.compile(r"\s+")
_YEAR_SUFFIX_RE = re.compile(r"^(?P<base>.+?)(?:\s*-\s*|\s+)(?P<year>\d{4})$")
_HYPHEN_TRANSLATION = str.maketrans(
    {
        "\u2010": "-",
        "\u2011": "-",
        "\u2012": "-",
        "\u2013": "-",
        "\u2014": "-",
        "\u2015": "-",
        "\u2212": "-",
        "\uff0d": "-",
        "\ufe58": "-",
        "\ufe63": "-",
        "\u2043": "-",
        "\uff0f": "/",
        "\u2215": "/",
        "\u2044": "/",
    }
)
_NAME_LEADING_SEPARATORS = " \t\r\n:：,，、;-－–—"
_NAME_TRAILING_SEPARATORS = " \t\r\n:：,，、;-－–—.。"
_SAME_TEXT_NAME_END_RE = re.compile(r"[;；\r\n]+")
_YEAR_CAPTURE_PATTERN = r"[\(\[（【]?\s*(?P<year>\d{4})\s*[\)\]）】]?"
_CODE_TRAILING_CLOSERS = ")]】"
_CODE_OPEN_YEAR_SUFFIX_RE = re.compile(r"[\(\[【]\s*(?P<year>\d{4})$")
_BRACKETED_STANDARD_NAME_RE = re.compile(
    r"(?:《|<<)(?P<name>.*?)(?:》|>>)"
)
_EDITION_MARKER_RE = re.compile(r"[\(\[]?\s*(?P<edition>\d{4}\s*年版)\s*[\)\]]?")


@dataclass(frozen=True, slots=True)
class StandardEntry:
    canonical_code: str
    code_without_year: str
    expected_year: str | None
    expected_name: str
    source_sheet: str
    source_row: int
    department: str | None = None
    major: str | None = None
    status: str | None = None


@dataclass(frozen=True, slots=True)
class _CodeCandidate:
    item: ScanTextItem
    entry: StandardEntry
    actual_code: str
    actual_year: str | None
    start: int
    end: int


class StandardLibraryLoader:
    """Read standard-code metadata from the structured standard library workbook."""

    def load(self, workbook_path: str | Path, *, sheet_name: str = "DatStdItem") -> list[StandardEntry]:
        path = Path(workbook_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[1]]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_values = next(rows)
            except StopIteration:
                return []
            headers = self._headers_from_values(header_values)
            entries: list[StandardEntry] = []
            for row_number, row_values in enumerate(rows, start=2):
                raw_code = self._cell_text(self._row_value(row_values, headers["code"]))
                raw_name = self._cell_text(self._row_value(row_values, headers["name"]))
                if not raw_code or not raw_name or raw_code == "标准号":
                    continue
                raw_version = self._cell_text(self._optional_row_value(row_values, headers, "version"))
                code_without_year, expected_year, canonical_code = _canonicalize_standard_code(
                    raw_code,
                    raw_version,
                )
                if not canonical_code:
                    continue
                entries.append(
                    StandardEntry(
                        canonical_code=canonical_code,
                        code_without_year=code_without_year,
                        expected_year=expected_year,
                        expected_name=raw_name,
                        source_sheet=worksheet.title,
                        source_row=row_number,
                        department=self._optional_cell(row_values, headers, "department"),
                        major=self._optional_cell(row_values, headers, "major"),
                        status=self._optional_cell(row_values, headers, "status"),
                    )
                )
            return entries
        finally:
            workbook.close()

    @staticmethod
    def _headers_from_values(header_values: tuple[Any, ...]) -> dict[str, int]:
        aliases = {
            "code": {"CodeStd", "标准号"},
            "name": {"NameStd", "标准名称"},
            "version": {"Version", "版本"},
            "department": {"Department", "部门"},
            "major": {"Major", "专业"},
            "status": {"Status", "状态"},
        }
        found: dict[str, int] = {}
        for column, raw in enumerate(header_values):
            if raw is None:
                continue
            value = str(raw).strip()
            for key, names in aliases.items():
                if value in names:
                    found[key] = column
        if "code" not in found or "name" not in found:
            raise ValueError("standard library requires CodeStd/NameStd columns")
        return found

    @staticmethod
    def _cell_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    def _optional_cell(
        self,
        row_values: tuple[Any, ...],
        headers: dict[str, int],
        key: str,
    ) -> str | None:
        return self._cell_text(self._optional_row_value(row_values, headers, key)) or None

    @staticmethod
    def _optional_row_value(
        row_values: tuple[Any, ...],
        headers: dict[str, int],
        key: str,
    ) -> object | None:
        column = headers.get(key)
        if column is None:
            return None
        return StandardLibraryLoader._row_value(row_values, column)

    @staticmethod
    def _row_value(row_values: tuple[Any, ...], column: int) -> object | None:
        if column >= len(row_values):
            return None
        return row_values[column]


class StandardReviewEngine:
    def __init__(
        self,
        entries: list[StandardEntry],
        *,
        same_line_y_tolerance: float,
        same_text_pairing_enabled: bool = True,
        format_variant_compatibility_enabled: bool = True,
        same_entity_name_before_code_enabled: bool = True,
        same_entity_code_before_name_enabled: bool = True,
        multiple_pairs_in_one_entity_enabled: bool = True,
        fallback_name_keywords: list[str] | None = None,
        fallback_min_name_length: int = 4,
        continuation_line_enabled: bool = True,
        continuation_line_y_height_factor: float = 2.2,
        continuation_line_x_height_factor: float = 1.0,
    ) -> None:
        self.entries = entries
        self.same_line_y_tolerance = float(same_line_y_tolerance)
        self.same_text_pairing_enabled = bool(same_text_pairing_enabled)
        self.format_variant_compatibility_enabled = bool(format_variant_compatibility_enabled)
        self.same_entity_name_before_code_enabled = bool(same_entity_name_before_code_enabled)
        self.same_entity_code_before_name_enabled = bool(same_entity_code_before_name_enabled)
        self.multiple_pairs_in_one_entity_enabled = bool(multiple_pairs_in_one_entity_enabled)
        self.fallback_name_keywords = tuple(
            _normalize_name_key(keyword)
            for keyword in (fallback_name_keywords or ["标准", "规范", "规程", "图集"])
            if _normalize_name_key(keyword)
        )
        self.fallback_min_name_length = max(1, int(fallback_min_name_length))
        self.continuation_line_enabled = bool(continuation_line_enabled)
        self.continuation_line_y_height_factor = max(
            1.0,
            float(continuation_line_y_height_factor),
        )
        self.continuation_line_x_height_factor = max(
            0.0,
            float(continuation_line_x_height_factor),
        )
        self._base_entries: dict[str, list[StandardEntry]] = {}
        for entry in entries:
            self._base_entries.setdefault(_normalize_code_key(entry.code_without_year), []).append(entry)
        self._base_patterns = [
            (
                entry,
                self._compile_code_pattern(
                    entry.code_without_year,
                    format_variant_compatibility_enabled=self.format_variant_compatibility_enabled,
                ),
            )
            for entry in entries
            if entry.expected_year
        ]

    def evaluate(self, items: list[ScanTextItem]) -> list[AuditFinding]:
        code_candidates = self._find_code_candidates(items)
        findings: list[AuditFinding] = []
        for candidate in code_candidates:
            findings.extend(self._year_findings(candidate))
            findings.extend(self._name_findings(candidate, items))
        return findings

    def _find_code_candidates(self, items: list[ScanTextItem]) -> list[_CodeCandidate]:
        candidates: list[_CodeCandidate] = []
        seen: set[tuple[str | None, str, str]] = set()
        for item in items:
            normalized = _normalize_code_text(item.raw_text)
            if not normalized:
                continue
            for pattern_entry, pattern in self._base_patterns:
                for match in pattern.finditer(normalized):
                    actual_year = str(match.group("year") or "")
                    entries = self._base_entries.get(_normalize_code_key(pattern_entry.code_without_year), [])
                    entry = self._select_expected_entry(entries, actual_year) or pattern_entry
                    actual_code = _normalize_actual_standard_code(match.group(0))
                    key = (item.entity_handle, actual_code, entry.canonical_code)
                    if key in seen:
                        continue
                    seen.add(key)
                    candidates.append(
                        _CodeCandidate(
                            item=item,
                            entry=entry,
                            actual_code=actual_code,
                            actual_year=actual_year,
                            start=match.start(),
                            end=match.end(),
                        )
                    )
        return candidates

    @staticmethod
    def _select_expected_entry(entries: list[StandardEntry], actual_year: str) -> StandardEntry | None:
        for entry in entries:
            if entry.expected_year == actual_year:
                return entry
        return entries[0] if entries else None

    def _year_findings(self, candidate: _CodeCandidate) -> list[AuditFinding]:
        expected_year = candidate.entry.expected_year
        if not expected_year or candidate.actual_year == expected_year:
            return []
        return [
            self._finding(
                candidate=candidate,
                context_kind="standard_review_year",
                details={
                    "issue_type": "year_mismatch",
                    "actual_code": candidate.actual_code,
                    "expected_code": candidate.entry.canonical_code,
                    "actual_year": candidate.actual_year or "",
                    "expected_year": expected_year,
                    "expected_name": candidate.entry.expected_name,
                },
            )
        ]

    def _name_findings(self, candidate: _CodeCandidate, items: list[ScanTextItem]) -> list[AuditFinding]:
        same_item_name: str | None = None
        if self.same_text_pairing_enabled:
            same_item_name = self._same_item_name(candidate)
            if (
                same_item_name is not None
                and _name_matches(same_item_name, candidate.entry.expected_name)
            ):
                return []

        continued_name = self._continued_name(candidate, items)
        if continued_name is not None:
            if _name_matches(continued_name, candidate.entry.expected_name):
                return []
            same_item_name = continued_name

        if same_item_name is not None:
            return [
                self._finding(
                    candidate=candidate,
                    context_kind="standard_review_name",
                    details={
                        "issue_type": "name_mismatch",
                        "actual_code": candidate.actual_code,
                        "expected_code": candidate.entry.canonical_code,
                        "actual_name": same_item_name,
                        "expected_name": candidate.entry.expected_name,
                    },
                )
            ]

        name_item = self._nearest_name_item(candidate, items)
        if name_item is None:
            return [
                self._finding(
                    candidate=candidate,
                    context_kind="standard_review_name",
                    details={
                        "issue_type": "name_missing",
                        "actual_code": candidate.actual_code,
                        "expected_code": candidate.entry.canonical_code,
                        "actual_name": "",
                        "expected_name": candidate.entry.expected_name,
                    },
                )
            ]
        actual_name = str(name_item.raw_text or "").strip()
        if _name_matches(actual_name, candidate.entry.expected_name):
            return []
        return [
            self._finding(
                candidate=candidate,
                context_kind="standard_review_name",
                details={
                    "issue_type": "name_mismatch",
                    "actual_code": candidate.actual_code,
                    "expected_code": candidate.entry.canonical_code,
                    "actual_name": actual_name,
                    "expected_name": candidate.entry.expected_name,
                },
            )
        ]

    def _nearest_name_item(
        self,
        candidate: _CodeCandidate,
        items: list[ScanTextItem],
    ) -> ScanTextItem | None:
        code_y = _item_y(candidate.item)
        if code_y is None:
            return None
        code_x = _item_x(candidate.item) or 0.0
        nearby: list[tuple[float, float, ScanTextItem]] = []
        for item in items:
            if item is candidate.item:
                continue
            text = str(item.raw_text or "").strip()
            if not text or self._looks_like_standard_code(text):
                continue
            if not self._looks_like_standard_name(text, candidate.entry.expected_name):
                continue
            item_y = _item_y(item)
            if item_y is None:
                continue
            y_distance = abs(item_y - code_y)
            if y_distance > self.same_line_y_tolerance:
                continue
            item_x = _item_x(item)
            x_distance = abs((item_x or 0.0) - code_x)
            nearby.append((y_distance, x_distance, item))
        if not nearby:
            return None
        return sorted(nearby, key=lambda row: (row[0], row[1]))[0][2]

    def _looks_like_standard_code(self, text: str) -> bool:
        normalized = _normalize_code_text(text)
        return any(pattern.search(normalized) for _, pattern in self._base_patterns)

    def _looks_like_standard_name(self, text: str, expected_name: str) -> bool:
        normalized = _normalize_name_key(text)
        if len(normalized) < self.fallback_min_name_length:
            return False
        if _name_matches(text, expected_name):
            return True
        stripped = text.strip()
        if _BRACKETED_STANDARD_NAME_RE.search(stripped):
            return True
        return any(keyword in normalized for keyword in self.fallback_name_keywords)

    @staticmethod
    def _compile_code_pattern(
        code_without_year: str,
        *,
        format_variant_compatibility_enabled: bool,
    ) -> re.Pattern[str]:
        normalized = _normalize_standard_code_display(code_without_year)
        if not format_variant_compatibility_enabled:
            escaped = re.escape(normalized).replace(r"\ ", r"\s*")
            return re.compile(rf"(?<![A-Z0-9]){escaped}\s*-\s*{_YEAR_CAPTURE_PATTERN}(?!\d)")
        parts = _split_standard_code_without_year(normalized)
        if parts is None:
            escaped = re.escape(normalized).replace(r"\ ", r"\s*")
            return re.compile(rf"(?<![A-Z0-9]){escaped}\s*-\s*{_YEAR_CAPTURE_PATTERN}(?!\d)")
        prefix, number = parts
        prefix_pattern = _standard_prefix_pattern(prefix)
        number_pattern = _loose_literal_pattern(number)
        separator_pattern = r"(?:\s*-\s*|\s+|/|\s*)"
        return re.compile(
            rf"(?<![A-Z0-9]){prefix_pattern}{number_pattern}{separator_pattern}{_YEAR_CAPTURE_PATTERN}(?!\d)"
        )

    def _same_item_name(self, candidate: _CodeCandidate) -> str | None:
        normalized = _normalize_code_text(candidate.item.raw_text)
        bracketed_name = self._bracketed_name_in_same_item(candidate, normalized)
        if bracketed_name is not None:
            return bracketed_name
        if not self.same_entity_code_before_name_enabled:
            return None
        return self._name_after_code_in_same_item(candidate, normalized)

    def _bracketed_name_in_same_item(
        self,
        candidate: _CodeCandidate,
        normalized: str,
    ) -> str | None:
        matches = list(_BRACKETED_STANDARD_NAME_RE.finditer(normalized))
        if not matches:
            return None
        if not self.multiple_pairs_in_one_entity_enabled:
            code_count = sum(1 for _, pattern in self._base_patterns for _ in pattern.finditer(normalized))
            if code_count > 1:
                return None

        if self.same_entity_name_before_code_enabled:
            preceding = [
                match
                for match in matches
                if match.end() <= candidate.start
                and not self._has_standard_code(normalized[match.end() : candidate.start])
            ]
            if preceding:
                match = preceding[-1]
                if self._bracketed_name_belongs_to_candidate(
                    candidate=candidate,
                    normalized=normalized,
                    name_match=match,
                ):
                    return self._name_with_edition_marker(
                        candidate=candidate,
                        normalized=normalized,
                        name=match.group("name").strip(),
                    )

        if self.same_entity_code_before_name_enabled:
            following = [
                match
                for match in matches
                if match.start() >= candidate.end
                and not self._has_standard_code(normalized[candidate.end : match.start()])
            ]
            if following:
                match = following[0]
                if self._bracketed_name_belongs_to_candidate(
                    candidate=candidate,
                    normalized=normalized,
                    name_match=match,
                ):
                    return self._name_with_edition_marker(
                        candidate=candidate,
                        normalized=normalized,
                        name=match.group("name").strip(),
                    )
        return None

    def _bracketed_name_belongs_to_candidate(
        self,
        *,
        candidate: _CodeCandidate,
        normalized: str,
        name_match: re.Match[str],
    ) -> bool:
        if name_match.end() <= candidate.start:
            candidate_distance = candidate.start - name_match.end()
        else:
            candidate_distance = name_match.start() - candidate.end

        other_distances: list[int] = []
        seen_spans: set[tuple[int, int]] = set()
        for _, pattern in self._base_patterns:
            for code_match in pattern.finditer(normalized):
                span = code_match.span()
                if span in seen_spans or span == (candidate.start, candidate.end):
                    continue
                seen_spans.add(span)
                if code_match.end() <= name_match.start():
                    other_distances.append(name_match.start() - code_match.end())
                elif code_match.start() >= name_match.end():
                    other_distances.append(code_match.start() - name_match.end())

        return not other_distances or candidate_distance < min(other_distances)

    def _continued_name(
        self,
        candidate: _CodeCandidate,
        items: list[ScanTextItem],
    ) -> str | None:
        if not self.continuation_line_enabled:
            return None
        candidate_y = _item_y(candidate.item)
        candidate_left = _item_left(candidate.item)
        if candidate_y is None or candidate_left is None:
            return None

        aligned: list[tuple[float, float, ScanTextItem]] = []
        for item in items:
            if item is candidate.item:
                continue
            if item.internal_code != candidate.item.internal_code:
                continue
            if item.layout_name != candidate.item.layout_name:
                continue
            item_y = _item_y(item)
            item_left = _item_left(item)
            if item_y is None or item_left is None:
                continue

            y_gap = candidate_y - item_y
            if y_gap <= self.same_line_y_tolerance:
                continue
            reference_height = max(
                _item_height(candidate.item) or self.same_line_y_tolerance,
                _item_height(item) or self.same_line_y_tolerance,
            )
            max_y_gap = reference_height * self.continuation_line_y_height_factor
            if y_gap > max_y_gap:
                continue
            x_gap = abs(candidate_left - item_left)
            max_x_gap = max(
                self.same_line_y_tolerance,
                reference_height * self.continuation_line_x_height_factor,
            )
            if x_gap > max_x_gap:
                continue
            aligned.append((y_gap, x_gap, item))

        for _, _, item in sorted(aligned, key=lambda row: (row[0], row[1])):
            current_text = _normalize_code_text(candidate.item.raw_text)
            continued_text = _normalize_code_text(item.raw_text)
            combined = f"{current_text[candidate.end:]}{continued_text}"
            match = _BRACKETED_STANDARD_NAME_RE.search(combined)
            if match is not None:
                return match.group("name").strip()
        return None

    def _has_standard_code(self, text: str) -> bool:
        return any(pattern.search(text) for _, pattern in self._base_patterns)

    @staticmethod
    def _name_with_edition_marker(
        *,
        candidate: _CodeCandidate,
        normalized: str,
        name: str,
    ) -> str:
        expected_key = _normalize_name_key(candidate.entry.expected_name)
        name_key = _normalize_name_key(name)
        if not name_key or not expected_key.startswith(name_key):
            return name
        expected_suffix = expected_key[len(name_key) :]
        if not re.fullmatch(r"\d{4}年版", expected_suffix):
            return name
        edition_match = _EDITION_MARKER_RE.search(normalized[candidate.end :])
        if edition_match is None:
            return name
        edition = edition_match.group("edition")
        if _normalize_name_key(edition) != expected_suffix:
            return name
        return f"{name}（{edition}）"

    def _name_after_code_in_same_item(self, candidate: _CodeCandidate, normalized: str) -> str | None:
        if candidate.end >= len(normalized):
            return None
        tail = normalized[candidate.end :]
        matched_segment = normalized[candidate.start : candidate.end]
        consumed_closing_parenthesis = matched_segment.endswith(tuple(_CODE_TRAILING_CLOSERS))
        if (
            not tail
            or (
                tail[0] not in _NAME_LEADING_SEPARATORS
                and not consumed_closing_parenthesis
            )
        ):
            return None
        tail = tail.lstrip(_NAME_LEADING_SEPARATORS)
        if not tail:
            return ""

        next_code_start: int | None = None
        for _, pattern in self._base_patterns:
            next_match = pattern.search(tail)
            if next_match is None:
                continue
            if next_code_start is None or next_match.start() < next_code_start:
                next_code_start = next_match.start()
        if next_code_start is not None:
            tail = tail[:next_code_start]

        end_match = _SAME_TEXT_NAME_END_RE.search(tail)
        if end_match is not None:
            tail = tail[: end_match.start()]
        return tail.strip(_NAME_TRAILING_SEPARATORS)

    @staticmethod
    def _finding(
        *,
        candidate: _CodeCandidate,
        context_kind: str,
        details: dict[str, Any],
    ) -> AuditFinding:
        return AuditFinding(
            raw_text=candidate.item.raw_text,
            matched_text=candidate.actual_code,
            matched_project_nos=[],
            context_kind=context_kind,
            confidence="high",
            entity_type=candidate.item.entity_type,
            field_context=candidate.item.field_context,
            internal_code=candidate.item.internal_code,
            layout_name=candidate.item.layout_name,
            entity_handle=candidate.item.entity_handle,
            block_path=candidate.item.block_path,
            position_x=candidate.item.position_x,
            position_y=candidate.item.position_y,
            text_bbox=candidate.item.text_bbox,
            details=details,
        )


def _canonicalize_standard_code(raw_code: str, raw_version: str) -> tuple[str, str | None, str]:
    code = _normalize_standard_code_display(raw_code)
    version = str(raw_version or "").strip()
    match = _YEAR_SUFFIX_RE.match(code)
    if match:
        base = _normalize_standard_code_display(match.group("base"))
        year = match.group("year")
        return base, year, f"{base}-{year}"
    if version and re.fullmatch(r"\d{4}", version):
        return code, version, f"{code}-{version}"
    return code, None, code


def _split_standard_code_without_year(value: str) -> tuple[str, str] | None:
    compact = re.sub(r"\s+", "", value or "")
    match = re.fullmatch(r"(?P<prefix>[A-Z]+(?:/[A-Z]+)?)(?P<number>\d[\dA-Z.]*)", compact)
    if match is None:
        return None
    return match.group("prefix"), match.group("number")


def _standard_prefix_pattern(prefix: str) -> str:
    pieces: list[str] = []
    for char in prefix:
        if char == "/":
            pieces.append(r"(?:\s*/\s*)?")
        else:
            pieces.append(re.escape(char) + r"\s*")
    return "".join(pieces)


def _loose_literal_pattern(value: str) -> str:
    compact = re.sub(r"\s+", "", value or "")
    return r"\s*".join(re.escape(char) for char in compact)


def _normalize_code_text(value: str) -> str:
    return _SPACE_RE.sub(" ", unicodedata.normalize("NFKC", value or "").translate(_HYPHEN_TRANSLATION)).strip().upper()


def _normalize_standard_code_display(value: str) -> str:
    return _normalize_code_text(value)


def _normalize_actual_standard_code(value: str) -> str:
    normalized = _normalize_standard_code_display(value).rstrip(_CODE_TRAILING_CLOSERS)
    return _CODE_OPEN_YEAR_SUFFIX_RE.sub(
        lambda match: f"-{match.group('year')}",
        normalized,
    )


def _normalize_code_key(value: str) -> str:
    return re.sub(r"\s+", "", _normalize_standard_code_display(value))


def _normalize_name_key(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").upper()
    text = text.translate(_HYPHEN_TRANSLATION)
    return re.sub(r"[\s,，、。.:：;；()（）\[\]【】《》<>_\-/\\]+", "", text)


def _name_matches(actual_name: str, expected_name: str) -> bool:
    actual = _normalize_name_key(actual_name)
    expected = _normalize_name_key(expected_name)
    return bool(expected and (actual == expected or expected in actual))


def _item_x(item: ScanTextItem) -> float | None:
    if item.text_bbox is not None:
        return float((item.text_bbox.xmin + item.text_bbox.xmax) / 2.0)
    if item.position_x is not None:
        return float(item.position_x)
    return None


def _item_left(item: ScanTextItem) -> float | None:
    if item.text_bbox is not None:
        return float(item.text_bbox.xmin)
    if item.position_x is not None:
        return float(item.position_x)
    return None


def _item_height(item: ScanTextItem) -> float | None:
    if item.text_bbox is None:
        return None
    return max(0.0, float(item.text_bbox.ymax - item.text_bbox.ymin))


def _item_y(item: ScanTextItem) -> float | None:
    if item.text_bbox is not None:
        return float((item.text_bbox.ymin + item.text_bbox.ymax) / 2.0)
    if item.position_y is not None:
        return float(item.position_y)
    return None
