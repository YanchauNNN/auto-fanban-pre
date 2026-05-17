"""
目录生成器 - Excel文档生成

职责：
1. 打开目录模板
2. 写入表头和明细行
3. 计算页数（优先Excel分页信息，兜底PDF计页）
4. 回填页数后导出PDF

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: catalog_bindings配置

测试要点：
- test_generate_catalog_common: 通用目录生成
- test_generate_catalog_1818: 1818目录（中英文标题同格）
- test_catalog_row_order: 行顺序（封面→目录→图纸）
- test_catalog_page_count: 页数计算
- test_catalog_upgrade_note: 升版标记
"""

from __future__ import annotations

import contextlib
import gc
import math
import os
import re
import shutil
from copy import copy
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter, range_boundaries

from ..config import load_spec
from ..interfaces import GenerationError, ICatalogGenerator, IPDFExporter
from .catalog_display_title import build_catalog_display_title
from .naming import make_document_output_name
from .office_automation import get_office_automation_limiter
from .pdf_engine import PDFExporter
from .upgrade_marking import (
    UpgradeEntryParseError,
    UpgradeSheetCodeParseError,
    get_added_note_text,
    get_upgrade_note_text,
    parse_upgrade_entries,
    parse_upgrade_sheet_codes,
)

if TYPE_CHECKING:
    from ..models import DocContext


class CatalogGenerator(ICatalogGenerator):
    """目录生成器实现"""

    BODY_ROW_HEIGHT = 36
    THREE_LINE_HEIGHT = 50
    FOUR_LINE_HEIGHT = 60
    EXTRA_LINE_STEP = 12
    DEFAULT_1818_HEADER_FONT_SIZES = (12, 9, 7, 6, 5)
    HEADER_WIDTH_PADDING_RATIO = 0.96
    HEADER_LINE_HEIGHT_RATIO = 1.18

    def __init__(
        self,
        spec_path: str | None = None,
        pdf_exporter: IPDFExporter | None = None,
    ):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
        self.pdf_exporter = pdf_exporter or PDFExporter()

    def generate(self, ctx: DocContext, output_dir: Path) -> tuple[Path, Path, int]:
        """生成目录文档"""
        output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 选择模板
        template_path = self._get_template_path(ctx)
        if not Path(template_path).exists():
            raise GenerationError(f"目录模板不存在: {template_path}")

        # 2. 获取落点配置
        bindings = self.spec.get_catalog_bindings()

        # 3. 写入Excel
        output_stem = self._build_output_stem(ctx)
        output_xlsx = output_dir / f"{output_stem}.xlsx"
        self._write_catalog(template_path, output_xlsx, bindings, ctx)

        # 4. 计算页数（优先Excel分页信息）
        page_count = self._count_pages(output_xlsx)

        # 5. 回填目录行页数
        self._backfill_page_count(output_xlsx, page_count, bindings)

        # 6. 导出PDF
        output_pdf = output_dir / f"{output_stem}.pdf"
        self.pdf_exporter.export_xlsx_to_pdf(output_xlsx, output_pdf)

        return output_xlsx, output_pdf, page_count

    def _build_output_stem(self, ctx: DocContext) -> str:
        return make_document_output_name(
            external_code=ctx.derived.catalog_external_code,
            revision=ctx.get_cover_catalog_revision(),
            status=ctx.params.doc_status,
            internal_code=ctx.derived.catalog_internal_code,
            fallback_name="目录",
        )

    def _get_template_path(self, ctx: DocContext) -> str:
        """获取模板路径"""
        return self.spec.get_template_path("catalog", ctx.params.project_no)

    def _write_catalog(
        self,
        template_path: str,
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入目录Excel"""
        wb = load_workbook(template_path)
        ws = wb.active
        if ws is None:
            raise GenerationError("目录模板缺少活动工作表")

        # 写入表头
        self._write_header(ws, bindings, ctx)

        # 写入明细行
        start_row = bindings.get("detail", {}).get("start_row", 9)
        current_row = start_row

        # 行顺序：封面 → 目录 → 图纸（按internal_code尾号升序）
        rows = self._build_detail_rows(ctx)

        for row_data in rows:
            self._write_detail_row(ws, current_row, row_data, bindings, ctx)
            current_row += 1

        # 动态设置打印区域，保证目录计页与实际行数一致
        last_row = max(start_row, current_row - 1)
        ws.print_area = f"$A$1:$I${last_row}"
        self._apply_detail_layout(ws, start_row, last_row)
        self._repair_detail_grid_border_holes(ws, start_row, last_row)

        # 保存
        wb.save(output_path)
        self._refine_detail_layout_via_com(output_path, start_row, last_row)

    def _write_header(self, ws, bindings: dict, ctx: DocContext) -> None:
        """写入表头"""
        header = bindings.get("header", {})
        derived = ctx.derived
        params = ctx.params

        # engineering_no → C1
        if "engineering_no" in header:
            cell = self._resolve_writable_cell(ws, header["engineering_no"].get("cell", "C1"))
            ws[cell] = params.engineering_no

        if "album_title_cn" in header and params.album_title_cn:
            cell = self._resolve_writable_cell(ws, header["album_title_cn"].get("cell", "D1:E1"))
            ws[cell] = params.album_title_cn

        if (
            ctx.is_1818
            and "album_title_en" in header
            and params.album_title_en
        ):
            title_binding = header["album_title_en"]
            cell_ref = title_binding.get("cell", "D2:E3")
            cell = self._resolve_writable_cell(ws, cell_ref)
            ws[cell] = params.album_title_en
            self._fit_1818_english_header_title(ws, title_binding, params.album_title_en)

        # catalog_internal_code → H1
        if "catalog_internal_code" in header:
            cell = self._resolve_writable_cell(ws, header["catalog_internal_code"].get("cell", "H1"))
            ws[cell] = derived.catalog_internal_code

        # catalog_external_code → H3
        if "catalog_external_code" in header:
            cell = self._resolve_writable_cell(ws, header["catalog_external_code"].get("cell", "H3"))
            ws[cell] = derived.catalog_external_code

        # subitem_no → C5
        if "subitem_no" in header:
            cell = self._resolve_writable_cell(ws, header["subitem_no"].get("cell", "C5"))
            ws[cell] = params.subitem_no

        # catalog_revision → H5
        if "catalog_revision" in header:
            cell = self._resolve_writable_cell(ws, header["catalog_revision"].get("cell", "H5"))
            ws[cell] = derived.catalog_revision

        if "album_code_title" in header:
            title_binding = header["album_code_title"]
            cell_ref = (
                title_binding.get("cell_1818")
                if ctx.is_1818 and title_binding.get("cell_1818")
                else title_binding.get("cell", "D3:E3")
            )
            cell = self._resolve_writable_cell(ws, cell_ref)
            template = header["album_code_title"].get(
                "template",
                "图纸(文件)目录",
            )
            ws[cell] = str(template).format(album_code=str(derived.album_code or "").strip())

    def _resolve_writable_cell(self, ws, cell_ref: str) -> str:
        anchor = cell_ref.split(":")[0]
        if not isinstance(ws[anchor], MergedCell):
            return anchor
        for merged_range in ws.merged_cells.ranges:
            if anchor in merged_range:
                return merged_range.start_cell.coordinate
        return anchor

    def _fit_1818_english_header_title(
        self,
        ws,
        binding: dict[str, Any],
        text: str,
    ) -> None:
        cell_ref = binding.get("cell", "D2:E3")
        cell = ws[self._resolve_writable_cell(ws, cell_ref)]
        font_sizes = self._resolve_header_font_sizes(binding.get("font_sizes"))
        shrink_to_fit_fallback = bool(binding.get("shrink_to_fit_fallback", False))

        chosen_size = font_sizes[-1]
        fits_in_range = False
        for font_size in font_sizes:
            if self._merged_text_fits(ws, cell_ref, text, font_size):
                chosen_size = font_size
                fits_in_range = True
                break

        font = copy(cell.font)
        font.sz = chosen_size
        cell.font = font

        alignment = copy(cell.alignment)
        alignment.wrapText = True
        alignment.shrinkToFit = shrink_to_fit_fallback and not fits_in_range
        cell.alignment = alignment

    def _resolve_header_font_sizes(self, raw_sizes: Any) -> tuple[int, ...]:
        if isinstance(raw_sizes, (list, tuple)):
            parsed = []
            for size in raw_sizes:
                with contextlib.suppress(TypeError, ValueError):
                    parsed.append(int(size))
            if parsed:
                return tuple(parsed)
        return self.DEFAULT_1818_HEADER_FONT_SIZES

    def _merged_text_fits(self, ws, cell_ref: str, text: str, font_size: float) -> bool:
        available_width_points = self._merged_range_width_points(ws, cell_ref)
        available_height_points = self._merged_range_height_points(ws, cell_ref)
        if available_width_points <= 0 or available_height_points <= 0:
            return True

        line_count = self._estimate_wrapped_line_count_points(
            text=text,
            available_width_points=available_width_points * self.HEADER_WIDTH_PADDING_RATIO,
            font_size=font_size,
        )
        available_lines = max(
            1,
            math.floor(available_height_points / (font_size * self.HEADER_LINE_HEIGHT_RATIO)),
        )
        return line_count <= available_lines

    def _merged_range_width_points(self, ws, cell_ref: str) -> float:
        min_col, _, max_col, _ = range_boundaries(cell_ref if ":" in cell_ref else f"{cell_ref}:{cell_ref}")
        default_width = ws.sheet_format.defaultColWidth or 8.43
        total_width = 0.0
        for column_idx in range(min_col, max_col + 1):
            column_letter = get_column_letter(column_idx)
            column_dimension = ws.column_dimensions[column_letter]
            column_width = column_dimension.width if column_dimension.width is not None else default_width
            total_width += self._column_width_to_points(column_width)
        return total_width

    def _merged_range_height_points(self, ws, cell_ref: str) -> float:
        _, min_row, _, max_row = range_boundaries(cell_ref if ":" in cell_ref else f"{cell_ref}:{cell_ref}")
        default_height = ws.sheet_format.defaultRowHeight or 15
        total_height = 0.0
        for row_idx in range(min_row, max_row + 1):
            row_dimension = ws.row_dimensions[row_idx]
            total_height += row_dimension.height if row_dimension.height is not None else default_height
        return total_height

    def _column_width_to_points(self, column_width: float) -> float:
        return max(column_width, 0.0) * 5.3

    def _estimate_wrapped_line_count_points(
        self,
        text: str,
        available_width_points: float,
        font_size: float,
    ) -> int:
        paragraphs = text.splitlines() or [text]
        return sum(
            self._wrap_word_line_count(
                text=paragraph if paragraph.strip() else " ",
                available_width_points=available_width_points,
                font_size=font_size,
            )
            for paragraph in paragraphs
        )

    def _wrap_word_line_count(
        self,
        text: str,
        available_width_points: float,
        font_size: float,
    ) -> int:
        stripped = text.strip()
        if not stripped:
            return 1
        if available_width_points <= 0:
            return len(stripped)

        words = re.findall(r"\S+", stripped)
        if not words:
            return 1

        space_width = self._char_display_width_points(" ", font_size)
        line_count = 1
        line_width = 0.0

        for word in words:
            word_width = sum(self._char_display_width_points(char, font_size) for char in word)
            if line_width == 0:
                extra_lines = max(1, math.ceil(word_width / available_width_points))
                line_count += extra_lines - 1
                line_width = word_width - ((extra_lines - 1) * available_width_points)
                continue

            projected_width = line_width + space_width + word_width
            if projected_width <= available_width_points:
                line_width = projected_width
                continue

            line_count += 1
            extra_lines = max(1, math.ceil(word_width / available_width_points))
            line_count += extra_lines - 1
            line_width = word_width - ((extra_lines - 1) * available_width_points)

        return line_count

    def _char_display_width_points(self, char: str, font_size: float) -> float:
        if char.isspace():
            return font_size * 0.28
        if "\u4e00" <= char <= "\u9fff":
            return font_size
        if char.isascii() and char.isalnum():
            return font_size * 0.58
        if char.isascii():
            return font_size * 0.38
        return font_size * 0.9

    def _build_detail_rows(self, ctx: DocContext) -> list[dict]:
        """构建明细行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params
        upgrade_note_text = get_upgrade_note_text(params.project_no) if params.is_upgrade else ""
        added_note_text = get_added_note_text(params.project_no) if params.is_upgrade else ""
        note_by_sheet_code: dict[str, str] = {}
        upgraded_sheet_codes: set[str] = set()

        if params.is_upgrade and params.upgrade_entries:
            try:
                for entry in parse_upgrade_entries(params.upgrade_entries):
                    note = added_note_text if entry.is_added else upgrade_note_text
                    for sheet_code in entry.sheet_codes:
                        note_by_sheet_code[sheet_code] = note
            except UpgradeEntryParseError as exc:
                raise GenerationError(f"升版规则格式错误: {exc.error_code}") from exc
        elif params.is_upgrade and params.upgrade_sheet_codes.strip():
            try:
                upgraded_sheet_codes = set(parse_upgrade_sheet_codes(params.upgrade_sheet_codes))
            except UpgradeSheetCodeParseError as exc:
                invalid = "、".join(exc.invalid_fragments)
                raise GenerationError(f"升版图纸编号格式错误: {invalid}") from exc

        # 1. 封面行
        rows.append({
            "type": "cover",
            "internal_code": derived.cover_internal_code,
            "external_code": derived.cover_external_code,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
            "revision": ctx.get_cover_catalog_revision(),
            "status": params.doc_status,
            "page_total": 1,
            "upgrade_note": upgrade_note_text,
        })

        # 2. 目录行
        rows.append({
            "type": "catalog",
            "internal_code": derived.catalog_internal_code,
            "external_code": derived.catalog_external_code,
            "title_cn": derived.catalog_title_cn,
            "title_en": derived.catalog_title_en,
            "revision": ctx.get_cover_catalog_revision(),
            "status": params.doc_status,
            "page_total": 0,  # 占位，后续回填
            "upgrade_note": upgrade_note_text,
        })

        # 3. 图纸行（按internal_code尾号升序）
        for frame in ctx.get_sorted_document_frames():
            tb = frame.titleblock
            seq_no = tb.get_seq_no()

            # 判断是否需要升版标记
            upgrade_note = ""
            if params.is_upgrade and seq_no is not None:
                sheet_code = f"{seq_no:03d}"
                if note_by_sheet_code:
                    upgrade_note = note_by_sheet_code.get(sheet_code, "")
                elif sheet_code in upgraded_sheet_codes:
                    upgrade_note = upgrade_note_text

            rows.append({
                "type": "drawing",
                "internal_code": tb.internal_code,
                "external_code": tb.external_code,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
                "revision": tb.revision,
                "status": tb.status,
                "page_total": ctx.get_page_total_for_frame(frame),
                "upgrade_note": upgrade_note,
            })

        return rows

    def _write_detail_row(
        self,
        ws,
        row: int,
        data: dict,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入单行明细"""
        columns = bindings.get("detail", {}).get("columns", {})

        # A: 序号
        ws[f"A{row}"] = row - bindings.get("detail", {}).get("start_row", 9) + 1

        # B: 图纸编号（internal_code）
        if "B" in columns:
            ws[f"B{row}"] = data.get("internal_code", "")

        # D: 文件编码（external_code）
        if "D" in columns:
            ws[f"D{row}"] = data.get("external_code", "")

        # E: 名称（1818需要中英文换行）
        if "E" in columns:
            title = data.get("title_cn", "")
            if data.get("type") == "catalog":
                title = build_catalog_display_title(ctx, self.spec)
            elif ctx.is_1818 and data.get("title_en"):
                title = f"{title}\n{data['title_en']}"
            ws[f"E{row}"] = title
            cell = ws[f"E{row}"]
            alignment = copy(cell.alignment)
            alignment.horizontal = "center"
            alignment.vertical = "center"
            alignment.wrapText = True
            cell.alignment = alignment

        # F: 版次
        if "F" in columns:
            ws[f"F{row}"] = data.get("revision", "")

        # G: 状态
        if "G" in columns:
            ws[f"G{row}"] = data.get("status", "")

        # H: 页数
        if "H" in columns:
            ws[f"H{row}"] = data.get("page_total", 1)

        # I: 附注（升版标记）
        if "I" in columns:
            ws[f"I{row}"] = data.get("upgrade_note", "")

    def _apply_detail_layout(self, ws, start_row: int, last_row: int) -> None:
        column_width = ws.column_dimensions["E"].width or 30
        for row in range(start_row, last_row + 1):
            text = str(ws[f"E{row}"].value or "")
            line_count = self._estimate_wrapped_line_count(text, column_width)
            ws.row_dimensions[row].height = self._bucket_row_height_for_line_count(line_count)

    def _repair_detail_grid_border_holes(self, ws, start_row: int, last_row: int) -> None:
        if last_row <= start_row:
            return

        for row in range(start_row + 1, last_row + 1):
            if row + 1 > ws.max_row:
                continue
            for column in range(1, 10):
                cell = ws.cell(row=row, column=column)
                if isinstance(cell, MergedCell):
                    continue
                if not self._border_is_empty(cell.border):
                    continue

                prev_cell = ws.cell(row=row - 1, column=column)
                next_cell = ws.cell(row=row + 1, column=column)
                if isinstance(prev_cell, MergedCell) or isinstance(next_cell, MergedCell):
                    continue
                if self._border_is_empty(prev_cell.border) or self._border_is_empty(next_cell.border):
                    continue
                if self._border_signature(prev_cell.border) != self._border_signature(next_cell.border):
                    continue
                cell.border = copy(prev_cell.border)

    @staticmethod
    def _border_is_empty(border) -> bool:
        return all(
            side.style is None
            for side in (border.left, border.right, border.top, border.bottom)
        )

    @staticmethod
    def _border_signature(border) -> tuple[str | None, str | None, str | None, str | None]:
        return (
            border.left.style,
            border.right.style,
            border.top.style,
            border.bottom.style,
        )

    def _refine_detail_layout_via_com(
        self,
        xlsx_path: Path,
        start_row: int,
        last_row: int,
    ) -> None:
        if not self._should_use_excel_com():
            return

        pythoncom = None
        try:
            import pythoncom  # type: ignore[import]
            import win32com.client
        except ImportError:
            return

        excel = None
        excel_owned = False
        workbook = None
        worksheet = None
        row_range = None
        temp_dir = None
        working_copy = xlsx_path
        should_copy_back = False
        with get_office_automation_limiter().excel_session():
            try:
                pythoncom.CoInitialize()
                excel, excel_owned = PDFExporter._create_excel_application(win32com)
                PDFExporter._prepare_excel_for_headless_run(excel)
                working_copy, temp_dir = PDFExporter._prepare_excel_path_for_com(
                    xlsx_path,
                    label=xlsx_path.stem,
                )
                workbook = PDFExporter._open_excel_workbook(excel, working_copy, read_only=False)
                workbook_com = cast(Any, workbook)
                worksheet = PDFExporter._retry_excel_com_call(
                    lambda: workbook_com.Worksheets(1),
                    "Workbook.Worksheets(1)",
                )
                worksheet_com = cast(Any, worksheet)
                row_range = PDFExporter._retry_excel_com_call(
                    lambda: worksheet_com.Rows(f"{start_row}:{last_row}"),
                    "Worksheet.Rows(range)",
                )
                row_range_com = cast(Any, row_range)
                PDFExporter._retry_excel_com_call(
                    lambda: row_range_com.AutoFit(),
                    "Rows.AutoFit",
                )
                row_range = None

                for row in range(start_row, last_row + 1):
                    current_row = row
                    row_ref = PDFExporter._retry_excel_com_call(
                        lambda current_row=current_row: worksheet_com.Rows(current_row),
                        f"Worksheet.Rows({row})",
                    )
                    row_ref_com = cast(Any, row_ref)
                    auto_height = float(
                        PDFExporter._retry_excel_com_call(
                            lambda row_ref_com=row_ref_com: row_ref_com.RowHeight or 0,
                            f"Rows({row}).RowHeight",
                        )
                    )
                    bucket_height = self._bucket_row_height_from_measured_height(auto_height)
                    if bucket_height:
                        PDFExporter._retry_excel_com_call(
                            lambda row_ref_com=row_ref_com, bucket_height=bucket_height: setattr(
                                row_ref_com,
                                "RowHeight",
                                bucket_height,
                            ),
                            f"Rows({row}).RowHeight=set",
                        )

                PDFExporter._retry_excel_com_call(
                    lambda: workbook_com.Save(),
                    "Workbook.Save",
                )
                should_copy_back = True
            except Exception:
                return
            finally:
                worksheet = None
                row_range = None
                if workbook:
                    with contextlib.suppress(Exception):
                        cast(Any, workbook).Close(False)
                workbook = None
                if excel and excel_owned:
                    with contextlib.suppress(Exception):
                        cast(Any, excel).Quit()
                excel = None
                if should_copy_back and temp_dir is not None:
                    with contextlib.suppress(Exception):
                        shutil.copy2(working_copy, xlsx_path)
                if temp_dir is not None:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                gc.collect()
                if pythoncom is not None:
                    with contextlib.suppress(Exception):
                        pythoncom.CoUninitialize()

    def _estimate_wrapped_line_count(self, text: str, column_width: float) -> int:
        normalized = text.replace("\r\n", "\n").replace("\r", "\n").strip("\n")
        if not normalized:
            return 1

        effective_width = max(8.0, float(column_width) * 0.9)
        total_lines = 0
        for raw_line in normalized.split("\n"):
            display_width = sum(self._char_display_width(ch) for ch in raw_line)
            wrapped_lines = max(1, math.ceil(display_width / effective_width))
            total_lines += wrapped_lines
        return max(1, total_lines)

    def _char_display_width(self, char: str) -> float:
        if not char:
            return 0
        if char.isspace():
            return 0.35
        if ord(char) > 127:
            return 1.0
        if char.isalnum():
            return 0.55
        return 0.65

    def _bucket_row_height_for_line_count(self, line_count: int) -> int:
        if line_count <= 2:
            return self.BODY_ROW_HEIGHT
        if line_count == 3:
            return self.THREE_LINE_HEIGHT
        if line_count == 4:
            return self.FOUR_LINE_HEIGHT
        return self.FOUR_LINE_HEIGHT + (line_count - 4) * self.EXTRA_LINE_STEP

    def _bucket_row_height_from_measured_height(self, measured_height: float) -> int:
        if measured_height <= 0:
            return self.BODY_ROW_HEIGHT
        if measured_height <= self.BODY_ROW_HEIGHT:
            return self.BODY_ROW_HEIGHT
        if measured_height <= self.THREE_LINE_HEIGHT:
            return self.THREE_LINE_HEIGHT
        if measured_height <= self.FOUR_LINE_HEIGHT:
            return self.FOUR_LINE_HEIGHT
        extra_steps = math.ceil((measured_height - self.FOUR_LINE_HEIGHT) / self.EXTRA_LINE_STEP)
        return self.FOUR_LINE_HEIGHT + max(1, extra_steps) * self.EXTRA_LINE_STEP

    def _count_pages(self, xlsx_path: Path) -> int:
        """计算目录页数"""
        # 优先尝试 Excel COM 的分页信息
        if self._should_use_excel_com():
            try:
                return self._count_pages_via_com(xlsx_path)
            except Exception:
                pass

        # 优先尝试Excel分页信息
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            if ws is None:
                return 1
            # 尝试通过分页符计算
            page_breaks = getattr(ws, "page_breaks", None)
            horizontal_breaks = getattr(page_breaks, "horizontalBreaks", None)
            h_breaks = len(horizontal_breaks) if horizontal_breaks is not None else 0
            if h_breaks > 0:
                return h_breaks + 1
        except Exception:
            pass

        # 兜底：导出PDF计页
        try:
            temp_pdf = xlsx_path.with_suffix(".temp.pdf")
            self.pdf_exporter.export_xlsx_to_pdf(xlsx_path, temp_pdf)
            count = self.pdf_exporter.count_pdf_pages(temp_pdf)
            temp_pdf.unlink(missing_ok=True)
            return count
        except Exception:
            return 1  # 默认1页

    def _count_pages_via_com(self, xlsx_path: Path) -> int:
        pythoncom = None
        try:
            import pythoncom  # type: ignore[import]
            import win32com.client
        except ImportError as exc:
            raise RuntimeError("pywin32 不可用") from exc

        excel = None
        excel_owned = False
        wb = None
        ws = None
        temp_dir = None
        working_copy = xlsx_path
        with get_office_automation_limiter().excel_session():
            try:
                pythoncom.CoInitialize()
                excel, excel_owned = PDFExporter._create_excel_application(win32com)
                PDFExporter._prepare_excel_for_headless_run(excel)
                working_copy, temp_dir = PDFExporter._prepare_excel_path_for_com(
                    xlsx_path,
                    label=xlsx_path.stem,
                )
                wb = PDFExporter._open_excel_workbook(excel, working_copy, read_only=True)
                workbook_com = cast(Any, wb)
                ws = PDFExporter._retry_excel_com_call(
                    lambda: workbook_com.Worksheets(1),
                    "Workbook.Worksheets(1)",
                )
                worksheet_com = cast(Any, ws)
                page_break_count = PDFExporter._retry_excel_com_call(
                    lambda: worksheet_com.HPageBreaks.Count,
                    "Worksheet.HPageBreaks.Count",
                )
                page_count = int(page_break_count) + 1
                return max(1, page_count)
            finally:
                ws = None
                if wb:
                    with contextlib.suppress(Exception):
                        cast(Any, wb).Close(False)
                wb = None
                if excel and excel_owned:
                    with contextlib.suppress(Exception):
                        cast(Any, excel).Quit()
                excel = None
                if temp_dir is not None:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                gc.collect()
                if pythoncom is not None:
                    with contextlib.suppress(Exception):
                        pythoncom.CoUninitialize()

    @staticmethod
    def _should_use_excel_com() -> bool:
        return "PYTEST_CURRENT_TEST" not in os.environ

    def _backfill_page_count(
        self,
        xlsx_path: Path,
        page_count: int,
        bindings: dict,
    ) -> None:
        """回填目录行页数"""
        wb = load_workbook(xlsx_path)
        ws = wb.active
        if ws is None:
            raise GenerationError("目录文件缺少活动工作表")

        # 目录行是第2行明细（封面后）
        start_row = bindings.get("detail", {}).get("start_row", 9)
        catalog_row = start_row + 1

        ws[f"H{catalog_row}"] = page_count

        wb.save(xlsx_path)
