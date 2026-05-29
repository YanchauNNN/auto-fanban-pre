"""
封面生成器 - Word文档生成

职责：
1. 打开封面模板（含内嵌Excel OLE）
2. 写入字段到指定单元格
3. 处理标题分割（中英文）
4. 导出PDF

依赖：
- python-docx: Word操作
- 参数规范.yaml: cover_bindings配置

测试要点：
- test_generate_cover_common: 通用封面生成
- test_generate_cover_1818: 1818封面生成（落点不同）
- test_title_split_cn: 中文标题分割
- test_title_split_en: 英文标题分割
- test_cover_revision_append: 版次追加模式
- test_external_code_19chars: 19位外部编码逐格写入
"""

from __future__ import annotations

import contextlib
import gc
import re
import shutil
import time
import zipfile
from collections.abc import Callable
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..config import load_spec
from ..config.spec_loader import CoverBinding
from ..interfaces import GenerationError, ICoverGenerator, IPDFExporter
from ..models import normalize_discipline_label
from .naming import make_document_output_name
from .office_automation import get_office_automation_limiter
from .pdf_engine import PDFExporter

if TYPE_CHECKING:
    from ..models import DocContext

_CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")
_CN_SPLIT_PROTECTED_PHRASES = ("标高", "厂房")
_XLSX_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_IGNORED_ERROR_FLAGS = {
    "evalError": "1",
    "twoDigitTextYear": "1",
    "numberStoredAsText": "1",
    "formula": "1",
    "formulaRange": "1",
    "unlockedFormula": "1",
    "emptyCellReference": "1",
    "listDataValidation": "1",
    "calculatedColumn": "1",
}
_IGNORED_ERRORS_INSERT_BEFORE = (
    "smartTags",
    "drawing",
    "legacyDrawing",
    "legacyDrawingHF",
    "picture",
    "oleObjects",
    "controls",
    "webPublishItems",
    "tableParts",
    "extLst",
)
_EXCEL_COM_ERROR_CHECKS = tuple(range(1, 10))
_EXCEL_ERROR_CHECKING_OPTIONS = (
    "BackgroundChecking",
    "EvaluateToError",
    "TextDate",
    "NumberAsText",
    "InconsistentFormula",
    "OmittedCells",
    "UnlockedFormulaCells",
    "EmptyCellReferences",
    "ListDataValidation",
    "InconsistentTableFormula",
)


class CoverGenerator(ICoverGenerator):
    """封面生成器实现"""

    def __init__(
        self,
        spec_path: str | None = None,
        pdf_exporter: IPDFExporter | None = None,
    ):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
        self.pdf_exporter = pdf_exporter or PDFExporter()

    def generate(self, ctx: DocContext, output_dir: Path) -> tuple[Path, Path]:
        """生成封面文档"""
        output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 选择模板
        template_path = self._get_template_path(ctx)
        if not Path(template_path).exists():
            raise GenerationError(f"封面模板不存在: {template_path}")

        # 2. 获取落点配置
        bindings = self.spec.get_cover_bindings(ctx.params.project_no)

        # 3. 准备写入数据
        data = self._prepare_data(ctx)

        # 4. 写入Word文档
        output_stem = self._build_output_stem(ctx)
        output_docx = output_dir / f"{output_stem}.docx"
        self._write_cover(template_path, output_docx, bindings, data, ctx)

        # 5. 导出PDF
        output_pdf = output_dir / f"{output_stem}.pdf"
        self.pdf_exporter.export_docx_to_pdf(output_docx, output_pdf)

        return output_docx, output_pdf

    def _build_output_stem(self, ctx: DocContext) -> str:
        return make_document_output_name(
            external_code=ctx.derived.cover_external_code,
            revision=ctx.get_cover_catalog_revision(),
            status=ctx.params.doc_status,
            internal_code=ctx.derived.cover_internal_code,
            fallback_name="封面",
        )

    def _get_template_path(self, ctx: DocContext) -> str:
        """获取模板路径"""
        variant = ""
        if ctx.params.cover_variant == "压力容器":
            variant = "压力容器" if ctx.is_1818 else "压力容器版"
        elif ctx.params.cover_variant == "核安全设备":
            variant = "核安全设备" if ctx.is_1818 else "核安全设备版"
        return self.spec.get_template_path(
            "cover",
            ctx.params.project_no,
            variant
        )

    def _prepare_data(self, ctx: DocContext) -> dict:
        """准备写入数据"""
        params = ctx.params
        derived = ctx.derived
        discipline = normalize_discipline_label(params.discipline, self.spec.get_mappings()) or (
            params.discipline or ""
        )

        return {
            "project_name": None if params.project_no == "1818" else self.spec.get_project_name(params.project_no),
            "engineering_no": params.engineering_no,
            "subitem_no": params.subitem_no,
            "subitem_name": params.subitem_name,
            "subitem_name_en": params.subitem_name_en,  # 仅1818
            "design_phase": derived.design_phase,
            "design_phase_en": derived.design_phase_en,  # 仅1818
            "discipline": discipline,
            "discipline_en": derived.discipline_en,  # 仅1818
            "album_title_cn": params.album_title_cn,
            "album_title_en": params.album_title_en,  # 仅1818
            "album_code": derived.album_code,
            "album_internal_code": derived.album_internal_code,
            "cover_revision": ctx.get_cover_catalog_revision(),
            "doc_status": params.doc_status,
            "cover_external_code": derived.cover_external_code,
        }

    def _write_cover(
        self,
        template_path: str,
        output_path: Path,
        bindings: dict,
        data: dict,
        ctx: DocContext,
    ) -> None:
        """写入封面文档"""
        shutil.copy(template_path, output_path)

        com_error: Exception | None = None
        try:
            self._write_cover_via_com(
                output_path=output_path,
                bindings=bindings,
                data=data,
            )
            suppress_cover_excel_error_indicators(output_path)
            return
        except Exception as exc:
            com_error = exc

        embedded_xlsx = self._find_embedded_xlsx(output_path)
        if embedded_xlsx:
            try:
                self._write_cover_via_embedded_xlsx(
                    output_path=output_path,
                    embedded_xlsx_path=embedded_xlsx,
                    bindings=bindings,
                    data=data,
                )
                return
            except Exception as embedded_exc:
                raise GenerationError(
                    f"封面写入失败: COM={com_error}; embedded_xlsx={embedded_exc}"
                ) from embedded_exc

        raise GenerationError(f"封面写入失败: {com_error}") from com_error

    def _write_cover_via_embedded_xlsx(
        self,
        *,
        output_path: Path,
        embedded_xlsx_path: str,
        bindings: dict[str, CoverBinding],
        data: dict[str, Any],
    ) -> None:
        with zipfile.ZipFile(output_path, "r") as zf:
            package = {name: zf.read(name) for name in zf.namelist()}

        workbook_bytes = package.get(embedded_xlsx_path)
        if workbook_bytes is None:
            raise GenerationError(f"嵌入工作簿不存在: {embedded_xlsx_path}")

        wb = load_workbook(BytesIO(workbook_bytes))
        ws = wb["封面"] if "封面" in wb.sheetnames else wb.active
        if ws is None:
            raise GenerationError("封面模板缺少活动工作表")

        def read_cell(cell: str) -> Any:
            return ws[cell].value

        def write_cell(cell: str, value: Any) -> None:
            if ":" in cell:
                start_cell, end_cell = self._split_range_ref(cell)
                # Embedded workbook fallback only needs the anchor cell for merged
                # or visually spanning title cells. Writing every cell breaks merged
                # ranges in openpyxl.
                ws[start_cell] = value
                return
            ws[cell] = value

        self._apply_bindings(bindings, data, read_cell, write_cell)

        buf = BytesIO()
        wb.save(buf)
        package[embedded_xlsx_path] = _suppress_excel_workbook_error_indicators(buf.getvalue())

        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for name, payload in package.items():
                zf.writestr(name, payload)

    def _write_cover_via_com(
        self,
        *,
        output_path: Path,
        bindings: dict[str, CoverBinding],
        data: dict[str, Any],
    ) -> None:
        pythoncom = None
        try:
            import pythoncom  # type: ignore[import]
            import win32com.client
        except ImportError as exc:
            raise GenerationError("缺少 pywin32，无法写入 OLE 封面模板") from exc

        word = None
        doc = None
        ws = None
        limiter = get_office_automation_limiter()
        with limiter.word_session():
            with limiter.excel_session():
                try:
                    pythoncom.CoInitialize()
                    word = win32com.client.DispatchEx("Word.Application")
                    word.Visible = False
                    word.DisplayAlerts = 0
                    doc = word.Documents.Open(str(output_path.absolute()))
                    time.sleep(1.0)
                    ws = self._get_embedded_excel_sheet(doc)
                    if ws is None:
                        raise GenerationError("未找到封面中的嵌入 Excel 对象")
                    worksheet = ws

                    def read_cell(cell: str) -> Any:
                        return self._com_call_with_retry(
                            lambda: worksheet.Range(cell).Value,
                            f"Range({cell}).Value",
                        )

                    def write_cell(cell: str, value: Any) -> None:
                        self._com_call_with_retry(
                            lambda: setattr(worksheet.Range(cell), "Value", value),
                            f"Range({cell}).Value={value}",
                        )

                    self._apply_bindings(bindings, data, read_cell, write_cell)
                    self._suppress_cover_error_indicators_via_com(worksheet)
                    self._com_call_with_retry(doc.Save, "Document.Save")
                finally:
                    ws = None
                    if doc is not None:
                        doc_obj = doc
                        self._mark_document_saved(doc)
                        self._close_com_object(lambda: doc_obj.Close(False), "Document.Close")
                    doc = None
                    if word is not None:
                        self._close_all_word_documents(word, keep=doc)
                        self._mark_normal_template_saved(word)
                        self._close_com_object(word.Quit, "Word.Quit")
                    word = None
                    gc.collect()
                    if pythoncom is not None:
                        with contextlib.suppress(Exception):
                            pythoncom.CoUninitialize()

    def _get_embedded_excel_sheet(self, doc: Any) -> Any | None:
        for collection_name in ("InlineShapes", "Shapes"):
            collection = getattr(doc, collection_name, None)
            if collection is None:
                continue
            collection_obj = collection
            try:
                count = int(
                    self._com_call_with_retry(
                        lambda collection_obj=collection_obj: collection_obj.Count,
                        f"{collection_name}.Count",
                    )
                )
            except Exception:
                continue

            for idx in range(1, count + 1):
                try:
                    shape = self._com_call_with_retry(
                        lambda collection_obj=collection_obj, idx=idx: collection_obj.Item(idx),
                        f"{collection_name}.Item({idx})",
                    )
                    ole_format = self._com_call_with_retry(
                        lambda shape=shape: shape.OLEFormat,
                        f"{collection_name}.Item({idx}).OLEFormat",
                    )
                    self._com_call_with_retry(
                        ole_format.Activate,
                        f"{collection_name}.Item({idx}).OLEFormat.Activate",
                    )
                    time.sleep(0.8)
                    ole_obj = self._com_call_with_retry(
                        lambda ole_format=ole_format: ole_format.Object,
                        f"{collection_name}.Item({idx}).OLEFormat.Object",
                    )
                except Exception:
                    continue

                sheet = self._to_excel_sheet(ole_obj)
                if sheet is not None:
                    return sheet
        return None

    def _to_excel_sheet(self, ole_obj: Any) -> Any | None:
        if ole_obj is None:
            return None

        try:
            parent = self._com_call_with_retry(
                lambda: getattr(ole_obj, "Parent", None),
                "OLEObject.Parent",
            )
            if parent is not None and hasattr(parent, "Worksheets"):
                return self._com_call_with_retry(
                    lambda: parent.Worksheets(1),
                    "OLEObject.Parent.Worksheets(1)",
                )
        except Exception:
            pass

        try:
            if hasattr(ole_obj, "Worksheets"):
                return self._com_call_with_retry(
                    lambda: ole_obj.Worksheets(1),
                    "OLEObject.Worksheets(1)",
                )
        except Exception:
            pass

        if hasattr(ole_obj, "Range"):
            return ole_obj

        return None

    def _suppress_cover_error_indicators_via_com(self, worksheet: Any) -> None:
        with contextlib.suppress(Exception):
            excel_app = self._com_call_with_retry(
                lambda: worksheet.Application,
                "Worksheet.Application",
                retries=3,
            )
            error_options = self._com_call_with_retry(
                lambda: excel_app.ErrorCheckingOptions,
                "Excel.ErrorCheckingOptions",
                retries=3,
            )
            for option_name in _EXCEL_ERROR_CHECKING_OPTIONS:
                with contextlib.suppress(Exception):
                    self._com_call_with_retry(
                        lambda option_name=option_name: setattr(
                            error_options,
                            option_name,
                            False,
                        ),
                        f"Excel.ErrorCheckingOptions.{option_name}=False",
                        retries=3,
                    )

        with contextlib.suppress(Exception):
            used_range = self._com_call_with_retry(
                lambda: worksheet.UsedRange,
                "Worksheet.UsedRange",
                retries=3,
            )
            for error_check_index in _EXCEL_COM_ERROR_CHECKS:
                with contextlib.suppress(Exception):
                    self._com_call_with_retry(
                        lambda used_range=used_range, error_check_index=error_check_index: setattr(
                            used_range.Errors(error_check_index),
                            "Ignore",
                            True,
                        ),
                        f"UsedRange.Errors({error_check_index}).Ignore=True",
                        retries=3,
                    )
        with contextlib.suppress(Exception):
            workbook = self._com_call_with_retry(
                lambda: worksheet.Parent,
                "Worksheet.Parent",
                retries=3,
            )
            self._com_call_with_retry(workbook.Save, "Workbook.Save", retries=3)

    def _find_embedded_xlsx(self, docx_path: Path) -> str | None:
        with zipfile.ZipFile(docx_path, "r") as zf:
            for name in zf.namelist():
                if name.startswith("word/embeddings/") and name.lower().endswith(".xlsx"):
                    return name
        return None

    def _apply_bindings(
        self,
        bindings: dict[str, CoverBinding],
        data: dict[str, Any],
        read_cell: Callable[[str], Any],
        write_cell: Callable[[str, Any], None],
    ) -> None:
        for key, binding in bindings.items():
            if key == "册次":
                continue

            cell_ref = binding.cell
            value = data.get(key)

            if key == "cover_external_code":
                self._write_external_code_chars(
                    cell_ref,
                    value if isinstance(value, str) else "",
                    write_cell,
                )
                continue

            if binding.split_rule and "+" in cell_ref:
                left_cell, right_cell = self._split_two_cells_ref(cell_ref)
                left, right = self._split_text_by_rule(str(value or ""), binding.split_rule)
                write_cell(left_cell, left)
                write_cell(right_cell, right)
                continue

            target_cell = cell_ref.strip() if ":" in cell_ref else self._first_cell(cell_ref)
            if binding.write_mode == "append_after_label":
                if self._is_empty(value):
                    continue
                current = read_cell(self._first_cell(cell_ref))
                merged = self._append_after_label(
                    current=str(current or ""),
                    label=binding.label or "",
                    value=str(value),
                )
                write_cell(target_cell, merged)
                continue

            if not self._is_empty(value):
                write_cell(target_cell, value)

    @staticmethod
    def _is_empty(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip() == ""
        return False

    def _split_text_by_rule(self, text: str, split_rule: str) -> tuple[str, str]:
        if split_rule.startswith("cn_split"):
            return self._split_cn_two_cells(text)
        if split_rule.startswith("en_split"):
            return self._split_en_two_cells(text)
        return text, ""

    def _split_cn_two_cells(self, text: str) -> tuple[str, str]:
        s = text.strip()
        if not s:
            return "", ""
        mid = len(s) // 2
        candidates = [
            idx
            for idx in range(1, len(s))
            if self._is_cjk(s[idx]) and not self._crosses_cn_protected_phrase(s, idx)
        ]
        if not candidates:
            return s, ""
        idx = min(candidates, key=lambda n: (self._cn_split_boundary_penalty(s, n), abs(n - mid)))
        return s[:idx].rstrip(), s[idx:].lstrip()

    def _split_en_two_cells(self, text: str) -> tuple[str, str]:
        s = self._normalize_cover_english_spacing(text)
        if not s:
            return "", ""
        mid = len(s) // 2
        candidates = [m.start() for m in re.finditer(r"\s+", s)]
        if not candidates:
            return s, ""

        def score(i: int) -> tuple[int, int]:
            right = s[i:].lstrip()
            right_ok = 0 if (right and right[0].isalpha()) else 1
            return right_ok, abs(i - mid)

        idx = min(candidates, key=score)
        return s[:idx].rstrip(), s[idx:].lstrip()

    @staticmethod
    def _is_cjk(ch: str) -> bool:
        if not ch:
            return False
        code = ord(ch)
        return 0x4E00 <= code <= 0x9FFF

    @staticmethod
    def _crosses_cn_protected_phrase(text: str, split_index: int) -> bool:
        for phrase in _CN_SPLIT_PROTECTED_PHRASES:
            start = text.find(phrase)
            while start >= 0:
                end = start + len(phrase)
                if start < split_index < end:
                    return True
                start = text.find(phrase, start + 1)
        return False

    def _cn_split_boundary_penalty(self, text: str, split_index: int) -> int:
        if split_index <= 0:
            return 1
        return 0 if self._is_cjk(text[split_index - 1]) else 1

    @staticmethod
    def _normalize_cover_english_spacing(text: str) -> str:
        s = re.sub(r"\s+", " ", text.strip())
        s = re.sub(r"\b(Level|Elevation)(?=[+-]?\d)", r"\1 ", s, flags=re.IGNORECASE)
        s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def _write_external_code_chars(
        self,
        range_ref: str,
        code: str,
        write_cell: Callable[[str, Any], None],
    ) -> None:
        start_ref, end_ref = self._split_range_ref(range_ref)
        start_col, start_row = self._parse_cell_ref(start_ref)
        end_col, end_row = self._parse_cell_ref(end_ref)
        if start_row != end_row:
            raise GenerationError(f"外部编码落点必须是单行范围: {range_ref}")

        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        if end_idx < start_idx:
            start_idx, end_idx = end_idx, start_idx

        chars = list((code or "")[:19].ljust(19))
        for i, col_idx in enumerate(range(start_idx, end_idx + 1)):
            char = chars[i] if i < len(chars) else ""
            write_cell(f"{get_column_letter(col_idx)}{start_row}", char)

    def _append_after_label(self, current: str, label: str, value: str) -> str:
        if not current:
            return f"{label}{value}" if label else value
        if label and label in current:
            prefix, _, _ = current.partition(label)
            return f"{prefix}{label}{value}"
        if current.endswith(value):
            return current
        return f"{current}{value}"

    def _first_cell(self, cell_ref: str) -> str:
        if "+" in cell_ref:
            return cell_ref.split("+", 1)[0].strip()
        if ":" in cell_ref:
            return cell_ref.split(":", 1)[0].strip()
        return cell_ref.strip()

    def _split_two_cells_ref(self, cell_ref: str) -> tuple[str, str]:
        left, right = cell_ref.split("+", 1)
        return left.strip(), right.strip()

    def _split_range_ref(self, ref: str) -> tuple[str, str]:
        if ":" not in ref:
            return ref.strip(), ref.strip()
        left, right = ref.split(":", 1)
        return left.strip(), right.strip()

    def _parse_cell_ref(self, ref: str) -> tuple[str, int]:
        m = _CELL_RE.match(ref)
        if m is None:
            raise GenerationError(f"非法单元格引用: {ref}")
        return m.group(1).upper(), int(m.group(2))

    def _com_call_with_retry(
        self,
        fn: Callable[[], Any],
        desc: str,
        *,
        retries: int = 10,
    ) -> Any:
        last_exc: Exception | None = None
        for _ in range(retries):
            try:
                return fn()
            except Exception as exc:
                last_exc = exc
                time.sleep(0.8 if self._is_call_rejected(exc) else 0.3)
        raise RuntimeError(f"COM 调用失败 {desc}: {last_exc}") from last_exc

    def _close_com_object(self, fn: Callable[[], Any], desc: str) -> None:
        with contextlib.suppress(Exception):
            self._com_call_with_retry(fn, desc, retries=6)

    def _mark_document_saved(self, doc: Any) -> None:
        with contextlib.suppress(Exception):
            self._com_call_with_retry(
                lambda: setattr(doc, "Saved", True),
                "Document.Saved=True",
                retries=3,
            )

    def _mark_normal_template_saved(self, word: Any) -> None:
        with contextlib.suppress(Exception):
            template = self._com_call_with_retry(
                lambda: getattr(word, "NormalTemplate", None),
                "Word.NormalTemplate",
                retries=3,
            )
            if template is not None:
                self._com_call_with_retry(
                    lambda: setattr(template, "Saved", True),
                    "Word.NormalTemplate.Saved=True",
                    retries=3,
                )

    def _close_all_word_documents(self, word: Any, *, keep: Any | None = None) -> None:
        try:
            documents = self._com_call_with_retry(
                lambda: getattr(word, "Documents", None),
                "Word.Documents",
                retries=3,
            )
            if documents is None:
                return

            count = int(
                self._com_call_with_retry(
                    lambda: documents.Count,
                    "Word.Documents.Count",
                    retries=3,
                )
            )
        except Exception:
            return

        for index in range(count, 0, -1):
            try:
                current = self._com_call_with_retry(
                    lambda index=index: documents.Item(index),
                    f"Word.Documents.Item({index})",
                    retries=3,
                )
            except Exception:
                continue

            if keep is not None and current is keep:
                continue

            self._mark_document_saved(current)
            self._close_com_object(
                lambda current=current: current.Close(False),
                f"Word.Documents.Item({index}).Close",
            )

    @staticmethod
    def _is_call_rejected(exc: Exception) -> bool:
        if getattr(exc, "hresult", None) == -2147418111:
            return True
        msg = str(exc).lower()
        return "call was rejected by callee" in msg or "拒绝接收呼叫" in msg


def suppress_cover_excel_error_indicators(docx_path: Path) -> bool:
    """Persist ignored Excel error indicators in embedded cover workbooks."""
    docx_path = Path(docx_path)
    with zipfile.ZipFile(docx_path, "r") as zf:
        entries = [(info, zf.read(info.filename)) for info in zf.infolist()]

    changed = False
    updated_entries: list[tuple[zipfile.ZipInfo, bytes]] = []
    for info, payload in entries:
        if info.filename.startswith("word/embeddings/") and info.filename.lower().endswith(".xlsx"):
            updated = _suppress_excel_workbook_error_indicators(payload)
            changed = changed or updated != payload
            payload = updated
        updated_entries.append((info, payload))

    if not changed:
        return False

    with zipfile.ZipFile(docx_path, "w") as zf:
        for info, payload in updated_entries:
            new_info = zipfile.ZipInfo(info.filename, info.date_time)
            new_info.compress_type = info.compress_type
            new_info.comment = info.comment
            new_info.extra = info.extra
            new_info.internal_attr = info.internal_attr
            new_info.external_attr = info.external_attr
            zf.writestr(new_info, payload)
    return True


def _suppress_excel_workbook_error_indicators(workbook_bytes: bytes) -> bytes:
    with zipfile.ZipFile(BytesIO(workbook_bytes), "r") as zf:
        entries = [(info, zf.read(info.filename)) for info in zf.infolist()]

    changed = False
    updated_entries: list[tuple[zipfile.ZipInfo, bytes]] = []
    for info, payload in entries:
        if (
            info.filename.startswith("xl/worksheets/sheet")
            and info.filename.endswith(".xml")
        ):
            updated = _suppress_sheet_error_indicators(payload)
            changed = changed or updated != payload
            payload = updated
        updated_entries.append((info, payload))

    if not changed:
        return workbook_bytes

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for info, payload in updated_entries:
            new_info = zipfile.ZipInfo(info.filename, info.date_time)
            new_info.compress_type = info.compress_type
            new_info.comment = info.comment
            new_info.extra = info.extra
            new_info.internal_attr = info.internal_attr
            new_info.external_attr = info.external_attr
            zf.writestr(new_info, payload)
    return buf.getvalue()


def _suppress_sheet_error_indicators(sheet_xml: bytes) -> bytes:
    ET.register_namespace("", _XLSX_SHEET_NS)
    root = ET.fromstring(sheet_xml)
    dimension = root.find(_xlsx_tag("dimension"))
    sqref = dimension.get("ref") if dimension is not None else None
    if not sqref:
        return sheet_xml

    ignored_errors = root.find(_xlsx_tag("ignoredErrors"))
    changed = False
    if ignored_errors is None:
        ignored_errors = ET.Element(_xlsx_tag("ignoredErrors"))
        root.insert(_ignored_errors_insert_index(root), ignored_errors)
        changed = True

    ignored_error = None
    for candidate in ignored_errors.findall(_xlsx_tag("ignoredError")):
        if candidate.get("sqref") == sqref:
            ignored_error = candidate
            break

    if ignored_error is None:
        ignored_error = ET.SubElement(
            ignored_errors,
            _xlsx_tag("ignoredError"),
            {"sqref": sqref},
        )
        changed = True

    for key, value in _IGNORED_ERROR_FLAGS.items():
        if ignored_error.get(key) != value:
            ignored_error.set(key, value)
            changed = True

    if not changed:
        return sheet_xml
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _ignored_errors_insert_index(root: ET.Element) -> int:
    for index, child in enumerate(list(root)):
        if _local_name(child.tag) in _IGNORED_ERRORS_INSERT_BEFORE:
            return index
    return len(root)


def _xlsx_tag(local_name: str) -> str:
    return f"{{{_XLSX_SHEET_NS}}}{local_name}"


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag
