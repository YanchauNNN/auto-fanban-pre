"""
IED计划生成器 - Excel文档生成（仅Excel，不导出PDF）

职责：
1. 打开IED计划模板
2. 写入所有行（封面+目录+图纸）
3. 单独输出Excel（不入package.zip）

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: ied_bindings配置

测试要点：
- test_generate_ied: IED计划生成
- test_ied_columns: 列映射正确性
- test_ied_fixed_values: 固定值列
- test_ied_no_pdf: 不导出PDF
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ..config import load_spec
from ..interfaces import GenerationError, IIEDGenerator

if TYPE_CHECKING:
    from ..models import DocContext


class IEDGenerator(IIEDGenerator):
    """IED计划生成器实现"""

    def __init__(self, spec_path: str | None = None):
        self.spec = load_spec(spec_path) if spec_path else load_spec()

    def generate(self, ctx: DocContext, output_dir: Path) -> Path:
        """生成IED计划（仅Excel）"""
        output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 获取模板路径
        template_path = self.spec.get_template_path("ied", ctx.params.project_no)
        if not Path(template_path).exists():
            raise GenerationError(f"IED计划模板不存在: {template_path}")

        # 2. 获取落点配置
        bindings = self.spec.get_ied_bindings()

        # 3. 写入Excel
        output_xlsx = output_dir / "IED计划.xlsx"
        self._write_ied(template_path, output_xlsx, bindings, ctx)

        # 注意：IED不导出PDF
        return output_xlsx

    def _write_ied(
        self,
        template_path: str,
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入IED计划Excel"""
        wb = load_workbook(template_path)

        # 使用指定的sheet
        sheet_name = bindings.get("sheet", "IED导入模板 (修改)")
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        start_row = bindings.get("start_row", 2)
        columns = bindings.get("columns", {})

        # 准备全局数据
        global_data = self._prepare_global_data(ctx)

        # 行顺序：封面 → 目录 → 图纸
        rows = self._build_rows(ctx)

        current_row = start_row
        for row_data in rows:
            self._write_row(ws, current_row, row_data, global_data, columns, ctx)
            current_row += 1

        wb.save(output_path)

    def _prepare_global_data(self, ctx: DocContext) -> dict:
        """准备全局数据"""
        params = ctx.params
        derived = ctx.derived
        discipline_office = self._normalize_discipline_office(params.ied_discipline_office)

        return {
            "ied_change_flag": params.ied_change_flag,
            "ied_doc_type": params.ied_doc_type,
            "ied_status": params.ied_status,
            "wbs_code": params.wbs_code,
            "album_internal_code": derived.album_internal_code,
            "ied_design_type": params.ied_design_type,
            "ied_responsible_unit": params.ied_responsible_unit,
            "ied_discipline_office": discipline_office,
            "ied_chief_designer": params.ied_chief_designer,
            "ied_person_qual_category": params.ied_person_qual_category,
            "ied_fu_flag": params.ied_fu_flag,
            "ied_internal_tag": params.ied_internal_tag,
            "ied_prepared_by": params.ied_prepared_by,
            "ied_prepared_by_2": params.ied_prepared_by_2,
            "ied_prepared_date": params.ied_prepared_date,
            "ied_checked_by": params.ied_checked_by,
            "ied_checked_date": params.ied_checked_date,
            "ied_discipline_leader": params.ied_discipline_leader,
            "ied_discipline_leader_date": params.ied_discipline_leader_date,
            "ied_reviewed_by": params.ied_reviewed_by,
            "ied_reviewed_date": params.ied_reviewed_date,
            "ied_approved_by": params.ied_approved_by,
            "ied_approved_date": params.ied_approved_date,
            "ied_submitted_plan_date": params.ied_submitted_plan_date,
            "ied_publish_plan_date": params.ied_publish_plan_date,
            "ied_external_plan_date": params.ied_external_plan_date,
            "ied_fu_plan_date": params.ied_fu_plan_date,
            "classification": params.classification,
            "work_hours": params.work_hours,
        }

    def _build_rows(self, ctx: DocContext) -> list[dict]:
        """构建行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params

        # 封面行
        rows.append({
            "type": "cover",
            "external_code": derived.cover_external_code,
            "internal_code": derived.cover_internal_code,
            "revision": params.cover_revision,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
        })

        # 目录行
        rows.append({
            "type": "catalog",
            "external_code": derived.catalog_external_code,
            "internal_code": derived.catalog_internal_code,
            "revision": derived.catalog_revision,
            "title_cn": derived.catalog_title_cn,
            "title_en": derived.catalog_title_en,
        })

        # 图纸行
        for frame in ctx.get_sorted_document_frames():
            tb = frame.titleblock
            rows.append({
                "type": "drawing",
                "external_code": tb.external_code,
                "internal_code": tb.internal_code,
                "revision": tb.revision,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
            })

        return rows

    def _write_row(
        self,
        ws,
        row: int,
        row_data: dict,
        global_data: dict,
        columns: dict,
        ctx: DocContext,
    ) -> None:
        """写入单行"""
        for col_letter, col_config in columns.items():
            # 固定值
            if "value" in col_config:
                ws[f"{col_letter}{row}"] = col_config["value"]
                continue

            source = col_config.get("source", "")
            is_global = col_config.get("global", False)

            value = self._resolve_value(
                source=source,
                is_global=is_global,
                row_data=row_data,
                global_data=global_data,
                ctx=ctx,
            )

            ws[f"{col_letter}{row}"] = value

    def _resolve_value(
        self,
        *,
        source: str,
        is_global: bool,
        row_data: dict,
        global_data: dict,
        ctx: DocContext,
    ) -> str:
        if is_global:
            return global_data.get(source, "") or ""

        if source in row_data:
            value = row_data.get(source, "")
            if source == "title_en" and not ctx.is_1818:
                return ""
            return value or ""

        if source in global_data:
            return global_data.get(source, "") or ""

        return ""

    def _normalize_discipline_office(self, office: str | None) -> str:
        if office is None:
            return ""
        text = office.strip()
        if text == "":
            return ""
        if "-" in text:
            return text.rsplit("-", 1)[-1].strip()
        return text
