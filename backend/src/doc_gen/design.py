"""
设计文件生成器 - Excel文档生成

职责：
1. 打开设计文件模板
2. 写入所有行（封面+目录+图纸）
3. 仅输出Excel（不导出PDF）

依赖：
- openpyxl: Excel操作
- 参数规范.yaml: design_bindings配置

测试要点：
- test_generate_design_file: 设计文件生成
- test_design_global_fields: 全局字段写入
- test_design_frame_fields: 图纸行字段
- test_design_cover_catalog_rows: 封面/目录行特化
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from ..cad.titleblock_consistency import TitleblockConsistencyService
from ..config import load_spec
from ..interfaces import GenerationError, IDesignFileGenerator
from ..models import normalize_discipline_label
from .catalog_display_title import build_catalog_single_line_titles, flatten_title_text

if TYPE_CHECKING:
    from ..models import DocContext


class DesignFileGenerator(IDesignFileGenerator):
    """设计文件生成器实现"""

    def __init__(
        self,
        spec_path: str | None = None,
        pdf_exporter=None,
    ):
        self.spec = load_spec(spec_path) if spec_path else load_spec()
        self.consistency = TitleblockConsistencyService()
        # 保留参数以兼容历史构造调用；设计文件已不再导出PDF。
        self.pdf_exporter = pdf_exporter

    def generate(self, ctx: DocContext, output_dir: Path) -> Path:
        """生成设计文件"""
        output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 获取模板路径
        template_path = self.spec.get_template_path("design", ctx.params.project_no)
        if not Path(template_path).exists():
            raise GenerationError(f"设计文件模板不存在: {template_path}")

        # 2. 获取落点配置
        bindings = self.spec.get_design_bindings()

        # 3. 写入Excel
        output_xlsx = output_dir / "设计文件.xlsx"
        self._write_design(template_path, output_xlsx, bindings, ctx)

        return output_xlsx

    def _write_design(
        self,
        template_path: str,
        output_path: Path,
        bindings: dict,
        ctx: DocContext,
    ) -> None:
        """写入设计文件Excel"""
        wb = load_workbook(template_path)
        ws = wb.active
        template_lookups = self._load_template_lookups(wb)

        start_row = bindings.get("start_row", 2)
        columns = bindings.get("columns", {})

        # 准备全局数据
        global_data = self._prepare_global_data(ctx)

        # 行顺序：封面 → 目录 → 图纸
        rows = self._build_rows(ctx)

        current_row = start_row
        for row_data in rows:
            self._write_row(
                ws,
                current_row,
                row_data,
                global_data,
                columns,
                template_lookups,
            )
            current_row += 1

        wb.save(output_path)

    def _prepare_global_data(self, ctx: DocContext) -> dict:
        """准备全局数据（所有行相同）"""
        params = ctx.params
        derived = ctx.derived
        mappings = self.spec.get_mappings()
        discipline = normalize_discipline_label(params.discipline, mappings) or ""

        # 专业代码映射
        discipline_code = mappings.get("discipline_to_code", {}).get(
            discipline, ""
        )

        return {
            "design_status": params.design_status,
            "wbs_code": params.wbs_code,
            "album_internal_code": derived.album_internal_code,
            "internal_tag": params.internal_tag,
            "subitem_name": params.subitem_name,
            "subitem_no": params.subitem_no,
            "system_code": params.system_code,
            "system_name": params.system_name,
            "discipline": discipline,
            "discipline_code": discipline_code,
            "discipline_office": params.discipline_office,
            "design_phase": derived.design_phase,
            "classification": params.classification,
            "file_category": params.file_category,
            "attachment_name": params.attachment_name,
            "qa_required": params.qa_required,
            "qa_engineer": params.qa_engineer,
            "work_hours": params.work_hours,
        }

    def _build_rows(self, ctx: DocContext) -> list[dict]:
        """构建行数据"""
        rows = []
        derived = ctx.derived
        params = ctx.params
        discipline = normalize_discipline_label(params.discipline, self.spec.get_mappings()) or ""
        catalog_title_cn, catalog_title_en = build_catalog_single_line_titles(ctx, self.spec)

        # 封面行
        rows.append({
            "type": "cover",
            "external_code": derived.cover_external_code,
            "internal_code": derived.cover_internal_code,
            "revision": params.cover_revision,
            "title_cn": derived.cover_title_cn,
            "title_en": derived.cover_title_en,
            "paper_size_text": self.consistency.cover_paper_text(),
            "page_total": 1,
            "status": params.doc_status,
            "discipline": discipline,
            "design_phase": derived.design_phase,
        })

        # 目录行
        rows.append({
            "type": "catalog",
            "external_code": derived.catalog_external_code,
            "internal_code": derived.catalog_internal_code,
            "revision": derived.catalog_revision,
            "title_cn": catalog_title_cn,
            "title_en": catalog_title_en,
            "paper_size_text": self.consistency.catalog_paper_text(),
            "page_total": derived.catalog_page_total or 1,
            "status": params.doc_status,
            "discipline": discipline,
            "design_phase": derived.design_phase,
        })

        # 图纸行
        for frame in ctx.get_sorted_document_frames():
            tb = frame.titleblock
            rows.append({
                "type": "drawing",
                "external_code": tb.external_code,
                "internal_code": tb.internal_code,
                "revision": tb.revision,
                "title_cn": tb.title_cn,
                "title_en": tb.title_en,
                "paper_size_text": self.consistency.drawing_paper_text(frame),
                "page_total": ctx.get_page_total_for_frame(frame),
                "status": tb.status,
                "discipline": normalize_discipline_label(
                    tb.discipline or params.discipline,
                    self.spec.get_mappings(),
                ) or "",
                "design_phase": derived.design_phase,
            })

        return rows

    def _write_row(
        self,
        ws,
        row: int,
        row_data: dict,
        global_data: dict,
        columns: dict,
        template_lookups: dict[str, list[str]],
    ) -> None:
        """写入单行"""
        # 遍历列配置写入
        for col_letter, col_config in columns.items():
            source = col_config.get("source", "")
            is_global = col_config.get("global", False)
            normalize = str(col_config.get("normalize", "") or "").strip()

            value = self._resolve_value(
                source=source,
                is_global=is_global,
                row_data=row_data,
                global_data=global_data,
                template_lookups=template_lookups,
            )

            if normalize == "cjk_only":
                value = self._keep_cjk_text(value)

            # 写入
            ws[f"{col_letter}{row}"] = value

    def _resolve_value(
        self,
        *,
        source: str,
        is_global: bool,
        row_data: dict,
        global_data: dict,
        template_lookups: dict[str, list[str]],
    ) -> str:
        if source == "discipline_code_map[discipline]":
            return global_data.get("discipline_code", "") or ""

        if is_global:
            value = global_data.get(source, "") or ""
        elif source in row_data:
            value = row_data.get(source, "") or ""
        elif source in global_data:
            value = global_data.get(source, "") or ""
        else:
            value = ""

        if source == "paper_size_text":
            return self._match_template_value(
                value,
                template_lookups.get("paper_sizes", []),
            )

        if source == "discipline":
            return self._match_template_value(
                value,
                template_lookups.get("disciplines", []),
            )

        if source in {"title_cn", "title_en"}:
            return flatten_title_text(value)

        return value

    def _load_template_lookups(self, workbook) -> dict[str, list[str]]:
        if len(workbook.worksheets) < 2:
            return {"paper_sizes": [], "disciplines": []}

        sheet = workbook.worksheets[1]
        paper_sizes = self._read_lookup_column(sheet, "B", 53, 79)
        disciplines = self._read_lookup_column(sheet, "E", 53, 101)
        return {
            "paper_sizes": paper_sizes,
            "disciplines": disciplines,
        }

    def _read_lookup_column(
        self,
        sheet,
        column: str,
        start_row: int,
        end_row: int,
    ) -> list[str]:
        values: list[str] = []
        for row in range(start_row, end_row + 1):
            value = sheet[f"{column}{row}"].value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                values.append(text)
        return values

    def _match_template_value(
        self,
        raw_value: Any,
        options: list[str],
    ) -> str:
        text = str(raw_value or "").strip()
        if not text:
            return ""

        normalized = self._normalize_lookup_value(text)

        for option in options:
            if self._normalize_lookup_value(option) == normalized:
                return option

        return text

    def _normalize_lookup_value(self, value: str) -> str:
        normalized = re.sub(r"\s+", "", str(value or "")).strip().lower()
        return normalized

    def _keep_cjk_text(self, value: Any) -> str:
        text = str(value or "")
        if not text:
            return ""
        text = re.sub(
            r"[^\u3400-\u4dbf\u4e00-\u9fff\s]+",
            "",
            text,
        )
        text = re.sub(r"\s+", " ", text).strip()
        return text
