from __future__ import annotations

import math
import re
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Literal, Protocol, cast
from xml.etree.ElementTree import ParseError, iterparse
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]
from openpyxl.worksheet.formula import (  # type: ignore[import-untyped]
    ArrayFormula,
    DataTableFormula,
)

from .reinforcement_input import (
    InvalidReinforcementWorkbook,
    _find_columns,
    _find_slab_columns,
    _header_text,
    _normalized_cell_text,
    _uses_standard_layout,
    normalize_slab_elevation,
    normalize_wall_id,
)


@dataclass(frozen=True)
class FormatReason:
    scope: Literal["wall", "slab"]
    code: str
    sheet: str | None
    message: str


@dataclass(frozen=True)
class WorkbookFormatInspection:
    requires_ai_normalization: bool
    reasons: tuple[FormatReason, ...]
    wall_sheet: str | None
    slab_sheet: str | None


class _CellLike(Protocol):
    row: int
    column: int
    coordinate: str
    value: object
    data_type: str


class _MergedRangeLike(Protocol):
    min_row: int
    min_col: int

    def __str__(self) -> str: ...


class _MergedCellsLike(Protocol):
    ranges: set[_MergedRangeLike]


class _WorksheetLike(Protocol):
    title: str
    max_row: int
    max_column: int
    merged_cells: _MergedCellsLike

    def cell(self, *, row: int, column: int) -> _CellLike: ...


class _SnapshotWorksheetLike(_WorksheetLike, Protocol):
    # openpyxl's supported Worksheet API exposes only rectangular iteration.
    # This private mapping is the authoritative sparse store populated from XML.
    _cells: dict[tuple[int, int], _CellLike]


class _WorkbookLike(Protocol):
    worksheets: Sequence[_WorksheetLike]

    def close(self) -> None: ...


type _WallColumns = tuple[int, int, int, int, int]
type _WallCandidate = tuple[_WorksheetLike, _WallColumns]
type _SlabColumns = tuple[int, dict[str, int]]
type _SlabCandidate = tuple[
    _WorksheetLike,
    _SlabColumns | None,
    int | None,
]


_STANDARD_LINEAR_SPEC = re.compile(
    r"(?:"
    r"(?P<layers_space>\d+)\s+(?P<diameter_space>\d+)"
    r"|(?P<layers_marked>\d+)\s*D\s*(?P<diameter_marked>\d+)"
    r"|D\s*(?P<diameter_only_marked>\d+)"
    r"|(?P<diameter_only>\d+)"
    r")\s*(?:@|间距)\s*(?P<spacing_primary>\d+)",
    re.IGNORECASE,
)
_STANDARD_GRID_SPEC = re.compile(
    r"(?:"
    r"(?P<layers_space>\d+)\s+(?P<diameter_space>\d+)"
    r"|(?P<layers_marked>\d+)\s*C\s*(?P<diameter_marked>\d+)"
    r"|C\s*(?P<diameter_only_marked>\d+)"
    r"|(?P<diameter_only>\d+)"
    r")\s*(?:@|间距)\s*(?P<spacing_primary>\d+)"
    r"\s*[*xX]\s*(?P<spacing_secondary>\d+)",
    re.IGNORECASE,
)
_STANDARD_WALL_TEXT = re.compile(
    r"[A-Za-z]+\d+[A-Za-z]?(?:-\d+)?\s*墙?",
    re.IGNORECASE,
)


def _is_standard_rebar(value: object, *, direction: str) -> bool:
    normalized = _normalized_cell_text(value)
    pattern = _STANDARD_GRID_SPEC if direction == "Z" else _STANDARD_LINEAR_SPEC
    match = pattern.fullmatch(normalized)
    if match is None:
        return False
    return all(
        int(component) > 0
        for component in match.groupdict().values()
        if component is not None
    )


def _is_standard_wall_text(value: object) -> bool:
    if normalize_wall_id(value) is None:
        return False
    return _STANDARD_WALL_TEXT.fullmatch(str(value).strip()) is not None


def _wall_candidates(workbook: _WorkbookLike) -> list[_WallCandidate]:
    candidates: list[_WallCandidate] = []
    for sheet in workbook.worksheets:
        columns = _find_columns(sheet)
        if columns is not None:
            candidates.append((sheet, columns))
    return candidates


def _wall_value_reason(
    sheet: _WorksheetLike,
    columns: _WallColumns,
) -> FormatReason | None:
    header_row, wall_column, x_column, y_column, z_column = columns
    for row in range(header_row + 1, sheet.max_row + 1):
        values = {
            "wall": sheet.cell(row=row, column=wall_column).value,
            "X": sheet.cell(row=row, column=x_column).value,
            "Y": sheet.cell(row=row, column=y_column).value,
            "Z": sheet.cell(row=row, column=z_column).value,
        }
        if not any(value is not None and str(value).strip() for value in values.values()):
            continue
        if not _is_standard_wall_text(values["wall"]) or any(
            not _is_standard_rebar(values[direction], direction=direction)
            for direction in ("X", "Y", "Z")
        ):
            return FormatReason(
                scope="wall",
                code="wall_value_nonstandard",
                sheet=sheet.title,
                message=f"{sheet.title} 第 {row} 行不是标准墙体配筋写法",
            )
    return None


def _find_slab_like_header(sheet: _WorksheetLike) -> int | None:
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0
    for row in range(1, min(max_row, 20) + 1):
        headers = [
            _header_text(sheet.cell(row=row, column=column).value)
            for column in range(1, max_column + 1)
        ]
        has_elevation = any("标高" in header or "楼层" in header for header in headers)
        has_top_x = any(
            ("顶层" in header or "上部" in header)
            and ("水平" in header or "X" in header.upper())
            for header in headers
        )
        has_top_y = any(
            ("顶层" in header or "上部" in header)
            and ("竖向" in header or "Y" in header.upper())
            for header in headers
        )
        has_bottom_x = any(
            ("底层" in header or "下部" in header)
            and ("水平" in header or "X" in header.upper())
            for header in headers
        )
        has_bottom_y = any(
            ("底层" in header or "下部" in header)
            and ("竖向" in header or "Y" in header.upper())
            for header in headers
        )
        has_tie = any("拉筋" in header for header in headers)
        if all(
            (
                has_elevation,
                has_top_x,
                has_top_y,
                has_bottom_x,
                has_bottom_y,
                has_tie,
            )
        ):
            return row
    return None


def _slab_candidates(workbook: _WorkbookLike) -> list[_SlabCandidate]:
    candidates: list[_SlabCandidate] = []
    for sheet in workbook.worksheets:
        columns = _find_slab_columns(sheet)
        like_header_row = _find_slab_like_header(sheet)
        if columns is not None or like_header_row is not None:
            candidates.append((sheet, columns, like_header_row))
    return candidates


def _uses_standard_slab_layout(
    sheet: _WorksheetLike,
    found: _SlabColumns | None,
) -> bool:
    if found is None:
        return False
    header_row, columns = found
    return (
        sheet.title == "楼板配筋"
        and header_row == 1
        and tuple(columns.values()) == tuple(range(1, 9))
    )


def _inspect_slab_values(
    sheet: _WorksheetLike,
    found: _SlabColumns,
) -> tuple[int, FormatReason | None]:
    header_row, columns = found
    valid_row_count = 0
    for row in range(header_row + 1, sheet.max_row + 1):
        values = {
            key: sheet.cell(row=row, column=column).value
            for key, column in columns.items()
        }
        if not any(value is not None and str(value).strip() for value in values.values()):
            continue
        try:
            normalize_slab_elevation(values["elevation"])
        except InvalidReinforcementWorkbook:
            return (
                valid_row_count,
                FormatReason(
                    scope="slab",
                    code="slab_value_nonstandard",
                    sheet=sheet.title,
                    message=f"{sheet.title} 第 {row} 行的楼板标高不是标准写法",
                ),
            )
        for key in (
            "top_x",
            "top_y",
            "middle_x",
            "middle_y",
            "bottom_x",
            "bottom_y",
            "z",
        ):
            value = values[key]
            if key in {"middle_x", "middle_y"} and value in (None, ""):
                continue
            if not _is_standard_rebar(value, direction="X"):
                return (
                    valid_row_count,
                    FormatReason(
                        scope="slab",
                        code="slab_value_nonstandard",
                        sheet=sheet.title,
                        message=f"{sheet.title} 第 {row} 行不是标准楼板配筋写法",
                    ),
                )
        valid_row_count += 1
    return valid_row_count, None


def inspect_reinforcement_workbook(
    path: Path,
    *,
    include_slab: bool,
) -> WorkbookFormatInspection:
    workbook = cast(
        _WorkbookLike,
        load_workbook(path, data_only=False, read_only=True),
    )
    reasons: list[FormatReason] = []
    wall_sheet_name: str | None = None
    slab_sheet_name: str | None = None
    try:
        wall_candidates = _wall_candidates(workbook)
        selected_wall = wall_candidates[0] if wall_candidates else None
        if selected_wall is None:
            reasons.append(
                FormatReason(
                    scope="wall",
                    code="wall_sheet_missing",
                    sheet=None,
                    message="未找到墙体配筋输入表",
                )
            )
        else:
            wall_sheet, wall_columns = selected_wall
            wall_sheet_name = wall_sheet.title
            if len(wall_candidates) > 1:
                reasons.append(
                    FormatReason(
                        scope="wall",
                        code="wall_sheet_ambiguous",
                        sheet=wall_sheet.title,
                        message=(
                            "找到多个墙体配筋候选表，按加载顺序选择首个："
                            + "、".join(
                                candidate[0].title
                                for candidate in wall_candidates
                            )
                        ),
                    )
                )
            if not _uses_standard_layout(wall_sheet, wall_columns):
                reasons.append(
                    FormatReason(
                        scope="wall",
                        code="wall_layout_nonstandard",
                        sheet=wall_sheet.title,
                        message=f"{wall_sheet.title} 不是标准四列墙体配筋模板",
                    )
                )
            else:
                wall_reason = _wall_value_reason(wall_sheet, wall_columns)
                if wall_reason is not None:
                    reasons.append(wall_reason)

        if include_slab:
            slab_candidates = _slab_candidates(workbook)
            selected_slab = slab_candidates[0] if slab_candidates else None
            if selected_slab is None:
                reasons.append(
                    FormatReason(
                        scope="slab",
                        code="slab_sheet_missing",
                        sheet=None,
                        message="已启用楼板计算，但未找到楼板配筋输入表",
                    )
                )
            else:
                slab_sheet, slab_columns, _ = selected_slab
                slab_sheet_name = slab_sheet.title
                if len(slab_candidates) > 1:
                    reasons.append(
                        FormatReason(
                            scope="slab",
                            code="slab_sheet_ambiguous",
                            sheet=slab_sheet.title,
                            message=(
                                "找到多个楼板配筋候选表，按工作簿顺序报告首个："
                                + "、".join(
                                    candidate[0].title
                                    for candidate in slab_candidates
                                )
                            ),
                        )
                    )
                if not _uses_standard_slab_layout(slab_sheet, slab_columns):
                    reasons.append(
                        FormatReason(
                            scope="slab",
                            code="slab_layout_nonstandard",
                            sheet=slab_sheet.title,
                            message=f"{slab_sheet.title} 不是标准八列楼板配筋模板",
                        )
                    )
                else:
                    assert slab_columns is not None
                    valid_rows, slab_reason = _inspect_slab_values(
                        slab_sheet,
                        slab_columns,
                    )
                    if slab_reason is not None:
                        reasons.append(slab_reason)
                    elif valid_rows == 0:
                        reasons.append(
                            FormatReason(
                                scope="slab",
                                code="slab_data_missing",
                                sheet=slab_sheet.title,
                                message="楼板配筋标准表没有有效数据行",
                            )
                        )
    finally:
        workbook.close()

    return WorkbookFormatInspection(
        requires_ai_normalization=bool(reasons),
        reasons=tuple(reasons),
        wall_sheet=wall_sheet_name,
        slab_sheet=slab_sheet_name,
    )


def _snapshot_resource_limits(
    max_non_empty_cells: int,
) -> tuple[int, int, int, int]:
    worksheet_xml_bytes = min(
        128 * 1024 * 1024,
        max(1024 * 1024, max_non_empty_cells * 16 * 1024),
    )
    cell_records = min(1_000_000, max(1024, max_non_empty_cells * 8))
    merge_records = min(100_000, max(256, max_non_empty_cells * 4))
    merged_span_cells = min(1_000_000, max(10_000, max_non_empty_cells * 32))
    return worksheet_xml_bytes, cell_records, merge_records, merged_span_cells


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _preflight_snapshot_xlsx(path: Path, *, max_non_empty_cells: int) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"workbook does not exist: {path}")
    if not is_zipfile(path):
        raise ValueError(f"workbook is not a valid XLSX zip: {path.name}")

    (
        max_worksheet_xml_bytes,
        max_cell_records,
        max_merge_records,
        max_merged_span_cells,
    ) = _snapshot_resource_limits(max_non_empty_cells)
    cell_record_count = 0
    non_empty_record_count = 0
    merge_record_count = 0
    merged_span_cell_count = 0
    try:
        with ZipFile(path) as archive:
            worksheet_infos = [
                info
                for info in archive.infolist()
                if (
                    not info.is_dir()
                    and info.filename.startswith("xl/worksheets/")
                    and info.filename.endswith(".xml")
                )
            ]
            if not worksheet_infos:
                raise ValueError("XLSX does not contain worksheet XML")
            if sum(info.file_size for info in worksheet_infos) > max_worksheet_xml_bytes:
                raise ValueError(
                    "worksheet XML size exceeds snapshot limit "
                    f"{max_worksheet_xml_bytes} bytes"
                )

            for info in worksheet_infos:
                with archive.open(info) as stream:
                    for _, element in iterparse(stream, events=("end",)):
                        local_name = _xml_local_name(element.tag)
                        if local_name == "c":
                            cell_record_count += 1
                            if cell_record_count > max_cell_records:
                                raise ValueError(
                                    "worksheet cell records exceed snapshot limit "
                                    f"{max_cell_records}"
                                )
                            if any(
                                _xml_local_name(child.tag) in {"f", "is", "v"}
                                for child in element
                            ):
                                non_empty_record_count += 1
                                if non_empty_record_count > max_non_empty_cells:
                                    raise ValueError(
                                        "workbook contains more non-empty cells than "
                                        f"max_non_empty_cells={max_non_empty_cells}"
                                    )
                        elif local_name == "mergeCell":
                            merge_record_count += 1
                            if merge_record_count > max_merge_records:
                                raise ValueError(
                                    "worksheet merge records exceed snapshot limit "
                                    f"{max_merge_records}"
                                )
                            reference = element.attrib.get("ref")
                            if not reference:
                                raise ValueError("worksheet merge is missing a range")
                            try:
                                min_col, min_row, max_col, max_row = cast(
                                    tuple[int, int, int, int],
                                    range_boundaries(reference),
                                )
                            except (TypeError, ValueError) as exc:
                                raise ValueError(
                                    f"worksheet merged range is invalid: {reference}"
                                ) from exc
                            range_span = (
                                (max_row - min_row + 1)
                                * (max_col - min_col + 1)
                            )
                            if range_span > max_merged_span_cells:
                                raise ValueError(
                                    f"worksheet merged range {reference} exceeds "
                                    "snapshot limit"
                                )
                            merged_span_cell_count += range_span
                            if merged_span_cell_count > max_merged_span_cells:
                                raise ValueError(
                                    "worksheet merged ranges exceed snapshot limit "
                                    f"{max_merged_span_cells} cells"
                                )
                        element.clear()
    except (BadZipFile, OSError, ParseError) as exc:
        raise ValueError(f"unable to inspect XLSX structure: {path.name}") from exc


def _json_value(value: object) -> object:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, Decimal):
        return str(value)
    value_type = type(value)
    raise TypeError(
        "unsupported workbook cell value type: "
        f"{value_type.__module__}.{value_type.__qualname__}"
    )


def _serialize_formula(value: object) -> str | dict[str, object]:
    if isinstance(value, str):
        return value
    if isinstance(value, ArrayFormula):
        return {
            "type": "array",
            "ref": _json_value(value.ref),
            "text": _json_value(value.text),
        }
    if isinstance(value, DataTableFormula):
        return {
            "type": "data_table",
            "ref": _json_value(value.ref),
            "ca": _formula_boolean(value.ca),
            "dt2D": _formula_boolean(value.dt2D),
            "dtr": _formula_boolean(value.dtr),
            "r1": _json_value(value.r1),
            "r2": _json_value(value.r2),
            "del1": _formula_boolean(value.del1),
            "del2": _formula_boolean(value.del2),
        }
    value_type = type(value)
    raise TypeError(
        "unsupported workbook formula type: "
        f"{value_type.__module__}.{value_type.__qualname__}"
    )


def _formula_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true"}:
            return True
        if normalized in {"0", "false"}:
            return False
    raise TypeError(f"unsupported workbook formula boolean: {value!r}")


def build_workbook_snapshot(
    path: Path,
    *,
    max_non_empty_cells: int,
) -> dict[str, object]:
    if max_non_empty_cells <= 0:
        raise ValueError("max_non_empty_cells must be greater than 0")

    _preflight_snapshot_xlsx(
        path,
        max_non_empty_cells=max_non_empty_cells,
    )
    workbook = cast(
        _WorkbookLike,
        load_workbook(path, data_only=False, read_only=False),
    )
    sheets: list[dict[str, object]] = []
    cells: list[dict[str, object]] = []
    try:
        for workbook_sheet in workbook.worksheets:
            sheet = cast(_SnapshotWorksheetLike, workbook_sheet)
            merged_cell_ranges = sorted(sheet.merged_cells.ranges, key=str)
            merged_ranges = [str(cell_range) for cell_range in merged_cell_ranges]
            merged_anchors = {
                (cell_range.min_row, cell_range.min_col): str(cell_range)
                for cell_range in merged_cell_ranges
            }
            sheets.append(
                {
                    "name": sheet.title,
                    "merged_ranges": merged_ranges,
                }
            )
            # Iterating Worksheet.iter_rows() expands the max-row/max-column
            # rectangle. Sorting the XML-populated sparse store is O(records).
            for (row, column), cell in sorted(sheet._cells.items()):
                if cell.value in (None, ""):
                    continue
                if len(cells) >= max_non_empty_cells:
                    raise ValueError(
                        "workbook contains more non-empty cells than "
                        f"max_non_empty_cells={max_non_empty_cells}"
                    )
                is_advanced_formula = isinstance(
                    cell.value,
                    (ArrayFormula, DataTableFormula),
                )
                formula = (
                    _serialize_formula(cell.value)
                    if cell.data_type == "f" or is_advanced_formula
                    else None
                )
                value = formula if is_advanced_formula else _json_value(cell.value)
                cells.append(
                    {
                        "sheet": sheet.title,
                        "row": row,
                        "column": column,
                        "address": cell.coordinate,
                        "value": value,
                        "formula": formula,
                        "merged_range": merged_anchors.get((row, column)),
                    }
                )
    finally:
        workbook.close()

    return {
        "workbook": path.name,
        "non_empty_cell_count": len(cells),
        "sheets": sheets,
        "cells": cells,
    }
