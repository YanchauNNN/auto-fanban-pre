from __future__ import annotations

import math
import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


class InsufficientRebarCapacity(ValueError):
    pass


@dataclass(frozen=True)
class RebarAreaRow:
    diameter: int
    area_by_bar_count: dict[int, float]


@dataclass(frozen=True)
class RebarSelection:
    specification: str
    target_area: int
    actual_area: int
    diameter: int
    row_count: int
    spacing: int
    margin_percent: float


def _number(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).strip())
    return float(match.group()) if match else None


def load_rebar_area_table(path: Path) -> list[RebarAreaRow]:
    if not path.is_file():
        raise FileNotFoundError(f"未找到钢筋面积表：{path}")
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    try:
        header_row: int | None = None
        bar_count_columns: dict[int, int] = {}
        for row_index in range(1, min(sheet.max_row, 20) + 1):
            current: dict[int, int] = {}
            for column_index in range(1, sheet.max_column + 1):
                numeric = _number(sheet.cell(row=row_index, column=column_index).value)
                if numeric is None or not numeric.is_integer():
                    continue
                bar_count = int(numeric)
                if 1 <= bar_count <= 20:
                    current[bar_count] = column_index
            if len(current) >= 3:
                header_row = row_index
                bar_count_columns = current
                break
        if header_row is None:
            raise ValueError("钢筋面积表缺少可识别的根数表头")

        rows: list[RebarAreaRow] = []
        for row_index in range(header_row + 1, sheet.max_row + 1):
            raw_diameter = _number(sheet.cell(row=row_index, column=1).value)
            if raw_diameter is None:
                continue
            areas: dict[int, float] = {}
            for bar_count, column_index in bar_count_columns.items():
                area = _number(sheet.cell(row=row_index, column=column_index).value)
                if area is not None and area > 0:
                    areas[bar_count] = area
            if areas:
                rows.append(
                    RebarAreaRow(
                        diameter=int(round(raw_diameter)),
                        area_by_bar_count=areas,
                    )
                )
    finally:
        workbook.close()
    rows.sort(key=lambda row: row.diameter)
    if not rows:
        raise ValueError("钢筋面积表未识别到有效数据")
    return rows


def select_rebar(
    sm_value: float | int | str,
    table: Iterable[RebarAreaRow],
    *,
    row_counts: tuple[int, ...] = (1, 2),
    spacings: tuple[int, ...] = (200, 250),
    max_diameter: int = 40,
    extra_ratio: float = 0.2,
) -> RebarSelection:
    try:
        sm = float(sm_value)
    except (TypeError, ValueError) as exc:
        raise ValueError("SM 配筋面积不是有效数字") from exc
    if not math.isfinite(sm) or sm <= 0:
        raise ValueError("SM 配筋面积必须大于 0")
    if extra_ratio < 0:
        raise ValueError("SM 安全放大系数不能小于 0")

    target = sm * (1 + extra_ratio)
    candidates: list[tuple[float, int, int, int]] = []
    for row in table:
        if row.diameter > max_diameter:
            continue
        for spacing in spacings:
            bars_per_meter = int(round(1000 / spacing))
            single_row_area = row.area_by_bar_count.get(bars_per_meter)
            if single_row_area is None:
                continue
            for row_count in row_counts:
                actual = single_row_area * row_count
                if actual >= target:
                    candidates.append((actual, row_count, row.diameter, spacing))

    if not candidates:
        raise InsufficientRebarCapacity(
            f"没有满足目标面积 {math.ceil(target)} mm²/m 的配筋组合"
        )

    actual, row_count, diameter, spacing = min(
        candidates,
        key=lambda item: (item[0], item[1], item[2], -item[3]),
    )
    return RebarSelection(
        specification=f"{row_count}排{diameter}@{spacing}",
        target_area=math.ceil(target),
        actual_area=int(round(actual)),
        diameter=diameter,
        row_count=row_count,
        spacing=spacing,
        margin_percent=(actual - sm) / sm * 100,
    )
