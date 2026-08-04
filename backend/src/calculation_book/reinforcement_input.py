from __future__ import annotations

import math
import re
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class InvalidReinforcementWorkbook(ValueError):
    pass


@dataclass(frozen=True)
class RebarConfiguration:
    layers: int
    diameter: int
    spacing_primary: int
    spacing_secondary: int | None
    canonical_specification: str
    narrative_specification: str
    actual_area: float
    is_parenthetical: bool


@dataclass(frozen=True)
class ParsedRebarCell:
    original_text: str
    normalized_text: str
    candidates: tuple[RebarConfiguration, ...]
    selected: RebarConfiguration


@dataclass(frozen=True)
class NormalizedReinforcementRow:
    wall_id: str
    x: ParsedRebarCell
    y: ParsedRebarCell
    z: ParsedRebarCell
    source_sheet: str
    source_row: int
    source_cells: dict[str, str]


@dataclass(frozen=True)
class ReinforcementRowIssue:
    source_sheet: str
    source_row: int
    source_cells: dict[str, str]
    original_values: dict[str, str]
    original_wall_text: str
    wall_id: str | None
    error: str


@dataclass(frozen=True)
class ReinforcementSchedule:
    rows: tuple[NormalizedReinforcementRow, ...]
    duplicate_wall_ids: tuple[str, ...]
    issues: tuple[ReinforcementRowIssue, ...] = ()
    source_row_count: int = 0
    normalization_triggered: bool = False

    def __post_init__(self) -> None:
        audited_count = len(self.rows) + len(self.issues)
        if self.source_row_count == 0 and audited_count:
            object.__setattr__(self, "source_row_count", audited_count)
        if self.source_row_count != audited_count:
            raise ValueError(
                "配筋表审计数量不守恒："
                f"{self.source_row_count} != {len(self.rows)} + {len(self.issues)}"
            )

    @property
    def requires_manual_confirmation(self) -> bool:
        return bool(self.duplicate_wall_ids or self.issues)

    @property
    def normalized_row_count(self) -> int:
        return len(self.rows)

    @property
    def issue_row_count(self) -> int:
        return len(self.issues)

    @property
    def unique_wall_count(self) -> int:
        return len({row.wall_id for row in self.rows})


def build_reinforcement_schedule(
    *,
    rows: tuple[NormalizedReinforcementRow, ...],
    issues: tuple[ReinforcementRowIssue, ...] = (),
    source_row_count: int | None = None,
    normalization_triggered: bool = False,
) -> ReinforcementSchedule:
    counts = Counter(
        [row.wall_id for row in rows]
        + [issue.wall_id for issue in issues if issue.wall_id is not None]
    )
    duplicates = tuple(
        sorted(wall_id for wall_id, count in counts.items() if count > 1)
    )
    audited_count = len(rows) + len(issues)
    return ReinforcementSchedule(
        rows=rows,
        duplicate_wall_ids=duplicates,
        issues=issues,
        source_row_count=(
            audited_count if source_row_count is None else source_row_count
        ),
        normalization_triggered=normalization_triggered,
    )


@dataclass(frozen=True)
class NormalizedSlabReinforcementRow:
    elevation: str
    top_x: ParsedRebarCell
    top_y: ParsedRebarCell
    middle_x: ParsedRebarCell | None
    middle_y: ParsedRebarCell | None
    bottom_x: ParsedRebarCell
    bottom_y: ParsedRebarCell
    z: ParsedRebarCell
    source_sheet: str
    source_row: int
    source_cells: dict[str, str]


@dataclass(frozen=True)
class SlabReinforcementSchedule:
    rows: tuple[NormalizedSlabReinforcementRow, ...]


_SPEC_PATTERN = re.compile(
    r"""
    (?:
        (?P<layers_marker>\d+)\s*(?:排\s*)?[DCA]\s*
        |
        (?P<layers_row>\d+)\s*排\s*
        |
        (?P<layers_space>\d+)\s+
    )?
    (?P<diameter>\d+)
    \s*(?:@|间距)\s*
    (?P<spacing_primary>\d+)
    (?:\s*[*xX×]\s*(?P<spacing_secondary>\d+))?
    """,
    re.VERBOSE | re.IGNORECASE,
)
_WALL_ID_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])(?P<wall>[A-Za-z]+\d+[A-Za-z]?(?:-\d+)?)(?![A-Za-z0-9])"
)


def _normalized_cell_text(value: object) -> str:
    return (
        str(value)
        .strip()
        .replace("（", "(")
        .replace("）", ")")
        .replace("＃", "#")
        .replace("×", "*")
    )


def _format_direction(direction: str) -> str:
    normalized = direction.strip().upper()
    if normalized not in {"X", "Y", "Z"}:
        raise ValueError(f"不支持的配筋方向：{direction}")
    return normalized


def _configuration(
    *,
    layers: int,
    diameter: int,
    spacing_primary: int,
    spacing_secondary: int | None,
    direction: str,
    is_parenthetical: bool,
) -> RebarConfiguration:
    if layers not in {1, 2}:
        raise InvalidReinforcementWorkbook("钢筋层数只允许 1 或 2")
    if layers <= 0 or diameter <= 0 or spacing_primary <= 0:
        raise InvalidReinforcementWorkbook("钢筋层数、直径和间距必须大于 0")

    bar_area = math.pi * (diameter / 2) ** 2
    actual_area = layers * bar_area * (1000 / spacing_primary)
    if direction == "Z":
        if spacing_secondary is None or spacing_secondary <= 0:
            raise InvalidReinforcementWorkbook("拉筋必须填写两个方向的网格间距")
        actual_area *= 1000 / spacing_secondary
        canonical = (
            f"{layers}C{diameter}间距{spacing_primary}*{spacing_secondary}"
        )
        narrative = f"{layers}排{diameter}@{spacing_primary}x{spacing_secondary}"
    else:
        if spacing_secondary is not None:
            raise InvalidReinforcementWorkbook("水平筋或竖向筋只能填写一个方向的间距")
        canonical = f"{layers}D{diameter}间距{spacing_primary}"
        narrative = f"{layers}排{diameter}@{spacing_primary}"

    return RebarConfiguration(
        layers=layers,
        diameter=diameter,
        spacing_primary=spacing_primary,
        spacing_secondary=spacing_secondary,
        canonical_specification=canonical,
        narrative_specification=narrative,
        actual_area=actual_area,
        is_parenthetical=is_parenthetical,
    )


def parse_rebar_cell(value: object, *, direction: str) -> ParsedRebarCell:
    normalized_direction = _format_direction(direction)
    original = "" if value is None else str(value).strip()
    normalized = _normalized_cell_text(original).replace("#", "")
    parenthetical_start = normalized.find("(")
    candidates: list[RebarConfiguration] = []

    for match in _SPEC_PATTERN.finditer(normalized):
        raw_layers = (
            match.group("layers_marker")
            or match.group("layers_row")
            or match.group("layers_space")
        )
        layers = int(raw_layers) if raw_layers is not None else 1
        spacing_secondary = match.group("spacing_secondary")
        candidates.append(
            _configuration(
                layers=layers,
                diameter=int(match.group("diameter")),
                spacing_primary=int(match.group("spacing_primary")),
                spacing_secondary=(
                    int(spacing_secondary) if spacing_secondary is not None else None
                ),
                direction=normalized_direction,
                is_parenthetical=(
                    parenthetical_start >= 0 and match.start() > parenthetical_start
                ),
            )
        )

    if not candidates:
        raise InvalidReinforcementWorkbook(f"无法识别配筋写法：{original or '<空>'}")

    parenthetical = [candidate for candidate in candidates if candidate.is_parenthetical]
    selection_pool = parenthetical or candidates
    selected = max(
        selection_pool,
        key=lambda candidate: (
            candidate.actual_area,
            candidate.layers,
            candidate.diameter,
            -candidate.spacing_primary,
            -(candidate.spacing_secondary or 0),
        ),
    )
    return ParsedRebarCell(
        original_text=original,
        normalized_text=normalized,
        candidates=tuple(candidates),
        selected=selected,
    )


def parse_linear_rebar_cell(value: object) -> ParsedRebarCell:
    return parse_rebar_cell(value, direction="X")


def normalize_slab_elevation(value: object) -> str:
    text = str(value).strip()
    if text.lower().endswith("m"):
        text = text[:-1].strip()
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise InvalidReinforcementWorkbook(f"无法识别楼板标高：{value}") from exc
    if not decimal_value.is_finite():
        raise InvalidReinforcementWorkbook(f"楼板标高必须是有限数值：{value}")
    normalized = format(decimal_value.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return "0" if normalized in {"-0", "+0"} else normalized


def normalize_wall_id(value: object) -> str | None:
    text = _normalized_cell_text(value).replace("墙", "").replace(" ", "")
    match = _WALL_ID_PATTERN.search(text)
    return match.group("wall").upper() if match is not None else None


def _header_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).replace("（", "(").replace("）", ")")


def _ensure_worksheet_dimensions(sheet) -> tuple[int, int]:
    if sheet.max_row is None or sheet.max_column is None:
        sheet.calculate_dimension(force=True)
    return sheet.max_row or 0, sheet.max_column or 0


def _wall_column_score(sheet, column: int, header_row: int) -> int:
    score = 0
    end_row = min(sheet.max_row, header_row + 100)
    for row in range(header_row + 1, end_row + 1):
        if normalize_wall_id(sheet.cell(row=row, column=column).value) is not None:
            score += 1
    return score


def _find_columns(sheet) -> tuple[int, int, int, int, int] | None:
    max_row, max_column = _ensure_worksheet_dimensions(sheet)
    for row in range(1, min(max_row, 20) + 1):
        headers = {
            column: _header_text(sheet.cell(row=row, column=column).value)
            for column in range(1, max_column + 1)
        }
        x_columns = [
            column
            for column, text in headers.items()
            if "水平筋" in text or "水平钢筋" in text
        ]
        y_columns = [
            column
            for column, text in headers.items()
            if "竖向筋" in text or "竖向钢筋" in text
        ]
        z_columns = [
            column
            for column, text in headers.items()
            if "拉筋" in text
        ]
        wall_columns = [
            column
            for column, text in headers.items()
            if "墙号" in text or "构件编号" in text
        ]
        if not (x_columns and y_columns and z_columns and wall_columns):
            continue
        wall_column = max(
            wall_columns,
            key=lambda column: (_wall_column_score(sheet, column, row), column),
        )
        return row, wall_column, x_columns[0], y_columns[0], z_columns[0]
    return None


def _uses_standard_layout(
    sheet,
    columns: tuple[int, int, int, int, int],
) -> bool:
    header_row, wall_column, x_column, y_column, z_column = columns
    expected = {
        wall_column: "构件编号及位置",
        x_column: "单侧水平钢筋(对称配筋)",
        y_column: "单侧竖向钢筋(对称配筋)",
        z_column: "拉筋",
    }
    return (
        (wall_column, x_column, y_column, z_column) == (1, 2, 3, 4)
        and all(
            _header_text(sheet.cell(row=header_row, column=column).value)
            == header
            for column, header in expected.items()
        )
    )


def load_reinforcement_schedule(path: Path) -> ReinforcementSchedule:
    if not path.is_file():
        raise FileNotFoundError(f"未找到墙体配筋表：{path}")
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InvalidReinforcementWorkbook(f"无法读取墙体配筋表：{path.name}") from exc

    rows: list[NormalizedReinforcementRow] = []
    issues: list[ReinforcementRowIssue] = []
    source_row_count = 0
    normalization_triggered = False
    try:
        selected_sheet = None
        selected_columns = None
        for sheet in workbook.worksheets:
            columns = _find_columns(sheet)
            if columns is not None:
                selected_sheet = sheet
                selected_columns = columns
                break
        if selected_sheet is None or selected_columns is None:
            raise InvalidReinforcementWorkbook(
                "未找到墙号、水平筋、竖向筋和拉筋表头"
            )

        header_row, wall_column, x_column, y_column, z_column = selected_columns
        normalization_triggered = not _uses_standard_layout(
            selected_sheet,
            selected_columns,
        )
        for row_number in range(header_row + 1, selected_sheet.max_row + 1):
            raw_wall = selected_sheet.cell(
                row=row_number,
                column=wall_column,
            ).value
            raw_values = {
                "X": selected_sheet.cell(row=row_number, column=x_column).value,
                "Y": selected_sheet.cell(row=row_number, column=y_column).value,
                "Z": selected_sheet.cell(row=row_number, column=z_column).value,
            }
            if not any(
                value is not None and str(value).strip()
                for value in (raw_wall, *raw_values.values())
            ):
                continue
            source_row_count += 1
            source_cells = {
                "wall": f"{get_column_letter(wall_column)}{row_number}",
                "X": f"{get_column_letter(x_column)}{row_number}",
                "Y": f"{get_column_letter(y_column)}{row_number}",
                "Z": f"{get_column_letter(z_column)}{row_number}",
            }
            original_wall_text = "" if raw_wall is None else str(raw_wall).strip()
            original_values = {
                "wall": original_wall_text,
                **{
                    direction: "" if value is None else str(value).strip()
                    for direction, value in raw_values.items()
                },
            }
            wall_id = normalize_wall_id(raw_wall)
            if wall_id is None:
                normalization_triggered = True
                issues.append(
                    ReinforcementRowIssue(
                        source_sheet=selected_sheet.title,
                        source_row=row_number,
                        source_cells=source_cells,
                        original_values=original_values,
                        original_wall_text=original_wall_text,
                        wall_id=None,
                        error="无法识别墙号",
                    )
                )
                continue
            try:
                parsed = {
                    direction: parse_rebar_cell(value, direction=direction)
                    for direction, value in raw_values.items()
                }
            except InvalidReinforcementWorkbook as exc:
                normalization_triggered = True
                issues.append(
                    ReinforcementRowIssue(
                        source_sheet=selected_sheet.title,
                        source_row=row_number,
                        source_cells=source_cells,
                        original_values=original_values,
                        original_wall_text=original_wall_text,
                        wall_id=wall_id,
                        error=str(exc),
                    )
                )
                continue
            if original_wall_text.upper() != wall_id:
                normalization_triggered = True
            if any(
                parsed[direction].original_text
                != parsed[direction].selected.canonical_specification
                for direction in ("X", "Y", "Z")
            ):
                normalization_triggered = True
            rows.append(
                NormalizedReinforcementRow(
                    wall_id=wall_id,
                    x=parsed["X"],
                    y=parsed["Y"],
                    z=parsed["Z"],
                    source_sheet=selected_sheet.title,
                    source_row=row_number,
                    source_cells=source_cells,
                )
            )
    finally:
        workbook.close()

    if source_row_count == 0:
        raise InvalidReinforcementWorkbook("墙体配筋表没有可识别的数据行")

    return build_reinforcement_schedule(
        rows=tuple(rows),
        issues=tuple(issues),
        source_row_count=source_row_count,
        normalization_triggered=normalization_triggered,
    )


_SLAB_HEADERS = {
    "elevation": "标高",
    "top_x": "顶层水平",
    "top_y": "顶层竖向",
    "middle_x": "中层水平",
    "middle_y": "中层竖向",
    "bottom_x": "底层水平",
    "bottom_y": "底层竖向",
    "z": "纵向拉筋",
}


def _find_slab_columns(sheet) -> tuple[int, dict[str, int]] | None:
    expected = set(_SLAB_HEADERS.values())
    max_row, max_column = _ensure_worksheet_dimensions(sheet)
    for row in range(1, min(max_row, 20) + 1):
        headers = {
            _header_text(sheet.cell(row=row, column=column).value): column
            for column in range(1, max_column + 1)
        }
        if not expected.issubset(headers):
            continue
        return row, {
            key: headers[header]
            for key, header in _SLAB_HEADERS.items()
        }
    return None


def load_slab_reinforcement_schedule(
    path: Path,
    *,
    required: bool,
) -> SlabReinforcementSchedule | None:
    if not path.is_file():
        raise FileNotFoundError(f"未找到配筋表：{path}")
    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InvalidReinforcementWorkbook(f"无法读取楼板配筋表：{path.name}") from exc

    try:
        if "楼板配筋" not in workbook.sheetnames:
            if required:
                raise InvalidReinforcementWorkbook("未找到“楼板配筋”Sheet")
            return None
        sheet = workbook["楼板配筋"]
        found = _find_slab_columns(sheet)
        if found is None:
            raise InvalidReinforcementWorkbook(
                "楼板配筋Sheet缺少标高、顶层/中层/底层水平竖向或纵向拉筋表头"
            )
        header_row, columns = found
        rows: list[NormalizedSlabReinforcementRow] = []
        seen_elevations: set[str] = set()
        for row_number in range(header_row + 1, sheet.max_row + 1):
            raw_elevation = sheet.cell(
                row=row_number,
                column=columns["elevation"],
            ).value
            if raw_elevation in (None, ""):
                continue
            elevation = normalize_slab_elevation(raw_elevation)
            if elevation in seen_elevations:
                raise InvalidReinforcementWorkbook(f"楼板配筋存在重复标高：{elevation}")
            seen_elevations.add(elevation)

            source_cells = {
                key: f"{get_column_letter(column)}{row_number}"
                for key, column in columns.items()
            }
            parsed: dict[str, ParsedRebarCell | None] = {}
            for key in (
                "top_x",
                "top_y",
                "middle_x",
                "middle_y",
                "bottom_x",
                "bottom_y",
                "z",
            ):
                value = sheet.cell(row=row_number, column=columns[key]).value
                if key in {"middle_x", "middle_y"} and value in (None, ""):
                    parsed[key] = None
                    continue
                try:
                    parsed[key] = parse_linear_rebar_cell(value)
                except InvalidReinforcementWorkbook as exc:
                    raise InvalidReinforcementWorkbook(
                        f"{sheet.title}!{source_cells[key]}（标高{elevation}）：{exc}"
                    ) from exc

            rows.append(
                NormalizedSlabReinforcementRow(
                    elevation=elevation,
                    top_x=cast(ParsedRebarCell, parsed["top_x"]),
                    top_y=cast(ParsedRebarCell, parsed["top_y"]),
                    middle_x=parsed["middle_x"],
                    middle_y=parsed["middle_y"],
                    bottom_x=cast(ParsedRebarCell, parsed["bottom_x"]),
                    bottom_y=cast(ParsedRebarCell, parsed["bottom_y"]),
                    z=cast(ParsedRebarCell, parsed["z"]),
                    source_sheet=sheet.title,
                    source_row=row_number,
                    source_cells=source_cells,
                )
            )
    finally:
        workbook.close()

    if not rows:
        raise InvalidReinforcementWorkbook("楼板配筋Sheet没有可识别的数据行")
    return SlabReinforcementSchedule(rows=tuple(rows))
