from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from src.doc_gen.ied import IEDGenerator
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    TitleblockFields,
)


def _make_frame(seq: int) -> FrameMeta:
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
    )
    titleblock = TitleblockFields(
        internal_code=f"1234567-JG001-{seq:03d}",
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"图纸{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        ied_status="发布",
        ied_doc_type="图册",
        ied_change_flag="MOD",
        ied_design_type="设计类",
        ied_responsible_unit="河北分公司-建筑结构所-结构一室",
        ied_discipline_office="河北分公司-建筑结构所-结构一室",
        ied_chief_designer="张三@A001",
        ied_person_qual_category="一般核安全物项-民用",
        ied_fu_flag="N",
        ied_internal_tag="否",
        ied_prepared_by="李四@A002",
        ied_prepared_by_2="王五@A003",
        ied_prepared_date="2026-03-01",
        ied_checked_by="赵六@A004",
        ied_checked_date="2026-03-02",
        ied_discipline_leader="孙七@A005",
        ied_discipline_leader_date="2026-03-03",
        ied_reviewed_by="周八@A006",
        ied_reviewed_date="2026-03-04",
        ied_approved_by="吴九@A007",
        ied_approved_date="2026-03-05",
        ied_submitted_plan_date="2026-03-06",
        ied_publish_plan_date="2026-03-07",
        ied_external_plan_date="2026-03-08",
        ied_fu_plan_date="2026-03-09",
        wbs_code="WBS-001",
        classification="非密",
        work_hours="66",
        cover_revision="A",
    )
    derived = DerivedFields(
        album_internal_code="1234567-JG001",
        cover_external_code="JD1NHT11F01B25C42SD",
        cover_internal_code="1234567-JG001-FM",
        cover_title_cn="测试图册封面",
        catalog_external_code="JD1NHT11T01B25C42SD",
        catalog_internal_code="1234567-JG001-TM",
        catalog_title_cn="测试图册目录",
        catalog_revision="B",
    )
    return DocContext(params=params, derived=derived, frames=[_make_frame(1)])


def test_ied_write_rows_with_bindings(temp_dir: Path) -> None:
    gen = IEDGenerator()
    ctx = _build_context()
    bindings = gen.spec.get_ied_bindings()
    output_xlsx = temp_dir / "IED计划.xlsx"

    gen._write_ied(
        template_path="documents_bin/IED计划模板文件.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    wb = load_workbook(output_xlsx)
    ws = wb[bindings.get("sheet", "IED导入模板 (修改)")]

    # 第2行：封面行
    assert ws["A2"].value == "A"
    assert ws["C2"].value == "图册"
    assert ws["D2"].value == "发布"
    assert ws["F2"].value == "1234567-JG001"
    assert ws["G2"].value == "JD1NHT11F01B25C42SD"
    assert ws["I2"].value == "1234567-JG001-FM"
    assert ws["AT2"].value == "66"
    assert ws["AY2"].value == "李四@A002"
    assert ws["BA2"].value == "2026-03-01"
    assert ws["BJ2"].value in ("", None)
    assert ws["BK2"].value == "非密"

    # 第4行：图纸行
    assert ws["G4"].value == "JD1NHT11001B25C42SD"
    assert ws["I4"].value == "1234567-JG001-001"
    assert ws["K4"].value == "图纸1"
