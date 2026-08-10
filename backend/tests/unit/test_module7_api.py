from __future__ import annotations

import json
import os
import sys
import time
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient as FastApiTestClient
from openpyxl import Workbook, load_workbook
from PIL import Image

from src.calculation_book.diagnostic_log import CalculationBookDiagnosticLog
from src.calculation_book.models import CalculationBookParams, ReinforcementSource
from src.config import MechanismSpecLoader, SpecLoader, reload_config
from src.models import BBox, FrameMeta, FrameRuntime, Job, JobStatus, JobType, PageInfo, SheetSet
from src.pipeline.shared_prep import SharedPrepArtifacts, SharedPrepService


class TestClient(FastApiTestClient):
    """Use the default administrator for tests of protected job endpoints."""

    def __enter__(self):
        client = super().__enter__()
        response = client.post(
            "/api/auth/login",
            json={"account_id": "hbjjswd", "password": "password"},
        )
        assert response.status_code == 200, response.text
        client.headers["Authorization"] = f"Bearer {response.json()['token']}"
        return client


class FakeFontPreflightService:
    def __init__(self) -> None:
        self.replacement_options = [
            {
                "label": "SimSun (simsun.ttc)",
                "value": "simsun.ttc",
                "family": "SimSun",
                "path": r"C:\Windows\Fonts\simsun.ttc",
                "kind": "ttf",
            },
            {
                "label": "simplex.shx (AutoCAD SHX)",
                "value": "simplex.shx",
                "family": "simplex",
                "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\simplex.shx",
                "kind": "shx",
            },
            {
                "label": "romans.shx (AutoCAD SHX)",
                "value": "romans.shx",
                "family": "romans",
                "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\romans.shx",
                "kind": "shx",
            },
        ]
        self.inspect_calls: list[dict[str, object]] = []
        self.list_requests: list[list[str]] = []

    def list_replacement_options(self, *, missing_kinds: list[str] | None = None) -> list[dict[str, str]]:
        requested = [str(kind or "").strip().lower() for kind in (missing_kinds or []) if str(kind or "").strip()]
        self.list_requests.append(requested)
        if not requested:
            return list(self.replacement_options)
        return [
            option
            for option in self.replacement_options
            if str(option.get("kind") or "").strip().lower() in requested
        ]

    def list_replacement_options_by_kind(
        self,
        *,
        missing_kinds: list[str] | None = None,
    ) -> dict[str, list[dict[str, str]]]:
        requested = [str(kind or "").strip().lower() for kind in (missing_kinds or []) if str(kind or "").strip()]
        return {
            kind: [
                option
                for option in self.replacement_options
                if str(option.get("kind") or "").strip().lower() == kind
            ]
            for kind in requested
        }

    def default_replacement_fonts(
        self,
        *,
        missing_kinds: list[str] | None = None,
        missing_fonts: list[dict[str, object]] | None = None,
    ) -> dict[str, str]:
        requested = [str(kind or "").strip().lower() for kind in (missing_kinds or []) if str(kind or "").strip()]
        defaults = {
            "ttf": "simsun.ttc",
            "shx": "simplex.shx",
            "bigfont": "romans.shx",
        }
        return {kind: defaults[kind] for kind in requested if kind in defaults}

    def validate_replacement_font(self, font_name: str, *, kind: str | None = None) -> bool:
        normalized_kind = str(kind or "").strip().lower()
        return any(
            option["value"] == font_name
            and (
                not normalized_kind
                or str(option.get("kind") or "").strip().lower() == normalized_kind
            )
            for option in self.replacement_options
        )

    def inspect_dwg(
        self,
        *,
        source_dwg: Path,
        replacement_policy: str = "none",
        replacement_font: str | None = None,
        replacement_fonts: dict[str, str] | None = None,
        font_compatibility_mode: bool = False,
        frames: list[object] | None = None,
        workspace_dir: Path | None = None,
        slot_runtime: dict[str, str] | None = None,
    ) -> dict[str, object]:
        self.inspect_calls.append(
            {
                "source_dwg": source_dwg,
                "replacement_policy": replacement_policy,
                "replacement_font": replacement_font,
                "replacement_fonts": replacement_fonts,
                "font_compatibility_mode": font_compatibility_mode,
                "frames": frames,
                "workspace_dir": workspace_dir,
                "slot_runtime": slot_runtime,
            }
        )
        filename = source_dwg.name
        if filename.startswith("empty-style"):
            if font_compatibility_mode:
                return {
                    "filename": filename,
                    "status": "ok",
                    "missing_fonts": [],
                    "detected_style_count": 4,
                    "missing_style_count": 0,
                    "font_replacement_applied": False,
                    "replacement_font": None,
                    "replacement_fonts": {},
                    "replaced_style_count": 0,
                    "font_compatibility_mode": True,
                    "empty_style_entity_replaced_count": 2,
                    "empty_style_target_regions_count": 3,
                    "empty_style_global_replaced_count": 0,
                }
            return {
                "filename": filename,
                "status": "ok",
                "missing_fonts": [],
                "detected_style_count": 4,
                "missing_style_count": 0,
                "font_replacement_applied": False,
                "replacement_font": None,
                "replacement_fonts": {},
                "replaced_style_count": 0,
            }

        if filename.startswith("patched-empty-style"):
            if font_compatibility_mode:
                return {
                    "filename": filename,
                    "status": "ok",
                    "missing_fonts": [],
                    "detected_style_count": 4,
                    "missing_style_count": 0,
                    "font_replacement_applied": False,
                    "replacement_font": None,
                    "replacement_fonts": {},
                    "replaced_style_count": 0,
                    "font_compatibility_mode": True,
                    "font_compatibility_required": True,
                    "empty_style_entity_replaced_count": 2,
                    "empty_style_style_patched_count": 1,
                    "empty_style_shared_skipped_count": 0,
                    "empty_style_shared_styles": [],
                    "empty_style_target_regions_count": 3,
                    "empty_style_global_replaced_count": 0,
                }
            return {
                "filename": filename,
                "status": "ok",
                "missing_fonts": [],
                "detected_style_count": 4,
                "missing_style_count": 0,
                "font_replacement_applied": False,
                "replacement_font": None,
                "replacement_fonts": {},
                "replaced_style_count": 0,
            }

        if filename.startswith("missing-font"):
            return {
                "filename": filename,
                "status": "missing_fonts",
                "missing_fonts": [
                    {
                        "style_name": "STYLE1",
                        "font_name": "missing.shx",
                        "bigfont_name": "",
                        "kind": "shx",
                        "used_in_block": True,
                    }
                ],
                "detected_style_count": 3,
                "missing_style_count": 1,
                "font_replacement_applied": replacement_policy == "replace_missing",
                "replacement_font": replacement_font,
                "replacement_fonts": replacement_fonts or {},
                "replaced_style_count": 1 if replacement_policy == "replace_missing" else 0,
            }

        return {
            "filename": filename,
            "status": "ok",
            "missing_fonts": [],
            "detected_style_count": 2,
            "missing_style_count": 0,
            "font_replacement_applied": False,
            "replacement_font": None,
            "replaced_style_count": 0,
        }


class FakeJobProcessor:
    def __call__(self, job: Job) -> None:
        job.work_dir = Path(job.work_dir or "")
        if job.job_type == JobType.AUDIT_REPLACE:
            mode = str(job.options.get("mode", "")).strip().lower()
            if mode == "replace":
                job.mark_running(stage="AUDIT_REPLACE")
                job.progress.message = "replacing"
                reports_dir = job.work_dir / "reports"
                reports_dir.mkdir(parents=True, exist_ok=True)
                report_xlsx = reports_dir / "report.xlsx"
                report_json = reports_dir / "report.json"
                replaced_dwg = job.work_dir / "replaced.dwg"
                workbook = Workbook()
                summary_sheet = workbook.active
                assert summary_sheet is not None
                summary_sheet.title = "Summary"
                summary_sheet.append(["source_filename", job.source_filename or "upload.dwg"])
                summary_sheet.append(["source_project_no", job.params.get("source_project_no", "")])
                summary_sheet.append(["target_project_no", job.params.get("target_project_no", "")])
                summary_sheet.append(["replacement_count", 2])
                workbook.save(report_xlsx)
                workbook.close()
                replaced_dwg.write_bytes(b"dwg-replaced")
                report_json.write_text(
                    json.dumps(
                        {
                            "replacement_count": 2,
                            "source_project_no": job.params.get("source_project_no", ""),
                            "target_project_no": job.params.get("target_project_no", ""),
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                job.artifacts.reports_dir = reports_dir
                job.artifacts.report_xlsx = report_xlsx
                job.artifacts.report_json = report_json
                job.artifacts.replaced_dwg = replaced_dwg
                job.progress.details["replacement_count"] = 2
                job.progress.details["factory_index_map"] = {
                    "applied": True,
                    "action_count": 1,
                    "report_json": str(reports_dir / "factory-index-map.json"),
                    "message": "",
                }
                job.mark_succeeded()
                return

            job.mark_running(stage="AUDIT_CHECK")
            job.progress.message = "auditing"
            reports_dir = job.work_dir / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            report_xlsx = reports_dir / "report.xlsx"
            report_json = reports_dir / "report.json"
            workbook = Workbook()
            summary_sheet = workbook.active
            assert summary_sheet is not None
            summary_sheet.title = "Summary"
            summary_sheet.append(["source_filename", job.source_filename or "upload.dwg"])
            summary_sheet.append(["project_no", job.project_no])
            summary_sheet.append(["findings_count", 2])
            summary_sheet.append(["affected_drawings_count", 1])
            workbook.save(report_xlsx)
            workbook.close()
            report_json.write_text(
                json.dumps(
                    {
                        "findings_count": 2,
                        "affected_drawings_count": 1,
                        "top_wrong_texts": ["2016", "JD"],
                        "top_internal_codes": ["1234567-JGS01-001"],
                        "finding_groups": [
                            {
                                "matched_text": "2016",
                                "count": 1,
                                "internal_codes": ["1234567-JGS01-001"],
                            },
                            {
                                "matched_text": "JD",
                                "count": 1,
                                "internal_codes": ["1234567-JGS01-001"],
                            },
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            preview_dir = job.work_dir / "output" / "preview"
            preview_dir.mkdir(parents=True, exist_ok=True)
            preview_pdf = preview_dir / "preview-annotated.pdf"
            preview_pdf.write_bytes(b"%PDF-annotated")
            job.artifacts.reports_dir = reports_dir
            job.artifacts.report_xlsx = report_xlsx
            job.artifacts.report_json = report_json
            job.artifacts.preview_pdf = preview_pdf
            job.artifacts.preview_mode = "annotated"
            job.progress.details["findings_count"] = 2
            job.progress.details["affected_drawings_count"] = 1
            job.progress.details["top_wrong_texts"] = ["2016", "JD"]
            job.progress.details["top_internal_codes"] = ["1234567-JGS01-001"]
            job.progress.details["workload"] = {
                "initial_workload_a1": 1.0,
                "final_workload_a1": 1.0,
                "one_review_factor": 1.0,
                "two_review_factor": 1.0,
                "three_review_factor": 1.0,
                "settlement_status": "pending",
                "settled_at": None,
                "contributor_entries": [],
            }
            job.progress.details["effective_workload"] = 1.0
            job.mark_succeeded()
            return

        if bool(job.options.get("split_only")):
            job.mark_running(stage="EXPORT_PDF_AND_DWG")
            job.progress.message = "split only"
            package_zip = job.work_dir / "package.zip"
            drawings_dir = job.work_dir / "output" / "drawings"
            drawings_dir.mkdir(parents=True, exist_ok=True)
            package_zip.write_bytes(b"PK\x03\x04split-only")
            (drawings_dir / "drawing-001.dwg").write_bytes(b"dwg")
            (drawings_dir / "drawing-001.pdf").write_bytes(b"pdf")
            (job.work_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "deliverable_outputs": {
                            "dwg_count": 1,
                            "pdf_count": 1,
                            "documents": [],
                            "drawings": [
                                {
                                    "name": "DRAW001 (20261RS-JGS65-001)",
                                    "internal_code": "20261RS-JGS65-001",
                                    "dwg_name": "drawing-001.dwg",
                                    "pdf_name": "drawing-001.pdf",
                                    "page_total": 1,
                                }
                            ],
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            job.artifacts.package_zip = package_zip
            job.artifacts.drawings_dir = drawings_dir
            job.progress.details["workload"] = {
                "initial_workload_a1": 0.25,
                "final_workload_a1": 0.25,
                "one_review_factor": 1.0,
                "two_review_factor": 1.0,
                "three_review_factor": 1.0,
                "settlement_status": "pending",
                "settled_at": None,
                "contributor_entries": [],
            }
            job.progress.details["effective_workload"] = 0.25
            job.mark_succeeded()
            return

        job.mark_running(stage="GENERATE_DOCS")
        job.progress.message = "processing"

        package_zip = job.work_dir / "package.zip"
        ied_xlsx = job.work_dir / "ied" / "IED计划.xlsx"
        drawings_dir = job.work_dir / "output" / "drawings"
        docs_dir = job.work_dir / "output" / "docs"
        drawings_dir.mkdir(parents=True, exist_ok=True)
        docs_dir.mkdir(parents=True, exist_ok=True)
        package_zip.write_bytes(b"PK\x03\x04test")
        include_ied_plan = job.params.get("include_ied_plan", True)
        if bool(include_ied_plan):
            ied_xlsx.parent.mkdir(parents=True, exist_ok=True)
            ied_xlsx.write_bytes(b"ied")
        (drawings_dir / "drawing-001.dwg").write_bytes(b"dwg")
        (drawings_dir / "drawing-001.pdf").write_bytes(b"pdf")
        (drawings_dir / "drawing-002.dwg").write_bytes(b"dwg")
        (drawings_dir / "drawing-002.pdf").write_bytes(b"pdf")
        preview_dir = job.work_dir / "output" / "preview"
        preview_dir.mkdir(parents=True, exist_ok=True)
        preview_pdf = preview_dir / "preview-plain.pdf"
        preview_pdf.write_bytes(b"%PDF-plain")
        (docs_dir / "cover.docx").write_text("cover", encoding="utf-8")
        (docs_dir / "cover.pdf").write_text("cover-pdf", encoding="utf-8")
        (docs_dir / "design.xlsx").write_text("design", encoding="utf-8")
        (job.work_dir / "manifest.json").write_text(
            json.dumps(
                {
                    "drawings": [
                        {
                            "name": "DRAW001 (20261RS-JGS65-001)",
                            "type": "single_frame",
                            "internal_code": "20261RS-JGS65-001",
                            "pdf_path": str(drawings_dir / "drawing-001.pdf"),
                            "dwg_path": str(drawings_dir / "drawing-001.dwg"),
                            "page_total": 1,
                            "flags": [
                                "PLOT_WINDOW_USED",
                                "PLOT_FROM_SOURCE_WINDOW",
                                "PAPER_SIZE_MISMATCH",
                                "PAPER_SIZE_AUTO_FIXED",
                            ],
                        },
                        {
                            "name": "DRAW002 (20261RS-JGS65-002)",
                            "type": "a4_sheet_set",
                            "internal_code": "20261RS-JGS65-002",
                            "pdf_path": str(drawings_dir / "drawing-002.pdf"),
                            "dwg_path": str(drawings_dir / "drawing-002.dwg"),
                            "page_total": 4,
                            "flags": ["PLOT_WINDOW_USED"],
                        },
                    ],
                    "deliverable_outputs": {
                        "dwg_count": 2,
                        "pdf_count": 2,
                        "documents": [
                            {"name": "cover.docx", "kind": "docx"},
                            {"name": "cover.pdf", "kind": "pdf"},
                            {"name": "design.xlsx", "kind": "xlsx"},
                        ],
                        "drawings": [
                            {
                                "name": "DRAW001 (20261RS-JGS65-001)",
                                "internal_code": "20261RS-JGS65-001",
                                "dwg_name": "drawing-001.dwg",
                                "pdf_name": "drawing-001.pdf",
                                "page_total": 1,
                            },
                            {
                                "name": "DRAW002 (20261RS-JGS65-002)",
                                "internal_code": "20261RS-JGS65-002",
                                "dwg_name": "drawing-002.dwg",
                                "pdf_name": "drawing-002.pdf",
                                "page_total": 4,
                            },
                        ],
                    },
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        job.artifacts.package_zip = package_zip
        job.artifacts.ied_xlsx = ied_xlsx if bool(include_ied_plan) else None
        job.artifacts.drawings_dir = drawings_dir
        job.artifacts.docs_dir = docs_dir
        job.artifacts.preview_pdf = preview_pdf
        job.artifacts.preview_mode = "plain"
        job.flags = [
            "[DRAW001 (20261RS-JGS65-001)] PLOT_WINDOW_USED",
            "[DRAW001 (20261RS-JGS65-001)] PLOT_FROM_SOURCE_WINDOW",
            "[DRAW001 (20261RS-JGS65-001)] PAPER_SIZE_MISMATCH",
            "[DRAW001 (20261RS-JGS65-001)] PAPER_SIZE_AUTO_FIXED",
        ]
        job.progress.details["workload"] = {
            "initial_workload_a1": 2.0,
            "final_workload_a1": 2.0,
            "one_review_factor": 1.0,
            "two_review_factor": 1.0,
            "three_review_factor": 1.0,
            "settlement_status": "pending",
            "settled_at": None,
            "contributor_entries": [],
        }
        job.progress.details["effective_workload"] = 2.0
        job.mark_succeeded()


def _configure_api_env(monkeypatch, tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[3]
    spec_path = repo_root / "documents" / "参数规范.yaml"
    runtime_spec_path = repo_root / "documents" / "参数规范_运行期.yaml"

    monkeypatch.setenv("FANBAN_SPEC_PATH", str(spec_path))
    monkeypatch.setenv("FANBAN_RUNTIME_SPEC_PATH", str(runtime_spec_path))
    monkeypatch.setenv("FANBAN_STORAGE_DIR", str(tmp_path / "storage"))
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_FILES", "3")
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_TOTAL_MB", "1")

    SpecLoader.clear_cache()
    reload_config()


def _create_client(monkeypatch, tmp_path: Path, processor=None, font_service=None) -> TestClient:
    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    app = create_app(
        job_processor=processor or FakeJobProcessor(),
        font_preflight_service=cast(Any, font_service),
    )
    return TestClient(app)


def _deliverable_params() -> dict[str, str]:
    return {
        "project_no": "2016",
        "classification": "非密",
        "subitem_name": "示例子项",
        "album_title_cn": "示例图册",
        "wbs_code": "WBS-001",
        "file_category": "1 总体文件",
        "ied_status": "编制",
        "ied_doc_type": "图册",
        "cover_variant": "通用",
    }


def _poll_job(client: TestClient, job_id: str, timeout_sec: float = 3.0) -> dict:
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        detail = client.get(f"/api/jobs/{job_id}")
        assert detail.status_code == 200
        payload = detail.json()
        if payload["status"] in {"succeeded", "failed"}:
            return payload
        time.sleep(0.05)
    raise AssertionError(f"job {job_id} did not finish within {timeout_sec}s")


def test_health_endpoint_returns_runtime_status(monkeypatch, tmp_path: Path) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.get("/api/system/health")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["ready"] is True
    assert "storage_writable" in payload
    assert "worker_alive" in payload
    assert "active_doc_jobs" in payload
    assert "active_total_jobs" in payload


def test_ping_endpoint_reports_http_process_without_runtime_health_probe(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        def _health_should_not_run() -> dict[str, object]:
            raise AssertionError("ping must not call runtime.health")

        client.app.state.runtime.health = _health_should_not_run
        response = client.get("/api/system/ping")

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert "server_time" in payload


def test_pipeline_job_processor_dispatches_audit_jobs(monkeypatch) -> None:
    from API.app.runtime import PipelineJobProcessor

    class DeliverableExecutor:
        def __init__(self) -> None:
            self.calls: list[str] = []

        def execute(self, job: Job) -> None:
            self.calls.append(job.job_id)

    class AuditExecutor:
        def __init__(self) -> None:
            self.calls: list[str] = []

        def execute(self, job: Job) -> None:
            self.calls.append(job.job_id)

    class ReplaceExecutor:
        def __init__(self) -> None:
            self.calls: list[str] = []

        def execute(self, job: Job) -> None:
            self.calls.append(job.job_id)

    deliverable = DeliverableExecutor()
    audit = AuditExecutor()
    replace = ReplaceExecutor()
    monkeypatch.setattr("API.app.runtime.PipelineExecutor", lambda: deliverable)
    monkeypatch.setattr("API.app.runtime.AuditCheckExecutor", lambda: audit)
    monkeypatch.setattr("API.app.runtime.AuditReplaceExecutor", lambda: replace)

    processor = PipelineJobProcessor()
    processor(Job(job_id="job-deliverable", job_type=JobType.DELIVERABLE, project_no="2016"))
    processor(
        Job(
            job_id="job-audit",
            job_type=JobType.AUDIT_REPLACE,
            project_no="2016",
            options={"mode": "check"},
        ),
    )
    processor(
        Job(
            job_id="job-replace",
            job_type=JobType.AUDIT_REPLACE,
            project_no="1818",
            options={"mode": "replace"},
        ),
    )

    assert deliverable.calls == ["job-deliverable"]
    assert audit.calls == ["job-audit"]
    assert replace.calls == ["job-replace"]


def test_pipeline_job_processor_exposes_slot_bound_phase_for_deliverables(monkeypatch) -> None:
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.runtime import PipelineJobProcessor

    class DeliverableExecutor:
        def __init__(self) -> None:
            self.phase_calls: list[str] = []

        def execute_slot_bound_phase(self, job: Job):
            self.phase_calls.append(job.job_id)
            return None

        def execute(self, job: Job) -> None:
            raise AssertionError("deliverable should use execute_slot_bound_phase")

    deliverable = DeliverableExecutor()
    monkeypatch.setattr("API.app.runtime.PipelineExecutor", lambda **_: deliverable)

    processor = PipelineJobProcessor()
    result = processor.execute_slot_bound_phase(
        Job(job_id="job-deliverable", job_type=JobType.DELIVERABLE, project_no="2016")
    )

    assert result is None
    assert deliverable.phase_calls == ["job-deliverable"]


def test_health_endpoint_allows_local_frontend_origin(monkeypatch, tmp_path: Path) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.get(
            "/api/system/health",
            headers={"Origin": "http://127.0.0.1:5175"},
        )

    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == "http://127.0.0.1:5175"


def test_health_reports_autocad_unready_when_runner_path_blank_and_autodetect_fails(
    monkeypatch,
    tmp_path: Path,
) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    monkeypatch.setenv("FANBAN_MODULE5_EXPORT__CAD_RUNNER__ACCORECONSOLE_EXE", "")
    monkeypatch.setenv("FANBAN_AUTOCAD_INSTALL_DIR", "")
    monkeypatch.setenv("FANBAN_AUTOCAD__CTB_PATH", "")
    reload_config()
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app import runtime as runtime_module
    from API.app.main import create_app

    monkeypatch.setattr(
        runtime_module,
        "resolve_autocad_paths",
        lambda configured_install_dir=None: SimpleNamespace(accoreconsole_exe=None),
        raising=False,
    )

    with TestClient(create_app(job_processor=FakeJobProcessor())) as client:
        response = client.get("/api/system/health")

    assert response.status_code == 200
    payload = response.json()
    assert payload["autocad_ready"] is False


def test_form_schema_returns_deliverable_fields_and_options(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.get("/api/meta/form-schema")

    assert response.status_code == 200
    payload = response.json()
    assert payload["upload_limits"]["max_files"] == 3
    assert payload["deliverable"]["sections"]
    assert "2016" in payload["audit_replace"]["project_options"]
    assert payload["audit_replace"]["batch_filename_identity_regex"]
    assert payload["audit_replace"]["factory_index_maps"]["source_variant_options"]["2016"] == [
        "1",
        "2",
    ]
    assert payload["audit_replace"]["factory_index_maps"]["target_variant_options"]["1916"] == [
        "3",
        "4",
    ]
    assert payload["audit_replace"]["project_units"]["1915"] == ["0", "1", "2", "7", "9"]
    assert payload["audit_replace"]["source_unit_options"]["2016"] == [
        {"value": "0", "label": "0号机组/岛"},
        {"value": "1", "label": "1号机组/岛"},
        {"value": "2", "label": "2号机组/岛"},
        {"value": "7", "label": "7号机组/岛"},
        {"value": "9", "label": "9号机组/岛"},
    ]
    assert payload["audit_replace"]["source_unit_options"]["1916"] == [
        {"value": "0", "label": "0号机组/岛"},
        {"value": "3", "label": "3号机组/岛"},
        {"value": "4", "label": "4号机组/岛"},
        {"value": "7", "label": "7号机组/岛"},
        {"value": "9", "label": "9号机组/岛"},
    ]
    assert payload["audit_replace"]["target_unit_options"]["1915"] == [
        {"value": "0", "label": "0号机组/岛"},
        {"value": "1", "label": "1号机组/岛"},
        {"value": "2", "label": "2号机组/岛"},
        {"value": "7", "label": "7号机组/岛"},
        {"value": "9", "label": "9号机组/岛"},
    ]
    assert payload["audit_replace"]["target_unit_options"]["2016"] == [
        {"value": "0", "label": "0号机组/岛"},
        {"value": "1", "label": "1号机组/岛"},
        {"value": "2", "label": "2号机组/岛"},
        {"value": "7", "label": "7号机组/岛"},
        {"value": "9", "label": "9号机组/岛"},
    ]
    assert payload["audit_check"]["unit_consistency"]["enabled"] is True
    assert payload["audit_check"]["unit_consistency"]["project_units"]["2016"] == [
        "0",
        "1",
        "2",
        "7",
        "9",
    ]
    assert payload["audit_check"]["unit_consistency"]["project_units"]["1916"] == [
        "0",
        "3",
        "4",
        "7",
        "9",
    ]

    project_section = next(
        section for section in payload["deliverable"]["sections"] if section["id"] == "project"
    )
    project_no = next(field for field in project_section["fields"] if field["key"] == "project_no")
    unit_no = next(field for field in project_section["fields"] if field["key"] == "unit_no")
    file_category = next(
        field
        for section in payload["deliverable"]["sections"]
        for field in section["fields"]
        if field["key"] == "file_category"
    )
    ied_design_type = next(
        field
        for section in payload["deliverable"]["sections"]
        for field in section["fields"]
        if field["key"] == "ied_design_type"
    )
    ied_responsible_unit = next(
        field
        for section in payload["deliverable"]["sections"]
        for field in section["fields"]
        if field["key"] == "ied_responsible_unit"
    )
    ied_person_qual_category = next(
        field
        for section in payload["deliverable"]["sections"]
        for field in section["fields"]
        if field["key"] == "ied_person_qual_category"
    )
    include_ied_plan = next(
        field
        for section in payload["deliverable"]["sections"]
        for field in section["fields"]
        if field["key"] == "include_ied_plan"
    )

    assert "2016" in project_no["options"]
    assert "1915" in project_no["options"]
    assert project_no["required"] is False
    assert "DWG" in project_no["desc"]
    assert "2016" in project_no["desc"]
    assert unit_no["required"] is False
    assert unit_no["options"] == ["0", "1", "2", "3", "4", "5", "6", "7", "9"]
    assert "1 总体文件" in file_category["options"]
    assert ied_design_type["required_when"] == "ied_status == '发布'"
    assert ied_design_type["type"] == "combobox"
    assert ied_design_type["allow_custom_input"] is True
    assert ied_design_type["filterable"] is True
    assert ied_design_type["options"][:3] == ["安装技术要求", "常规岛厂房设计", "初步设计"]
    assert ied_responsible_unit["required_when"] == "ied_status == '发布'"
    assert ied_responsible_unit["type"] == "combobox"
    assert ied_responsible_unit["allow_custom_input"] is True
    assert ied_responsible_unit["filterable"] is True
    assert ied_responsible_unit["options"][:3] == [
        "河北分公司-核工程研究设计所-电仪室",
        "公用系统所-水工工艺二室",
        "河北分公司-电气自动化所-仪控一室",
    ]
    assert ied_person_qual_category["options"] == [
        "非核安全物项",
        "非核压力容器",
        "非核压力管道",
        "一般核安全物项-军工",
        "一般核安全物项-民用",
        "核安全承压机械设备-军工-甲级",
        "核安全承压机械设备-军工-乙级",
        "核安全承压机械设备-民用-甲级",
        "核安全承压机械设备-民用-乙级",
    ]
    assert include_ied_plan["type"] == "checkbox"
    assert include_ied_plan["default"] is True


def test_create_batch_requires_ied_publish_fields_when_status_is_publish(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params["ied_status"] = "发布"

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["ied_submitted_plan_date"] == ["required"]
    assert payload["detail"]["param_errors"]["ied_publish_plan_date"] == ["required"]
    assert payload["detail"]["param_errors"]["ied_external_plan_date"] == ["required"]
    assert payload["detail"]["param_errors"]["ied_design_type"] == ["required"]
    assert payload["detail"]["param_errors"]["ied_chief_designer"] == ["required"]
    assert payload["detail"]["param_errors"]["ied_responsible_unit"] == ["required"]


def test_create_batch_rejects_non_dwg_upload(monkeypatch, tmp_path: Path) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(_deliverable_params(), ensure_ascii=False)},
            files=[("files[]", ("bad.txt", b"nope", "text/plain"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["upload_errors"]["files"] == ["only .dwg files are allowed"]


def test_preflight_fonts_returns_missing_fonts_and_replacement_options(monkeypatch, tmp_path: Path) -> None:
    font_service = FakeFontPreflightService()

    with _create_client(monkeypatch, tmp_path, font_service=font_service) as client:
        response = client.post(
            "/api/jobs/preflight-fonts",
            files=[
                ("files[]", ("missing-font.dwg", b"dwg-a", "application/acad")),
                ("files[]", ("ok-font.dwg", b"dwg-b", "application/acad")),
            ],
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requires_confirmation"] is True
    assert payload["replacement_options"] == [
        {
            "label": "simplex.shx (AutoCAD SHX)",
            "value": "simplex.shx",
            "family": "simplex",
            "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\simplex.shx",
            "kind": "shx",
        },
        {
            "label": "romans.shx (AutoCAD SHX)",
            "value": "romans.shx",
            "family": "romans",
            "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\romans.shx",
            "kind": "shx",
        },
    ]
    assert payload["replacement_options_by_kind"] == {
        "shx": [
            {
                "label": "simplex.shx (AutoCAD SHX)",
                "value": "simplex.shx",
                "family": "simplex",
                "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\simplex.shx",
                "kind": "shx",
            },
            {
                "label": "romans.shx (AutoCAD SHX)",
                "value": "romans.shx",
                "family": "romans",
                "path": r"D:\Program Files\AUTOCAD\AutoCAD 2022\Fonts\romans.shx",
                "kind": "shx",
            },
        ]
    }
    assert payload["default_replacement_fonts"] == {"shx": "simplex.shx"}
    assert font_service.list_requests[-1] == ["shx"]
    assert payload["default_replacement_font"] == "simplex.shx"
    assert payload["files"] == [
        {
            "filename": "missing-font.dwg",
            "status": "missing_fonts",
            "missing_fonts": [
                {
                    "style_name": "STYLE1",
                    "font_name": "missing.shx",
                    "bigfont_name": "",
                    "kind": "shx",
                    "used_in_block": True,
                }
            ],
            "detected_style_count": 3,
            "missing_style_count": 1,
            "font_replacement_applied": False,
            "replacement_font": None,
            "replacement_fonts": {},
            "replaced_style_count": 0,
        },
        {
            "filename": "ok-font.dwg",
            "status": "ok",
            "missing_fonts": [],
            "detected_style_count": 2,
            "missing_style_count": 0,
            "font_replacement_applied": False,
            "replacement_font": None,
            "replaced_style_count": 0,
        },
    ]


def test_preflight_fonts_does_not_block_system_ping(monkeypatch, tmp_path: Path) -> None:
    import threading

    class SlowFontPreflightService(FakeFontPreflightService):
        def __init__(self) -> None:
            super().__init__()
            self.started = threading.Event()

        def inspect_dwg(self, **kwargs: Any) -> dict[str, object]:
            self.started.set()
            time.sleep(0.45)
            return {
                "filename": Path(kwargs["source_dwg"]).name,
                "status": "missing_fonts",
                "missing_fonts": [
                    {
                        "style_name": "STYLE1",
                        "font_name": "missing.shx",
                        "bigfont_name": "",
                        "kind": "shx",
                        "used_in_block": True,
                    }
                ],
                "detected_style_count": 1,
                "missing_style_count": 1,
                "font_replacement_applied": False,
                "replacement_font": None,
                "replacement_fonts": {},
                "replaced_style_count": 0,
            }

    font_service = SlowFontPreflightService()
    result: dict[str, Any] = {}

    with _create_client(monkeypatch, tmp_path, font_service=font_service) as client:
        def run_preflight() -> None:
            result["response"] = client.post(
                "/api/jobs/preflight-fonts",
                files=[("files[]", ("slow-font.dwg", b"dwg-a", "application/acad"))],
            )

        thread = threading.Thread(target=run_preflight, name="slow-preflight-test")
        thread.start()
        assert font_service.started.wait(timeout=1.0)

        ping_started = time.perf_counter()
        ping_response = client.get("/api/system/ping")
        ping_elapsed = time.perf_counter() - ping_started

        thread.join(timeout=2.0)
        assert not thread.is_alive()

    preflight_response = result["response"]
    assert preflight_response.status_code == 200
    assert ping_response.status_code == 200
    assert ping_elapsed < 0.25


def test_preflight_fonts_keeps_file_results_when_replacement_inventory_fails(
    monkeypatch,
    tmp_path: Path,
) -> None:
    class FailingReplacementInventoryFontService(FakeFontPreflightService):
        def list_replacement_options(self, *, missing_kinds: list[str] | None = None):  # noqa: ANN201
            raise RuntimeError("font inventory unavailable")

        def list_replacement_options_by_kind(  # noqa: ANN201
            self,
            *,
            missing_kinds: list[str] | None = None,
        ):
            raise RuntimeError("font inventory unavailable")

        def default_replacement_fonts(  # noqa: ANN201
            self,
            *,
            missing_kinds: list[str] | None = None,
            missing_fonts: list[dict[str, object]] | None = None,
        ):
            raise RuntimeError("font inventory unavailable")

    with _create_client(
        monkeypatch,
        tmp_path,
        font_service=FailingReplacementInventoryFontService(),
    ) as client:
        response = client.post(
            "/api/jobs/preflight-fonts",
            files=[("files[]", ("missing-font.dwg", b"dwg-a", "application/acad"))],
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requires_confirmation"] is True
    assert payload["replacement_options"] == []
    assert payload["replacement_options_by_kind"] == {}
    assert payload["default_replacement_fonts"] == {}
    assert payload["files"][0]["status"] == "missing_fonts"


def test_preflight_fonts_reports_empty_style_compatibility_risk(
    monkeypatch,
    tmp_path: Path,
) -> None:
    font_service = FakeFontPreflightService()

    with _create_client(monkeypatch, tmp_path, font_service=font_service) as client:
        runtime = client.app.state.runtime

        class FakeOda:
            def dwg_to_dxf(self, dwg_path: Path, output_dir: Path) -> Path:
                output_dir.mkdir(parents=True, exist_ok=True)
                dxf_path = output_dir / f"{dwg_path.stem}.dxf"
                dxf_path.write_text("0\nEOF\n", encoding="utf-8")
                return dxf_path

        class FakeFrameDetector:
            def __init__(self) -> None:
                self.project_no: str | None = None

            def set_project_no(self, project_no: str | None) -> None:
                self.project_no = project_no

            def detect_frames(self, dxf_path: Path) -> list[object]:
                return [SimpleNamespace(frame_id="frame-1")]

        runtime.font_preflight_oda = FakeOda()
        runtime.font_preflight_frame_detector = FakeFrameDetector()

        response = client.post(
            "/api/jobs/preflight-fonts",
            files=[("files[]", ("empty-style-font.dwg", b"dwg-a", "application/acad"))],
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requires_confirmation"] is True
    assert payload["files"][0]["status"] == "ok"
    assert payload["files"][0]["font_compatibility_required"] is True
    assert payload["files"][0]["empty_style_entity_replaced_count"] == 2
    assert payload["files"][0]["empty_style_target_regions_count"] == 3
    assert [call["font_compatibility_mode"] for call in font_service.inspect_calls] == [False, True]


def test_preflight_fonts_reports_empty_style_patch(
    monkeypatch,
    tmp_path: Path,
) -> None:
    font_service = FakeFontPreflightService()

    with _create_client(monkeypatch, tmp_path, font_service=font_service) as client:
        runtime = client.app.state.runtime

        class FakeOda:
            def dwg_to_dxf(self, dwg_path: Path, output_dir: Path) -> Path:
                output_dir.mkdir(parents=True, exist_ok=True)
                dxf_path = output_dir / f"{dwg_path.stem}.dxf"
                dxf_path.write_text("0\nEOF\n", encoding="utf-8")
                return dxf_path

        class FakeFrameDetector:
            def set_project_no(self, project_no: str | None) -> None:
                self.project_no = project_no

            def detect_frames(self, dxf_path: Path) -> list[object]:
                return [SimpleNamespace(frame_id="frame-1")]

        runtime.font_preflight_oda = FakeOda()
        runtime.font_preflight_frame_detector = FakeFrameDetector()

        response = client.post(
            "/api/jobs/preflight-fonts",
            files=[
                (
                    "files[]",
                    ("patched-empty-style-font.dwg", b"dwg-a", "application/acad"),
                )
            ],
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requires_confirmation"] is True
    assert payload["files"][0]["status"] == "ok"
    assert payload["files"][0]["font_compatibility_required"] is True
    assert payload["files"][0]["empty_style_entity_replaced_count"] == 2
    assert payload["files"][0]["empty_style_style_patched_count"] == 1
    assert payload["files"][0]["empty_style_shared_skipped_count"] == 0
    assert payload["files"][0]["empty_style_shared_styles"] == []


def test_preflight_fonts_uses_ascii_working_copy_for_non_ascii_upload_names(
    monkeypatch,
    tmp_path: Path,
) -> None:
    font_service = FakeFontPreflightService()

    with _create_client(monkeypatch, tmp_path, font_service=font_service) as client:
        response = client.post(
            "/api/jobs/preflight-fonts",
            files=[("files[]", ("20261NH-JGS51-B合并版.dwg", b"dwg-a", "application/acad"))],
        )

    assert response.status_code == 200
    assert len(font_service.inspect_calls) == 1
    source_path = font_service.inspect_calls[0]["source_dwg"]
    assert isinstance(source_path, Path)
    assert source_path.suffix.lower() == ".dwg"
    assert source_path.name != "20261NH-JGS51-B合并版.dwg"
    assert all(ord(ch) < 128 for ch in source_path.name)


def test_create_batch_preserves_source_filename_but_stores_ascii_upload_copy(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(_deliverable_params(), ensure_ascii=False)},
            files=[("files[]", ("20261NH-JGS51-B合并版.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        job_id = payload["jobs"][0]["job_id"]
        assert payload["jobs"][0]["source_filename"] == "20261NH-JGS51-B合并版.dwg"

        detail = _poll_job(client, job_id)
        assert detail["workload"]["initial_workload_a1"] == 2.0
        assert detail["workload"]["final_workload_a1"] == 2.0
        assert detail["effective_workload"] == 2.0
        assert detail["source_filename"] == "20261NH-JGS51-B合并版.dwg"

    uploads_dir = tmp_path / "storage" / "jobs" / job_id / "uploads"
    stored_files = list(uploads_dir.iterdir())
    assert len(stored_files) == 1
    assert stored_files[0].suffix.lower() == ".dwg"
    assert stored_files[0].name != "20261NH-JGS51-B合并版.dwg"
    assert all(ord(ch) < 128 for ch in stored_files[0].name)


def test_create_batch_preserves_cjk_params_in_job_payload(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("A01.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        job_id = payload["jobs"][0]["job_id"]
        _poll_job(client, job_id)

    job_payload = json.loads(
        (tmp_path / "storage" / "jobs" / job_id / "job.json").read_text(encoding="utf-8")
    )

    assert job_payload["params"]["classification"] == params["classification"]
    assert job_payload["params"]["subitem_name"] == params["subitem_name"]
    assert job_payload["params"]["album_title_cn"] == params["album_title_cn"]
    assert job_payload["params"]["file_category"] == params["file_category"]
    assert job_payload["params"]["ied_status"] == params["ied_status"]
    assert job_payload["params"]["ied_doc_type"] == params["ied_doc_type"]
    assert job_payload["params"]["cover_variant"] == params["cover_variant"]


def test_create_batch_rejects_replace_missing_without_replacement_font(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path, font_service=FakeFontPreflightService()) as client:
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(
                    {
                        **_deliverable_params(),
                        "font_replace_policy": "replace_missing",
                    },
                    ensure_ascii=False,
                )
            },
            files=[("files[]", ("missing-font.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["font_replacement_font"] == [
        "required_when_font_replace_policy_is_replace_missing"
    ]


def test_create_batch_accepts_kind_specific_font_replacements(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path, font_service=FakeFontPreflightService()) as client:
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(
                    {
                        **_deliverable_params(),
                        "font_replace_policy": "replace_missing",
                        "font_replacement_fonts": {"ttf": "simsun.ttc"},
                    },
                    ensure_ascii=False,
                )
            },
            files=[("files[]", ("missing-font.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201


def test_create_batch_rejects_missing_required_param(monkeypatch, tmp_path: Path) -> None:
    params = _deliverable_params()
    params.pop("album_title_cn")

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["album_title_cn"] == ["required"]


def test_create_batch_split_only_skips_document_param_validation(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps({"project_no": ""}, ensure_ascii=False),
                "split_only": "true",
            },
            files=[("files[]", ("20261RS-JGS65.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        job_id = payload["jobs"][0]["job_id"]
        assert payload["jobs"][0]["job_mode"] == "split_only"
        assert payload["jobs"][0]["task_role"] == "仅拆图"
        detail = _poll_job(client, job_id)

    job_payload = json.loads(
        (tmp_path / "storage" / "jobs" / job_id / "job.json").read_text(encoding="utf-8")
    )
    assert job_payload["project_no"] == "2026"
    assert job_payload["options"]["split_only"] is True
    assert detail["artifacts"]["package_available"] is True
    assert detail["workload"]["initial_workload_a1"] == 0.25
    assert detail["workload"]["final_workload_a1"] == 0.25
    assert detail["effective_workload"] == 0.25


def test_create_batch_infers_project_no_from_uploaded_filename_when_blank(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params["project_no"] = ""

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("2026-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["jobs"][0]["project_no"] == "2026"


def test_create_batch_uses_inferred_project_no_for_required_when_validation(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params["project_no"] = ""

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("1818-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["subitem_name_en"] == ["required"]
    assert payload["detail"]["param_errors"]["album_title_en"] == ["required"]


def test_create_batch_falls_back_to_default_project_no_when_not_inferable(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params.pop("project_no")

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("sample-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["jobs"][0]["project_no"] == "2016"


def test_create_audit_check_rejects_when_project_no_cannot_be_inferred(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = {"project_no": ""}

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps(params, ensure_ascii=False),
            },
            files=[("files[]", ("sample-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["project_no"] == ["required_for_audit_check"]


def test_create_audit_check_requires_unit_no_for_unit_consistency_project(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps({"project_no": "2026"}, ensure_ascii=False),
            },
            files=[("files[]", ("sample-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["detail"]["param_errors"]["unit_no"] == ["required_for_unit_consistency"]


def test_create_audit_check_infers_unit_no_from_filename(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps({"project_no": "2026"}, ensure_ascii=False),
            },
            files=[("files[]", ("20261NS-JGS01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["jobs"][0]["project_no"] == "2026"


def test_create_batch_with_audit_check_infers_project_and_unit_per_upload(
    monkeypatch,
    tmp_path: Path,
) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    with TestClient(
        create_app(
            job_processor=FakeJobProcessor(),
            shared_prep_service=FakeSharedPrepService(),
        ),
    ) as client:
        params = _deliverable_params()
        params["project_no"] = ""
        params["unit_no"] = ""
        params["subitem_name_en"] = "NR Building"
        params["album_title_en"] = "Test Album"
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(params, ensure_ascii=False),
                "run_audit_check": "true",
            },
            files=[
                ("files[]", ("20261NS-JGS01.dwg", b"dwg", "application/acad")),
                ("files[]", ("出图版--18185NR-JGS50-A.dwg", b"dwg", "application/acad")),
            ],
        )

        assert response.status_code == 201, response.json()
        payload = response.json()
        projects = [job["project_no"] for job in payload["jobs"]]
        assert projects == ["2026", "1818"]

        stored_params = []
        for group in payload["jobs"]:
            group_json = json.loads(
                (tmp_path / "storage" / "groups" / group["job_id"] / "group.json").read_text(
                    encoding="utf-8"
                )
            )
            child_job_id = group_json["child_job_ids"][0]
            child_json = (
                tmp_path / "storage" / "jobs" / child_job_id / "job.json"
            ).read_text(encoding="utf-8")
            stored_params.append(json.loads(child_json)["params"])

    assert [(item["project_no"], item["unit_no"]) for item in stored_params] == [
        ("2026", "1"),
        ("1818", "5"),
    ]


def test_create_audit_check_accepts_unlisted_unit_no_for_configured_project(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps(
                    {"project_no": "1907", "unit_no": "7"},
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("19077NH-JGS01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["jobs"][0]["project_no"] == "1907"


def test_create_audit_check_processes_job_and_exposes_report_download(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps({"project_no": "2016", "unit_no": "1"}, ensure_ascii=False),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 1
        assert payload["jobs"][0]["task_kind"] == "audit_check"
        assert payload["jobs"][0]["job_mode"] == "check"

        job_id = payload["jobs"][0]["job_id"]
        detail = _poll_job(client, job_id)
        assert detail["status"] == "succeeded"
        assert detail["artifacts"]["report_available"] is True
        assert detail["artifacts"]["package_available"] is False
        assert detail["artifacts"]["preview_available"] is True
        assert detail["artifacts"]["preview_mode"] == "annotated"
        assert detail["artifacts"]["preview_download_url"] == f"/api/jobs/{job_id}/download/preview"
        assert detail["findings_count"] == 2
        assert detail["affected_drawings_count"] == 1
        assert detail["workload"]["initial_workload_a1"] == 1.0
        assert detail["workload"]["final_workload_a1"] == 1.0
        assert detail["effective_workload"] == 1.0
        assert detail["top_wrong_texts"] == ["2016", "JD"]
        assert detail["top_internal_codes"] == ["1234567-JGS01-001"]
        assert detail["finding_groups"] == [
            {
                "matched_text": "2016",
                "count": 1,
                "internal_codes": ["1234567-JGS01-001"],
            },
            {
                "matched_text": "JD",
                "count": 1,
                "internal_codes": ["1234567-JGS01-001"],
            },
        ]
        assert detail["slot_id"] is not None
        assert detail["profile_arg"] is not None
        assert detail["plot_style_key"] == "red_wider"
        assert detail["plot_resource_mode"] == "slot_private_with_shared_mirror"

        report_download = client.get(f"/api/jobs/{job_id}/download/report")
        assert report_download.status_code == 200
        workbook = load_workbook(filename=BytesIO(report_download.content))
        assert workbook.sheetnames[0] == "Summary"

        preview_download = client.get(f"/api/jobs/{job_id}/download/preview")
        assert preview_download.status_code == 200
        assert preview_download.content == b"%PDF-annotated"


def test_audit_check_detail_preserves_standard_review_finding_group(
    monkeypatch,
    tmp_path: Path,
) -> None:
    class FakeStandardReviewProcessor(FakeJobProcessor):
        def __call__(self, job: Job) -> None:
            super().__call__(job)
            if job.job_type != JobType.AUDIT_REPLACE or str(job.options.get("mode", "")).lower() != "check":
                return
            assert job.artifacts.report_json is not None
            payload = json.loads(job.artifacts.report_json.read_text(encoding="utf-8"))
            payload["finding_groups"].append(
                {
                    "matched_text": "GB 51058-2011",
                    "count": 1,
                    "internal_codes": ["18185NF-JGS19-003"],
                    "category": "规范审查",
                    "context_kind": "standard_review_year",
                    "issue_type": "year_mismatch",
                    "summary": "标准号年限不一致：GB 51058-2011 应为 GB 51058-2014",
                    "details": [
                        "实际标准号：GB 51058-2011",
                        "期望标准号：GB 51058-2014",
                        "期望标准名称：核电厂抗震设计标准",
                    ],
                }
            )
            job.artifacts.report_json.write_text(
                json.dumps(payload, ensure_ascii=False),
                encoding="utf-8",
            )

    with _create_client(monkeypatch, tmp_path, processor=FakeStandardReviewProcessor()) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps({"project_no": "2016", "unit_no": "1"}, ensure_ascii=False),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        job_id = response.json()["jobs"][0]["job_id"]
        detail = _poll_job(client, job_id)
        assert detail["finding_groups"][-1] == {
            "matched_text": "GB 51058-2011",
            "count": 1,
            "internal_codes": ["18185NF-JGS19-003"],
            "category": "规范审查",
            "context_kind": "standard_review_year",
            "issue_type": "year_mismatch",
            "summary": "标准号年限不一致：GB 51058-2011 应为 GB 51058-2014",
            "details": [
                "实际标准号：GB 51058-2011",
                "期望标准号：GB 51058-2014",
                "期望标准名称：核电厂抗震设计标准",
            ],
        }


def test_create_audit_check_reuses_explicit_batch_id_when_provided(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "check",
                "params_json": json.dumps(
                    {"project_no": "2016", "unit_no": "1", "batch_id": "batch-shared-1"},
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch_id"] == "batch-shared-1"
    assert payload["jobs"][0]["batch_id"] == "batch-shared-1"


def test_create_audit_replace_rejects_missing_or_same_project_pair(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        missing_source = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {"source_project_no": "", "target_project_no": "1818", "run_deliverable": False},
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("1818-A01.dwg", b"dwg", "application/acad"))],
        )
        missing_target = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {"source_project_no": "2016", "target_project_no": "", "run_deliverable": False},
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )
        same_pair = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {"source_project_no": "2016", "target_project_no": "2016", "run_deliverable": False},
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

    assert missing_source.status_code == 422
    assert missing_source.json()["detail"]["param_errors"]["source_project_no"] == ["required_for_replace"]
    assert missing_target.status_code == 422
    assert missing_target.json()["detail"]["param_errors"]["target_project_no"] == ["required_for_replace"]
    assert same_pair.status_code == 422
    assert same_pair.json()["detail"]["param_errors"]["target_project_no"] == [
        "must_differ_from_source_project_no"
    ]


def test_create_audit_replace_allows_same_project_when_unit_changes(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {
                        "source_project_no": "2026",
                        "source_island_no": "1",
                        "target_project_no": "2026",
                        "target_island_no": "2",
                        "unit_factory_codes": ["RB"],
                        "run_deliverable": False,
                    },
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("20261RB-JGS11-A.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["jobs"][0]["task_kind"] == "audit_replace"


def test_create_audit_replace_rejects_mixed_source_projects_and_factory_codes(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {
                        "source_project_no": "2016",
                        "source_island_no": "1",
                        "target_project_no": "2026",
                        "target_island_no": "2",
                        "unit_factory_codes": ["RC"],
                        "run_deliverable": False,
                    },
                    ensure_ascii=False,
                ),
            },
            files=[
                ("files[]", ("20161RC-JGS01-A.dwg", b"dwg-1", "application/acad")),
                ("files[]", ("18185RB-JGS02-A.dwg", b"dwg-2", "application/acad")),
            ],
        )

    assert response.status_code == 422
    errors = response.json()["detail"]["param_errors"]
    assert errors["source_project_no"] == ["mixed_source_projects"]
    assert errors["source_island_no"] == ["mixed_source_units"]
    assert errors["unit_factory_codes"] == ["mixed_factory_codes"]


def test_create_audit_replace_requires_exactly_one_factory_code(
    monkeypatch,
    tmp_path: Path,
) -> None:
    for factory_codes, expected_error in (
        ([], "required_for_replace"),
        (["RC", "RB"], "single_factory_code_required"),
    ):
        with _create_client(monkeypatch, tmp_path) as client:
            response = client.post(
                "/api/jobs/audit-replace",
                data={
                    "mode": "replace",
                    "params_json": json.dumps(
                        {
                            "source_project_no": "2016",
                            "source_island_no": "1",
                            "target_project_no": "2026",
                            "target_island_no": "2",
                            "unit_factory_codes": factory_codes,
                            "run_deliverable": False,
                        },
                        ensure_ascii=False,
                    ),
                },
                files=[("files[]", ("20161RC-JGS01-A.dwg", b"dwg", "application/acad"))],
            )

        assert response.status_code == 422
        errors = response.json()["detail"]["param_errors"]
        assert errors["unit_factory_codes"] == [expected_error]


def test_create_audit_replace_processes_job_without_deliverable(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {
                        "source_project_no": "2016",
                        "source_island_no": "1",
                        "target_project_no": "1916",
                        "target_island_no": "3",
                        "unit_factory_codes": ["RC"],
                        "run_deliverable": False,
                    },
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 1
        assert payload["jobs"][0]["task_kind"] == "audit_replace"
        assert payload["jobs"][0]["job_mode"] == "replace"

        job_id = payload["jobs"][0]["job_id"]
        detail = _poll_job(client, job_id)
        assert detail["status"] == "succeeded"
        assert detail["artifacts"]["report_available"] is True
        assert detail["artifacts"]["replaced_dwg_available"] is True
        assert detail["artifacts"]["package_available"] is False
        assert detail["artifacts"]["replaced_dwg_download_url"] == f"/api/jobs/{job_id}/download/replaced"
        assert detail["replace_summary"]["replacement_count"] == 2
        assert detail["replace_summary"]["source_project_no"] == "2016"
        assert detail["replace_summary"]["source_island_no"] == "1"
        assert detail["replace_summary"]["target_project_no"] == "1916"
        assert detail["replace_summary"]["target_island_no"] == "3"
        assert detail["factory_index_map"]["applied"] is True
        assert detail["factory_index_map"]["action_count"] == 1

        replaced_download = client.get(f"/api/jobs/{job_id}/download/replaced")
        assert replaced_download.status_code == 200
        assert replaced_download.content == b"dwg-replaced"


def test_meta_remembers_audit_replace_factory_codes_in_mechanism_yaml(
    monkeypatch,
    tmp_path: Path,
) -> None:
    mechanism_spec = tmp_path / "documents" / "参数规范-3.yaml"
    mechanism_spec.parent.mkdir(parents=True, exist_ok=True)
    mechanism_spec.write_text(
        "schema_version: '1.0'\n"
        "backend_mechanism:\n"
        "  audit_replace:\n"
        "    unit_factory_codes:\n"
        "      - RC\n",
        encoding="utf-8",
    )
    monkeypatch.setenv("FANBAN_MECHANISM_SPEC_PATH", str(mechanism_spec))
    MechanismSpecLoader.clear_cache()

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/meta/audit-replace/factory-codes",
            json={"codes": ["hl", "RC", "16mm"]},
        )
        schema_response = client.get("/api/meta/form-schema")

    assert response.status_code == 200
    assert response.json()["factory_codes"] == ["RC", "HL"]
    assert schema_response.status_code == 200
    assert schema_response.json()["audit_replace"]["unit_factory_codes"] == ["RC", "HL"]
    assert "HL" in mechanism_spec.read_text(encoding="utf-8")


def test_create_audit_replace_creates_group_when_run_deliverable_enabled(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params.pop("project_no")
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {
                        "source_project_no": "2016",
                        "source_island_no": "1",
                        "target_project_no": "1818",
                        "unit_factory_codes": ["RC"],
                        "run_deliverable": True,
                        "deliverable_params": params,
                    },
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

    assert response.status_code == 201
    payload = response.json()
    assert len(payload["jobs"]) == 1
    group_summary = payload["jobs"][0]
    assert group_summary["is_group"] is True
    assert len(group_summary["child_job_ids"]) == 2


def test_create_audit_replace_with_deliverable_runs_replace_before_deliverable(
    monkeypatch,
    tmp_path: Path,
) -> None:
    class ReplaceThenDeliverableProcessor:
        def __call__(self, job: Job) -> None:
            job.work_dir = Path(job.work_dir or "")
            if job.job_type == JobType.AUDIT_REPLACE and str(job.options.get("mode")) == "replace":
                job.mark_running(stage="AUDIT_REPLACE")
                replaced_dwg = job.work_dir / "replaced.dwg"
                replaced_dwg.write_bytes(b"replaced-dwg")
                reports_dir = job.work_dir / "reports"
                reports_dir.mkdir(parents=True, exist_ok=True)
                report_xlsx = reports_dir / "report.xlsx"
                report_json = reports_dir / "report.json"
                workbook = Workbook()
                summary_sheet = workbook.active
                assert summary_sheet is not None
                summary_sheet.title = "Summary"
                summary_sheet.append(["replacement_count", 1])
                workbook.save(report_xlsx)
                workbook.close()
                report_json.write_text(
                    json.dumps(
                        {
                            "replacement_count": 1,
                            "source_project_no": "2016",
                            "target_project_no": "1818",
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                job.artifacts.replaced_dwg = replaced_dwg
                job.artifacts.report_xlsx = report_xlsx
                job.artifacts.report_json = report_json
                job.mark_succeeded()
                return

            assert job.job_type == JobType.DELIVERABLE
            assert len(job.input_files) == 1
            assert Path(job.input_files[0]).read_bytes() == b"replaced-dwg"
            job.mark_running(stage="GENERATE_DOCS")
            package_zip = job.work_dir / "package.zip"
            package_zip.write_bytes(b"PK\x03\x04deliverable")
            job.artifacts.package_zip = package_zip
            job.mark_succeeded()

    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    params = _deliverable_params()
    params.pop("project_no")
    with TestClient(
        create_app(
            job_processor=ReplaceThenDeliverableProcessor(),
            shared_prep_service=FakeSharedPrepService(),
        ),
    ) as client:
        response = client.post(
            "/api/jobs/audit-replace",
            data={
                "mode": "replace",
                "params_json": json.dumps(
                    {
                        "source_project_no": "2016",
                        "source_island_no": "1",
                        "target_project_no": "1818",
                        "unit_factory_codes": ["RC"],
                        "run_deliverable": True,
                        "deliverable_params": params,
                    },
                    ensure_ascii=False,
                ),
            },
            files=[("files[]", ("2016-A01.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        group_id = response.json()["jobs"][0]["job_id"]
        detail = _poll_job(client, group_id, timeout_sec=5.0)
        assert detail["status"] == "succeeded"
        children = {child["task_role"]: child for child in detail["children"]}
        assert children["audit_replace"]["artifacts"]["replaced_dwg_available"] is True
        assert children["deliverable_main"]["artifacts"]["package_available"] is True


def test_create_batch_processes_jobs_and_exposes_downloads(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(_deliverable_params(), ensure_ascii=False)},
            files=[
                ("files[]", ("A01.dwg", b"dwg-a", "application/acad")),
                ("files[]", ("A02.dwg", b"dwg-b", "application/acad")),
            ],
        )

        assert response.status_code == 201
        payload = response.json()
        assert payload["batch_id"]
        assert len(payload["jobs"]) == 2
        assert {item["source_filename"] for item in payload["jobs"]} == {"A01.dwg", "A02.dwg"}

        job_id = payload["jobs"][0]["job_id"]
        final_detail = _poll_job(client, job_id)
        assert final_detail["status"] == "succeeded"
        assert final_detail["artifacts"]["package_available"] is True
        assert final_detail["artifacts"]["ied_available"] is True
        assert final_detail["artifacts"]["preview_available"] is True
        assert final_detail["artifacts"]["preview_mode"] == "plain"
        assert final_detail["artifacts"]["preview_download_url"] == f"/api/jobs/{job_id}/download/preview"
        assert final_detail["deliverable_outputs"] == {
            "dwg_count": 2,
            "pdf_count": 2,
            "documents": [
                {"name": "cover.docx", "kind": "docx"},
                {"name": "cover.pdf", "kind": "pdf"},
                {"name": "design.xlsx", "kind": "xlsx"},
            ],
            "drawings": [
                {
                    "name": "DRAW001 (20261RS-JGS65-001)",
                    "internal_code": "20261RS-JGS65-001",
                    "dwg_name": "drawing-001.dwg",
                    "pdf_name": "drawing-001.pdf",
                    "page_total": 1,
                },
                {
                    "name": "DRAW002 (20261RS-JGS65-002)",
                    "internal_code": "20261RS-JGS65-002",
                    "dwg_name": "drawing-002.dwg",
                    "pdf_name": "drawing-002.pdf",
                    "page_total": 4,
                },
            ],
        }
        assert final_detail["flags"] == [
            "[DRAW001 (20261RS-JGS65-001)] PAPER_SIZE_AUTO_FIXED",
        ]
        assert final_detail["slot_id"] is not None
        assert final_detail["profile_arg"] is not None
        assert final_detail["plot_style_key"] == "red_wider"
        assert final_detail["plot_resource_mode"] == "slot_private_with_shared_mirror"

        listing = client.get("/api/jobs")
        assert listing.status_code == 200
        list_payload = listing.json()
        assert list_payload["total"] == 2
        assert list_payload["items"][0]["task_kind"] == "deliverable"

        package_download = client.get(f"/api/jobs/{job_id}/download/package")
        assert package_download.status_code == 200
        assert package_download.content.startswith(b"PK")

        ied_download = client.get(f"/api/jobs/{job_id}/download/ied")
        assert ied_download.status_code == 200
        assert ied_download.content == b"ied"

        preview_download = client.get(f"/api/jobs/{job_id}/download/preview")
        assert preview_download.status_code == 200
        assert preview_download.content == b"%PDF-plain"


def test_create_calculation_book_uses_job_flow_without_a_cad_slot(
    monkeypatch,
    tmp_path: Path,
) -> None:
    observed_ai_options: list[tuple[bool, int | None]] = []
    observed_job_options: list[dict[str, object]] = []

    class CalculationProcessor:
        def __call__(self, job: Job) -> None:
            assert job.job_type == JobType.CALCULATION_BOOK
            assert job.slot_id is None
            assert "project_number" not in job.params
            validated = CalculationBookParams.model_validate(job.params)
            assert validated.project_no == "JQ"
            assert validated.include_slab_stress is True
            raw_expected_count = job.options.get(
                "ai_reinforcement_expected_source_row_count"
            )
            observed_ai_options.append(
                (
                    bool(job.options.get("ai_reinforcement_normalization")),
                    (
                        raw_expected_count
                        if isinstance(raw_expected_count, int)
                        else None
                    ),
                )
            )
            observed_job_options.append(dict(job.options))
            job.mark_running(stage="VALIDATE_ARCHIVE")
            job.progress.percent = 80
            job.progress.details.update(
                {
                    "figure_count": 3,
                    "template_type": "internal_structure",
                    "output_filename": "JQ计算书.docx",
                }
            )
            if job.options.get("ai_reinforcement_normalization") is True:
                job.progress.details.update(
                    {
                        "ai_reinforcement_normalization": {
                            "skill_id": "reinforcement_table_normalizer",
                            "model": "structured-test",
                            "profile": "intranet-test",
                            "call_count": 1,
                            "source_row_count": 40,
                            "normalized_wall_count": 38,
                            "normalized_slab_count": 2,
                            "review_warning_count": 2,
                            "duration_ms": 125,
                            "validation": "passed",
                            "prompt": "must-not-leak",
                        },
                        "calculation_book_warnings": [
                            {
                                "code": "duplicate_reinforcement_rows",
                                "scope": "wall",
                                "identity": "S7157",
                                "direction": None,
                                "source_sheet": "Sheet1",
                                "source_row": 28,
                                "source_cells": {
                                    "wall": "A28",
                                    "X": "B28",
                                    "Y": "C28",
                                    "Z": "D28",
                                },
                                "reason": "model supplied secret reason must-not-leak",
                                "blank_fields": ["X", "Y", "Z"],
                                "original_values": {"X": "must-not-leak"},
                            },
                            {
                                "code": "image_only_wall",
                                "scope": "wall",
                                "identity": "N5012",
                                "direction": None,
                                "source_sheet": None,
                                "source_row": None,
                                "source_cells": {},
                                "reason": "应力图中存在该墙体，但配筋表没有对应数据，相关配筋字段已留空",
                                "blank_fields": ["X", "Y", "Z"],
                            },
                        ],
                    }
                )
            if job.options.get("ai_rebar_suggestion") is True:
                calculation_log = (
                    cast(Path, job.work_dir)
                    / "calculation-book"
                    / "logs"
                    / f"calculation-book-{job.job_id}.log"
                )
                calculation_log.parent.mkdir(parents=True, exist_ok=True)
                with CalculationBookDiagnosticLog.create(
                    calculation_log,
                    job_id=job.job_id,
                    correlation_id=job.job_id,
                    max_bytes=8_192,
                ) as diagnostic_log:
                    diagnostic_log.write(
                        "task_completed",
                        duration_ms=1,
                        figure_count=3,
                        warning_count=1,
                        output_filename="result.docx",
                    )
                job.artifacts.calculation_log = calculation_log
                job.progress.details.update(
                    {
                        "ai_rebar_suggestion": {
                            "skill_id": "recommend-rebar-from-smx",
                            "skill_version": "1.0.0",
                            "skill_sha256": "a" * 64,
                            "model": "structured-test",
                            "call_count": 6,
                            "suggested_direction_count": 181,
                            "blank_direction_count": 1,
                            "repair_round_count": 2,
                            "validation": "passed_with_warnings",
                            "prompt": "must-not-leak",
                            "candidates": ["must-not-leak"],
                        },
                        "calculation_book_warnings": [
                            {
                                "code": "AI_BASE_FAILURE_LIMIT",
                                "scope": "wall",
                                "identity": "N5012",
                                "direction": "Z",
                                "source_sheet": None,
                                "source_row": None,
                                "source_cells": {},
                                "reason": "raw model failure must-not-leak",
                                "blank_fields": ["Z"],
                            }
                        ],
                    }
                )
            output_path = cast(Path, job.work_dir) / "JQ计算书.docx"
            output_path.write_bytes(b"docx")
            job.artifacts.calculation_docx = output_path
            job.mark_succeeded()

    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    def fake_calculation_book_preflight(**kwargs):
        assert kwargs["include_slab_stress"] is True
        reinforcement_source = kwargs["reinforcement_source"]
        assert isinstance(reinforcement_source, ReinforcementSource)
        assert kwargs["archive_path"].suffix == ".rar"
        archive_name = kwargs["archive_path"].name
        uncountable = archive_name.startswith("uncountable-")
        manual_review = archive_name.startswith("manual-")
        requires_ai = (
            reinforcement_source is ReinforcementSource.PROVIDED
            and (archive_name.startswith("nonstandard-") or uncountable)
        )
        requires_ai_recommendation = (
            reinforcement_source is ReinforcementSource.AI_SUGGESTED
        )
        return {
            "reinforcement_source": reinforcement_source.value,
            "figure_count": 3,
            "zero_figure_count": 1,
            "wall_count": 1,
            "reinforcement_workbook": (
                None
                if requires_ai_recommendation
                else "计算书模板文件.xlsx"
            ),
            "reinforcement_issue_row_count": 1 if requires_ai else 0,
            "requires_ai_normalization": requires_ai,
            "requires_ai_recommendation": requires_ai_recommendation,
            "ai_reinforcement_expected_source_row_count": (
                40 if requires_ai and not uncountable else None
            ),
            "ai_confirmation_message": (
                "您上传的墙体配筋表非标准格式，程序将启动人工智能。"
                if requires_ai
                else None
            ),
            "format_inspection": {
                "wall_sheet": "非标准墙表" if requires_ai else "Sheet1",
                "slab_sheet": "楼板配筋",
                "reasons": (
                    [
                        {
                            "scope": "wall",
                            "code": "wall_layout_nonstandard",
                            "sheet": "非标准墙表",
                            "message": "非标准墙表不是标准四列墙体配筋模板",
                        }
                    ]
                    if requires_ai
                    else []
                ),
            },
            "requires_manual_confirmation": manual_review,
            "requires_wall_count_confirmation": not requires_ai,
            "confirmation_candidates": (
                {"N5012": [2, 3]} if manual_review else {}
            ),
            "confirmations": (
                [
                    {
                        "wall_id": "N5012",
                        "base_wall_id": "N5012",
                        "reasons": ["duplicate_reinforcement_rows"],
                        "suggested_source_row": 2,
                        "candidates": [
                            {"source_row": 2},
                            {"source_row": 3},
                        ],
                    }
                ]
                if manual_review
                else []
            ),
            "walls": [],
            "warnings": [],
        }

    monkeypatch.setattr(
        "API.app.runtime.run_calculation_book_preflight",
        fake_calculation_book_preflight,
    )
    params = {
        "template_type": "internal_structure",
        "project_no": "JQ",
        "project_name": "浙江金七门核电厂1、2号机组",
        "internal_code": "JQ00-NN-001",
        "version": "A",
        "subproject_code": "RX",
        "subproject_name": "内部结构",
        "design_phase": "施工图设计",
        "document_name": "0.000m~15.000m配筋计算书",
        "workshop_length": 72.5,
        "workshop_width": 48.0,
        "raft_slab_top_elevation": -8.5,
        "roof_top_elevation": 31.2,
        "factory_extreme_min_temperature": -18.0,
        "factory_extreme_max_temperature": 39.0,
        "site_soil_temperature": 15.0,
        "include_slab_stress": True,
    }
    with TestClient(
        create_app(
            job_processor=CalculationProcessor(),
            shared_prep_service=FakeSharedPrepService(),
            font_preflight_service=FakeFontPreflightService(),
        )
    ) as client:
        archive_bytes = b"Rar!\x1a\x07\x01\x00test"
        wrong_token_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {**params, "preflight_token": "forged-token"},
                    ensure_ascii=False,
                )
            },
        )
        assert wrong_token_response.status_code == 422
        assert wrong_token_response.json()["detail"]["param_errors"][
            "preflight_token"
        ] == ["请先完成计算书文件预检"]

        expired_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        expired_token = expired_preflight.json()["preflight_token"]
        runtime = client.app.state.runtime
        expired_entry = runtime._calculation_preflight_tokens[expired_token]
        expired_archive = Path(expired_entry["archive_path"])
        expired_entry["created_at"] -= 1801
        expired_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "preflight_token": expired_token,
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert expired_response.status_code == 422
        assert expired_response.json()["detail"]["param_errors"][
            "preflight_token"
        ] == ["请先完成计算书文件预检"]
        assert not expired_archive.exists()

        uncountable_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "uncountable-calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert uncountable_preflight.status_code == 422
        assert uncountable_preflight.json()["detail"]["upload_errors"] == {
            "archive": ["无法可靠统计非标准配筋表数据行"]
        }
        assert runtime._calculation_preflight_tokens == {}

        slab_mismatch_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "nonstandard-slab-mismatch.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        slab_mismatch_token = slab_mismatch_preflight.json()["preflight_token"]
        slab_mismatch_archive = Path(
            runtime._calculation_preflight_tokens[slab_mismatch_token][
                "archive_path"
            ]
        )
        slab_mismatch_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "include_slab_stress": False,
                        "preflight_token": slab_mismatch_token,
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert slab_mismatch_response.status_code == 422
        assert slab_mismatch_response.json()["detail"]["param_errors"][
            "include_slab_stress"
        ] == ["楼板应力选项已变化，请重新预检"]
        assert slab_mismatch_token not in runtime._calculation_preflight_tokens
        assert not slab_mismatch_archive.exists()

        nonstandard_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "nonstandard-calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert nonstandard_preflight.status_code == 200, nonstandard_preflight.text
        assert nonstandard_preflight.json()["requires_ai_normalization"] is True
        assert (
            nonstandard_preflight.json()[
                "ai_reinforcement_expected_source_row_count"
            ]
            == 40
        )
        assert nonstandard_preflight.json()["ai_confirmation_message"] == (
            "您上传的墙体配筋表非标准格式，程序将启动人工智能。"
        )
        nonstandard_params = {
            **params,
            "preflight_token": nonstandard_preflight.json()["preflight_token"],
        }
        nonstandard_token = nonstandard_params["preflight_token"]
        nonstandard_archive = Path(
            runtime._calculation_preflight_tokens[nonstandard_token][
                "archive_path"
            ]
        )
        forged_count_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **nonstandard_params,
                        "ai_reinforcement_expected_source_row_count": 999,
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert forged_count_response.status_code == 422
        assert "ai_reinforcement_expected_source_row_count" in (
            forged_count_response.json()["detail"]["param_errors"]
        )
        assert nonstandard_token in runtime._calculation_preflight_tokens
        unconfirmed_ai_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    nonstandard_params,
                    ensure_ascii=False,
                )
            },
        )
        assert unconfirmed_ai_response.status_code == 422
        assert unconfirmed_ai_response.json()["detail"]["param_errors"][
            "confirm_ai_normalization"
        ] == ["请确认启动人工智能规范化非标准配筋表"]
        assert nonstandard_token in runtime._calculation_preflight_tokens
        assert nonstandard_archive.is_file()

        confirmed_ai_params = {
            **nonstandard_params,
            "confirm_ai_normalization": True,
        }
        confirmed_ai_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    confirmed_ai_params,
                    ensure_ascii=False,
                )
            },
        )
        assert confirmed_ai_response.status_code == 201, confirmed_ai_response.text
        confirmed_ai_detail = _poll_job(
            client,
            confirmed_ai_response.json()["jobs"][0]["job_id"],
        )
        assert confirmed_ai_detail["status"] == "succeeded"
        assert confirmed_ai_detail["calculation_book_output"] == {
            "figure_count": 3,
            "template_type": "internal_structure",
            "output_filename": "JQ计算书.docx",
            "reinforcement_source": "provided",
            "ai_normalized": True,
            "warning_count": 2,
            "warnings": [
                {
                    "code": "duplicate_reinforcement_rows",
                    "scope": "wall",
                    "identity": "S7157",
                    "direction": None,
                    "source_sheet": "Sheet1",
                    "source_row": 28,
                    "source_cells": {
                        "wall": "A28",
                        "X": "B28",
                        "Y": "C28",
                        "Z": "D28",
                    },
                    "reason": "同一墙体存在重复配筋行，相关配筋字段已留空",
                    "blank_fields": ["X", "Y", "Z"],
                },
                {
                    "code": "image_only_wall",
                    "scope": "wall",
                    "identity": "N5012",
                    "direction": None,
                    "source_sheet": None,
                    "source_row": None,
                    "source_cells": {},
                    "reason": "应力图中存在该墙体，但配筋表没有对应数据，相关配筋字段已留空",
                    "blank_fields": ["X", "Y", "Z"],
                },
            ],
            "ai_normalization": {
                "skill_id": "reinforcement_table_normalizer",
                "model": "structured-test",
                "profile": "intranet-test",
                "call_count": 1,
                "source_row_count": 40,
                "normalized_wall_count": 38,
                "normalized_slab_count": 2,
                "review_warning_count": 2,
                "duration_ms": 125,
                "validation": "passed",
            },
            "ai_rebar_suggestion": None,
        }
        assert observed_ai_options == [(True, 40)]
        assert nonstandard_token not in runtime._calculation_preflight_tokens
        assert not nonstandard_archive.exists()

        replay_ai_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    confirmed_ai_params,
                    ensure_ascii=False,
                )
            },
        )
        assert replay_ai_response.status_code == 422
        assert replay_ai_response.json()["detail"]["param_errors"][
            "preflight_token"
        ] == ["请先完成计算书文件预检"]

        manual_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "manual-calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert manual_preflight.status_code == 200, manual_preflight.text
        assert manual_preflight.json()["requires_manual_confirmation"] is True
        no_row_confirmation = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "preflight_token": manual_preflight.json()[
                            "preflight_token"
                        ],
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert no_row_confirmation.status_code == 201, no_row_confirmation.text

        preflight_response = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert preflight_response.status_code == 200, preflight_response.text
        params["preflight_token"] = preflight_response.json()["preflight_token"]
        assert runtime._calculation_preflight_tokens[params["preflight_token"]][
            "include_slab_stress"
        ] is True
        cached_archive = Path(
            runtime._calculation_preflight_tokens[params["preflight_token"]][
                "archive_path"
            ]
        )
        assert cached_archive.is_file()
        assert cached_archive.suffix == ".rar"

        mismatch_params = {
            **params,
            "include_slab_stress": False,
        }
        mismatch_response = client.post(
            "/api/jobs/calculation-books",
            data={"params_json": json.dumps(mismatch_params, ensure_ascii=False)},
        )
        assert mismatch_response.status_code == 422
        assert mismatch_response.json()["detail"]["param_errors"][
            "include_slab_stress"
        ] == ["楼板应力选项已变化，请重新预检"]
        assert not cached_archive.exists()

        preflight_response = client.post(
            "/api/jobs/calculation-books/preflight",
            data={"include_slab_stress": "true"},
            files={
                "archive": (
                    "calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert preflight_response.status_code == 200, preflight_response.text
        assert preflight_response.json()["requires_ai_normalization"] is False
        assert preflight_response.json()["requires_wall_count_confirmation"] is True
        params["preflight_token"] = preflight_response.json()["preflight_token"]
        params["confirm_ai_normalization"] = True
        cached_archive = Path(
            runtime._calculation_preflight_tokens[params["preflight_token"]][
                "archive_path"
            ]
        )

        response = client.post(
            "/api/jobs/calculation-books",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
        )

        assert response.status_code == 201, response.text
        assert not cached_archive.exists()
        job_id = response.json()["jobs"][0]["job_id"]
        detail = _poll_job(client, job_id)
        assert observed_ai_options[-1] == (False, None)
        assert detail["status"] == "succeeded"
        assert detail["task_kind"] == "calculation_book"
        assert detail["slot_id"] is None
        assert detail["calculation_book_output"]["figure_count"] == 3
        assert detail["calculation_book_output"]["ai_normalized"] is False
        assert detail["calculation_book_output"]["warning_count"] == 0
        assert detail["calculation_book_output"]["warnings"] == []
        assert detail["calculation_book_output"]["ai_normalization"] is None
        assert detail["artifacts"]["calculation_docx_available"] is True
        assert detail["artifacts"]["calculation_docx_download_url"] == (
            f"/api/jobs/{job_id}/download/calculation-book"
        )

        download = client.get(f"/api/jobs/{job_id}/download/calculation-book")
        assert download.status_code == 200
        assert download.content == b"docx"

        replay = client.post(
            "/api/jobs/calculation-books",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
        )
        assert replay.status_code == 422
        assert replay.json()["detail"]["param_errors"]["preflight_token"] == [
            "请先完成计算书文件预检"
        ]

        ai_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={
                "include_slab_stress": "true",
                "reinforcement_source": "ai_suggested",
            },
            files={
                "archive": (
                    "ai-calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        assert ai_preflight.status_code == 200, ai_preflight.text
        assert ai_preflight.json()["requires_ai_recommendation"] is True
        ai_token = ai_preflight.json()["preflight_token"]
        assert runtime._calculation_preflight_tokens[ai_token][
            "reinforcement_source"
        ] == "ai_suggested"

        source_mismatch_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "preflight_token": ai_token,
                        "confirm_ai_normalization": False,
                        "reinforcement_source": "provided",
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert source_mismatch_response.status_code == 422
        assert source_mismatch_response.json()["detail"]["param_errors"][
            "reinforcement_source"
        ] == ["配筋来源已变化，请重新预检"]

        ai_preflight = client.post(
            "/api/jobs/calculation-books/preflight",
            data={
                "include_slab_stress": "true",
                "reinforcement_source": "ai_suggested",
            },
            files={
                "archive": (
                    "ai-calculation-images.rar",
                    archive_bytes,
                    "application/vnd.rar",
                )
            },
        )
        ai_token = ai_preflight.json()["preflight_token"]
        forged_options_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "preflight_token": ai_token,
                        "confirm_ai_normalization": False,
                        "reinforcement_source": "ai_suggested",
                        "options": {
                            "candidates": ["client-forged"],
                            "skill_id": "client-forged",
                            "failure_count": 999,
                        },
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert forged_options_response.status_code == 422
        assert "options" in forged_options_response.json()["detail"][
            "param_errors"
        ]
        assert ai_token in runtime._calculation_preflight_tokens

        ai_response = client.post(
            "/api/jobs/calculation-books",
            data={
                "params_json": json.dumps(
                    {
                        **params,
                        "preflight_token": ai_token,
                        "confirm_ai_normalization": False,
                        "reinforcement_source": "ai_suggested",
                    },
                    ensure_ascii=False,
                )
            },
        )
        assert ai_response.status_code == 201, ai_response.text
        ai_detail = _poll_job(
            client,
            ai_response.json()["jobs"][0]["job_id"],
        )
        assert ai_detail["status"] == "succeeded"
        ai_job_id = ai_response.json()["jobs"][0]["job_id"]
        assert ai_detail["calculation_book_output"] == {
            "figure_count": 3,
            "template_type": "internal_structure",
            "output_filename": "JQ计算书.docx",
            "reinforcement_source": "ai_suggested",
            "ai_normalized": False,
            "warning_count": 1,
            "warnings": [
                {
                    "code": "AI_BASE_FAILURE_LIMIT",
                    "scope": "wall",
                    "identity": "N5012",
                    "direction": "Z",
                    "source_sheet": None,
                    "source_row": None,
                    "source_cells": {},
                    "reason": "人工智能连续三次调用或协议失败，当前方向已留空，请人工复核",
                    "blank_fields": ["Z"],
                }
            ],
            "ai_normalization": None,
            "ai_rebar_suggestion": {
                "skill_id": "recommend-rebar-from-smx",
                "skill_version": "1.0.0",
                "skill_sha256": "a" * 64,
                "model": "structured-test",
                "call_count": 6,
                "suggested_direction_count": 181,
                "blank_direction_count": 1,
                "repair_round_count": 2,
                "validation": "passed_with_warnings",
            },
        }
        assert ai_detail["artifacts"]["calculation_log_available"] is True
        assert ai_detail["artifacts"]["calculation_log_download_url"] == (
            f"/api/jobs/{ai_job_id}/download/calculation-book-log"
        )
        log_download = client.get(
            f"/api/jobs/{ai_job_id}/download/calculation-book-log"
        )
        assert log_download.status_code == 200
        assert json.loads(log_download.content.splitlines()[-1])["event"] == (
            "task_completed"
        )
        assert log_download.headers["content-type"] == (
            "text/plain; charset=utf-8"
        )
        assert log_download.headers["cache-control"] == "no-store"
        assert log_download.headers["x-content-type-options"] == "nosniff"
        authorization = client.headers.pop("Authorization")
        unauthenticated_log = client.get(
            f"/api/jobs/{ai_job_id}/download/calculation-book-log"
        )
        assert unauthenticated_log.status_code == 401
        client.headers["Authorization"] = authorization

        ai_options = next(
            item
            for item in observed_job_options
            if item.get("reinforcement_source") == "ai_suggested"
        )
        assert ai_options == {
            "mode": "calculation_book",
            "reinforcement_source": "ai_suggested",
            "ai_reinforcement_normalization": False,
            "ai_rebar_suggestion": True,
        }
        provided_options = [
            item
            for item in observed_job_options
            if item.get("reinforcement_source") == "provided"
        ]
        assert any(
            item.get("ai_reinforcement_normalization") is True
            and "ai_rebar_suggestion" not in item
            for item in provided_options
        )
        assert any(
            item["ai_reinforcement_normalization"] is False
            and "ai_rebar_suggestion" not in item
            for item in provided_options
        )


def test_download_standard_reinforcement_template_is_authenticated_and_fixed(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template_dir = tmp_path / "calculation-book-templates"
    template_dir.mkdir()
    template = template_dir / "计算书模板文件.xlsx"
    expected = b"PK\x03\x04standard-reinforcement-template"
    template.write_bytes(expected)
    (template_dir / "other.xlsx").write_bytes(b"must-not-be-downloaded")

    with _create_client(monkeypatch, tmp_path) as client:
        runtime = client.app.state.runtime
        runtime.config.calculation_book.template_dir = template_dir
        runtime.config.calculation_book.standard_reinforcement_template = template

        response = client.get(
            "/api/jobs/calculation-books/reinforcement-template",
            params={"filename": "other.xlsx"},
        )

        assert response.status_code == 200, response.text
        assert response.content == expected
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "标准配筋模板.xlsx" in unquote(
            response.headers["content-disposition"]
        )
        assert response.headers["cache-control"] == "no-store"
        assert response.headers["x-content-type-options"] == "nosniff"

        client.headers.pop("Authorization")
        unauthenticated = client.get(
            "/api/jobs/calculation-books/reinforcement-template"
        )

    assert unauthenticated.status_code == 401


def test_failed_ai_calculation_hides_word_but_keeps_terminal_log_download(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        runtime = client.app.state.runtime
        job = runtime.job_manager.create_job(
            job_type=JobType.CALCULATION_BOOK.value,
            project_no="JQ",
            options={
                "reinforcement_source": "ai_suggested",
                "ai_rebar_suggestion": True,
            },
        )
        work_dir = tmp_path / "failed-ai-calculation"
        runtime.config.calculation_book.ai_suggestion.log_dir = (
            tmp_path / "central-ai-audit"
        )
        log_path = (
            runtime.config.calculation_book.ai_suggestion.log_dir
            / f"calculation-book-{job.job_id}.log"
        )
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with CalculationBookDiagnosticLog.create(
            log_path,
            job_id=job.job_id,
            correlation_id=job.job_id,
            max_bytes=8_192,
        ) as diagnostic_log:
            diagnostic_log.write(
                "task_failed",
                stage="render_document",
                duration_ms=1,
                error_code="RuntimeError",
            )
        docx_path = work_dir / "calculation-book" / "failed.docx"
        docx_path.parent.mkdir(parents=True, exist_ok=True)
        docx_path.write_bytes(b"PK\x03\x04failed-word")
        job.work_dir = work_dir
        job.artifacts.calculation_log = log_path
        job.artifacts.calculation_docx = docx_path
        job.status = JobStatus.FAILED
        runtime.job_manager.update_job(job)

        detail = client.get(f"/api/jobs/{job.job_id}")
        assert detail.status_code == 200
        artifacts = detail.json()["artifacts"]
        assert artifacts["calculation_docx_available"] is False
        assert artifacts["calculation_docx_download_url"] is None
        assert artifacts["calculation_log_available"] is True
        assert artifacts["calculation_log_download_url"] == (
            f"/api/jobs/{job.job_id}/download/calculation-book-log"
        )

        word = client.get(
            f"/api/jobs/{job.job_id}/download/calculation-book"
        )
        assert word.status_code == 404
        log = client.get(
            f"/api/jobs/{job.job_id}/download/calculation-book-log"
        )
        assert log.status_code == 200
        assert json.loads(log.content.splitlines()[-1])["event"] == (
            "task_failed"
        )


@pytest.mark.parametrize("invalid_kind", ["missing", "directory", "escape"])
def test_download_standard_reinforcement_template_hides_invalid_paths(
    monkeypatch,
    tmp_path: Path,
    invalid_kind: str,
) -> None:
    template_dir = tmp_path / "calculation-book-templates"
    template_dir.mkdir()
    outside = tmp_path / "outside.xlsx"
    outside.write_bytes(b"outside")
    if invalid_kind == "missing":
        configured = template_dir / "missing.xlsx"
    elif invalid_kind == "directory":
        configured = template_dir / "nested"
        configured.mkdir()
    else:
        configured = outside

    with _create_client(monkeypatch, tmp_path) as client:
        runtime = client.app.state.runtime
        runtime.config.calculation_book.template_dir = template_dir
        runtime.config.calculation_book.standard_reinforcement_template = configured

        response = client.get(
            "/api/jobs/calculation-books/reinforcement-template"
        )

    assert response.status_code == 404
    assert response.json() == {"detail": "标准配筋模板不可用"}
    assert str(configured) not in response.text


def test_calculation_preflight_cache_cleanup_is_scoped_and_age_limited(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from API.app.runtime import _cleanup_calculation_preflight_cache

    cache_root = tmp_path / "calculation-preflight"
    cache_root.mkdir()
    now = time.time()
    old_zip = cache_root / "calculation-preflight-old.zip"
    old_rar = cache_root / "calculation-preflight-old.rar"
    old_7z = cache_root / "calculation-preflight-old.7z"
    fresh_zip = cache_root / "calculation-preflight-fresh.zip"
    unrelated = cache_root / "user-upload.zip"
    wrong_suffix = cache_root / "calculation-preflight-old.txt"
    matching_directory = cache_root / "calculation-preflight-folder.zip"
    for path in (old_zip, old_rar, old_7z, fresh_zip, unrelated, wrong_suffix):
        path.write_bytes(b"payload")
    matching_directory.mkdir()
    old_mtime = now - 1801
    for path in (old_zip, old_rar, old_7z, unrelated, wrong_suffix):
        os.utime(path, (old_mtime, old_mtime))
    os.utime(matching_directory, (old_mtime, old_mtime))

    symlink_path = cache_root / "calculation-preflight-link.zip"
    symlink_target = tmp_path / "outside.zip"
    symlink_target.write_bytes(b"outside")
    symlink_created = False
    try:
        symlink_path.symlink_to(symlink_target)
        symlink_created = True
    except OSError:
        pass

    _cleanup_calculation_preflight_cache(cache_root, now=now)

    assert not old_zip.exists()
    assert not old_rar.exists()
    assert not old_7z.exists()
    assert fresh_zip.is_file()
    assert unrelated.is_file()
    assert wrong_suffix.is_file()
    assert matching_directory.is_dir()
    assert symlink_target.is_file()
    if symlink_created:
        assert symlink_path.is_symlink()

    outside_root = tmp_path / "outside-cache"
    outside_root.mkdir()
    outside_old = outside_root / "calculation-preflight-outside.zip"
    outside_old.write_bytes(b"outside")
    os.utime(outside_old, (old_mtime, old_mtime))
    linked_root = tmp_path / "linked-cache"
    try:
        linked_root.symlink_to(outside_root, target_is_directory=True)
    except OSError:
        linked_root.mkdir()
        original_is_symlink = Path.is_symlink
        monkeypatch.setattr(
            Path,
            "is_symlink",
            lambda self: (
                True
                if self == linked_root
                else original_is_symlink(self)
            ),
        )
    with pytest.raises(
        RuntimeError,
        match="unsafe calculation preflight cache root",
    ):
        _cleanup_calculation_preflight_cache(linked_root, now=now)
    assert outside_old.is_file()


def test_calculation_preflight_accepts_7z_forwards_config_and_fixes_mime(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    observed: dict[str, object] = {}

    def fake_preflight(**kwargs):
        observed.update(kwargs)
        return {
            "confirmations": [],
            "requires_ai_normalization": False,
            "requires_wall_count_confirmation": False,
            "format_inspection": {},
        }

    monkeypatch.setattr(
        "API.app.runtime.run_calculation_book_preflight",
        fake_preflight,
    )
    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/calculation-books/preflight",
            files={
                "archive": (
                    "calculation-images.7z",
                    b"7z\xbc\xaf'\x1cpayload",
                    "text/plain",
                )
            },
        )
        assert response.status_code == 200, response.text
        runtime = client.app.state.runtime
        token = response.json()["preflight_token"]
        cached = runtime._calculation_preflight_tokens[token]

        assert (
            observed["archive_extractor"]
            is runtime.config.calculation_book.archive_extractor
        )
        assert cached["content_type"] == "application/x-7z-compressed"
        assert Path(str(cached["archive_path"])).suffix == ".7z"


def test_calculation_preflight_rejects_all_invalid_editable_params_before_archive_scan(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    def unexpected_preflight(**_kwargs):
        raise AssertionError("archive preflight must not run for invalid params")

    monkeypatch.setattr(
        "API.app.runtime.run_calculation_book_preflight",
        unexpected_preflight,
    )
    params = {
        "template_type": "internal_structure",
        "project_no": "JQ",
        "project_name": "测试项目",
        "internal_code": "JQ00-NN-001",
        "version": "A",
        "subproject_code": "RX",
        "subproject_name": "内部结构",
        "design_phase": "施工图设计",
        "document_name": "11111",
        "workshop_length": 15,
        "workshop_width": 15,
        "raft_slab_top_elevation": 15,
        "roof_top_elevation": 15,
        "factory_extreme_min_temperature": 15,
        "factory_extreme_max_temperature": 15,
        "site_soil_temperature": 15,
        "include_slab_stress": True,
        "reinforcement_source": "ai_suggested",
    }

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/calculation-books/preflight",
            data={
                "include_slab_stress": "true",
                "reinforcement_source": "ai_suggested",
                "params_json": json.dumps(params, ensure_ascii=False),
            },
            files={
                "archive": (
                    "calculation-images.rar",
                    b"Rar!\x1a\x07\x01\x00payload",
                    "application/vnd.rar",
                )
            },
        )

    assert response.status_code == 422
    errors = response.json()["detail"]["param_errors"]
    assert set(errors) >= {
        "document_name",
        "roof_top_elevation",
        "factory_extreme_max_temperature",
    }
    assert "标高范围" in errors["document_name"][0]
    assert "筏板顶标高 15m" in errors["roof_top_elevation"][0]
    assert "历史最低温度 15℃" in errors["factory_extreme_max_temperature"][0]


@pytest.mark.parametrize(
    ("filename", "payload", "content_type"),
    [
        ("input.rar", b"Rar!\x1a\x07\x01\x00payload", "application/vnd.rar"),
        ("input.7z", b"7z\xbc\xaf'\x1cpayload", "application/x-7z-compressed"),
    ],
)
def test_calculation_preflight_missing_private_extractor_is_stable_and_uncached(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    filename: str,
    payload: bytes,
    content_type: str,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        runtime = client.app.state.runtime
        extractor = runtime.config.calculation_book.archive_extractor
        runtime.config.calculation_book.archive_extractor = extractor.model_copy(
            update={"executable": (tmp_path / "missing-7z.exe").resolve()}
        )

        response = client.post(
            "/api/jobs/calculation-books/preflight",
            files={"archive": (filename, payload, content_type)},
        )

        assert response.status_code == 422
        assert response.json()["detail"]["upload_errors"] == {
            "archive": ["RAR/7z 私有解包器不存在"]
        }
        assert runtime._calculation_preflight_tokens == {}
        cache_root = (
            runtime.config.storage_dir
            / "runtime"
            / "calculation-preflight"
        )
        assert not tuple(cache_root.glob("calculation-preflight-*"))


def _calculation_archive_bytes(
    tmp_path: Path,
    *,
    malicious_workbook: bool,
) -> bytes:
    source = tmp_path / "malicious-calculation-source"
    source.mkdir()
    for direction in ("X", "Y", "Z"):
        Image.new("RGB", (32, 32), "white").save(
            source / f"S7159-{direction}.png"
        )
    for folder_name in ("01", "02"):
        folder = source / folder_name
        folder.mkdir()
        Image.new("RGB", (32, 32), "white").save(folder / "figure.png")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    sheet.append(
        ["S7159墙", "1D28间距200", "1D28间距200", "1C12间距200*400"]
    )
    workbook_path = source / "计算书模板文件.xlsx"
    if malicious_workbook:
        safe_workbook = source / "safe.xlsx"
        workbook.save(safe_workbook)
        with zipfile.ZipFile(safe_workbook) as source_archive, zipfile.ZipFile(
            workbook_path,
            "w",
            zipfile.ZIP_DEFLATED,
        ) as target_archive:
            for info in source_archive.infolist():
                target_archive.writestr(info, source_archive.read(info.filename))
            target_archive.writestr(
                "xl/sharedStrings.xml",
                b"<sst><si><t>"
                + (b"A" * (8 * 1024 * 1024))
                + b"</t></si></sst>",
            )
        safe_workbook.unlink()
    else:
        workbook.save(workbook_path)

    archive_buffer = BytesIO()
    with zipfile.ZipFile(
        archive_buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as archive:
        for path in source.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source).as_posix())
    return archive_buffer.getvalue()


def test_calculation_preflight_rejects_embedded_xlsx_bomb_before_ocr_ai_and_cache(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizer

    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe workbook must fail before OCR or AI")

    monkeypatch.setattr(
        "API.app.runtime.recognize_stress_legend",
        fail_if_called,
    )
    monkeypatch.setattr(
        ReinforcementTaskNormalizer,
        "normalize",
        fail_if_called,
    )
    with TestClient(
        create_app(
            shared_prep_service=FakeSharedPrepService(),
            font_preflight_service=FakeFontPreflightService(),
        )
    ) as client:
        response = client.post(
            "/api/jobs/calculation-books/preflight",
            files={
                "archive": (
                    "malicious.zip",
                    _calculation_archive_bytes(
                        tmp_path,
                        malicious_workbook=True,
                    ),
                    "application/zip",
                )
            },
        )

        assert response.status_code == 422
        assert "XLSX internal resource limit" in response.text
        assert client.app.state.runtime._calculation_preflight_tokens == {}


def test_calculation_preflight_rejects_linked_cache_root_before_ocr_or_write(
    monkeypatch,
    tmp_path: Path,
) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    cache_parent = tmp_path / "storage" / "runtime"
    cache_parent.mkdir(parents=True)
    outside_root = tmp_path / "outside-preflight-cache"
    outside_root.mkdir()
    cache_root = cache_parent / "calculation-preflight"
    try:
        cache_root.symlink_to(outside_root, target_is_directory=True)
    except OSError:
        cache_root.mkdir()
        original_is_symlink = Path.is_symlink
        monkeypatch.setattr(
            Path,
            "is_symlink",
            lambda self: (
                True
                if self == cache_root
                else original_is_symlink(self)
            ),
        )

    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe cache root must fail before OCR")

    monkeypatch.setattr(
        "API.app.runtime.recognize_stress_legend",
        fail_if_called,
    )
    with TestClient(
        create_app(
            shared_prep_service=FakeSharedPrepService(),
            font_preflight_service=FakeFontPreflightService(),
        )
    ) as client:
        response = client.post(
            "/api/jobs/calculation-books/preflight",
            files={
                "archive": (
                    "safe.zip",
                    _calculation_archive_bytes(
                        tmp_path,
                        malicious_workbook=False,
                    ),
                    "application/zip",
                )
            },
        )

        assert response.status_code == 422
        assert "unsafe calculation preflight cache root" in response.text
        assert list(outside_root.iterdir()) == []
        assert client.app.state.runtime._calculation_preflight_tokens == {}


def test_create_batch_without_ied_plan_hides_ied_artifact_and_download(
    monkeypatch,
    tmp_path: Path,
) -> None:
    params = _deliverable_params()
    params["include_ied_plan"] = False

    with _create_client(monkeypatch, tmp_path) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(params, ensure_ascii=False)},
            files=[("files[]", ("A01.dwg", b"dwg-a", "application/acad"))],
        )

        assert response.status_code == 201
        job_id = response.json()["jobs"][0]["job_id"]
        final_detail = _poll_job(client, job_id)
        assert final_detail["status"] == "succeeded"
        assert final_detail["artifacts"]["ied_available"] is False
        assert final_detail["artifacts"]["ied_download_url"] is None

        ied_download = client.get(f"/api/jobs/{job_id}/download/ied")
        assert ied_download.status_code == 404


def test_list_jobs_supports_offset_limit_and_status_filtered_total(
    monkeypatch,
    tmp_path: Path,
) -> None:
    with _create_client(monkeypatch, tmp_path) as client:
        runtime = client.app.state.runtime
        base_time = datetime(2026, 4, 27, 9, 0, 0)
        created_ids: list[str] = []

        for index in range(5):
            job = runtime.job_manager.create_job(
                job_type=JobType.DELIVERABLE.value,
                project_no="2016",
                source_filename=f"job-{index + 1}.dwg",
            )
            job.created_at = base_time + timedelta(minutes=index)
            job.status = JobStatus.SUCCEEDED if index < 3 else JobStatus.FAILED
            runtime.job_manager.update_job(job)
            created_ids.append(job.job_id)

        paged = client.get("/api/jobs", params={"offset": 1, "limit": 2})
        assert paged.status_code == 200
        paged_payload = paged.json()
        assert paged_payload["total"] == 5
        assert [item["source_filename"] for item in paged_payload["items"]] == [
            "job-4.dwg",
            "job-3.dwg",
        ]

        filtered = client.get("/api/jobs", params={"status": "succeeded", "offset": 1, "limit": 1})
        assert filtered.status_code == 200
        filtered_payload = filtered.json()
        assert filtered_payload["total"] == 3
        assert [item["source_filename"] for item in filtered_payload["items"]] == ["job-2.dwg"]


class FakeSharedPrepService(SharedPrepService):
    def prepare(
        self,
        *,
        group_id: str,
        project_no: str | None = None,
        source_dwg: Path,
        shared_dir: Path,
        font_replace_policy: str = "none",
        font_replacement_font: str | None = None,
        font_replacement_fonts: dict[str, str] | None = None,
        font_compatibility_mode: bool = False,
        slot_runtime: dict[str, str] | None = None,
    ) -> SharedPrepArtifacts:
        shared_dir.mkdir(parents=True, exist_ok=True)
        staged_source = shared_dir / source_dwg.name
        staged_source.write_bytes(source_dwg.read_bytes())
        converted_dxf = shared_dir / "source_converted.dxf"
        (shared_dir / "prep_summary.json").write_text(
            json.dumps(
                {
                    "group_id": group_id,
                    "source_input_dwg": str(staged_source),
                    "source_converted_dxf": str(converted_dxf),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        (shared_dir / "frames.json").write_text("[]", encoding="utf-8")
        (shared_dir / "sheet_sets.json").write_text("[]", encoding="utf-8")
        (shared_dir / "titleblock_extracts.json").write_text("[]", encoding="utf-8")
        (shared_dir / "audit_roi_context.json").write_text("{}", encoding="utf-8")
        converted_dxf.write_text("0\nEOF\n", encoding="utf-8")
        return SharedPrepArtifacts(
            shared_dir=shared_dir,
            source_input_dwg=staged_source,
            source_converted_dxf=converted_dxf,
            font_preflight_summary={
                "filename": staged_source.name,
                "status": "ok",
                "missing_fonts": [],
                "detected_style_count": 0,
                "missing_style_count": 0,
                "font_replacement_applied": font_replace_policy == "replace_missing",
                "replacement_font": font_replacement_font,
                "replacement_fonts": font_replacement_fonts or {},
                "font_compatibility_mode": font_compatibility_mode,
                "replaced_style_count": 0,
            },
            frames=[],
            sheet_sets=[],
        )


class FakeSharedPrepServiceWithWorkload(FakeSharedPrepService):
    def prepare(self, **kwargs: Any) -> SharedPrepArtifacts:
        artifacts = super().prepare(**kwargs)
        bbox = BBox(xmin=0, ymin=0, xmax=100, ymax=100)
        frame = FrameMeta(
            runtime=FrameRuntime(
                frame_id="frame-a1",
                source_file=artifacts.source_converted_dxf,
                cad_source_file=artifacts.source_input_dwg,
                outer_bbox=bbox,
                paper_variant_id="CNPE_A1",
            ),
        )
        sheet_frame = FrameMeta(
            runtime=FrameRuntime(
                frame_id="frame-a4-master",
                source_file=artifacts.source_converted_dxf,
                cad_source_file=artifacts.source_input_dwg,
                outer_bbox=bbox,
                paper_variant_id="CNPE_A4",
            ),
        )
        pages = [
            PageInfo(page_index=1, outer_bbox=bbox, has_titleblock=True, frame_meta=sheet_frame),
            PageInfo(page_index=2, outer_bbox=bbox, has_titleblock=False, frame_meta=None),
        ]
        sheet_set = SheetSet(
            cluster_id="sheet-a4",
            paper="A4",
            page_total=2,
            pages=pages,
            master_page=pages[0],
        )
        (artifacts.shared_dir / "frames.json").write_text(
            json.dumps([frame.model_dump(mode="json")], ensure_ascii=False),
            encoding="utf-8",
        )
        (artifacts.shared_dir / "sheet_sets.json").write_text(
            json.dumps([sheet_set.model_dump(mode="json")], ensure_ascii=False),
            encoding="utf-8",
        )
        return SharedPrepArtifacts(
            shared_dir=artifacts.shared_dir,
            source_input_dwg=artifacts.source_input_dwg,
            source_converted_dxf=artifacts.source_converted_dxf,
            font_preflight_summary=artifacts.font_preflight_summary,
            frames=[frame],
            sheet_sets=[sheet_set],
        )


def test_grouped_batch_exposes_workload_from_shared_prep(
    monkeypatch,
    tmp_path: Path,
) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    with TestClient(
        create_app(
            job_processor=FakeJobProcessor(),
            shared_prep_service=FakeSharedPrepServiceWithWorkload(),
        ),
    ) as client:
        params = _deliverable_params()
        params["unit_no"] = "1"
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(params, ensure_ascii=False),
                "run_audit_check": "true",
            },
            files=[("files[]", ("20261RS-JGS65.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        group_summary = response.json()["jobs"][0]
        detail = _poll_job(client, group_summary["job_id"], timeout_sec=5.0)

        assert detail["status"] == "succeeded"
        assert detail["workload"]["initial_workload_a1"] == 1.25
        assert detail["workload"]["final_workload_a1"] == 1.25
        assert detail["effective_workload"] == 1.25


def test_create_batch_with_run_audit_check_returns_group_detail_and_children(
    monkeypatch,
    tmp_path: Path,
) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    with TestClient(
        create_app(
            job_processor=FakeJobProcessor(),
            shared_prep_service=FakeSharedPrepService(),
        ),
    ) as client:
        params = _deliverable_params()
        params["unit_no"] = "1"
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(params, ensure_ascii=False),
                "run_audit_check": "true",
            },
            files=[("files[]", ("20261RS-JGS65.dwg", b"dwg", "application/acad"))],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 1
        group_summary = payload["jobs"][0]
        assert group_summary["is_group"] is True
        assert group_summary["run_audit_check"] is True
        assert len(group_summary["child_job_ids"]) == 2

        detail = _poll_job(client, group_summary["job_id"], timeout_sec=5.0)
        assert detail["is_group"] is True
        assert detail["run_audit_check"] is True
        assert detail["flags"] == ["[DRAW001 (20261RS-JGS65-001)] PAPER_SIZE_AUTO_FIXED"]
        assert detail["workload"]["initial_workload_a1"] == 0.0
        assert detail["workload"]["final_workload_a1"] == 0.0
        assert detail["effective_workload"] == 0.0
        assert detail["artifacts"]["preview_available"] is True
        assert detail["artifacts"]["preview_mode"] == "annotated"
        assert detail["artifacts"]["preview_download_url"] == f"/api/jobs/{group_summary['job_id']}/download/preview"
        assert len(detail["children"]) == 2
        assert {child["task_role"] for child in detail["children"]} == {
            "deliverable_main",
            "audit_check",
        }
        assert {child["plot_style_key"] for child in detail["children"]} == {"red_wider"}
        assert {child["plot_resource_mode"] for child in detail["children"]} == {
            "slot_private_with_shared_mirror",
        }
        assert all(child["slot_id"] is not None for child in detail["children"])
        assert all(child["ctb_path"] for child in detail["children"])
        children = {child["task_role"]: child for child in detail["children"]}
        assert children["deliverable_main"]["artifacts"]["preview_available"] is True
        assert children["deliverable_main"]["artifacts"]["preview_mode"] == "plain"
        assert children["audit_check"]["artifacts"]["preview_available"] is True
        assert children["audit_check"]["artifacts"]["preview_mode"] == "annotated"

        preview_download = client.get(f"/api/jobs/{group_summary['job_id']}/download/preview")
        assert preview_download.status_code == 200
        assert preview_download.content == b"%PDF-annotated"

        listing = client.get("/api/jobs")
        assert listing.status_code == 200
        items = listing.json()["items"]
        assert len(items) == 1
        assert items[0]["job_id"] == group_summary["job_id"]


def test_grouped_batches_can_run_children_concurrently(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import threading

    class ConcurrentTrackingProcessor:
        def __init__(self) -> None:
            self.inner = FakeJobProcessor()
            self.current = 0
            self.max_seen = 0
            self.lock = threading.Lock()

        def __call__(self, job: Job) -> None:
            with self.lock:
                self.current += 1
                self.max_seen = max(self.max_seen, self.current)
            try:
                time.sleep(0.15)
                self.inner(job)
            finally:
                with self.lock:
                    self.current -= 1

    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    processor = ConcurrentTrackingProcessor()
    params = _deliverable_params()
    params["subitem_name_en"] = "Example Subitem"
    params["album_title_en"] = "Example Album"
    params["unit_no"] = "1"

    with TestClient(
        create_app(
            job_processor=processor,
            shared_prep_service=FakeSharedPrepService(),
        ),
    ) as client:
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(params, ensure_ascii=False),
                "run_audit_check": "true",
            },
            files=[
                ("files[]", ("18185NE-JGS11.dwg", b"dwg-a", "application/acad")),
                ("files[]", ("20261RS-JGS65.dwg", b"dwg-b", "application/acad")),
            ],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 2
        for item in payload["jobs"]:
            detail = _poll_job(client, item["job_id"], timeout_sec=8.0)
            assert detail["status"] == "succeeded"

    assert processor.max_seen >= 2


def test_multi_file_batch_keeps_backlog_visible_until_worker_capacity_frees(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import threading

    class SlowTrackingProcessor:
        def __init__(self) -> None:
            self.inner = FakeJobProcessor()
            self.current = 0
            self.max_seen = 0
            self.lock = threading.Lock()

        def __call__(self, job: Job) -> None:
            with self.lock:
                self.current += 1
                self.max_seen = max(self.max_seen, self.current)
            try:
                time.sleep(0.4)
                self.inner(job)
            finally:
                with self.lock:
                    self.current -= 1

    def _wait_for_runtime_health(
        client: TestClient,
        *,
        predicate,
        timeout_sec: float = 3.0,
    ) -> dict:
        deadline = time.time() + timeout_sec
        last_payload: dict | None = None
        while time.time() < deadline:
            response = client.get("/api/system/health")
            assert response.status_code == 200
            payload = response.json()
            last_payload = payload
            if predicate(payload):
                return payload
            time.sleep(0.05)
        raise AssertionError(f"runtime health did not satisfy predicate: {last_payload}")

    _configure_api_env(monkeypatch, tmp_path)
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_FILES", "20")
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_TOTAL_MB", "50")
    monkeypatch.setenv("FANBAN_CAD_RUNTIME__SLOT_COUNT", "4")
    monkeypatch.setenv("FANBAN_CONCURRENCY__MAX_JOBS", "4")
    SpecLoader.clear_cache()
    reload_config()

    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    processor = SlowTrackingProcessor()
    with TestClient(
        create_app(
            job_processor=processor,
            font_preflight_service=cast(Any, FakeFontPreflightService()),
        )
    ) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(_deliverable_params(), ensure_ascii=False)},
            files=[
                ("files[]", (f"2026-A{index:02d}.dwg", f"dwg-{index}".encode(), "application/acad"))
                for index in range(1, 9)
            ],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 8

        health = _wait_for_runtime_health(
            client,
            predicate=lambda item: item["active_jobs"] >= 4,
        )
        assert health["active_jobs"] == 4
        assert health["queue_depth"] == 4

        details = [client.get(f"/api/jobs/{job['job_id']}").json() for job in payload["jobs"]]
        assert sum(detail["slot_id"] is not None for detail in details) == 4
        assert sum(detail["slot_id"] is None for detail in details) == 4

        for job in payload["jobs"]:
            detail = _poll_job(client, job["job_id"], timeout_sec=8.0)
            assert detail["status"] == "succeeded"

    assert processor.max_seen == 4


def test_slot_bound_phase_allows_next_wave_to_start_before_docs_finish(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import threading

    class PhaseAwareProcessor:
        def __init__(self) -> None:
            self.lock = threading.Lock()
            self.cad_current = 0
            self.max_cad_seen = 0
            self.cad_start: dict[str, float] = {}
            self.doc_end: dict[str, float] = {}

        def __call__(self, job: Job) -> None:
            self._run_cad(job)
            self._run_docs(job)

        def execute_slot_bound_phase(self, job: Job):
            self._run_cad(job)

            def finish() -> None:
                self._run_docs(job)

            return finish

        def _run_cad(self, job: Job) -> None:
            with self.lock:
                self.cad_current += 1
                self.max_cad_seen = max(self.max_cad_seen, self.cad_current)
            job.mark_running(stage="EXPORT_PDF_AND_DWG")
            self.cad_start[job.job_id] = time.monotonic()
            time.sleep(0.2)
            with self.lock:
                self.cad_current -= 1

        def _run_docs(self, job: Job) -> None:
            job.progress.stage = "GENERATE_DOCS"
            time.sleep(0.6)
            job.mark_succeeded()
            self.doc_end[job.job_id] = time.monotonic()

    _configure_api_env(monkeypatch, tmp_path)
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_FILES", "20")
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_TOTAL_MB", "50")
    monkeypatch.setenv("FANBAN_CONCURRENCY__MAX_JOBS", "2")
    SpecLoader.clear_cache()
    reload_config()

    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    processor = PhaseAwareProcessor()
    with TestClient(create_app(job_processor=processor)) as client:
        response = client.post(
            "/api/jobs/batch",
            data={"params_json": json.dumps(_deliverable_params(), ensure_ascii=False)},
            files=[
                ("files[]", (f"2026-P{index:02d}.dwg", f"dwg-{index}".encode(), "application/acad"))
                for index in range(1, 5)
            ],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 4

        for job in payload["jobs"]:
            detail = _poll_job(client, job["job_id"], timeout_sec=8.0)
            assert detail["status"] == "succeeded"

    assert processor.max_cad_seen == 2
    cad_starts = sorted(processor.cad_start.values())
    doc_ends = sorted(processor.doc_end.values())
    assert len(cad_starts) == 4
    assert len(doc_ends) == 4
    assert cad_starts[2] < doc_ends[0]


def test_grouped_batch_keeps_pending_groups_in_external_queue(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import threading

    class SlowTrackingProcessor:
        def __init__(self) -> None:
            self.inner = FakeJobProcessor()
            self.current = 0
            self.max_seen = 0
            self.lock = threading.Lock()

        def __call__(self, job: Job) -> None:
            with self.lock:
                self.current += 1
                self.max_seen = max(self.max_seen, self.current)
            try:
                time.sleep(0.4)
                self.inner(job)
            finally:
                with self.lock:
                    self.current -= 1

    def _wait_for_runtime_health(
        client: TestClient,
        *,
        predicate,
        timeout_sec: float = 3.0,
    ) -> dict:
        deadline = time.time() + timeout_sec
        last_payload: dict | None = None
        while time.time() < deadline:
            response = client.get("/api/system/health")
            assert response.status_code == 200
            payload = response.json()
            last_payload = payload
            if predicate(payload):
                return payload
            time.sleep(0.05)
        raise AssertionError(f"runtime health did not satisfy predicate: {last_payload}")

    _configure_api_env(monkeypatch, tmp_path)
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_FILES", "20")
    monkeypatch.setenv("FANBAN_UPLOAD_LIMITS__MAX_TOTAL_MB", "50")
    monkeypatch.setenv("FANBAN_CAD_RUNTIME__SLOT_COUNT", "4")
    monkeypatch.setenv("FANBAN_CONCURRENCY__MAX_JOBS", "4")
    SpecLoader.clear_cache()
    reload_config()

    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    processor = SlowTrackingProcessor()
    with TestClient(
        create_app(
            job_processor=processor,
            shared_prep_service=FakeSharedPrepService(),
            font_preflight_service=cast(Any, FakeFontPreflightService()),
        )
    ) as client:
        params = _deliverable_params()
        params["subitem_name_en"] = "Example Subitem"
        params["album_title_en"] = "Example Album"
        params["unit_no"] = "1"
        response = client.post(
            "/api/jobs/batch",
            data={
                "params_json": json.dumps(params, ensure_ascii=False),
                "run_audit_check": "true",
            },
            files=[
                ("files[]", (f"2026-G{index:02d}.dwg", f"dwg-{index}".encode(), "application/acad"))
                for index in range(1, 5)
            ],
        )

        assert response.status_code == 201
        payload = response.json()
        assert len(payload["jobs"]) == 4

        health = _wait_for_runtime_health(
            client,
            predicate=lambda item: item["active_groups"] >= 2 and item["active_jobs"] >= 4,
        )
        assert health["active_groups"] == 2
        assert health["active_jobs"] == 4
        assert health["queue_depth"] == 2

        details = [client.get(f"/api/jobs/{item['job_id']}").json() for item in payload["jobs"]]
        assert sum(detail["status"] == "running" for detail in details) == 2
        assert sum(detail["status"] == "queued" for detail in details) == 2

        for item in payload["jobs"]:
            detail = _poll_job(client, item["job_id"], timeout_sec=8.0)
            assert detail["status"] == "succeeded"

    assert processor.max_seen == 4


def test_startup_recovery_marks_stale_jobs_failed(monkeypatch, tmp_path: Path) -> None:
    _configure_api_env(monkeypatch, tmp_path)
    repo_root = Path(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    from API.app.main import create_app

    storage_root = tmp_path / "storage"
    stale_job = Job(
        job_id="job-stale-1",
        job_type=JobType.DELIVERABLE,
        project_no="2016",
        status=JobStatus.RUNNING,
        params=_deliverable_params(),
    )
    stale_job.progress.stage = "A4_MULTIPAGE_GROUPING"
    stale_job.progress.message = "完成阶段: A4_MULTIPAGE_GROUPING"
    job_dir = storage_root / "jobs" / stale_job.job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    (job_dir / "job.json").write_text(
        stale_job.model_dump_json(indent=2),
        encoding="utf-8",
    )

    with TestClient(create_app(job_processor=FakeJobProcessor())) as client:
        response = client.get(f"/api/jobs/{stale_job.job_id}")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "failed"
    assert "service_restarted_before_completion" in payload["errors"]
    assert payload["failure_reason"] == "服务重启/中断，任务未完成"
    assert payload["stage_context"] == "中断前最后完成阶段：A4 多页合并"

    list_response = client.get("/api/jobs")
    assert list_response.status_code == 200
    list_payload = list_response.json()
    [summary] = list_payload["items"]
    assert summary["failure_reason"] == "服务重启/中断，任务未完成"
    assert summary["stage_context"] == "中断前最后完成阶段：A4 多页合并"
