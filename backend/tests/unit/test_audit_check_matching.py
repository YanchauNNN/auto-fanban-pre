from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.audit_check.lexicon import AuditLexiconLoader
from src.audit_check.matcher import AuditMatchEngine
from src.audit_check.models import AuditLexicon, ScanTextItem


def _build_lexicon_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["project", "1418", "2016", "2026", "note"])
    ws.append(["lexicon", "CHANGJIANG", "JINQIMEN", "XUWEI", "note"])
    ws.append([None, "1418YNI-JGS01", "20161NH-JGS01-002", "20261NH-JGS01-002", None])
    ws.append([None, "SHARED", "SHARED", "OTHER", None])
    ws.append([None, "HL", "JD", "XZ", None])
    wb.save(path)
    return path


def _build_project_name_lexicon_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["project", "2016", "2026"])
    ws.append(["lexicon", "浙江金七门核电厂 1、2 号 机 组", "江苏徐圩核能供热发电厂一期工程"])
    wb.save(path)
    return path


def test_lexicon_loader_includes_row1_and_row2_and_ignores_note_columns(tmp_path: Path) -> None:
    workbook = _build_lexicon_workbook(tmp_path / "lexicon.xlsx")

    lexicon = AuditLexiconLoader().load(workbook)

    assert lexicon.project_options == ["1418", "2016", "2026"]
    assert "1418" in lexicon.allowed_texts["1418"]
    assert "CHANGJIANG" in lexicon.allowed_texts["1418"]
    assert "JINQIMEN" in lexicon.foreign_texts["1418"]
    assert "2026" in lexicon.foreign_texts["1418"]
    assert "SHARED" not in lexicon.foreign_texts["1418"]


def test_lexicon_loader_streams_worksheet_once_and_closes_workbook(monkeypatch) -> None:
    class _StreamingWorksheet:
        def __init__(self) -> None:
            self.iteration_count = 0

        def iter_rows(self, *, values_only: bool):
            assert values_only is True
            self.iteration_count += 1
            yield ("备注", "2016", "2026")
            yield ("项目名称", "旧项目", "新项目")

        def cell(self, *args, **kwargs):
            raise AssertionError("read-only worksheets must not be accessed cell by cell")

    class _StreamingWorkbook:
        sheetnames = ["词库"]

        def __init__(self) -> None:
            self.worksheet = _StreamingWorksheet()
            self.closed = False

        def __getitem__(self, name: str):
            assert name == "词库"
            return self.worksheet

        def close(self) -> None:
            self.closed = True

    workbook = _StreamingWorkbook()
    monkeypatch.setattr(
        "src.audit_check.lexicon.load_workbook",
        lambda *args, **kwargs: workbook,
    )

    lexicon = AuditLexiconLoader().load("unused.xlsx")

    assert workbook.worksheet.iteration_count == 1
    assert workbook.closed is True
    assert lexicon.project_options == ["2016", "2026"]
    assert lexicon.allowed_texts["2016"] == {"2016", "旧项目"}
    assert lexicon.allowed_texts["2026"] == {"2026", "新项目"}


def test_match_engine_reports_code_like_project_no_and_short_code_but_suppresses_noise(
    tmp_path: Path,
) -> None:
    workbook = _build_lexicon_workbook(tmp_path / "lexicon.xlsx")
    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1418",
        items=[
            ScanTextItem(raw_text="20161NH-JGS01-002", entity_type="TEXT"),
            ScanTextItem(raw_text="JD1NHT11001B25C42SD", entity_type="TEXT"),
            ScanTextItem(raw_text="2026.03.12", entity_type="TEXT"),
            ScanTextItem(raw_text="2026.04", entity_type="TEXT"),
            ScanTextItem(raw_text="645X600X2016", entity_type="TEXT"),
            ScanTextItem(raw_text="RVV2016P", entity_type="TEXT"),
            ScanTextItem(raw_text="ABCD2016X", entity_type="TEXT"),
            ScanTextItem(raw_text="smooth", entity_type="TEXT"),
        ],
    )

    matched = {(item.matched_text, item.context_kind, item.confidence) for item in findings}
    assert ("2016", "code_like", "high") in matched
    assert ("JD", "code_like", "high") in matched
    assert all(item.raw_text != "2026.03.12" for item in findings)
    assert all(item.raw_text != "2026.04" for item in findings)
    assert all(item.raw_text != "645X600X2016" for item in findings)
    assert all(item.raw_text != "RVV2016P" for item in findings)
    assert all(item.raw_text != "smooth" for item in findings)
    assert any(item.raw_text == "ABCD2016X" and item.matched_text == "2016" for item in findings)


def test_match_engine_uses_field_context_to_promote_project_sensitive_hits(tmp_path: Path) -> None:
    workbook = _build_lexicon_workbook(tmp_path / "lexicon.xlsx")
    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1418",
        items=[
            ScanTextItem(
                raw_text="6TT2016GX",
                entity_type="ATTRIB",
                field_context="titleblock_internal_code",
            ),
        ],
    )

    assert len(findings) == 1
    assert findings[0].matched_text == "2016"
    assert findings[0].context_kind == "titleblock_internal_code"
    assert findings[0].confidence == "high"


def test_match_engine_suppresses_project_number_inside_titleblock_date(tmp_path: Path) -> None:
    workbook = _build_lexicon_workbook(tmp_path / "lexicon.xlsx")
    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1418",
        items=[
            ScanTextItem(raw_text="2026.02", entity_type="ATTRIB", field_context="titleblock_date"),
            ScanTextItem(raw_text="2026/02", entity_type="ATTRIB", field_context="titleblock_date"),
        ],
    )

    assert findings == []


def test_match_engine_suppresses_project_no_inside_long_numeric_run_before_han_text(
    tmp_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["project", "1907", "2016", "note"])
    ws.append(["lexicon", "SANMEN", "JINQIMEN", "note"])
    ws.append([None, "1907", "2016", None])
    workbook = tmp_path / "lexicon-1907.xlsx"
    wb.save(workbook)

    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="2016",
        items=[
            ScanTextItem(
                raw_text="7788991907一一二二",
                entity_type="TEXT",
            ),
        ],
    )

    assert all(item.matched_text != "1907" for item in findings)


def test_match_engine_only_whitelists_exact_three_letters_plus_project_no_plus_one_letter(
    tmp_path: Path,
) -> None:
    workbook = _build_lexicon_workbook(tmp_path / "lexicon.xlsx")
    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1418",
        items=[
            ScanTextItem(raw_text="ABC2016X", entity_type="TEXT"),
            ScanTextItem(raw_text="ABCD2016X", entity_type="TEXT"),
            ScanTextItem(raw_text="12ABC2016X", entity_type="TEXT"),
        ],
    )

    assert all(item.raw_text != "ABC2016X" for item in findings)
    assert any(item.raw_text == "ABCD2016X" and item.matched_text == "2016" for item in findings)
    assert any(item.raw_text == "12ABC2016X" and item.matched_text == "2016" for item in findings)


def test_match_engine_whitelists_prefixed_embed_identifier_for_all_scanned_text_kinds() -> None:
    lexicon = AuditLexicon(
        project_options=["1907", "2026"],
        allowed_texts={"1907": {"1907"}, "2026": {"2026"}},
        foreign_texts={"1907": {"2026"}, "2026": {"1907"}},
        token_projects={"1907": {"1907"}, "2026": {"2026"}},
    )
    engine = AuditMatchEngine(lexicon)

    protected_types = [
        "DBText",
        "MText",
        "AttributeReference",
        "AttributeDefinition",
        "Dimension",
        "MLeader",
        "TableCell",
    ]
    findings = engine.evaluate(
        project_no="1907",
        items=[
            *[
                ScanTextItem(
                    raw_text="2RCFVV2026P," if index == 0 else "2RCFVV2026P",
                    entity_type=entity_type,
                    entity_handle=f"PROTECTED-{index}",
                )
                for index, entity_type in enumerate(protected_types)
            ],
            ScanTextItem(
                raw_text="20262RC-JGS38-001",
                entity_type="DBText",
                entity_handle="INTERNAL-CODE",
            ),
        ],
    )

    assert all("FVV2026P" not in finding.raw_text for finding in findings)
    assert any(
        finding.raw_text == "20262RC-JGS38-001" and finding.matched_text == "2026"
        for finding in findings
    )


def test_match_engine_ignores_yaml_project_no_context_whitelist(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["project", "1907", "2026"])
    ws.append(["lexicon", "SANMEN", "XUWEI"])
    ws.append([None, "1907", "2026"])
    workbook = tmp_path / "lexicon-project-whitelist.xlsx"
    wb.save(workbook)

    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="2026",
        items=[
            ScanTextItem(raw_text="资料单1907", entity_type="TEXT"),
            ScanTextItem(raw_text="提资1907", entity_type="TEXT"),
            ScanTextItem(raw_text="提资单号1907", entity_type="TEXT"),
            ScanTextItem(raw_text="提资单号：1907", entity_type="TEXT"),
            ScanTextItem(raw_text="设计依据1907", entity_type="TEXT"),
        ],
    )

    assert [finding.raw_text for finding in findings if finding.matched_text == "1907"] == [
        "设计依据1907",
    ]


def test_match_engine_ignores_spaces_when_matching_project_name(tmp_path: Path) -> None:
    workbook = _build_project_name_lexicon_workbook(tmp_path / "project-name-lexicon.xlsx")
    lexicon = AuditLexiconLoader().load(workbook)
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="2026",
        items=[
            ScanTextItem(
                raw_text="3.根据浙江金七门核电厂1、2号机组2016-J01ZHC04《厂址设计参数》",
                entity_type="TEXT",
            ),
        ],
    )

    assert any(
        item.matched_text == "浙江金七门核电厂 1、2 号 机 组"
        and item.raw_text.startswith("3.根据浙江金七门核电厂1、2号机组")
        for item in findings
    )


def test_match_engine_suppresses_project_number_inside_long_numeric_run_but_keeps_album_code() -> None:
    lexicon = AuditLexicon(
        project_options=["1907", "1818", "1915", "2026", "2035"],
        allowed_texts={"1907": set()},
        foreign_texts={"1907": {"1818", "1915", "2026", "2035"}},
        token_projects={
            "1818": {"1818"},
            "1915": {"1915"},
            "2026": {"2026"},
            "2035": {"2035"},
        },
    )
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1907",
        items=[
            ScanTextItem(raw_text="18187", entity_type="DBText"),
            ScanTextItem(raw_text="20350", entity_type="DBText"),
            ScanTextItem(raw_text="120261", entity_type="DBText"),
            ScanTextItem(raw_text="20261NH-JGS03-001", entity_type="DBText"),
        ],
    )

    assert [(finding.raw_text, finding.matched_text) for finding in findings] == [
        ("20261NH-JGS03-001", "2026"),
    ]


def test_match_engine_suppresses_plain_four_digit_project_token_on_configured_annotation_layer() -> None:
    lexicon = AuditLexicon(
        project_options=["1907", "1915", "2035"],
        allowed_texts={"1907": set()},
        foreign_texts={"1907": {"1915", "2035"}},
        token_projects={"1915": {"1915"}, "2035": {"2035"}},
    )
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1907",
        items=[
            ScanTextItem(
                raw_text="2035",
                entity_type="DBText",
                layer_name="NHJTTT_OpenLine",
            ),
            ScanTextItem(
                raw_text="1915",
                entity_type="DBText",
                layer_name="NHJTTT_OpenLine",
            ),
            ScanTextItem(raw_text="2035", entity_type="MText", layer_name="NHJTTT_OpenLine"),
            ScanTextItem(raw_text="1915", entity_type="DBText", layer_name="0"),
        ],
    )

    assert [(finding.raw_text, finding.entity_type) for finding in findings] == [
        ("2035", "MText"),
        ("1915", "DBText"),
    ]


def test_match_engine_uses_frame_and_roi_semantics_for_exact_four_digit_project_tokens() -> None:
    lexicon = AuditLexicon(
        project_options=["1907", "1915", "2026", "2035"],
        allowed_texts={"1907": set()},
        foreign_texts={"1907": {"1915", "2026", "2035"}},
        token_projects={"1915": {"1915"}, "2026": {"2026"}, "2035": {"2035"}},
    )
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1907",
        items=[
            ScanTextItem(
                raw_text="2035",
                entity_type="DBText",
                layer_name="0",
                internal_code="19076NH-JGS03-013",
            ),
            ScanTextItem(
                raw_text="1915",
                entity_type="MText",
                layer_name="任意图层",
                internal_code="19076NH-JGS03-023",
            ),
            ScanTextItem(
                raw_text="2026",
                entity_type="MText",
                layer_name="0",
                field_context="titleblock_engineering_no",
                internal_code="19076NH-JGS03-024",
            ),
            ScanTextItem(
                raw_text="设计依据1915",
                entity_type="MText",
                layer_name="0",
                internal_code="19076NH-JGS03-025",
            ),
            ScanTextItem(
                raw_text="20261NH-JGS03-001",
                entity_type="MText",
                layer_name="0",
                internal_code="19076NH-JGS03-026",
            ),
        ],
    )

    assert [(finding.raw_text, finding.matched_text) for finding in findings] == [
        ("2026", "2026"),
        ("设计依据1915", "1915"),
        ("20261NH-JGS03-001", "2026"),
    ]


def test_match_engine_keeps_layer_fallback_when_frame_semantics_are_unavailable() -> None:
    lexicon = AuditLexicon(
        project_options=["1907", "1915", "2035"],
        allowed_texts={"1907": set()},
        foreign_texts={"1907": {"1915", "2035"}},
        token_projects={"1915": {"1915"}, "2035": {"2035"}},
    )
    engine = AuditMatchEngine(lexicon)

    findings = engine.evaluate(
        project_no="1907",
        items=[
            ScanTextItem(
                raw_text="2035",
                entity_type="DBText",
                layer_name="NHJTTT_OpenLine",
            ),
            ScanTextItem(
                raw_text="1915",
                entity_type="DBText",
                layer_name="0",
            ),
        ],
    )

    assert [(finding.raw_text, finding.matched_text) for finding in findings] == [
        ("1915", "1915"),
    ]
