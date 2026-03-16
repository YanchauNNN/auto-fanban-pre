from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from src.doc_gen.design import DesignFileGenerator
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    PageInfo,
    SheetSet,
    TitleblockFields,
)


def _make_frame(
    seq: int,
    *,
    paper_size_text: str,
    discipline: str,
    paper_variant_id: str | None = None,
) -> FrameMeta:
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
        paper_variant_id=paper_variant_id,
    )
    titleblock = TitleblockFields(
        internal_code=f"1234567-JG001-{seq:03d}",
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"\u56fe\u7eb8{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
        page_total=1,
        paper_size_text=paper_size_text,
        discipline=discipline,
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        engineering_no="1234",
        subitem_no="JG001",
        subitem_name="\u5b50\u9879\u540d\u79f0",
        discipline="\u7ed3\u6784",
        album_title_cn="\u6d4b\u8bd5\u56fe\u518c",
        cover_revision="A",
        doc_status="CFC",
        wbs_code="WBS-001",
        file_category="\u56fe\u7eb8",
        classification="\u975e\u5bc6",
        work_hours="88",
    )
    derived = DerivedFields(
        album_internal_code="1234567-JG001",
        cover_external_code="JD1NHT11F01B25C42SD",
        cover_internal_code="1234567-JG001-FM",
        cover_title_cn="\u6d4b\u8bd5\u56fe\u518c\u5c01\u9762",
        catalog_external_code="JD1NHT11T01B25C42SD",
        catalog_internal_code="1234567-JG001-TM",
        catalog_title_cn="\u6d4b\u8bd5\u56fe\u518c\u76ee\u5f55",
        catalog_revision="A",
        catalog_page_total=2,
        design_phase="\u65bd\u5de5\u56fe\u8bbe\u8ba1",
    )
    frame_001 = _make_frame(
        1,
        paper_size_text="A4",
        discipline="\u7ed3\u6784",
        paper_variant_id="CNPE_A4",
    )
    master_page = PageInfo(
        page_index=1,
        outer_bbox=frame_001.runtime.outer_bbox,
        has_titleblock=True,
        frame_meta=frame_001,
    )
    return DocContext(
        params=params,
        derived=derived,
        frames=[
            _make_frame(
                2,
                paper_size_text="A 0",
                discipline="\u7ed3\u6784",
                paper_variant_id="CNPE_A0",
            ),
            _make_frame(
                3,
                paper_size_text="1\nA",
                discipline="\u7ed3\u6784",
                paper_variant_id="CNPE_A4H",
            ),
        ],
        sheet_sets=[
            SheetSet(
                cluster_id="sheet-set-001",
                page_total=1,
                pages=[master_page],
                master_page=master_page,
            ),
        ],
    )


def test_design_includes_001_and_maps_paper_size_and_discipline(temp_dir: Path) -> None:
    gen = DesignFileGenerator()
    ctx = _build_context()
    output_xlsx = temp_dir / "\u8bbe\u8ba1\u6587\u4ef6.xlsx"

    gen._write_design(
        template_path="documents_bin/\u8bbe\u8ba1\u6587\u4ef6\u6a21\u677f.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    assert ws["S2"].value == "A4\u56fe\u7eb8"
    assert ws["S3"].value == "A4\u6587\u4ef6"
    assert ws["E4"].value == "1234567-JG001-001"
    assert ws["S4"].value == "A4"
    assert ws["N4"].value == "\u7ed3\u6784"
    assert ws["E5"].value == "1234567-JG001-002"
    assert ws["S5"].value == "A0"
    assert ws["N5"].value == "\u7ed3\u6784"
    assert ws["E6"].value == "1234567-JG001-003"
    assert ws["S6"].value == "A4"
