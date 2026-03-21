from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from src.doc_gen.design import DesignFileGenerator
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    PageInfo,
    SheetSet,
    TitleblockFields,
)


class DummyPDFExporter:
    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")


def _make_frame(seq: int, discipline: str = "结构") -> FrameMeta:
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
    )
    titleblock = TitleblockFields(
        internal_code=f"1234567-JG001-{seq:03d}",
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"图纸{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
        page_total=1,
        paper_size_text="A1",
        discipline=discipline,
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        engineering_no="1234",
        subitem_no="JG001",
        subitem_name="子项名称",
        discipline="结构",
        album_title_cn="测试图册",
        album_title_en="Test Album",
        cover_revision="A",
        doc_status="CFC",
        wbs_code="WBS-001",
        file_category="图纸",
        classification="非密",
        work_hours="88",
    )
    derived = DerivedFields(
        album_code="01",
        album_internal_code="1234567-JG001",
        cover_external_code="JD1NHT11F01B25C42SD",
        cover_internal_code="1234567-JG001-FM",
        cover_title_cn="测试图册封面",
        catalog_external_code="JD1NHT11T01B25C42SD",
        catalog_internal_code="1234567-JG001-TM",
        catalog_title_cn="测试图册目录",
        catalog_revision="B",
        catalog_page_total=3,
        design_phase="施工图设计",
    )
    return DocContext(params=params, derived=derived, frames=[_make_frame(1)])


def _build_context_with_sheet_set_001() -> DocContext:
    ctx = _build_context()
    frame_001 = deepcopy(_make_frame(1, discipline="结 构"))
    frame_001.titleblock.paper_size_text = "A4"
    frame_001.titleblock.page_total = 1

    frame_002 = deepcopy(_make_frame(2, discipline="结 构"))
    frame_002.titleblock.paper_size_text = "A 0"

    master_page = PageInfo(
        page_index=1,
        outer_bbox=frame_001.runtime.outer_bbox,
        has_titleblock=True,
        frame_meta=frame_001,
    )

    ctx.frames = [frame_002]
    ctx.sheet_sets = [
        SheetSet(
            cluster_id="sheet-set-001",
            page_total=7,
            pages=[master_page],
            master_page=master_page,
        ),
    ]
    return ctx


def test_design_write_rows_with_bindings(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    bindings = gen.spec.get_design_bindings()
    output_xlsx = temp_dir / "设计文件.xlsx"

    gen._write_design(
        template_path="documents_bin/设计文件模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    # 第2行：封面
    assert ws["D2"].value == "JD1NHT11F01B25C42SD"
    assert ws["E2"].value == "1234567-JG001-FM"
    assert ws["G2"].value == "测试图册封面"
    assert ws["O2"].value == "JG"
    assert ws["Q2"].value == "施工图设计"
    assert ws["T2"].value == 1
    assert ws["U2"].value == "图纸"

    # 第3行：目录
    assert ws["D3"].value == "JD1NHT11T01B25C42SD"
    assert ws["E3"].value == "1234567-JG001-TM"
    assert ws["G3"].value == "测试图册第01图册图纸(文件)目录"
    assert ws["H3"].value in ("", None)
    assert ws["T3"].value == 3
    assert ws["U3"].value == "图纸"

    # 第4行：图纸
    assert ws["D4"].value == "JD1NHT11001B25C42SD"
    assert ws["E4"].value == "1234567-JG001-001"
    assert ws["G4"].value == "图纸1"
    assert ws["N4"].value == "结构"
    assert ws["Z4"].value == "88"
    assert ws["U4"].value == "图纸"


def test_design_cover_and_catalog_rows_share_document_revision() -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    ctx.derived.document_revision = "C"
    ctx.derived.catalog_revision = "C"
    rows = gen._build_rows(ctx)

    assert rows[0]["revision"] == "C"
    assert rows[1]["revision"] == "C"


def test_design_generate_xlsx_only_without_pdf(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()

    output_xlsx = gen.generate(ctx, temp_dir)

    assert output_xlsx == temp_dir / "设计文件.xlsx"
    assert output_xlsx.exists()
    assert not (temp_dir / "设计文件.pdf").exists()

def test_design_uses_sheet_set_page_total_for_001(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context_with_sheet_set_001()

    output_xlsx = temp_dir / "design-sheetset-001.xlsx"
    gen._write_design(
        template_path=gen.spec.get_template_path("design", ctx.params.project_no),
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["E4"].value == "1234567-JG001-001"
    assert ws["T4"].value == 7


def test_design_column_n_keeps_only_chinese(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    ctx.params.discipline = "结构 Structural Engineering"
    ctx.frames = []

    output_xlsx = temp_dir / "设计文件.xlsx"
    gen._write_design(
        template_path="documents_bin/设计文件模板.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["N2"].value == "结构"
    assert ws["N3"].value == "结构"

def test_design_column_o_maps_discipline_code_from_1818_structure_hint(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    ctx.params.discipline = "\uc368\ubbd0\nStructure"
    ctx.frames = []

    output_xlsx = temp_dir / "设计文件.xlsx"
    gen._write_design(
        template_path="documents_bin/设计文件模板.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["N2"].value == "结构"
    assert ws["O2"].value == "JG"


def test_design_non_1818_catalog_row_title_uses_single_line_cn_only(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    expected_cn = f"{ctx.params.album_title_cn}第{ctx.derived.album_code}图册图纸(文件)目录"

    output_xlsx = temp_dir / "design-2016.xlsx"
    gen._write_design(
        template_path=gen.spec.get_template_path("design", ctx.params.project_no),
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["G3"].value == expected_cn
    assert ws["H3"].value in ("", None)


def test_design_catalog_row_title_matches_catalog_e10_for_1818(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    ctx.params.project_no = "1818"
    expected_cn = f"{ctx.params.album_title_cn}第{ctx.derived.album_code}图册图纸(文件)目录"

    output_xlsx = temp_dir / "design-1818.xlsx"
    gen._write_design(
        template_path=gen.spec.get_template_path("design", ctx.params.project_no),
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["G3"].value == expected_cn
    assert ws["H3"].value == "Test AlbumDOCUMENT CONTENTS"


def test_design_normalizes_multiline_drawing_titles_to_single_line(temp_dir: Path) -> None:
    gen = DesignFileGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    ctx.params.project_no = "1818"
    ctx.frames[0].titleblock.title_cn = "Alpha\nBeta"
    ctx.frames[0].titleblock.title_en = "Drawing\nTitle"

    output_xlsx = temp_dir / "design-title.xlsx"
    gen._write_design(
        template_path=gen.spec.get_template_path("design", ctx.params.project_no),
        output_path=output_xlsx,
        bindings=gen.spec.get_design_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["G4"].value == "AlphaBeta"
    assert ws["H4"].value == "DrawingTitle"
