from __future__ import annotations

import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import load_workbook

import src.doc_gen.cover as cover_module
from src.doc_gen.cover import CoverGenerator
from src.interfaces import GenerationError, IPDFExporter
from src.models import DerivedFields, DocContext, GlobalDocParams


class DummyPDFExporter:
    def export_docx_to_pdf(self, docx_path: Path, pdf_path: Path) -> None:
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")


def _build_context(project_no: str = "2016") -> DocContext:
    params = GlobalDocParams(
        project_no=project_no,
        cover_variant="通用",
        engineering_no="1234",
        subitem_no="JG001",
        subitem_name="这是一个很长的子项名称用于测试",
        subitem_name_en="Secondary Steel Shop",
        discipline="结构",
        doc_status="CFC",
        album_title_cn="这是一个很长的中文图册标题用于分割测试",
        album_title_en="Secondary steel shop drawings at elevation -8.800m",
        cover_revision="B",
    )
    derived = DerivedFields(
        album_internal_code="1234567-JG001",
        album_code="01",
        cover_external_code="JD1NHT11F01B25C42SD",
        design_phase="施工图设计",
        design_phase_en="Constructing Design",
        discipline_en="Structural Engineering",
    )
    return DocContext(params=params, derived=derived, frames=[])


def _read_cover_embedded_wb(docx_path: Path) -> Any:
    with zipfile.ZipFile(docx_path, "r") as zf:
        payload = zf.read("word/embeddings/Microsoft_Excel_Worksheet.xlsx")
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    return ws


def _read_cover_embedded_sheet_xmls(docx_path: Path) -> list[str]:
    with zipfile.ZipFile(docx_path, "r") as zf:
        embedded_names = [
            name
            for name in zf.namelist()
            if name.startswith("word/embeddings/") and name.lower().endswith(".xlsx")
        ]
        assert embedded_names
        payloads = [zf.read(name) for name in embedded_names]

    sheet_xmls: list[str] = []
    for payload in payloads:
        with zipfile.ZipFile(BytesIO(payload), "r") as zf:
            sheet_xmls.extend(
                zf.read(name).decode("utf-8")
                for name in zf.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            )
    assert sheet_xmls
    return sheet_xmls


def _assert_suppresses_excel_error_indicators(docx_path: Path) -> None:
    sheet_xmls = _read_cover_embedded_sheet_xmls(docx_path)
    assert any("<ignoredErrors" in xml for xml in sheet_xmls)
    assert any('numberStoredAsText="1"' in xml for xml in sheet_xmls)
    assert any('twoDigitTextYear="1"' in xml for xml in sheet_xmls)


def _assert_keeps_source_template_error_indicators_unpersisted(docx_path: Path) -> None:
    sheet_xmls = _read_cover_embedded_sheet_xmls(docx_path)
    assert all("<ignoredErrors" not in xml for xml in sheet_xmls)


def test_suppress_sheet_error_indicators_preserves_excel_namespace_prefixes() -> None:
    sheet_xml = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="x14ac xr xr2 xr3" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" xr:uid="{00000000-0001-0000-0000-000000000000}"><dimension ref="A1:B2"/><sheetFormatPr x14ac:dyDescent="0.15"/><sheetData><row r="1" x14ac:dyDescent="0.25"><c r="A1" t="s"><v>0</v></c></row></sheetData><pageSetup r:id="rId1"/></worksheet>'''

    updated = cover_module._suppress_sheet_error_indicators(sheet_xml).decode("utf-8")

    assert "<ignoredErrors>" in updated
    assert 'numberStoredAsText="1"' in updated
    assert 'mc:Ignorable="x14ac xr xr2 xr3"' in updated
    assert 'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"' in updated
    assert 'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"' in updated
    assert 'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"' in updated
    assert 'x14ac:dyDescent="0.15"' in updated
    assert 'xr:uid="{00000000-0001-0000-0000-000000000000}"' in updated
    assert "ns1:" not in updated
    assert "ns2:" not in updated
    assert "ns3:" not in updated


def test_cover_variant_template_mapping() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()

    ctx.params.cover_variant = "压力容器"
    assert gen._get_template_path(ctx).endswith("封面模板文件压力容器版.docx")

    ctx.params.cover_variant = "核安全设备"
    assert gen._get_template_path(ctx).endswith("封面模板文件核安全设备版.docx")

    ctx.params.project_no = "1818"
    ctx.params.cover_variant = "通用"
    assert gen._get_template_path(ctx).endswith("1818图册封面模板.docx")

    ctx.params.cover_variant = "压力容器"
    assert gen._get_template_path(ctx).endswith("1818图册压力容器封面模板.docx")

    ctx.params.cover_variant = "核安全设备"
    assert gen._get_template_path(ctx).endswith("1818图册核安全设备封面模板.docx")


def test_1818_cover_binding_expands_discipline_en_range() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))

    bindings = gen.spec.get_cover_bindings("1818")

    assert bindings["discipline_en"].cell == "I19:S19"


def test_1818_cover_apply_bindings_writes_discipline_en_range() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    bindings = gen.spec.get_cover_bindings("1818")
    writes: list[tuple[str, object]] = []

    gen._apply_bindings(
        bindings,
        {"discipline_en": "Structural Engineering"},
        read_cell=lambda cell: "",
        write_cell=lambda cell, value: writes.append((cell, value)),
    )

    assert ("I19:S19", "Structural Engineering") in writes


def test_prepare_data_1818_includes_discipline_en() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")

    data = gen._prepare_data(ctx)

    assert data["discipline_en"] == "Structural Engineering"


def test_cover_uses_cover_catalog_revision_for_output_and_binding() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    ctx.derived.cover_catalog_revision = "C"
    ctx.derived.catalog_revision = "C"
    ctx.derived.cover_internal_code = "1234567-JG001-FM"

    data = gen._prepare_data(ctx)
    output_stem = gen._build_output_stem(ctx)

    assert data["cover_revision"] == "C"
    assert output_stem == "JD1NHT11F01B25C42SDCCFC (1234567-JG001-FM)"


def test_write_cover_with_embedded_xlsx(temp_dir: Path) -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="2016")
    bindings = gen.spec.get_cover_bindings(ctx.params.project_no)
    data = gen._prepare_data(ctx)

    output_docx = temp_dir / "封面.docx"
    def force_embedded_fallback(*, output_path, bindings, data):  # noqa: ANN001
        raise RuntimeError("force embedded fallback")

    gen._write_cover_via_com = force_embedded_fallback  # type: ignore[method-assign]
    refresh_calls: list[Path] = []
    gen._refresh_cover_ole_preview_via_com = lambda output_path: refresh_calls.append(output_path)  # type: ignore[method-assign]

    gen._write_cover(
        template_path="documents_bin/封面模板文件.docx",
        output_path=output_docx,
        bindings=bindings,
        data=data,
        ctx=ctx,
    )

    ws = _read_cover_embedded_wb(output_docx)
    assert str(ws["A7"].value or "").strip() == "浙江金七门核电厂1、2号机组"
    assert ws["I11"].value == "1234"
    assert ws["I13"].value == "JG001"
    assert ws["I21"].value
    assert ws["I22"].value
    assert str(ws["N5"].value).strip().endswith("：B")

    chars = [str(ws[f"{col}29"].value or "") for col in "BCDEFGHIJKLMNOPQRST"]
    assert "".join(chars) == "JD1NHT11F01B25C42SD"
    assert refresh_calls == [output_docx]
    _assert_suppresses_excel_error_indicators(output_docx)


def test_write_cover_via_com_does_not_reopen_ole_after_zip_error_indicator_suppression(
    temp_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="2026")
    bindings = gen.spec.get_cover_bindings(ctx.params.project_no)
    data = gen._prepare_data(ctx)
    output_docx = temp_dir / "cover-com-refresh.docx"
    refresh_calls: list[Path] = []

    def fake_write_via_com(*, output_path, bindings, data):  # noqa: ANN001
        return None

    gen._write_cover_via_com = fake_write_via_com  # type: ignore[method-assign]
    gen._refresh_cover_ole_preview_via_com = lambda output_path: refresh_calls.append(output_path)  # type: ignore[method-assign]
    monkeypatch.setattr(cover_module, "suppress_cover_excel_error_indicators", lambda output_path: True)

    gen._write_cover(
        template_path="documents_bin/封面模板文件.docx",
        output_path=output_docx,
        bindings=bindings,
        data=data,
        ctx=ctx,
    )

    assert refresh_calls == []


def test_cover_templates_keep_error_indicator_suppression_out_of_source_templates() -> None:
    templates = sorted(Path("documents_bin").glob("*封面*.docx"))
    assert templates

    for template in templates:
        _assert_keeps_source_template_error_indicators_unpersisted(template)


def test_write_cover_updates_common_project_name_from_yaml(temp_dir: Path) -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="2026")
    bindings = gen.spec.get_cover_bindings(ctx.params.project_no)
    data = gen._prepare_data(ctx)

    output_docx = temp_dir / "封面-2026.docx"

    def force_embedded_fallback(*, output_path, bindings, data):  # noqa: ANN001
        raise RuntimeError("force embedded fallback")

    gen._write_cover_via_com = force_embedded_fallback  # type: ignore[method-assign]
    gen._refresh_cover_ole_preview_via_com = lambda output_path: None  # type: ignore[method-assign]

    gen._write_cover(
        template_path="documents_bin/封面模板文件.docx",
        output_path=output_docx,
        bindings=bindings,
        data=data,
        ctx=ctx,
    )

    ws = _read_cover_embedded_wb(output_docx)
    assert str(ws["A7"].value or "").strip() == "江苏徐圩核能供热发电厂一期工程"


def test_write_cover_embedded_fallback_fails_when_ole_preview_refresh_fails(temp_dir: Path) -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="2026")
    bindings = gen.spec.get_cover_bindings(ctx.params.project_no)
    data = gen._prepare_data(ctx)
    output_docx = temp_dir / "cover-preview-stale.docx"

    def force_embedded_fallback(*, output_path, bindings, data):  # noqa: ANN001
        raise RuntimeError("word ole write failed")

    def fail_refresh(output_path: Path) -> None:
        raise RuntimeError("preview cache stayed stale")

    gen._write_cover_via_com = force_embedded_fallback  # type: ignore[method-assign]
    gen._refresh_cover_ole_preview_via_com = fail_refresh  # type: ignore[method-assign]

    with pytest.raises(GenerationError, match="封面OLE预览刷新失败"):
        gen._write_cover(
            template_path="documents_bin/封面模板文件.docx",
            output_path=output_docx,
            bindings=bindings,
            data=data,
            ctx=ctx,
        )


def test_write_cover_1818_uses_com_when_no_embedded_xlsx(
    temp_dir: Path,
    monkeypatch,
) -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    bindings = gen.spec.get_cover_bindings("1818")
    data = gen._prepare_data(ctx)
    output_docx = temp_dir / "封面1818.docx"

    called = {"hit": False}

    def fake_write_cover_via_com(self, *, output_path, bindings, data):  # noqa: ANN001
        called["hit"] = True

    monkeypatch.setattr(CoverGenerator, "_write_cover_via_com", fake_write_cover_via_com)
    monkeypatch.setattr(cover_module, "suppress_cover_excel_error_indicators", lambda output_path: False)

    gen._write_cover(
        template_path="documents_bin/1818图册封面模板.docx",
        output_path=output_docx,
        bindings=bindings,
        data=data,
        ctx=ctx,
    )

    assert called["hit"] is True


def test_1818_cover_binding_writes_external_code_on_row_30() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    bindings = gen.spec.get_cover_bindings("1818")

    assert "cover_external_code" in bindings
    assert bindings["cover_external_code"].cell == "B30:T30"


def test_cn_title_split_keeps_protected_phrases_together() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))

    left, right = gen._split_cn_two_cells("NP厂房标高3.900m~屋面模板图")

    assert left == "NP厂房"
    assert right == "标高3.900m~屋面模板图"
    assert not (left.endswith("标") and right.startswith("高"))


def test_en_title_split_restores_common_missing_spaces() -> None:
    gen = CoverGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))

    left, right = gen._split_en_two_cells("NP Building Level3.900m~RoofFormwork")

    assert left == "NP Building"
    assert right == "Level 3.900m~Roof Formwork"
