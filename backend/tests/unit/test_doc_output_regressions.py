from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from src.doc_gen.catalog import CatalogGenerator
from src.doc_gen.cover import CoverGenerator
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    PageInfo,
    SheetSet,
    TitleblockFields,
)


class DummyCatalogPDFExporter:
    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")

    def count_pdf_pages(self, pdf_path: Path) -> int:
        return 2


def _make_frame(seq: int) -> FrameMeta:
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
    )
    titleblock = TitleblockFields(
        internal_code=f"1234567-JG001-{seq:03d}",
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"图纸{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
        page_total=1,
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_catalog_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        engineering_no="1234",
        subitem_no="JG001",
        album_title_cn="测试图册",
        cover_revision="A",
        doc_status="CFC",
    )
    derived = DerivedFields(
        album_code="01",
        cover_internal_code="1234567-JG001-FM",
        catalog_internal_code="1234567-JG001-TM",
        cover_external_code="JD1NHT11F01B25C42SD",
        catalog_external_code="JD1NHT11T01B25C42SD",
        cover_title_cn="测试图册封面",
        catalog_title_cn="测试图册目录",
        cover_title_en="Test Album Cover",
        catalog_title_en="Test Album Contents",
        catalog_revision="A",
    )
    frame_001 = _make_frame(1)
    master_page = PageInfo(
        page_index=1,
        outer_bbox=frame_001.runtime.outer_bbox,
        has_titleblock=True,
        frame_meta=frame_001,
    )
    return DocContext(
        params=params,
        derived=derived,
        frames=[_make_frame(2), _make_frame(3)],
        sheet_sets=[
            SheetSet(
                cluster_id="sheet-set-001",
                page_total=1,
                pages=[master_page],
                master_page=master_page,
            ),
        ],
    )


def _build_cover_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        cover_variant="通用",
        engineering_no="1234",
        subitem_no="JG001",
        subitem_name="子项名称",
        discipline="结构",
        doc_status="CFC",
        album_title_cn="测试图册",
        cover_revision="A",
    )
    derived = DerivedFields(
        album_internal_code="1234567-JG001",
        album_code="01",
        cover_internal_code="1234567-JG001-FM",
        cover_external_code="JD1NHT11F01B25C42SD",
        design_phase="施工图设计",
    )
    return DocContext(params=params, derived=derived, frames=[])


def test_catalog_detail_rows_center_title_and_use_uniform_body_height(
    temp_dir: Path,
) -> None:
    gen = CatalogGenerator(pdf_exporter=DummyCatalogPDFExporter())
    ctx = _build_catalog_context()
    output_xlsx = temp_dir / "目录.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_catalog_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active

    assert ws["E9"].alignment.horizontal == "center"
    assert ws["E9"].alignment.vertical == "center"
    assert ws["E9"].alignment.wrapText is True
    assert ws["E10"].alignment.horizontal == "center"
    assert ws["E11"].alignment.horizontal == "center"
    assert ws.row_dimensions[9].height == 36
    assert ws.row_dimensions[10].height == 36
    assert ws.row_dimensions[11].height == 36


def test_catalog_detail_rows_raise_height_for_three_and_four_line_titles(
    temp_dir: Path,
) -> None:
    gen = CatalogGenerator(pdf_exporter=DummyCatalogPDFExporter())
    ctx = _build_catalog_context()
    ctx.sheet_sets[0].master_page.frame_meta.titleblock.title_cn = "第一行\n第二行\n第三行"
    ctx.frames[0].titleblock.title_cn = "甲\n乙\n丙\n丁"
    output_xlsx = temp_dir / "目录.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_catalog_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active

    assert ws.row_dimensions[11].height == pytest.approx(50, abs=0.2)
    assert ws.row_dimensions[12].height == pytest.approx(60, abs=0.2)


def test_generate_catalog_uses_external_code_revision_status_filename(
    temp_dir: Path,
    monkeypatch,
) -> None:
    gen = CatalogGenerator(pdf_exporter=DummyCatalogPDFExporter())
    ctx = _build_catalog_context()
    recorded: dict[str, Path] = {}

    def fake_write_catalog(self, template_path, output_path, bindings, ctx):  # noqa: ANN001
        output_path.write_bytes(b"xlsx")
        recorded["xlsx"] = output_path

    monkeypatch.setattr(CatalogGenerator, "_write_catalog", fake_write_catalog)
    monkeypatch.setattr(CatalogGenerator, "_count_pages", lambda self, path: 2)
    monkeypatch.setattr(
        CatalogGenerator,
        "_backfill_page_count",
        lambda self, output_xlsx, page_count, bindings: None,
    )

    class FakePDFExporter:
        def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
            pdf_path.write_bytes(b"pdf")
            recorded["pdf"] = pdf_path

        def count_pdf_pages(self, pdf_path: Path) -> int:  # noqa: ARG002
            return 2

    gen.pdf_exporter = FakePDFExporter()

    output_xlsx, output_pdf, page_count = gen.generate(ctx, temp_dir)

    expected_stem = "JD1NHT11T01B25C42SDACFC (1234567-JG001-TM)"
    assert output_xlsx == temp_dir / f"{expected_stem}.xlsx"
    assert output_pdf == temp_dir / f"{expected_stem}.pdf"
    assert page_count == 2
    assert recorded["xlsx"] == output_xlsx
    assert recorded["pdf"] == output_pdf


def test_generate_cover_uses_external_code_revision_status_filename(
    temp_dir: Path,
    monkeypatch,
) -> None:
    gen = CoverGenerator()
    ctx = _build_cover_context()
    recorded: dict[str, Path] = {}

    def fake_write_cover(self, template_path, output_path, bindings, data, ctx):  # noqa: ANN001
        output_path.write_bytes(b"docx")
        recorded["docx"] = output_path

    monkeypatch.setattr(CoverGenerator, "_write_cover", fake_write_cover)

    class FakePDFExporter:
        def export_docx_to_pdf(self, docx_path: Path, pdf_path: Path) -> None:
            pdf_path.write_bytes(b"pdf")
            recorded["pdf"] = pdf_path

    gen.pdf_exporter = FakePDFExporter()

    output_docx, output_pdf = gen.generate(ctx, temp_dir)

    expected_stem = "JD1NHT11F01B25C42SDACFC (1234567-JG001-FM)"
    assert output_docx == temp_dir / f"{expected_stem}.docx"
    assert output_pdf == temp_dir / f"{expected_stem}.pdf"
    assert recorded["docx"] == output_docx
    assert recorded["pdf"] == output_pdf
