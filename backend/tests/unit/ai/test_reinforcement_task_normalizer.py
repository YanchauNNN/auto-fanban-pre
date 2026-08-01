from __future__ import annotations

import json
import math
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
import yaml
from openpyxl import Workbook


class FakeClient:
    def __init__(self, content: str, *, error: Exception | None = None) -> None:
        self.content = content
        self.error = error
        self.calls: list[dict[str, Any]] = []
        self.api_key = "unit-test-secret-api-key"

    def complete(
        self,
        messages: list[dict[str, Any]],
        *,
        tools: list[dict[str, Any]] | None = None,
    ) -> SimpleNamespace:
        self.calls.append({"messages": messages, "tools": tools})
        if self.error is not None:
            raise self.error
        return SimpleNamespace(content=self.content, usage={})

    def __repr__(self) -> str:
        return f"FakeClient(api_key={self.api_key!r})"


def _skill_root(tmp_path: Path) -> Path:
    root = tmp_path / "reinforcement-table-normalizer"
    (root / "references").mkdir(parents=True)
    (root / "SKILL.md").write_text(
        "---\n"
        "name: reinforcement-table-normalizer\n"
        "description: Use when a confirmed calculation-book task contains a non-standard reinforcement table.\n"
        "---\n"
        "# 完整 Skill 规则\n"
        "只返回 schema v1 JSON 对象。\n",
        encoding="utf-8",
    )
    (root / "references" / "normalization-rules.md").write_text(
        "# 完整规范规则\n不得返回/计算实际配筋面积。\n",
        encoding="utf-8",
    )
    return root


def _workbook_path(tmp_path: Path, *, extra_cells: int = 0) -> Path:
    path = tmp_path / "非标准配筋表.xlsx"
    workbook = Workbook()
    wall = workbook.active
    wall.title = "墙体配筋"
    wall.append(["墙号", "水平筋", "竖向筋", "拉筋"])
    wall.append(["S7157", "原表水平筋", "原表竖向筋", "原表拉筋"])
    for offset in range(extra_cells):
        wall.cell(row=3, column=offset + 1, value=f"extra-{offset}")

    slab = workbook.create_sheet("楼板配筋")
    slab.append(["标高", "顶层X", "顶层Y", "底层X", "底层Y", "拉筋"])
    slab.append(
        ["11.20m", "顶层水平原文", "顶层竖向原文", "底层水平原文", "底层竖向原文", "纵向拉筋原文"]
    )
    workbook.save(path)
    workbook.close()
    return path


def _wall_row(*, source_x: str = "B2") -> dict[str, Any]:
    return {
        "kind": "wall",
        "status": "normalized",
        "wall_id": "S7157",
        "X": "1D36间距200",
        "Y": "1D32间距200",
        "Z": "1C14间距400*400",
        "source_sheet": "墙体配筋",
        "source_row": 2,
        "source_cells": {
            "wall": "A2",
            "X": source_x,
            "Y": "C2",
            "Z": "D2",
        },
    }


def _slab_row() -> dict[str, Any]:
    return {
        "kind": "slab",
        "status": "normalized",
        "elevation": "11.2",
        "top_x": "1D36间距200",
        "top_y": "1D40间距200",
        "middle_x": None,
        "middle_y": None,
        "bottom_x": "1D30间距200",
        "bottom_y": "1D28间距200",
        "z": "1D16间距200",
        "source_sheet": "楼板配筋",
        "source_row": 2,
        "source_cells": {
            "elevation": "A2",
            "top_x": "B2",
            "top_y": "C2",
            "middle_x": None,
            "middle_y": None,
            "bottom_x": "D2",
            "bottom_y": "E2",
            "z": "F2",
        },
    }


def _payload(*, include_slab: bool = True, source_x: str = "B2") -> dict[str, Any]:
    rows = [_wall_row(source_x=source_x)]
    if include_slab:
        rows.append(_slab_row())
    return {"schema_version": "1", "source_row_count": len(rows), "rows": rows}


def _normalizer(
    tmp_path: Path,
    client: FakeClient,
    **limit_overrides: int,
) -> Any:
    from src.ai.reinforcement_task_normalizer import (
        ReinforcementTaskNormalizer,
        ReinforcementTaskNormalizerLimits,
    )

    defaults = {
        "max_non_empty_cells": 100,
        "max_snapshot_chars": 100_000,
        "max_skill_chars": 10_000,
    }
    defaults.update(limit_overrides)
    return ReinforcementTaskNormalizer(
        client=client,
        skill_root=_skill_root(tmp_path),
        limits=ReinforcementTaskNormalizerLimits(**defaults),
    )


@pytest.mark.parametrize("fenced", [False, True])
def test_normalizes_plain_or_fenced_json_once_with_complete_skill_and_snapshot(
    tmp_path: Path,
    fenced: bool,
) -> None:
    payload = json.dumps(_payload(), ensure_ascii=False)
    content = f"```json\n{payload}\n```" if fenced else payload
    client = FakeClient(content)
    normalizer = _normalizer(tmp_path, client)

    result = normalizer.normalize(
        _workbook_path(tmp_path),
        include_slab=True,
        expected_source_row_count=2,
    )

    assert len(client.calls) == 1
    assert client.calls[0]["tools"] is None
    prompt = "\n".join(
        str(message["content"]) for message in client.calls[0]["messages"]
    )
    assert "skill_id=reinforcement_table_normalizer" in prompt
    assert "完整 Skill 规则" in prompt
    assert "完整规范规则" in prompt
    assert "不得返回/计算实际配筋面积" in prompt
    assert '"sheet":"墙体配筋"' in prompt
    assert '"address":"B2"' in prompt
    assert result.wall_schedule.rows[0].wall_id == "S7157"
    assert result.wall_schedule.rows[0].x.selected.actual_area == pytest.approx(
        math.pi * 18**2 * 5
    )
    assert result.slab_schedule is not None
    assert result.slab_schedule.rows[0].elevation == "11.2"


def test_include_slab_false_requests_and_accepts_wall_only_payload(tmp_path: Path) -> None:
    client = FakeClient(json.dumps(_payload(include_slab=False), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client)

    result = normalizer.normalize(
        _workbook_path(tmp_path),
        include_slab=False,
        expected_source_row_count=1,
    )

    assert result.slab_schedule is None
    prompt = "\n".join(
        str(message["content"]) for message in client.calls[0]["messages"]
    )
    assert "include_slab=false" in prompt
    assert "不得杜撰 slab" in prompt


def test_include_slab_false_rejects_model_invented_slab(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=False)

    assert exc_info.value.code == "model_schema_invalid"


def test_expected_source_row_count_is_enforced_by_deterministic_validator(
    tmp_path: Path,
) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(
            _workbook_path(tmp_path),
            include_slab=True,
            expected_source_row_count=40,
        )

    assert exc_info.value.code == "model_schema_invalid"
    assert "40" not in str(exc_info.value)
    prompt = "\n".join(
        str(message["content"]) for message in client.calls[0]["messages"]
    )
    assert "source_row_count 与 rows 数量都必须等于 40" in prompt


def test_prompt_without_expected_source_count_forbids_inventing_fixed_count(
    tmp_path: Path,
) -> None:
    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client)

    normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    prompt = "\n".join(
        str(message["content"]) for message in client.calls[0]["messages"]
    )
    assert "expected_source_row_count 未提供" in prompt
    assert "不得杜撰固定源行数" in prompt
    assert "必须等于 40" not in prompt


@pytest.mark.parametrize(
    "content, expected_code",
    [
        ("not json", "model_output_invalid"),
        ('说明文字\n{"schema_version":"1"}', "model_output_invalid"),
        ('[{"schema_version":"1"}]', "model_output_invalid"),
        ("```json\n{}\n```\n```json\n{}\n```", "model_output_invalid"),
        ('{"schema_version":"1","source_row_count":1,"rows":[]}', "model_schema_invalid"),
    ],
)
def test_rejects_invalid_model_output_without_echoing_it(
    tmp_path: Path,
    content: str,
    expected_code: str,
) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    secret = "unit-test-secret-api-key"
    client = FakeClient(content + secret if content == "not json" else content)
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert exc_info.value.code == expected_code
    assert secret not in str(exc_info.value)


def test_rejects_invalid_source_evidence_as_model_schema_error(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(source_x="Z99"), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert exc_info.value.code == "model_schema_invalid"
    assert "Z99" not in str(exc_info.value)


@pytest.mark.parametrize("error_type", ["timeout", "gateway"])
def test_preserves_diagnostic_gateway_errors(tmp_path: Path, error_type: str) -> None:
    from src.ai.chat_client import ChatClientTimeout, ChatGatewayError

    error = (
        ChatClientTimeout("model gateway timed out")
        if error_type == "timeout"
        else ChatGatewayError("model gateway request failed")
    )
    client = FakeClient("", error=error)
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(type(error), match="model gateway"):
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert len(client.calls) == 1


def test_missing_skill_fails_before_model_call(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import (
        ReinforcementTaskNormalizationError,
        ReinforcementTaskNormalizer,
        ReinforcementTaskNormalizerLimits,
    )

    root = _skill_root(tmp_path)
    (root / "references" / "normalization-rules.md").unlink()
    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = ReinforcementTaskNormalizer(
        client=client,
        skill_root=root,
        limits=ReinforcementTaskNormalizerLimits(),
    )

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert exc_info.value.code == "skill_missing"
    assert client.calls == []


def test_snapshot_cell_limit_fails_before_model_call(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client, max_non_empty_cells=3)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(
            _workbook_path(tmp_path, extra_cells=4),
            include_slab=True,
        )

    assert exc_info.value.code == "snapshot_too_large"
    assert client.calls == []


def test_snapshot_character_limit_fails_before_model_call(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client, max_snapshot_chars=50)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert exc_info.value.code == "snapshot_too_large"
    assert client.calls == []


def test_skill_size_limit_fails_without_reading_unbounded_content(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient(json.dumps(_payload(), ensure_ascii=False))
    normalizer = _normalizer(tmp_path, client, max_skill_chars=20)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    assert exc_info.value.code == "skill_missing"
    assert client.calls == []


def test_prompt_exception_and_normalizer_repr_do_not_expose_api_key(tmp_path: Path) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizationError

    client = FakeClient("invalid unit-test-secret-api-key")
    normalizer = _normalizer(tmp_path, client)

    with pytest.raises(ReinforcementTaskNormalizationError) as exc_info:
        normalizer.normalize(_workbook_path(tmp_path), include_slab=True)

    prompt = json.dumps(client.calls[0]["messages"], ensure_ascii=False)
    assert client.api_key not in prompt
    assert client.api_key not in str(exc_info.value)
    assert client.api_key not in repr(normalizer)


def test_repository_skill_documents_define_worker_model_schema_and_nonblocking_reviews() -> None:
    repo_root = Path(__file__).resolve().parents[4]
    skill_path = repo_root / "tools" / "ai" / "reinforcement-table-normalizer" / "SKILL.md"
    rules_path = skill_path.parent / "references" / "normalization-rules.md"
    skill_text = skill_path.read_text(encoding="utf-8")
    rules_text = rules_path.read_text(encoding="utf-8")

    _, frontmatter_text, _ = skill_text.split("---", 2)
    frontmatter = yaml.safe_load(frontmatter_text)
    assert set(frontmatter) == {"name", "description"}
    assert frontmatter["description"].startswith("Use when")

    combined = f"{skill_text}\n{rules_text}"
    for required in (
        "预检判定为非标准",
        "用户确认",
        "Worker",
        "标准表不调用",
        "schema_version",
        "source_row_count",
        "wall",
        "slab",
        "normalized",
        "needs_review",
        "source_sheet",
        "source_row",
        "source_cells",
        "blank_fields",
        "reason",
        "5 组",
        "7 组",
        "40",
        "物理来源",
        "duplicate",
        "-1/-2",
        "图片墙号数量不一致",
        "A→C",
        "D/C",
        "括号内值",
        "S7157A",
    ):
        assert required in combined
    assert "actual_area" in combined
    assert "模型禁止返回" in combined
    assert "后端" in combined and "精确公式" in combined
    assert "模型不生成 Excel" in combined
    assert "不得让语言模型自行识别" not in combined
    assert "正式计算书生成必须等待" not in combined
    assert "修正 Excel 后重新预检" not in combined
