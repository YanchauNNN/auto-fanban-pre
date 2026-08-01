from __future__ import annotations

import math
import re
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from src.calculation_book.reinforcement_input import (
    InvalidReinforcementWorkbook,
    NormalizedReinforcementRow,
    ReinforcementRowIssue,
    build_reinforcement_schedule,
    load_reinforcement_schedule,
    load_slab_reinforcement_schedule,
    parse_linear_rebar_cell,
    parse_rebar_cell,
)


def test_normalizes_standard_horizontal_and_tie_notation() -> None:
    horizontal = parse_rebar_cell("1   22@200", direction="X")
    tie = parse_rebar_cell("12@200x400", direction="Z")

    assert horizontal.selected.canonical_specification == "1D22间距200"
    assert horizontal.selected.narrative_specification == "1排22@200"
    assert horizontal.selected.actual_area == pytest.approx(
        math.pi * 11**2 * 5
    )
    assert tie.selected.canonical_specification == "1C12间距200*400"
    assert tie.selected.narrative_specification == "1排12@200x400"
    assert tie.selected.actual_area == pytest.approx(
        math.pi * 6**2 * 5 * 2.5
    )


def test_normalizes_a_marker_to_c_and_preserves_two_tie_layers() -> None:
    single = parse_rebar_cell("1A8间距400*400#", direction="Z")
    double = parse_rebar_cell("2 12@200x400", direction="Z")

    assert single.selected.canonical_specification == "1C8间距400*400"
    assert "#" not in single.selected.canonical_specification
    assert double.selected.canonical_specification == "2C12间距200*400"
    assert double.selected.actual_area == pytest.approx(
        2 * math.pi * 6**2 * 5 * 2.5
    )


def test_parenthetical_configuration_is_selected_as_actual_reinforcement() -> None:
    parsed = parse_rebar_cell(
        "1D28间距200(N5057-N5059:1D40间距200#)",
        direction="Y",
    )

    assert parsed.selected.canonical_specification == "1D40间距200"
    assert parsed.selected.actual_area == pytest.approx(
        math.pi * 20**2 * 5
    )
    assert parsed.original_text.endswith("#)")


def test_largest_parenthetical_configuration_is_selected() -> None:
    parsed = parse_rebar_cell(
        "1D28间距200（局部:1D32间距200；洞口:1D40间距200）",
        direction="Y",
    )

    assert parsed.selected.canonical_specification == "1D40间距200"
    assert len(parsed.candidates) == 3


def test_rejects_tie_without_two_direction_spacings() -> None:
    with pytest.raises(InvalidReinforcementWorkbook, match="两个方向"):
        parse_rebar_cell("1C14间距400", direction="Z")


def test_linear_rebar_uses_exact_per_meter_area_and_rejects_grid_spacing() -> None:
    parsed = parse_linear_rebar_cell("1D16@200")

    assert parsed.selected.canonical_specification == "1D16间距200"
    assert parsed.selected.actual_area == pytest.approx(math.pi * 8**2 * 5)

    with pytest.raises(InvalidReinforcementWorkbook, match="一个方向"):
        parse_linear_rebar_cell("1C16@200x400")


def _add_slab_sheet(
    workbook: Workbook,
    *rows: list[object],
) -> None:
    sheet = workbook.create_sheet("楼板配筋")
    sheet.append(
        [
            "标高",
            "顶层水平",
            "顶层竖向",
            "中层水平",
            "中层竖向",
            "底层水平",
            "底层竖向",
            "纵向拉筋",
        ]
    )
    for row in rows:
        sheet.append(row)


def test_loads_slab_reinforcement_sheet_with_optional_middle(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    _add_slab_sheet(
        workbook,
        [
            "11.20m",
            "1D36@200",
            "1D40@200",
            None,
            None,
            "1D36@200",
            "1D40@200",
            "1D16@200",
        ],
    )
    path = tmp_path / "slab-schedule.xlsx"
    workbook.save(path)

    schedule = load_slab_reinforcement_schedule(path, required=True)

    assert schedule is not None
    assert len(schedule.rows) == 1
    row = schedule.rows[0]
    assert row.elevation == "11.2"
    assert row.middle_x is None
    assert row.middle_y is None
    assert row.top_x.selected.canonical_specification == "1D36间距200"
    assert row.z.selected.actual_area == pytest.approx(math.pi * 8**2 * 5)
    assert row.source_cells["top_x"] == "B2"
    assert row.source_cells["bottom_y"] == "G2"
    assert row.source_cells["z"] == "H2"


def test_missing_slab_sheet_is_optional_until_slab_stress_is_enabled(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    path = tmp_path / "wall-only.xlsx"
    workbook.save(path)

    assert load_slab_reinforcement_schedule(path, required=False) is None
    with pytest.raises(InvalidReinforcementWorkbook, match="楼板配筋"):
        load_slab_reinforcement_schedule(path, required=True)


def test_rejects_duplicate_normalized_slab_elevation(tmp_path: Path) -> None:
    workbook = Workbook()
    values = [
        "1D36@200",
        "1D40@200",
        None,
        None,
        "1D36@200",
        "1D40@200",
        "1D16@200",
    ]
    _add_slab_sheet(
        workbook,
        ["11.2", *values],
        ["11.20m", *values],
    )
    path = tmp_path / "duplicate-slab.xlsx"
    workbook.save(path)

    with pytest.raises(InvalidReinforcementWorkbook, match="重复标高.*11.2"):
        load_slab_reinforcement_schedule(path, required=True)


def test_loads_standard_four_column_workbook_and_flags_duplicate_wall(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    sheet.append(["S7157 墙", "1 36@200", "1 36@200", "12@200x400"])
    sheet.append(["S7157 墙", "1 32@200", "1 32@200", "12@200x400"])
    sheet.append(["S7157a 墙", "1 28@200", "1 28@200", "1A8间距400*400"])
    path = tmp_path / "schedule.xlsx"
    workbook.save(path)

    schedule = load_reinforcement_schedule(path)

    assert [row.wall_id for row in schedule.rows] == ["S7157", "S7157", "S7157A"]
    assert schedule.duplicate_wall_ids == ("S7157",)
    assert schedule.requires_manual_confirmation
    assert schedule.source_row_count == 3
    assert schedule.normalized_row_count == 3
    assert schedule.issue_row_count == 0
    assert schedule.normalization_triggered is True
    assert schedule.rows[2].z.selected.canonical_specification == "1C8间距400*400"


def test_preserves_every_nonempty_source_row_in_normalization_audit(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["构件编号及位置", "水平筋", "竖向筋", "拉筋"])
    sheet.append(["S7157", "1D36间距200", "1D36间距200", "1C14间距400*400"])
    sheet.append(["待确认墙", "1D32间距200", "1D32间距200", "1C14间距400*400"])
    sheet.append(["S7158", "无法识别", "1D28间距200", "1C14间距400*400"])
    path = tmp_path / "issues.xlsx"
    workbook.save(path)

    schedule = load_reinforcement_schedule(path)

    assert schedule.source_row_count == 3
    assert schedule.normalized_row_count == 1
    assert schedule.issue_row_count == 2
    assert schedule.source_row_count == (
        schedule.normalized_row_count + schedule.issue_row_count
    )
    assert [issue.source_row for issue in schedule.issues] == [3, 4]
    assert schedule.issues[0].wall_id is None
    assert schedule.issues[0].original_wall_text == "待确认墙"
    assert schedule.issues[0].original_values == {
        "wall": "待确认墙",
        "X": "1D32间距200",
        "Y": "1D32间距200",
        "Z": "1C14间距400*400",
    }
    assert schedule.issues[1].wall_id == "S7158"
    assert schedule.issues[1].original_values["X"] == "无法识别"
    assert schedule.requires_manual_confirmation is True


def test_canonical_standard_template_does_not_trigger_normalization(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    sheet.append(["S7157", "1D36间距200", "1D36间距200", "1C14间距400*400"])
    path = tmp_path / "canonical.xlsx"
    workbook.save(path)

    schedule = load_reinforcement_schedule(path)

    assert schedule.source_row_count == 1
    assert schedule.normalization_triggered is False


def test_build_reinforcement_schedule_preserves_issues_and_duplicate_ids() -> None:
    parsed_x = parse_rebar_cell("1D36间距200", direction="X")
    parsed_y = parse_rebar_cell("1D32间距200", direction="Y")
    parsed_z = parse_rebar_cell("1C14间距400*400", direction="Z")
    rows = tuple(
        NormalizedReinforcementRow(
            wall_id="S7157",
            x=parsed_x,
            y=parsed_y,
            z=parsed_z,
            source_sheet="墙体配筋",
            source_row=source_row,
            source_cells={"wall": f"A{source_row}", "X": f"B{source_row}", "Y": f"C{source_row}", "Z": f"D{source_row}"},
        )
        for source_row in (2, 3)
    )
    issue = ReinforcementRowIssue(
        source_sheet="墙体配筋",
        source_row=4,
        source_cells={"wall": "A4", "X": "B4", "Y": "C4", "Z": "D4"},
        original_values={"wall": "待确认墙", "X": "?", "Y": "", "Z": ""},
        original_wall_text="待确认墙",
        wall_id=None,
        error="墙号不明确",
    )

    schedule = build_reinforcement_schedule(
        rows=rows,
        issues=(issue,),
        source_row_count=3,
        normalization_triggered=True,
    )

    assert schedule.duplicate_wall_ids == ("S7157",)
    assert schedule.issues == (issue,)
    assert schedule.source_row_count == 3
    assert schedule.normalization_triggered is True


def test_loads_workbook_without_optional_worksheet_dimensions(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    wall_sheet = workbook.active
    wall_sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    wall_sheet.append(["S7157 墙", "1 36@200", "1 36@200", "12@200x400"])
    _add_slab_sheet(
        workbook,
        [
            "11.20m",
            "1D36@200",
            "1D40@200",
            None,
            None,
            "1D36@200",
            "1D40@200",
            "1D16@200",
        ],
    )
    source_path = tmp_path / "with-dimensions.xlsx"
    dimensionless_path = tmp_path / "without-dimensions.xlsx"
    workbook.save(source_path)

    with ZipFile(source_path) as source, ZipFile(
        dimensionless_path,
        "w",
        ZIP_DEFLATED,
    ) as target:
        for entry in source.infolist():
            content = source.read(entry.filename)
            if entry.filename.startswith("xl/worksheets/"):
                content = re.sub(br"<dimension\b[^>]*/>", b"", content)
            target.writestr(entry, content)

    wall_schedule = load_reinforcement_schedule(dimensionless_path)
    slab_schedule = load_slab_reinforcement_schedule(
        dimensionless_path,
        required=True,
    )

    assert [row.wall_id for row in wall_schedule.rows] == ["S7157"]
    assert slab_schedule is not None
    assert [row.elevation for row in slab_schedule.rows] == ["11.2"]


def test_selects_canonical_wall_number_column_in_wide_workbook(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2016墙体配筋"
    sheet.append(
        [
            "层位",
            "墙号",
            "墙号",
            "墙厚",
            "水平筋",
            "竖向筋",
            "拉筋",
        ]
    )
    sheet.append(
        [
            "N5",
            "012墙",
            "N5012",
            600,
            "1D32间距200",
            "1D28间距200",
            "1C14间距400*400",
        ]
    )
    path = tmp_path / "wide.xlsx"
    workbook.save(path)

    schedule = load_reinforcement_schedule(path)

    assert len(schedule.rows) == 1
    assert schedule.rows[0].wall_id == "N5012"
    assert schedule.rows[0].z.selected.actual_area == pytest.approx(
        math.pi * 7**2 * 2.5 * 2.5
    )
