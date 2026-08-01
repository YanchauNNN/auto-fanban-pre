from __future__ import annotations

import json
from dataclasses import FrozenInstanceError
from importlib import import_module
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet


def _subject():
    return import_module("src.calculation_book.reinforcement_workbook")


def _add_standard_wall_sheet(workbook: Workbook, *, title: str = "Sheet1"):
    sheet = workbook.create_sheet(title)
    sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    sheet.append(["S7159 墙", "1 28@200", "1D28间距200", "12@200x400"])
    sheet.append(["S7160", "D25@200", "1D25间距200", "1C10间距200*400"])
    return sheet


def _add_standard_slab_sheet(workbook: Workbook, *, with_data: bool = True):
    sheet = workbook.create_sheet("楼板配筋")
    sheet.append(
        [
            "标高",
            "顶层水平",
            "顶层竖向",
            "中层水平",
            "中层竖向",
            "底层水平",
            "底层竖向",
            "纵向拉筋",
        ]
    )
    if with_data:
        sheet.append(
            [
                11.45,
                "D20@200",
                "D20@200",
                None,
                None,
                "D18@200",
                "D18@200",
                "D12@200",
            ]
        )
    return sheet


def _rewrite_first_worksheet_xml(
    source: Path,
    target: Path,
    transform,
) -> None:
    _rewrite_xlsx_entry(
        source,
        target,
        "xl/worksheets/sheet1.xml",
        transform,
    )


def _rewrite_xlsx_entry(
    source: Path,
    target: Path,
    entry_name: str,
    transform,
) -> None:
    with ZipFile(source) as input_archive, ZipFile(
        target,
        "w",
        ZIP_DEFLATED,
    ) as output_archive:
        for info in input_archive.infolist():
            payload = input_archive.read(info.filename)
            if info.filename == entry_name:
                payload = transform(payload)
            output_archive.writestr(info, payload)


def _add_xlsx_entry(
    source: Path,
    target: Path,
    entry_name: str,
    payload: bytes,
) -> None:
    with ZipFile(source) as input_archive, ZipFile(
        target,
        "w",
        ZIP_DEFLATED,
    ) as output_archive:
        for info in input_archive.infolist():
            output_archive.writestr(info, input_archive.read(info.filename))
        output_archive.writestr(entry_name, payload)


def test_standard_template_format_does_not_require_ai(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    margin_sheet = workbook.active
    margin_sheet.title = "配筋裕度"
    margin_sheet.append(["构件", "方向", "裕度"])
    margin_sheet.append(["S7159", "X", 0.15])
    _add_standard_wall_sheet(workbook)
    _add_standard_slab_sheet(workbook)
    path = tmp_path / "计算书模板文件.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=True)

    assert inspection == subject.WorkbookFormatInspection(
        requires_ai_normalization=False,
        reasons=(),
        wall_sheet="Sheet1",
        slab_sheet="楼板配筋",
    )
    with pytest.raises(FrozenInstanceError):
        inspection.wall_sheet = "changed"


def test_format_inspection_does_not_invoke_area_calculation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    from src.calculation_book import reinforcement_input

    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    _add_standard_slab_sheet(workbook)
    path = tmp_path / "syntax-only.xlsx"
    workbook.save(path)

    def fail_if_called(**_kwargs):
        raise AssertionError("format inspection must not calculate reinforcement area")

    monkeypatch.setattr(reinforcement_input, "_configuration", fail_if_called)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=True)

    assert inspection.requires_ai_normalization is False


def test_format_inspection_rejects_compressed_shared_strings_before_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    source = tmp_path / "safe.xlsx"
    workbook.save(source)
    path = tmp_path / "shared-strings-bomb.xlsx"
    _add_xlsx_entry(
        source,
        path,
        "xl/sharedStrings.xml",
        b"<sst><si><t>" + (b"A" * (8 * 1024 * 1024)) + b"</t></si></sst>",
    )

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe XLSX must be rejected before openpyxl load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="XLSX internal resource limit"):
        subject.inspect_reinforcement_workbook(path, include_slab=False)


def test_format_inspection_rejects_compressed_worksheet_before_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    source = tmp_path / "safe.xlsx"
    workbook.save(source)
    path = tmp_path / "worksheet-bomb.xlsx"

    def inflate_worksheet(payload: bytes) -> bytes:
        inflated_comment = b"<!--" + (b"A" * (8 * 1024 * 1024)) + b"-->"
        return payload.replace(
            b"</worksheet>",
            inflated_comment + b"</worksheet>",
        )

    _rewrite_first_worksheet_xml(source, path, inflate_worksheet)

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe XLSX must be rejected before openpyxl load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="XLSX internal resource limit"):
        subject.inspect_reinforcement_workbook(path, include_slab=False)


@pytest.mark.parametrize(
    ("address", "value"),
    [
        ("B2", "0D28@200"),
        ("B2", "1D0@200"),
        ("B2", "1D28@0"),
        ("B2", "1D28@-200"),
        ("B2", "1D28@200*400"),
        ("D2", "1C12@200"),
        ("D2", "1C12@200*0"),
        ("D2", "1A8间距400*400"),
    ],
)
def test_nonpositive_or_directionally_invalid_specs_require_ai(
    tmp_path: Path,
    address: str,
    value: str,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = _add_standard_wall_sheet(workbook)
    sheet[address] = value
    path = tmp_path / "invalid-standard-cell.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=False)

    assert inspection.requires_ai_normalization is True
    assert inspection.reasons[0].code == "wall_value_nonstandard"


def test_first_nonstandard_wall_sheet_is_not_hidden_by_later_standard_sheet(
    tmp_path: Path,
) -> None:
    subject = _subject()
    workbook = Workbook()
    first = workbook.active
    first.title = "前置非标墙表"
    first.append(["墙号", "水平筋(X)", "竖向筋(Y)", "拉筋(Z)"])
    first.append(["S7159", "D28@200", "D28@200", "C12@200x400"])
    _add_standard_wall_sheet(workbook)
    path = tmp_path / "first-wall-wins.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=False)

    assert inspection.requires_ai_normalization is True
    assert inspection.wall_sheet == "前置非标墙表"
    assert "wall_layout_nonstandard" in {
        reason.code for reason in inspection.reasons
    }


def test_multiple_wall_candidates_are_ambiguous(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook, title="墙表一")
    _add_standard_wall_sheet(workbook, title="墙表二")
    path = tmp_path / "multiple-wall-sheets.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=False)

    assert inspection.requires_ai_normalization is True
    assert inspection.wall_sheet == "墙表一"
    assert "wall_sheet_ambiguous" in {reason.code for reason in inspection.reasons}


def test_empty_standard_slab_sheet_requires_ai(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    _add_standard_slab_sheet(workbook, with_data=False)
    path = tmp_path / "empty-standard-slab.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=True)

    assert inspection.requires_ai_normalization is True
    assert inspection.slab_sheet == "楼板配筋"
    assert "slab_data_missing" in {reason.code for reason in inspection.reasons}


def test_empty_standard_and_custom_slab_sheets_are_ambiguous(
    tmp_path: Path,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    _add_standard_slab_sheet(workbook, with_data=False)
    custom = workbook.create_sheet("楼板自定义表")
    custom.append(["楼层", "上部X", "上部Y", "下部X", "下部Y", "拉筋"])
    custom.append(
        [11.45, "D20@200", "D20@200", "D18@200", "D18@200", "C12@200x400"]
    )
    path = tmp_path / "ambiguous-slab.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=True)

    assert inspection.requires_ai_normalization is True
    assert inspection.slab_sheet == "楼板配筋"
    assert "slab_sheet_ambiguous" in {reason.code for reason in inspection.reasons}


def test_nonstandard_wall_and_selected_slab_require_ai(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    wall_sheet = workbook.active
    wall_sheet.title = "墙体自定义表"
    wall_sheet.append(["墙号", "水平筋(X)", "竖向筋(Y)", "拉筋(Z)"])
    wall_sheet.append(
        ["S7159", "双层D28@200", "D28@200", "C12@200x400"]
    )
    slab_sheet = workbook.create_sheet("楼板自定义表")
    slab_sheet.append(["楼层", "上部X", "上部Y", "下部X", "下部Y", "拉筋"])
    slab_sheet.append(
        [11.45, "D20@200", "D20@200", "D18@200", "D18@200", "C12@200x400"]
    )
    path = tmp_path / "nonstandard.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=True)

    assert inspection.requires_ai_normalization is True
    assert inspection.wall_sheet == "墙体自定义表"
    assert inspection.slab_sheet == "楼板自定义表"
    assert {reason.scope for reason in inspection.reasons} == {"wall", "slab"}
    assert all(reason.sheet is not None for reason in inspection.reasons)


def test_unselected_nonstandard_slab_does_not_require_ai(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_standard_wall_sheet(workbook)
    slab_sheet = workbook.create_sheet("楼板自定义表")
    slab_sheet.append(["楼层", "上部X", "上部Y", "下部X", "下部Y", "拉筋"])
    slab_sheet.append(
        [11.45, "双层D20@200", "D20@200", "D18@200", "D18@200", "C12@200x400"]
    )
    path = tmp_path / "wall-only-selection.xlsx"
    workbook.save(path)

    inspection = subject.inspect_reinforcement_workbook(path, include_slab=False)

    assert inspection.requires_ai_normalization is False
    assert inspection.reasons == ()
    assert inspection.wall_sheet == "Sheet1"
    assert inspection.slab_sheet is None


def test_snapshot_preserves_nonempty_cells_formulas_and_merged_membership(
    tmp_path: Path,
) -> None:
    subject = _subject()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "证据"
    sheet["A1"] = "普通值"
    sheet["B2"] = "=SUM(1,2)"
    sheet.merge_cells("C3:D3")
    sheet["C3"] = "合并值"
    path = tmp_path / "snapshot.xlsx"
    workbook.save(path)

    snapshot = subject.build_workbook_snapshot(path, max_non_empty_cells=3)

    assert snapshot["workbook"] == "snapshot.xlsx"
    assert snapshot["non_empty_cell_count"] == 3
    assert snapshot["sheets"] == [
        {"name": "证据", "merged_ranges": ["C3:D3"]}
    ]
    assert snapshot["cells"] == [
        {
            "sheet": "证据",
            "row": 1,
            "column": 1,
            "address": "A1",
            "value": "普通值",
            "formula": None,
            "merged_range": None,
        },
        {
            "sheet": "证据",
            "row": 2,
            "column": 2,
            "address": "B2",
            "value": "=SUM(1,2)",
            "formula": "=SUM(1,2)",
            "merged_range": None,
        },
        {
            "sheet": "证据",
            "row": 3,
            "column": 3,
            "address": "C3",
            "value": "合并值",
            "formula": None,
            "merged_range": "C3:D3",
        },
    ]
    json.dumps(snapshot, ensure_ascii=False)


def test_snapshot_limit_allows_exact_boundary_and_rejects_overflow(
    tmp_path: Path,
) -> None:
    subject = _subject()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["A", "B", "C"])
    path = tmp_path / "limit.xlsx"
    workbook.save(path)

    assert subject.build_workbook_snapshot(
        path,
        max_non_empty_cells=3,
    )["non_empty_cell_count"] == 3
    with pytest.raises(ValueError, match="max_non_empty_cells=2"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=2)
    with pytest.raises(ValueError, match="max_non_empty_cells must be greater than 0"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=0)


def test_snapshot_does_not_enumerate_sparse_worksheet_rectangle(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "first"
    sheet["XFD1048576"] = "last"
    path = tmp_path / "sparse.xlsx"
    workbook.save(path)

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("snapshot must not call Worksheet.iter_rows")

    monkeypatch.setattr(Worksheet, "iter_rows", fail_if_called)

    snapshot = subject.build_workbook_snapshot(path, max_non_empty_cells=2)

    assert [cell["address"] for cell in snapshot["cells"]] == [
        "A1",
        "XFD1048576",
    ]


def test_snapshot_rejects_huge_merge_before_loading_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "value"
    workbook.save(source)
    path = tmp_path / "huge-merge.xlsx"

    def add_huge_merge(payload: bytes) -> bytes:
        return payload.replace(
            b"</worksheet>",
            b'<mergeCells count="1"><mergeCell ref="A1:XFD1048576"/>'
            b"</mergeCells></worksheet>",
        )

    _rewrite_first_worksheet_xml(source, path, add_huge_merge)

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe workbook must be rejected before openpyxl load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="merged range .* exceeds snapshot limit"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=10)


def test_snapshot_rejects_highly_compressed_xlsx_resource_before_loading(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "value"
    workbook.save(source)
    path = tmp_path / "inflated-styles.xlsx"

    def inflate_styles(payload: bytes) -> bytes:
        inflated_comment = b"<!--" + (b"A" * (8 * 1024 * 1024)) + b"-->"
        return payload.replace(
            b"</styleSheet>",
            inflated_comment + b"</styleSheet>",
        )

    _rewrite_xlsx_entry(source, path, "xl/styles.xml", inflate_styles)
    with ZipFile(path) as archive:
        styles_info = archive.getinfo("xl/styles.xml")
        assert styles_info.file_size > 8 * 1024 * 1024
        assert styles_info.compress_size < 64 * 1024

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe XLSX must be rejected before openpyxl load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="XLSX internal resource limit"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=1)


def test_snapshot_allows_normal_xlsx_resources_with_large_budget(
    tmp_path: Path,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.active["A1"] = "value"
    path = tmp_path / "normal.xlsx"
    workbook.save(path)

    snapshot = subject.build_workbook_snapshot(
        path,
        max_non_empty_cells=10_000,
    )

    assert snapshot["non_empty_cell_count"] == 1


def test_snapshot_rejects_excessive_cell_records_before_loading_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "value"
    workbook.save(source)
    path = tmp_path / "excessive-cell-records.xlsx"

    def add_cell_records(payload: bytes) -> bytes:
        records = b"".join(
            f'<c r="A{index}" s="1"></c>'.encode()
            for index in range(2, 2050)
        )
        return payload.replace(b"</sheetData>", records + b"</sheetData>")

    _rewrite_first_worksheet_xml(source, path, add_cell_records)

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("unsafe workbook must be rejected before openpyxl load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="cell records exceed snapshot limit"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=1)


def test_snapshot_rejects_nonempty_overflow_before_loading_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    workbook = Workbook()
    workbook.active.append(["A", "B", "C"])
    path = tmp_path / "nonempty-overflow.xlsx"
    workbook.save(path)

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("non-empty overflow must be rejected before load")

    monkeypatch.setattr(subject, "load_workbook", fail_if_called)

    with pytest.raises(ValueError, match="max_non_empty_cells=2"):
        subject.build_workbook_snapshot(path, max_non_empty_cells=2)


def test_snapshot_merge_lookup_is_not_cells_times_ranges(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    subject = _subject()
    workbook = Workbook()
    sheet = workbook.active
    for row in range(1, 41):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row=row, column=1, value=f"merge-{row}")
    path = tmp_path / "many-merges.xlsx"
    workbook.save(path)

    contains_calls = 0
    original_contains = CellRange.__contains__

    def count_contains(self, coord):
        nonlocal contains_calls
        contains_calls += 1
        return original_contains(self, coord)

    monkeypatch.setattr(CellRange, "__contains__", count_contains)

    snapshot = subject.build_workbook_snapshot(path, max_non_empty_cells=40)

    assert snapshot["non_empty_cell_count"] == 40
    assert contains_calls <= 40


def test_snapshot_serializes_advanced_formula_objects_stably(tmp_path: Path) -> None:
    subject = _subject()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = ArrayFormula(ref="A1:A2", text="=ROW(A1:A2)")
    sheet["B1"] = DataTableFormula(
        ref="B1:C2",
        dt2D=True,
        r1="A1",
        r2="A2",
    )
    path = tmp_path / "advanced-formulas.xlsx"
    workbook.save(path)

    first = subject.build_workbook_snapshot(path, max_non_empty_cells=2)
    second = subject.build_workbook_snapshot(path, max_non_empty_cells=2)

    assert first == second
    assert first["cells"][0]["formula"] == {
        "type": "array",
        "ref": "A1:A2",
        "text": "=ROW(A1:A2)",
    }
    assert first["cells"][1]["formula"] == {
        "type": "data_table",
        "ref": "B1:C2",
        "ca": False,
        "dt2D": True,
        "dtr": False,
        "r1": "A1",
        "r2": "A2",
        "del1": False,
        "del2": False,
    }
    assert first["cells"][0]["value"] == first["cells"][0]["formula"]
    assert first["cells"][1]["value"] == first["cells"][1]["formula"]
    json.dumps(first, ensure_ascii=False, allow_nan=False)


def test_snapshot_rejects_unknown_value_types() -> None:
    subject = _subject()

    with pytest.raises(TypeError, match="unsupported workbook cell value type"):
        subject._json_value(object())
