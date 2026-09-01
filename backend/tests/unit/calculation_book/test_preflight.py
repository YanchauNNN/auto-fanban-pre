from __future__ import annotations

import zipfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from PIL import Image

from src.calculation_book.archive import InvalidCalculationArchive
from src.calculation_book.models import ReinforcementSource
from src.calculation_book.ocr import StressLegendReading
from src.calculation_book.preflight import run_calculation_book_preflight


def test_preflight_forwards_private_archive_extractor(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    import src.calculation_book.preflight as preflight_module

    extractor = SimpleNamespace(
        executable=tmp_path / "7z.exe",
        list_timeout_seconds=120,
        extract_timeout_seconds=300,
        max_list_output_bytes=1024,
    )
    observed: dict[str, object] = {}

    def fake_extract(*_args, **kwargs):
        observed.update(kwargs)
        raise InvalidCalculationArchive("stop after forwarding")

    monkeypatch.setattr(
        preflight_module,
        "validate_and_extract_archive",
        fake_extract,
    )

    with pytest.raises(InvalidCalculationArchive, match="stop after forwarding"):
        run_calculation_book_preflight(
            archive_path=tmp_path / "input.rar",
            extraction_root=tmp_path / "extracted",
            archive_extractor=extractor,
        )

    assert observed["archive_extractor"] is extractor


def _build_duplicate_archive(
    tmp_path: Path,
    *,
    standard_layout: bool = True,
    include_workbook: bool = True,
) -> Path:
    source = tmp_path / "source"
    source.mkdir()
    for group, _smx in ((1, 3000), (2, 5000)):
        for direction in ("X", "Y", "Z"):
            Image.new("RGB", (1200, 800), "white").save(
                source / f"S7157-{group}-{direction}.png"
            )
    (source / "01").mkdir()
    (source / "02").mkdir()
    Image.new("RGB", (1200, 800), "white").save(source / "01" / "layout.png")
    Image.new("RGB", (1200, 800), "white").save(source / "02" / "model.png")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "构件编号\n及位置" if standard_layout else "墙号",
            "单侧水平钢筋\n(对称配筋)" if standard_layout else "水平筋(X)",
            "单侧竖向钢筋\n(对称配筋)" if standard_layout else "竖向筋(Y)",
            "拉筋" if standard_layout else "拉筋(Z)",
        ]
    )
    sheet.append(["S7157墙", "1D36间距200", "1D36间距200", "1C14间距400*400"])
    sheet.append(["S7157墙", "1D32间距200", "1D32间距200", "1C14间距400*400"])
    if include_workbook:
        workbook.save(source / "计算书模板文件.xlsx")

    archive_path = tmp_path / "input.zip"
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source).as_posix())
    return archive_path


def _build_slab_archive(
    tmp_path: Path,
    *,
    include_middle: bool = False,
    include_slab_sheet: bool = True,
) -> Path:
    source = tmp_path / "source"
    source.mkdir()
    for direction in ("X", "Y", "Z"):
        Image.new("RGB", (1200, 800), "white").save(
            source / f"S7159-{direction}.png"
        )
    slab_names = [
        "11.45-TOP-X",
        "11.45-BOTTOM-X",
        "11.45-TOP-Y",
        "11.45-BOTTOM-Y",
        "11.45-Z",
    ]
    if include_middle:
        slab_names.extend(("11.45-MIDDLE-X", "11.45-MIDDLE-Y"))
    for stem in slab_names:
        Image.new("RGB", (1200, 800), "white").save(source / f"{stem}.png")

    (source / "01").mkdir()
    (source / "02").mkdir()
    Image.new("RGB", (1200, 800), "white").save(source / "01" / "layout.png")
    Image.new("RGB", (1200, 800), "white").save(source / "02" / "model.png")

    workbook = Workbook()
    wall_sheet = workbook.active
    wall_sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    wall_sheet.append(
        ["S7159墙", "1D28间距200", "1D28间距200", "1C12间距200*400"]
    )
    if include_slab_sheet:
        slab_sheet = workbook.create_sheet("楼板配筋")
        slab_sheet.append(
            [
                "标高",
                "顶层水平",
                "顶层竖向",
                "中层水平",
                "中层竖向",
                "底层水平",
                "底层竖向",
                "纵向拉筋",
            ]
        )
        slab_sheet.append(
            [
                11.45,
                "1D36@200",
                "1D40@200",
                "1D32@200" if include_middle else None,
                "1D34@200" if include_middle else None,
                "1D30@200",
                "1D28@200",
                "1D16@200",
            ]
        )
    workbook.save(source / "计算书模板文件.xlsx")

    archive_path = tmp_path / "input.zip"
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source).as_posix())
    return archive_path


def _build_mismatch_archive(tmp_path: Path) -> Path:
    source = tmp_path / "source"
    source.mkdir()
    for wall_id in ("S7159", "NDTJ1"):
        for direction in ("X", "Y", "Z"):
            Image.new("RGB", (1200, 800), "white").save(
                source / f"{wall_id}-{direction}.png"
            )
    (source / "01").mkdir()
    (source / "02").mkdir()
    Image.new("RGB", (1200, 800), "white").save(source / "01" / "layout.png")
    Image.new("RGB", (1200, 800), "white").save(source / "02" / "model.png")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "构件编号\n及位置",
            "单侧水平钢筋\n(对称配筋)",
            "单侧竖向钢筋\n(对称配筋)",
            "拉筋",
        ]
    )
    for wall_id in ("S7159", "S7160"):
        sheet.append(
            [wall_id, "1D28间距200", "1D28间距200", "1C12间距200*400"]
        )
    workbook.save(source / "计算书模板文件.xlsx")

    archive_path = tmp_path / "mismatch.zip"
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source).as_posix())
    return archive_path


def _build_ai_suggested_archive(
    tmp_path: Path,
    *,
    include_slab_figures: bool = True,
) -> Path:
    source = tmp_path / "source"
    source.mkdir()
    for wall_id in ("S7157", "S7158"):
        for direction in ("X", "Y", "Z"):
            Image.new("RGB", (1200, 800), "white").save(
                source / f"{wall_id}-{direction}.png"
            )
    if include_slab_figures:
        for stem in (
            "11.45-TOP-X",
            "11.45-TOP-Y",
            "11.45-BOTTOM-X",
            "11.45-BOTTOM-Y",
            "11.45-Z",
        ):
            Image.new("RGB", (1200, 800), "white").save(
                source / f"{stem}.png"
            )
    Image.new("RGB", (1200, 800), "white").save(source / "说明图.png")
    (source / "01").mkdir()
    (source / "02").mkdir()
    Image.new("RGB", (1200, 800), "white").save(source / "01" / "layout.png")
    Image.new("RGB", (1200, 800), "white").save(source / "02" / "model.png")

    archive_path = tmp_path / "ai-suggested.zip"
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in source.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(source).as_posix())
    return archive_path


def test_preflight_returns_blank_duplicate_and_split_groups_without_confirmation(
    tmp_path: Path,
) -> None:
    archive_path = _build_duplicate_archive(tmp_path)

    def recognize(path: Path, _direction: str) -> StressLegendReading:
        smx = 3000.0 if "-1-" in path.name else 5000.0
        return StressLegendReading(
            smn=0,
            smx=smx,
            legend_values=tuple(smx * index / 9 for index in range(10)),
        )

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        ocr_recognizer=recognize,
    )

    assert result["figure_count"] == 6
    assert result["wall_count"] == 2
    assert result["requires_manual_confirmation"] is False
    assert result["confirmations"] == []
    assert all(item["suggested_source_row"] is None for item in result["walls"])
    assert all(
        direction["canonical_specification"] == ""
        and direction["actual_area"] == ""
        for wall in result["walls"]
        for direction in wall["directions"].values()
    )
    assert {warning["code"] for warning in result["warnings"]} == {
        "duplicate_reinforcement_rows",
        "split_image_group",
    }
    assert result["walls"][0]["directions"]["X"]["legend_values"][0] == 0
    assert result["walls"][0]["directions"]["X"]["source_cell"] == ""
    assert result["normalization_triggered"] is True
    assert (
        result["normalization_skill_id"]
        == "reinforcement_table_normalizer"
    )
    assert result["requires_ai_normalization"] is False
    assert result["requires_ai_recommendation"] is False
    assert result["reinforcement_source"] == "provided"
    assert result["ai_confirmation_message"] is None


def test_nonstandard_preflight_returns_ai_confirmation_without_ocr(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from src.ai.reinforcement_task_normalizer import ReinforcementTaskNormalizer

    archive_path = _build_duplicate_archive(
        tmp_path,
        standard_layout=False,
    )
    ocr_calls: list[Path] = []

    def fail_if_normalizer_runs(*_args, **_kwargs):
        raise AssertionError("HTTP preflight must not invoke the AI normalizer")

    monkeypatch.setattr(
        ReinforcementTaskNormalizer,
        "normalize",
        fail_if_normalizer_runs,
    )

    def fail_if_ocr_runs(path: Path, _direction: str) -> StressLegendReading:
        ocr_calls.append(path)
        raise AssertionError("nonstandard preflight must not run OCR")

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        ocr_recognizer=fail_if_ocr_runs,
    )

    assert ocr_calls == []
    assert result["requires_ai_normalization"] is True
    assert result["requires_ai_recommendation"] is False
    assert result["reinforcement_source"] == "provided"
    assert result["ai_reinforcement_expected_source_row_count"] == 2
    assert result["ai_confirmation_message"] == (
        "您上传的墙体配筋表非标准格式，程序将启动人工智能。"
    )
    assert result["format_inspection"] == {
        "wall_sheet": "Sheet",
        "slab_sheet": None,
        "reasons": [
            {
                "scope": "wall",
                "code": "wall_layout_nonstandard",
                "sheet": "Sheet",
                "message": "Sheet 不是标准四列墙体配筋模板",
            }
        ],
    }
    assert result["confirmations"] == []
    assert result["walls"] == []


@pytest.mark.parametrize(
    "reinforcement_source",
    [ReinforcementSource.PROVIDED, ReinforcementSource.AI_SUGGESTED],
)
def test_preflight_rejects_enabled_slab_stress_without_slab_figures_before_ocr(
    tmp_path: Path,
    reinforcement_source: ReinforcementSource,
) -> None:
    archive_path = (
        _build_duplicate_archive(tmp_path)
        if reinforcement_source is ReinforcementSource.PROVIDED
        else _build_ai_suggested_archive(
            tmp_path,
            include_slab_figures=False,
        )
    )
    ocr_calls: list[str] = []

    def recognize(path: Path, _direction: str) -> StressLegendReading:
        ocr_calls.append(path.name)
        return StressLegendReading(smn=0, smx=1000, legend_values=())

    with pytest.raises(
        InvalidCalculationArchive,
        match="启用楼板应力.*没有.*楼板",
    ):
        run_calculation_book_preflight(
            archive_path=archive_path,
            extraction_root=tmp_path / "extracted",
            reinforcement_source=reinforcement_source,
            include_slab_stress=True,
            ocr_recognizer=recognize,
        )

    assert ocr_calls == []


def test_ai_suggested_preflight_needs_no_workbook_and_keeps_ocr_review_evidence(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import src.calculation_book.preflight as preflight_module

    archive_path = _build_ai_suggested_archive(tmp_path)
    recognized_names: list[str] = []

    monkeypatch.setattr(
        preflight_module,
        "inspect_reinforcement_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("AI suggested preflight must not inspect a workbook")
        ),
    )
    monkeypatch.setattr(
        preflight_module,
        "load_reinforcement_schedule",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("AI suggested preflight must not load a wall schedule")
        ),
    )
    monkeypatch.setattr(
        preflight_module,
        "load_slab_reinforcement_schedule",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("AI suggested preflight must not load a slab schedule")
        ),
    )

    def recognize(path: Path, direction: str) -> StressLegendReading:
        recognized_names.append(path.name)
        if path.name == "S7158-Z.png":
            raise RuntimeError("SMX missing")
        return StressLegendReading(
            smn=0,
            smx=0 if direction == "Z" else 1000,
            legend_values=(
                () if direction == "Z" else tuple(1000 * index / 9 for index in range(10))
            ),
            is_zero_result=direction == "Z",
        )

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        reinforcement_source="ai_suggested",
        include_slab_stress=False,
        ocr_recognizer=recognize,
    )

    assert recognized_names == [
        "S7157-X.png",
        "S7157-Y.png",
        "S7157-Z.png",
        "S7158-X.png",
        "S7158-Y.png",
        "S7158-Z.png",
    ]
    assert result["reinforcement_source"] == "ai_suggested"
    assert result["reinforcement_workbook"] is None
    assert result["requires_ai_normalization"] is False
    assert result["requires_ai_recommendation"] is True
    assert result["image_wall_group_count"] == 2
    assert result["figure_count"] == 6
    assert result["slab_figure_count"] == 0
    assert result["slab_zero_figure_count"] == 0
    assert result["slab_elevation_count"] == 0
    assert result["z_zero_or_missing_smx_count"] == 2
    assert result["ignored_root_images"] == ["说明图.png"]
    assert result["review_items"] == [
        {
            "scope": "wall",
            "identity": "S7158",
            "direction": "Z",
            "image_filename": "S7158-Z.png",
            "reason": "OCR 识别失败：SMX missing",
        }
    ]
    assert len(result["walls"]) == 2


def test_ai_suggested_preflight_ocrs_slab_figures_only_when_enabled(
    tmp_path: Path,
) -> None:
    archive_path = _build_ai_suggested_archive(tmp_path)
    recognized_names: list[str] = []

    def recognize(path: Path, direction: str) -> StressLegendReading:
        recognized_names.append(path.name)
        return StressLegendReading(
            smn=0,
            smx=0 if direction == "Z" else 1000,
            legend_values=(),
            is_zero_result=direction == "Z",
        )

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        reinforcement_source=ReinforcementSource.AI_SUGGESTED,
        include_slab_stress=True,
        ocr_recognizer=recognize,
    )

    assert recognized_names[-5:] == [
        "11.45-TOP-X.png",
        "11.45-TOP-Y.png",
        "11.45-BOTTOM-X.png",
        "11.45-BOTTOM-Y.png",
        "11.45-Z.png",
    ]
    assert result["slab_figure_count"] == 5
    assert result["slab_elevation_count"] == 1
    assert result["zero_figure_count"] == 2
    assert result["slab_zero_figure_count"] == 1
    assert len(result["slabs"]) == 5


def test_ai_suggested_split_groups_emit_stable_review_evidence_without_blocking(
    tmp_path: Path,
) -> None:
    archive_path = _build_duplicate_archive(
        tmp_path,
        include_workbook=False,
    )

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        reinforcement_source=ReinforcementSource.AI_SUGGESTED,
        ocr_recognizer=lambda _path, _direction: StressLegendReading(
            smn=0,
            smx=1000,
            legend_values=(),
        ),
    )

    expected_reviews = [
        {
            "code": "split_image_group",
            "scope": "wall",
            "identity": f"S7157-{group_index}",
            "direction": direction,
            "image_filename": f"S7157-{group_index}-{direction}.png",
            "reason": "-1/-2 应力图组需在任务完成后确认配筋",
        }
        for group_index in (1, 2)
        for direction in ("X", "Y", "Z")
    ]
    assert result["review_items"] == expected_reviews
    assert result["warnings"] == expected_reviews
    assert result["requires_ocr_review"] is False
    assert result["requires_manual_confirmation"] is False
    assert result["confirmations"] == []
    assert [wall["wall_id"] for wall in result["walls"]] == [
        "S7157-1",
        "S7157-2",
    ]
    assert all(
        evidence["image_filename"]
        for wall in result["walls"]
        for evidence in wall["directions"].values()
    )


def test_preflight_rejects_unknown_reinforcement_source(tmp_path: Path) -> None:
    archive_path = _build_duplicate_archive(tmp_path)

    with pytest.raises(ValueError, match="unsupported.*ReinforcementSource"):
        run_calculation_book_preflight(
            archive_path=archive_path,
            extraction_root=tmp_path / "extracted",
            reinforcement_source="unsupported",
            ocr_recognizer=lambda _path, _direction: StressLegendReading(
                smn=0,
                smx=1000,
                legend_values=(),
            ),
        )


def test_preflight_reports_normalization_and_wall_count_audit_without_blocking(
    tmp_path: Path,
) -> None:
    archive_path = _build_mismatch_archive(tmp_path)

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        ocr_recognizer=lambda _path, _direction: StressLegendReading(
            smn=0,
            smx=1000,
            legend_values=tuple(1000 * index / 9 for index in range(10)),
        ),
    )

    assert result["reinforcement_source_row_count"] == 2
    assert result["reinforcement_normalized_row_count"] == 2
    assert result["reinforcement_issue_row_count"] == 0
    assert result["reinforcement_unique_wall_count"] == 2
    assert result["normalization_triggered"] is False
    assert result["normalization_skill_id"] is None
    assert result["image_unique_wall_count"] == 2
    assert result["image_wall_group_count"] == 2
    assert result["matched_unique_wall_count"] == 1
    assert result["wall_count"] == 2
    assert result["image_only_wall_ids"] == ["NDTJ1"]
    assert result["workbook_only_wall_ids"] == ["S7160"]
    assert result["requires_wall_count_confirmation"] is True
    assert result["requires_manual_confirmation"] is False
    image_only = next(
        wall for wall in result["walls"] if wall["wall_id"] == "NDTJ1"
    )
    assert all(
        direction["canonical_specification"] == ""
        for direction in image_only["directions"].values()
    )
    assert {warning["code"] for warning in result["warnings"]} == {
        "image_only_wall",
        "workbook_only_wall",
    }


def test_preflight_ignores_slab_ocr_when_option_is_disabled(
    tmp_path: Path,
) -> None:
    archive_path = _build_slab_archive(tmp_path)
    recognized_names: list[str] = []

    def recognize(path: Path, _direction: str) -> StressLegendReading:
        recognized_names.append(path.name)
        return StressLegendReading(
            smn=0,
            smx=1000,
            legend_values=tuple(1000 * index / 9 for index in range(10)),
        )

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        include_slab_stress=False,
        ocr_recognizer=recognize,
    )

    assert recognized_names == ["S7159-X.png", "S7159-Y.png", "S7159-Z.png"]
    assert result["slabs"] == []
    assert result["slab_figure_count"] == 0
    assert result["warnings"] == [
        {
            "code": "slab_ignored_by_choice",
            "filenames": [
                "11.45-TOP-X.png",
                "11.45-TOP-Y.png",
                "11.45-BOTTOM-X.png",
                "11.45-BOTTOM-Y.png",
                "11.45-Z.png",
            ],
        }
    ]


def test_preflight_returns_seven_ordered_slab_evidence_items(
    tmp_path: Path,
) -> None:
    archive_path = _build_slab_archive(tmp_path, include_middle=True)

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        include_slab_stress=True,
        ocr_recognizer=lambda _path, direction: StressLegendReading(
            smn=0,
            smx=0 if direction == "Z" else 1000,
            legend_values=(
                () if direction == "Z" else tuple(1000 * index / 9 for index in range(10))
            ),
            is_zero_result=direction == "Z",
        ),
    )

    assert result["slab_figure_count"] == 7
    assert result["slab_elevation_count"] == 1
    assert result["slab_zero_figure_count"] == 1
    assert [item["key"] for item in result["slabs"]] == [
        "top_x",
        "middle_x",
        "bottom_x",
        "top_y",
        "middle_y",
        "bottom_y",
        "z",
    ]
    assert result["slabs"][0]["source_cell"] == "B2"
    assert result["slabs"][0]["canonical_specification"] == "1D36间距200"
    assert result["slabs"][-1]["is_zero_result"] is True


def test_preflight_missing_slab_sheet_starts_ai_with_wall_row_audit(
    tmp_path: Path,
) -> None:
    archive_path = _build_slab_archive(tmp_path, include_slab_sheet=False)

    result = run_calculation_book_preflight(
        archive_path=archive_path,
        extraction_root=tmp_path / "extracted",
        include_slab_stress=True,
        ocr_recognizer=lambda _path, _direction: StressLegendReading(
            smn=0,
            smx=1000,
            legend_values=tuple(1000 * index / 9 for index in range(10)),
        ),
    )

    assert result["requires_ai_normalization"] is True
    assert result["ai_reinforcement_expected_source_row_count"] == 1
    assert result["slab_figure_count"] == 5
    assert "slab_sheet_missing" in {
        reason["code"] for reason in result["format_inspection"]["reasons"]
    }
