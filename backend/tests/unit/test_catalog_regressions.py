from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from src.doc_gen.catalog import CatalogGenerator
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    PageInfo,
    SheetSet,
    TitleblockFields,
)


class DummyPDFExporter:
    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")

    def count_pdf_pages(self, pdf_path: Path) -> int:
        return 2


def _make_frame(seq: int) -> FrameMeta:
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
    )
    titleblock = TitleblockFields(
        internal_code=f"1234567-JG001-{seq:03d}",
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"图纸{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
        page_total=1,
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_context() -> DocContext:
    params = GlobalDocParams(
        project_no="2016",
        engineering_no="1234",
        subitem_no="JG001",
        album_title_cn="测试图册",
        cover_revision="A",
        doc_status="CFC",
    )
    derived = DerivedFields(
        album_code="01",
        cover_internal_code="1234567-JG001-FM",
        catalog_internal_code="1234567-JG001-TM",
        cover_external_code="JD1NHT11F01B25C42SD",
        catalog_external_code="JD1NHT11T01B25C42SD",
        cover_title_cn="测试图册封面",
        catalog_title_cn="测试图册目录",
        cover_title_en="Test Album Cover",
        catalog_title_en="Test Album Contents",
        catalog_revision="A",
    )
    frame_001 = _make_frame(1)
    master_page = PageInfo(
        page_index=1,
        outer_bbox=frame_001.runtime.outer_bbox,
        has_titleblock=True,
        frame_meta=frame_001,
    )
    return DocContext(
        params=params,
        derived=derived,
        frames=[_make_frame(2), _make_frame(3)],
        sheet_sets=[
            SheetSet(
                cluster_id="sheet-set-001",
                page_total=1,
                pages=[master_page],
                master_page=master_page,
            ),
        ],
    )


def test_catalog_writes_album_code_into_merged_title_cell_and_includes_001(
    temp_dir: Path,
) -> None:
    gen = CatalogGenerator(pdf_exporter=DummyPDFExporter())
    ctx = _build_context()
    output_xlsx = temp_dir / "目录.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=gen.spec.get_catalog_bindings(),
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active

    assert ws["D3"].value == "第01图册图纸(文件)目录"
    assert ws["B11"].value == "1234567-JG001-001"
    assert ws["D11"].value == "JD1NHT11001B25C42SD"
