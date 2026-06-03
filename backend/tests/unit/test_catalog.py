from __future__ import annotations

import json
import shutil
import sys
import tempfile
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace
from typing import cast
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from src.doc_gen.catalog import CatalogGenerator
from src.doc_gen.pdf_engine import PDFExporter
from src.interfaces import IPDFExporter
from src.models import (
    BBox,
    DerivedFields,
    DocContext,
    FrameMeta,
    FrameRuntime,
    GlobalDocParams,
    PageInfo,
    SheetSet,
    TitleblockFields,
)


class DummyPDFExporter:
    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")

    def count_pdf_pages(self, pdf_path: Path) -> int:
        return 2


class FailingPDFExporter:
    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        raise RuntimeError("Workbook.ExportAsFixedFormat failed")

    def count_pdf_pages(self, pdf_path: Path) -> int:
        return 2


class MismatchedPageCountPDFExporter:
    def __init__(self) -> None:
        self.export_count = 0

    def export_xlsx_to_pdf(self, xlsx_path: Path, pdf_path: Path) -> None:
        self.export_count += 1
        pdf_path.write_bytes(b"%PDF-1.4\n%dummy\n")

    def count_pdf_pages(self, pdf_path: Path) -> int:
        return 16


class _FakeOfficeLimiter:
    def __init__(self) -> None:
        self.entries = 0
        self.active = False

    @contextmanager
    def excel_session(self):
        self.entries += 1
        self.active = True
        try:
            yield
        finally:
            self.active = False


class _FakeExcelRange:
    def __init__(self, sheet: "_FakeExcelWorksheet", cell: str) -> None:
        self._sheet = sheet
        self._cell = cell

    @property
    def Value(self) -> object | None:  # noqa: N802 - COM API spelling
        return self._sheet.cells.get(self._cell)

    @Value.setter
    def Value(self, value: object) -> None:  # noqa: N802 - COM API spelling
        self._sheet.cells[self._cell] = value


class _FakeExcelWorksheet:
    def __init__(self) -> None:
        self.cells: dict[str, object] = {}

    def Range(self, cell: str) -> _FakeExcelRange:  # noqa: N802 - COM API spelling
        return _FakeExcelRange(self, cell)


class _FakeExcelWorkbook:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.sheet = _FakeExcelWorksheet()
        self.exports: list[Path] = []
        self.saved = 0
        self.closed = False

    def Worksheets(self, index: int) -> _FakeExcelWorksheet:  # noqa: N802 - COM API spelling, ARG002
        return self.sheet

    def ExportAsFixedFormat(self, export_type: int, pdf_path: str) -> None:  # noqa: N802 - COM API spelling
        self.exports.append(Path(pdf_path))
        Path(pdf_path).write_bytes(b"%PDF-1.4\n%fake catalog pdf\n")

    def Save(self) -> None:  # noqa: N802 - COM API spelling
        self.saved += 1
        wb = load_workbook(self.workbook_path)
        ws = wb.active
        assert ws is not None
        for cell, value in self.sheet.cells.items():
            ws[cell] = value
        wb.save(self.workbook_path)

    def Close(self, save_changes: bool) -> None:  # noqa: N802 - COM API spelling, ARG002
        self.closed = True


def _make_frame(seq: int) -> FrameMeta:
    code = f"1234567-JG001-{seq:03d}"
    runtime = FrameRuntime(
        frame_id=str(uuid4()),
        source_file=Path("demo.dxf"),
        outer_bbox=BBox(xmin=0, ymin=0, xmax=100, ymax=100),
    )
    titleblock = TitleblockFields(
        internal_code=code,
        external_code=f"JD1NHT11{seq:03d}B25C42SD",
        title_cn=f"图纸{seq}",
        title_en=f"Drawing {seq}",
        revision="A",
        status="CFC",
        page_total=1,
    )
    return FrameMeta(runtime=runtime, titleblock=titleblock)


def _build_context(project_no: str = "2016") -> DocContext:
    params = GlobalDocParams(
        project_no=project_no,
        engineering_no="1234",
        subitem_no="JG001",
        album_title_cn="测试图册",
        album_title_en="Test Album",
        cover_revision="A",
        doc_status="CFC",
        is_upgrade=False,
        upgrade_sheet_codes="",
    )
    derived = DerivedFields(
        album_code="01",
        cover_internal_code="1234567-JG001-FM",
        catalog_internal_code="1234567-JG001-TM",
        cover_external_code="JD1NHT11F01B25C42SD",
        catalog_external_code="JD1NHT11T01B25C42SD",
        cover_title_cn="测试图册封面",
        catalog_title_cn="测试图册目录",
        cover_title_en="Test Album Cover",
        catalog_title_en="Test Album Contents",
        catalog_revision="A",
        cover_catalog_revision="A",
    )
    frames = [_make_frame(3), _make_frame(1), _make_frame(2)]
    return DocContext(params=params, derived=derived, frames=frames)


def _build_context_with_sheet_set_001() -> DocContext:
    ctx = _build_context()
    frame_001 = _make_frame(1)
    frame_001.titleblock.paper_size_text = "A4"
    frame_001.titleblock.page_total = 1
    master_page = PageInfo(
        page_index=1,
        outer_bbox=frame_001.runtime.outer_bbox,
        has_titleblock=True,
        frame_meta=frame_001,
    )
    ctx.frames = [_make_frame(3), _make_frame(2)]
    ctx.sheet_sets = [
        SheetSet(
            cluster_id="sheet-set-001",
            page_total=7,
            pages=[master_page],
            master_page=master_page,
        ),
    ]
    return ctx


def test_catalog_row_order_and_upgrade_note_defaults_blank() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    rows = gen._build_detail_rows(ctx)

    assert rows[0]["type"] == "cover"
    assert rows[1]["type"] == "catalog"
    assert [r["internal_code"] for r in rows[2:]] == [
        "1234567-JG001-001",
        "1234567-JG001-002",
        "1234567-JG001-003",
    ]
    assert rows[2]["upgrade_note"] == ""
    assert rows[3]["upgrade_note"] == ""
    assert rows[4]["upgrade_note"] == ""


def test_catalog_marks_only_catalog_row_when_upgrade_enabled_without_sheet_codes() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    ctx.params.is_upgrade = True
    ctx.params.upgrade_sheet_codes = ""

    rows = gen._build_detail_rows(ctx)

    assert rows[0]["upgrade_note"] == "升版"
    assert rows[1]["upgrade_note"] == "升版"
    assert all(row["upgrade_note"] == "" for row in rows[2:])


def test_catalog_marks_matching_drawing_rows_for_upgrade_sheet_codes() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    ctx.params.is_upgrade = True
    ctx.params.upgrade_sheet_codes = "001、3"

    rows = gen._build_detail_rows(ctx)

    assert rows[1]["upgrade_note"] == "升版"
    assert rows[2]["upgrade_note"] == "升版"
    assert rows[3]["upgrade_note"] == ""
    assert rows[4]["upgrade_note"] == "升版"


def test_catalog_1818_uses_upgrade_upgrade_label() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    ctx.params.is_upgrade = True
    ctx.params.upgrade_sheet_codes = "2"

    rows = gen._build_detail_rows(ctx)

    assert rows[0]["upgrade_note"] == "升版 upgrade"
    assert rows[1]["upgrade_note"] == "升版 upgrade"
    assert rows[2]["upgrade_note"] == ""
    assert rows[3]["upgrade_note"] == "升版 upgrade"
    assert rows[4]["upgrade_note"] == ""


def test_catalog_uses_structured_upgrade_entries_for_upgrade_and_added_notes() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    ctx.params.is_upgrade = True
    ctx.params.upgrade_entries = json.dumps(
        [
            {"revision": "B", "sheet_codes": "001~002", "is_added": False},
            {"revision": "D", "sheet_codes": "003", "is_added": True},
        ],
        ensure_ascii=False,
    )

    rows = gen._build_detail_rows(ctx)

    assert rows[0]["upgrade_note"] == "升版"
    assert rows[1]["upgrade_note"] == "升版"
    assert rows[2]["upgrade_note"] == "升版"
    assert rows[3]["upgrade_note"] == "升版"
    assert rows[4]["upgrade_note"] == "新增"
    assert rows[4]["revision"] == "A"


def test_catalog_1818_uses_added_add_label_for_structured_added_entries() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    ctx.params.is_upgrade = True
    ctx.params.upgrade_entries = json.dumps(
        [
            {"revision": "B", "sheet_codes": "001", "is_added": False},
            {"revision": "C", "sheet_codes": "003", "is_added": True},
        ],
        ensure_ascii=False,
    )

    rows = gen._build_detail_rows(ctx)

    assert rows[0]["upgrade_note"] == "升版 upgrade"
    assert rows[1]["upgrade_note"] == "升版 upgrade"
    assert rows[2]["upgrade_note"] == "升版 upgrade"
    assert rows[3]["upgrade_note"] == ""
    assert rows[4]["upgrade_note"] == "新增Add"


def test_catalog_cover_and_catalog_rows_use_cover_catalog_revision() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    ctx.derived.document_revision = "A"
    ctx.derived.catalog_revision = "C"
    ctx.derived.cover_catalog_revision = "C"

    rows = gen._build_detail_rows(ctx)

    assert rows[0]["revision"] == "C"
    assert rows[1]["revision"] == "C"
    assert rows[2]["revision"] == "A"


def test_catalog_1818_title_in_same_cell_with_newline() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    bindings = gen.spec.get_catalog_bindings()

    row_data = {
        "internal_code": "1234567-JG001-001",
        "external_code": "JD1NHT11001B25C42SD",
        "title_cn": "中文标题",
        "title_en": "English Title",
        "revision": "A",
        "status": "CFC",
        "page_total": 1,
        "upgrade_note": "",
    }
    gen._write_detail_row(ws, 9, row_data, bindings, ctx)

    assert ws["E9"].value == "中文标题\nEnglish Title"


def test_catalog_backfill_page_count(temp_dir: Path) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context()
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "目录.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )
    gen._backfill_page_count(output_xlsx, 3, bindings)

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["H10"].value == 3


def test_catalog_diagnostics_preserves_xlsx_page_count_when_pdf_export_fails(
    temp_dir: Path,
    monkeypatch,
) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, FailingPDFExporter()))
    ctx = _build_context()
    monkeypatch.setattr(CatalogGenerator, "_count_pages", lambda self, path: 3)

    result = gen.generate_with_diagnostics(ctx, temp_dir)

    assert result.xlsx_path.exists()
    assert result.page_count == 3
    assert result.pdf_path is not None
    assert not result.pdf_path.exists()
    assert result.pdf_export_error is not None
    assert "Workbook.ExportAsFixedFormat failed" in str(result.pdf_export_error)

    ws = load_workbook(result.xlsx_path).active
    assert ws is not None
    assert ws["H10"].value == 3


def test_catalog_generate_reconciles_page_count_with_exported_pdf(
    temp_dir: Path,
    monkeypatch,
) -> None:
    exporter = MismatchedPageCountPDFExporter()
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, exporter))
    ctx = _build_context()
    monkeypatch.setattr(CatalogGenerator, "_count_pages", lambda self, path: 6)

    output_xlsx, output_pdf, page_count = gen.generate(ctx, temp_dir)

    assert output_pdf.exists()
    assert page_count == 16
    assert exporter.export_count == 2

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["H10"].value == 16


def test_catalog_single_excel_session_exports_probe_and_final_pdf(
    temp_dir: Path,
    monkeypatch,
) -> None:
    exporter = PDFExporter(preferred_engine="office_com")
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, exporter))
    ctx = _build_context()
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "catalog.xlsx"
    output_pdf = temp_dir / "catalog.pdf"
    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )
    limiter = _FakeOfficeLimiter()
    fake_pythoncom = SimpleNamespace(CoInitialize=lambda: None, CoUninitialize=lambda: None)
    fake_win32com = SimpleNamespace(client=SimpleNamespace())
    monkeypatch.setitem(sys.modules, "pythoncom", fake_pythoncom)
    monkeypatch.setitem(sys.modules, "win32com.client", fake_win32com.client)
    monkeypatch.setitem(sys.modules, "win32com", fake_win32com)
    monkeypatch.setattr("src.doc_gen.catalog.get_office_automation_limiter", lambda: limiter)
    monkeypatch.setattr(PDFExporter, "_terminate_stale_excel_automation_processes", classmethod(lambda cls: None))
    monkeypatch.setattr(PDFExporter, "_snapshot_process_ids_by_image", staticmethod(lambda image_name: {100}))
    recycled: list[tuple[str, set[int]]] = []
    monkeypatch.setattr(
        PDFExporter,
        "_terminate_new_processes",
        classmethod(lambda cls, image_name, baseline: recycled.append((image_name, set(baseline)))),
    )
    monkeypatch.setattr(PDFExporter, "_create_excel_application", classmethod(lambda cls, module: (object(), True)))
    monkeypatch.setattr(PDFExporter, "_prepare_excel_for_headless_run", staticmethod(lambda excel: None))

    temp_dirs: list[Path] = []

    def fake_prepare_path(xlsx_path: Path, *, label: str) -> tuple[Path, Path]:  # noqa: ARG001
        temp_root = Path(tempfile.mkdtemp(prefix="catalog-single-session-", dir=temp_dir))
        temp_dirs.append(temp_root)
        working_copy = temp_root / xlsx_path.name
        shutil.copy2(xlsx_path, working_copy)
        return working_copy, temp_root

    workbooks: list[_FakeExcelWorkbook] = []

    def fake_open_workbook(excel: object, workbook_path: Path, *, read_only: bool) -> _FakeExcelWorkbook:  # noqa: ARG001
        assert read_only is False
        workbook = _FakeExcelWorkbook(workbook_path)
        workbooks.append(workbook)
        return workbook

    monkeypatch.setattr(PDFExporter, "_prepare_excel_path_for_com", classmethod(lambda cls, *args, **kwargs: fake_prepare_path(*args, **kwargs)))
    monkeypatch.setattr(PDFExporter, "_open_excel_workbook", classmethod(lambda cls, *args, **kwargs: fake_open_workbook(*args, **kwargs)))
    monkeypatch.setattr(PDFExporter, "_retry_excel_com_call", classmethod(lambda cls, fn, desc, retries=10: fn()))
    monkeypatch.setattr(exporter, "count_pdf_pages", lambda pdf_path: 4)

    page_count = gen._export_catalog_pdf_via_single_excel_session(output_xlsx, output_pdf, bindings)

    assert page_count == 4
    assert limiter.entries == 1
    assert recycled == []
    assert len(workbooks) == 1
    workbook = workbooks[0]
    assert workbook.saved == 1
    assert workbook.closed is True
    assert len(workbook.exports) == 2
    assert workbook.exports[0].name.endswith(".probe.pdf")
    assert workbook.exports[1] == output_pdf
    assert output_pdf.exists()
    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["H10"].value == 4
    assert all(not temp_path.exists() for temp_path in temp_dirs)


def test_catalog_writes_album_code_into_merged_title_cell_and_includes_sheet_set_001(
    temp_dir: Path,
) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context_with_sheet_set_001()
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "目录.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    assert ws["D3"].value == "图纸(文件)目录"
    assert ws["B11"].value == "1234567-JG001-001"
    assert ws["D11"].value == "JD1NHT11001B25C42SD"
    assert ws["D1"].value == "测试图册"
    assert ws["H11"].value == 7


def test_catalog_writes_1818_album_titles_into_header_cells(temp_dir: Path) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "目录-1818.xlsx"

    gen._write_catalog(
        template_path="documents_bin/1818图册目录模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    assert ws["D1"].value == "测试图册"
    assert ws["D2"].value == "Test Album"
    assert ws["D4"].value == "图纸(文件)目录"


def test_catalog_writes_non_1818_catalog_row_title_from_header_segments(temp_dir: Path) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="2016")
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "目录-2016.xlsx"

    gen._write_catalog(
        template_path="documents_bin/目录模板文件.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["E10"].value == "测试图册\n图纸(文件)目录"


def test_catalog_writes_1818_catalog_row_title_from_header_segments(temp_dir: Path) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    bindings = gen.spec.get_catalog_bindings()
    output_xlsx = temp_dir / "目录-1818-row-title.xlsx"

    gen._write_catalog(
        template_path="documents_bin/1818图册目录模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["E10"].value == "测试图册\n图纸(文件)目录\nTest Album\nDOCUMENT CONTENTS"


def test_catalog_excel_com_paths_use_pdf_exporter_retry_helpers() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    source_text = (repo_root / "backend" / "src" / "doc_gen" / "catalog.py").read_text(
        encoding="utf-8",
    )

    assert "PDFExporter._prepare_excel_path_for_com(" in source_text
    assert "PDFExporter._open_excel_workbook(" in source_text
    assert "PDFExporter._retry_excel_com_call(" in source_text


def test_catalog_preserves_1818_english_header_merge_and_default_font() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    bindings = gen.spec.get_catalog_bindings()
    output_dir = Path(__file__).resolve().parents[3] / "tmp" / f"catalog-test-{uuid4().hex}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / "catalog-1818-merge.xlsx"

    gen._write_catalog(
        template_path="documents_bin/1818图册目录模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    assert "D2:E3" in merged_ranges
    assert "D2:E2" not in merged_ranges
    assert "D3:E3" not in merged_ranges
    assert ws["D2"].value == "Test Album"
    assert ws["D2"].font.sz == 12


def test_catalog_shrinks_1818_english_header_font_for_long_title() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    ctx.params.album_title_en = (
        "Secondary steel shop drawings at elevation -8.800m and associated "
        "embedded parts for turbine building area with supplementary structural "
        "interfaces and platform support details for construction release"
    )
    bindings = gen.spec.get_catalog_bindings()
    output_dir = Path(__file__).resolve().parents[3] / "tmp" / f"catalog-test-{uuid4().hex}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / "catalog-1818-long-title.xlsx"

    gen._write_catalog(
        template_path="documents_bin/1818图册目录模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    assert ws["D2"].value == ctx.params.album_title_en
    assert ws["D2"].font.sz in {9, 7, 6, 5}


def test_catalog_enables_shrink_to_fit_when_1818_header_still_overflows() -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    ctx.params.album_title_en = " ".join(["SUPPLEMENTARY"] * 48)
    bindings = gen.spec.get_catalog_bindings()
    output_dir = Path(__file__).resolve().parents[3] / "tmp" / f"catalog-test-{uuid4().hex}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / "catalog-1818-shrink-to-fit.xlsx"

    gen._write_catalog(
        template_path="documents_bin/1818图册目录模板.xlsx",
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None

    assert ws["D2"].font.sz == 5
    assert ws["D2"].alignment.shrinkToFit is True


def test_catalog_repairs_missing_detail_grid_borders_from_template_holes(tmp_path: Path) -> None:
    gen = CatalogGenerator(pdf_exporter=cast(IPDFExporter, DummyPDFExporter()))
    ctx = _build_context(project_no="1818")
    ctx.frames = [_make_frame(seq) for seq in (5, 1, 3, 2, 4)]
    bindings = gen.spec.get_catalog_bindings()
    template_path = next((Path(__file__).resolve().parents[3] / "documents_bin").glob("*1818*目录*.xlsx"))
    output_xlsx = tmp_path / "catalog-1818-border-repair.xlsx"

    gen._write_catalog(
        template_path=str(template_path),
        output_path=output_xlsx,
        bindings=bindings,
        ctx=ctx,
    )

    ws = load_workbook(output_xlsx).active
    assert ws is not None
    assert ws["A15"].border.left.style == "thin"
    assert ws["A15"].border.right.style == "thin"
    assert ws["A15"].border.top.style == "thin"
    assert ws["A15"].border.bottom.style == "thin"
    assert ws["A15"].alignment.horizontal == ws["A14"].alignment.horizontal
    assert ws["A15"].alignment.vertical == ws["A14"].alignment.vertical
    assert ws["A15"].alignment.wrapText == ws["A14"].alignment.wrapText
    assert ws["A15"].font.name == ws["A14"].font.name
    assert ws["A15"].font.sz == ws["A14"].font.sz
    assert ws["A15"].fill.fill_type == ws["A14"].fill.fill_type

